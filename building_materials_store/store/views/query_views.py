from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Q, Sum, Count
from datetime import datetime, timedelta
from django.utils.dateparse import parse_date

from django.db.models import F
from django.contrib.postgres.search import TrigramSimilarity
from ..models import *
from icecream import ic
import json
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import FileResponse
import os
from io import BytesIO
from django.core.files.base import ContentFile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.page import PageMargins

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ..my_func.get_unit_map import get_unit_map 
from ..my_func.get_unit_and_cf import get_unit_and_cf 

def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)








# query agent
@require_GET
def search_agents_view(request):
    query = request.GET.get('q', '')
    agents = (
        Agent.objects.annotate(similarity=TrigramSimilarity('name', query))
        .filter(similarity__gt=0.1)
        .order_by('-similarity')
    )
    data = [{'id': a.id, 'name': a.name} for a in agents]
    return JsonResponse(data, safe=False)


# query account 
@require_GET
def search_accounts_view(request):
    query = request.GET.get('q', '')
    accounts = (
        Account.objects.annotate(similarity=TrigramSimilarity('number', query))
        .filter(similarity__gt=0.1, is_active=True)
        .order_by('-similarity')
    )
    data = [{'id': a.id, 'number': a.number, "type":a.type} for a in accounts]
    return JsonResponse(data, safe=False)

# poisk partnera po imeni (get)
@require_GET
def get_partner_by_name_view(request):
    name = request.GET.get('name', '').strip()

    if not name:
        return JsonResponse({'error': 'Empty name'}, status=400)

    try:
        partner = Partner.objects.get(name__iexact=name)
        data = {
            'id': partner.id,
            'name': partner.name,
            'type': partner.type,
            'balance': float(partner.balance),  # если decimal
            # добавь сюда любые поля, которые нужны фронту
        }
        return JsonResponse({'exists': True, 'partner': data})
    except Partner.DoesNotExist:
        return JsonResponse({'exists': False, 'partner': None})
    
    
# query partner 
@require_GET
def search_partners_view(request):
    query = request.GET.get('q', '')
    partners = (
        Partner.objects.annotate(similarity=TrigramSimilarity('name', query))
        .filter(similarity__gt=0.1, is_active=True)
        .order_by('-similarity')
    )
    data = [{'id': a.id, 'name': a.name, "type":a.type, "balance":a.balance, "is_active":a.is_active} for a in partners]
    return JsonResponse(data, safe=False)






def get_trial_balance(date_from, date_to):
    accounts = Account.objects.all().order_by("number")
    ACCOUNT_FOR_START_NEW_PROJECT = CustomePostingRule.objects.filter(operation__code="ACCOUNT_FOR_START_NEW_PROJECT").first()

    report = []
    detail_report = {}

    for acc in accounts:
        if acc.number == ACCOUNT_FOR_START_NEW_PROJECT.debit_account.number:  
            continue
        # Начальное сальдо (все проводки ДО начала периода)
        opening_debit = (
            Entry.objects.filter(account=acc, transaction__date__lt=date_from)
            .aggregate(total=Sum("debit"))
            .get("total") or Decimal("0.00")
        )
        opening_credit = (
            Entry.objects.filter(account=acc, transaction__date__lt=date_from)
            .aggregate(total=Sum("credit"))
            .get("total") or Decimal("0.00")
        )
        opening_balance = opening_debit - opening_credit

        # Обороты за период
        debit_turnover = (
            Entry.objects.filter(account=acc, transaction__date__range=(date_from, date_to))
            .aggregate(total=Sum("debit"))
            .get("total") or Decimal("0.00")
        )
        credit_turnover = (
            Entry.objects.filter(account=acc, transaction__date__range=(date_from, date_to))
            .aggregate(total=Sum("credit"))
            .get("total") or Decimal("0.00")
        )

        # Конечное сальдо
        closing_balance = opening_balance + debit_turnover - credit_turnover

        report.append({
            "account": acc.number,
            "name": acc.name,
            "opening_balance": opening_balance,
            "debit_turnover": debit_turnover,
            "credit_turnover": credit_turnover,
            "closing_balance": closing_balance,
        })
        
        # get_account_number_klient = rule = CustomePostingRule.objects.filter(operation__code="sale", amount_type="revenue").first()
        # if acc.number == "60":
        # allEntryes = Entry.objects.filter(account=acc)
        # detail_report[acc.number] = {}
        # for e in allEntryes:
        #     if e.transaction.partner.name not in detail_report[acc.number]:
        #         detail_report[acc.number][e.transaction.partner.name] = {
        #             'debit_start':Decimal(0), 
        #             'credit_start':Decimal(0), 
        #             'debit_oborot':Decimal(0), 
        #             'credit_oborot':Decimal(0),
        #             'debit_end':Decimal(0),
        #             'credit_end':Decimal(0),
        #             }
        #     else:
        #         detail_report[acc.number]['subkonto']['debit_start'] = Decimal(0)
               
        
        
        
        detail_report[acc.number] = {}
        entry_start = Entry.objects.filter(account=acc, transaction__date__lt=date_from)
        
        if entry_start:
            for e in entry_start:
                if e.transaction.partner:
                    if e.transaction.partner.name not in detail_report[acc.number]:
                        detail_report[acc.number][e.transaction.partner.name] = {
                            'debit_start': Decimal(e.debit),
                            'credit_start': Decimal(e.credit),
                            'debit_oborot': Decimal(0),
                            'credit_oborot': Decimal(0),
                            'debit_end': Decimal(0),
                            'credit_end': Decimal(0),
                            }
                    else:
                        detail_report[acc.number][e.transaction.partner.name]['debit_start'] += Decimal(e.debit)
                        detail_report[acc.number][e.transaction.partner.name]['credit_start'] += Decimal(e.credit)
                else:
                    if 'prochee' not in detail_report[acc.number]:
                        detail_report[acc.number]['prochee'] = {
                                'debit_start': Decimal(e.debit),
                                'credit_start': Decimal(e.credit),
                                'debit_oborot': Decimal(0),
                                'credit_oborot': Decimal(0),
                                'debit_end': Decimal(0),
                                'credit_end': Decimal(0),
                                }
                    else:
                        detail_report[acc.number]['prochee']['debit_start'] += Decimal(e.debit)
                        detail_report[acc.number]['prochee']['credit_start'] += Decimal(e.credit)
                        
                    
                    
                
                    
            
        entry_oborot = Entry.objects.filter(account=acc, transaction__date__range=(date_from, date_to))
        if entry_oborot:
            for e in entry_oborot:
                if e.transaction.partner:
                    if e.transaction.partner.name not in detail_report[acc.number]:
                        detail_report[acc.number][e.transaction.partner.name] = {
                            'debit_start': Decimal(0),
                            'credit_start': Decimal(0),
                            'debit_oborot': Decimal(e.debit),
                            'credit_oborot': Decimal(e.credit),
                            'debit_end': Decimal(0),
                            'credit_end': Decimal(0),
                            }
                    else:
                        detail_report[acc.number][e.transaction.partner.name]['debit_oborot'] += Decimal(e.debit)
                        detail_report[acc.number][e.transaction.partner.name]['credit_oborot'] += Decimal(e.credit)
            else:
                if 'prochee' not in detail_report[acc.number]:
                    detail_report[acc.number]['prochee'] = {
                        'debit_start': Decimal(0),
                        'credit_start': Decimal(0),
                        'debit_oborot': Decimal(e.debit),
                        'credit_oborot': Decimal(e.credit),
                        'debit_end': Decimal(0),
                        'credit_end': Decimal(0),
                        }
                else:
                    detail_report[acc.number]['prochee']['debit_oborot'] += Decimal(e.debit)
                    detail_report[acc.number]['prochee']['credit_oborot'] += Decimal(e.credit)
            
     
    
    detail_report_total = {}
    saldo_total = {}
    for acc_number, partners in detail_report.items():
        start_summ_debit = Decimal(0)
        start_summ_credit = Decimal(0)
        
        oborot_summ_debit = Decimal(0)
        oborot_summ_credit = Decimal(0)
        
        end_summ_debit = Decimal(0)
        end_summ_credit = Decimal(0)
        
        
        
        

        for partner, data in partners.items():
            debit_end = data['debit_start'] + data['debit_oborot']
            credit_end = data['credit_start'] + data['credit_oborot']
            data['debit_end'] = debit_end
            data['credit_end'] = credit_end
            
            start_summ_debit += data['debit_start']
            start_summ_credit += data['credit_start']
            
            oborot_summ_debit += data['debit_oborot']
            oborot_summ_credit += data['credit_oborot']
            
            end_summ_debit += debit_end
            end_summ_credit += credit_end
         
        detail_report_total[acc_number] = {
            'start_summ_debit': start_summ_debit,
            'start_summ_credit': start_summ_credit,
            'oborot_summ_debit': oborot_summ_debit,
            'oborot_summ_credit': oborot_summ_credit,
            'end_summ_debit': end_summ_debit,
            'end_summ_credit': end_summ_credit,
        }
        
        
        start_saldo = start_summ_debit - start_summ_credit
        start_saldo_debit = start_saldo if start_saldo > 0 else Decimal(0)
        start_saldo_credit = abs(start_saldo) if start_saldo < 0 else Decimal(0)
        
        oborot_saldo = oborot_summ_debit - oborot_summ_credit
        oborot_saldo_debit = oborot_saldo if oborot_saldo > 0 else Decimal(0)
        oborot_saldo_credit = abs(oborot_saldo) if oborot_saldo < 0 else Decimal(0)
        
        end_saldo = end_summ_debit - end_summ_credit
        end_saldo_debit = end_saldo if end_saldo > 0 else Decimal(0)
        end_saldo_credit = abs(end_saldo) if end_saldo < 0 else Decimal(0)
        
        saldo_total[acc_number] = {
            'start_saldo_debit': start_saldo_debit,
            'start_saldo_credit': start_saldo_credit,
            'oborot_saldo_debit': oborot_saldo_debit,
            'oborot_saldo_credit': oborot_saldo_credit,
            'end_saldo_debit': end_saldo_debit,
            'end_saldo_credit': end_saldo_credit,
        }
        
        
    return {"report":report, "detail_report":detail_report, "detail_report_total":detail_report_total, "saldo_total":saldo_total}
# OSW
@require_GET
def get_osw(request):
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    reports = get_trial_balance(date_from, date_to)  # твоя функция
    report = reports['report']
    detail_report = reports['detail_report']
    detail_report_total=reports["detail_report_total"]
    saldo_total=reports["saldo_total"]

    return JsonResponse({"report": report, "detail_report":detail_report, "detail_report_total":detail_report_total, "saldo_total":saldo_total}, safe=False)


########################################################################################################################################################################## START
# get saldo       ishem saldo dlya partnera na wybrannuyu daty, saleInvoice

def get_saldo(partner_obj, getDate):
    # USD
    rule = CustomePostingRule.objects.filter(operation__code="sale", directory_type=partner_obj.type, amount_type="revenue", currency__code="USD").first()
    
    if not rule:
        value = [Decimal('0.00'), Decimal('0.00')]
        return {"start": value, "oborot": value, "final": value, "saldo": value}
    
    account = rule.debit_account
    
    # ⭐ ИЗМЕНЕНИЕ: используем partner из Entry вместо transaction__partner
    entries_start = Entry.objects.filter(partner=partner_obj, transaction__date__lt=getDate).filter(account=account)
    debit_start = entries_start.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
    credit_start = entries_start.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
    
    entries_oborot = Entry.objects.filter(partner=partner_obj, transaction__date__date=getDate).filter(account=account)
    today_entries = []
    desc = ''
    already_have_pks = []
    
    if entries_oborot:
        count = 0
        for e in entries_oborot:
            str_date = e.transaction.date.strftime("%d-%m-%Y %H:%M")
            if e.transaction.pk in already_have_pks:
                for c in today_entries:
                    if c[4] == e.transaction.pk:
                        c[2] += e.debit
                        c[3] += e.credit
            else:
                today_entries.append([str_date, e.transaction.description, e.debit, e.credit, e.transaction.pk])
                already_have_pks.append(e.transaction.pk)
            count += 1
            desc += f"{count}) {e.transaction.description}"
            desc += "\n"

    debit_oborot = entries_oborot.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
    credit_oborot = entries_oborot.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
    
    debit_end = debit_start + debit_oborot
    credit_end = credit_start + credit_oborot
    
    saldo = debit_end - credit_end
    saldo_debit = abs(saldo) if saldo > 0 else 0
    saldo_credit = abs(saldo) if saldo < 0 else 0
    
    # TMT
    rule = CustomePostingRule.objects.filter(operation__code="sale", directory_type=partner_obj.type, amount_type="revenue", currency__code="TMT").first()
   
    if not rule:
        value = [Decimal('0.00'), Decimal('0.00')]
        return {"start": value, "oborot": value, "final": value, "saldo": value}
    
    account = rule.debit_account
    
    # ⭐ ИЗМЕНЕНИЕ: используем partner из Entry вместо transaction__partner
    entries_start = Entry.objects.filter(partner=partner_obj, transaction__date__lt=getDate).filter(account=account)
    debit_start_tmt = entries_start.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
    credit_start_tmt = entries_start.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
    
    entries_oborot = Entry.objects.filter(partner=partner_obj, transaction__date__date=getDate).filter(account=account)
    today_entries_tmt = []
    desc_tmt = ''
    already_have_pks = []
    
    if entries_oborot:
        count = 0
        for e in entries_oborot:
            str_date = e.transaction.date.strftime("%d-%m-%Y %H:%M")
            if e.transaction.pk in already_have_pks:
                for c in today_entries_tmt:
                    if c[4] == e.transaction.pk:
                        c[2] += e.debit
                        c[3] += e.credit
            else:
                today_entries_tmt.append([str_date, e.transaction.description, e.debit, e.credit, e.transaction.pk])
                already_have_pks.append(e.transaction.pk)
            count += 1
            desc_tmt += f"{count}) {e.transaction.description}"
            desc_tmt += "\n"

    debit_oborot_tmt = entries_oborot.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
    credit_oborot_tmt = entries_oborot.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
    
    debit_end_tmt = debit_start_tmt + debit_oborot_tmt
    credit_end_tmt = credit_start_tmt + credit_oborot_tmt
    
    saldo = debit_end_tmt - credit_end_tmt
    saldo_debit_tmt = abs(saldo) if saldo > 0 else 0
    saldo_credit_tmt = abs(saldo) if saldo < 0 else 0
    
    return {
        "start": [debit_start, credit_start], "oborot": [debit_oborot, credit_oborot, desc], "final": [debit_end, credit_end], "saldo": [saldo_debit, saldo_credit], "today_entries": today_entries,
        "start_tmt": [debit_start_tmt, credit_start_tmt], "oborot_tmt": [debit_oborot_tmt, credit_oborot_tmt, desc_tmt], "final_tmt": [debit_end_tmt, credit_end_tmt], "saldo_tmt": [saldo_debit_tmt, saldo_credit_tmt], "today_entries_tmt": today_entries_tmt,
    }


def get_saldo2(partner_obj, getDate):
    ic("tut saldo2")
    # WarehouseProduct.objects.all().update(quantity=0)
    results = {}
    
    # Для каждого счета отдельно
    accounts_to_check = [
        ('60', 'USD'),  # Клиент USD
        ('62', 'TMT'),  # Клиент TMT  
        ('75', 'USD'),  # Учредитель USD
        ('76', 'TMT')   # Учредитель TMT
    ]
    
    for account_number, currency in accounts_to_check:
        account = Account.objects.get(number=account_number)
        
        # ⭐ ИЗМЕНЕНИЕ: используем partner из Entry вместо transaction__partner
        # Начальное сальдо (до выбранной даты)
        entries_start = Entry.objects.filter(
            partner=partner_obj,  # ← ИЗМЕНИЛИ
            transaction__date__lt=getDate
        ).filter(account=account)
        
        debit_start = entries_start.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
        credit_start = entries_start.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
        
        start_saldo_debit = Decimal("0.00")
        start_saldo_credit = Decimal("0.00")
        if (debit_start - credit_start) > 0:
            start_saldo_debit = debit_start - credit_start
        elif (debit_start - credit_start) < 0:
            start_saldo_credit = abs(debit_start - credit_start)
            
        # if account.number == "60":
        #     ic(debit_start)
        #     ic(credit_start)
        #     ic(debit_start - credit_start)
        
        # Обороты за выбранную дату
        entries_oborot = Entry.objects.filter(
            partner=partner_obj,  # ← ИЗМЕНИЛИ
            transaction__date__date=getDate  
        ).filter(account=account)
        
        debit_oborot = entries_oborot.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
        credit_oborot = entries_oborot.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
        # if account.number == "60":
        #     ic(debit_oborot)
        #     ic(credit_oborot)
        #     ic(debit_oborot - credit_oborot)
        
        # Детали операций за день (аналогично оригинальной функции)
        today_entries = []
        desc = ''
        already_have_pks = []
        
        if entries_oborot:
            count = 0
            for e in entries_oborot:
                str_date = e.transaction.date.strftime("%d-%m-%Y %H:%M")
                if e.transaction.pk in already_have_pks:
                    for c in today_entries:
                        if c[4] == e.transaction.pk:
                            c[2] += e.debit
                            c[3] += e.credit
                else:
                    today_entries.append([str_date, e.transaction.description, e.debit, e.credit, e.transaction.pk])
                    already_have_pks.append(e.transaction.pk)
                count += 1
                desc += f"{count}) {e.transaction.description}"
                desc += "\n"
        
        # Итоговые расчеты
        
        # debit_end = debit_start + debit_oborot
        debit_end = start_saldo_debit + debit_oborot
        if account.number == "60":
            ic(debit_end)
            # ic(credit_oborot)
            # ic(debit_oborot - credit_oborot)
        credit_end = start_saldo_credit + credit_oborot
        saldo = debit_end - credit_end
        saldo_debit = abs(saldo) if saldo > 0 else 0
        saldo_credit = abs(saldo) if saldo < 0 else 0
        
        results[f"{account_number}_{currency}"] = {
            "account_name": f"{account.number} {account.name}",
            "start": [debit_start, credit_start],
            "oborot": [debit_oborot, credit_oborot, desc], 
            "final": [debit_end, credit_end],
            "saldo": [saldo_debit, saldo_credit],
            "today_entries": today_entries
        }
        # ic(results)
    
    return results

@require_GET
def get_saldo_for_partner_for_selected_date2(request):
    getDate = request.GET.get('date')
    partnerId = request.GET.get('partnerId')
    partner_obj = Partner.objects.get(pk=partnerId)
    saldo = get_saldo2(partner_obj, getDate)
    return JsonResponse({"saldo": saldo}, safe=False)


@require_GET
def get_saldo_for_partner_for_selected_date(request):
    getDate = request.GET.get('date')
    partnerId = request.GET.get('partnerId')
    partner_obj = Partner.objects.get(pk=partnerId)
    saldo = get_saldo(partner_obj, getDate)
    return JsonResponse({"saldo": saldo}, safe=False)
# get saldo
########################################################################################################################################################################## END


@require_GET
def delete_data(request):
    models_name = request.GET.get('models_name')
    password = request.GET.get('password')

    if password != "373839145654356":
        return JsonResponse({"success": False, "error": "Неверный пароль"})
    
    
    if models_name == "delete_products":
        SalesInvoiceItem.objects.all().delete()
        Product.objects.all().delete()
        Category.objects.all().delete()
        Tag.objects.all().delete()
        UnitOfMeasurement.objects.all().delete()
        Brand.objects.all().delete()
        WarehouseProduct.objects.all().delete()
        Model.objects.all().delete()
        PriceChangeHistory.objects.all().delete()
    # Логика удаления
    
    if models_name == "delete_agents":
        return JsonResponse({"success": True, "models_name": models_name, "date_focus": True})
        
    return JsonResponse({"success": True, "models_name": models_name})




    
    
@require_GET
def check_day_closed(request):
    date_str = request.GET.get("date")  # формат YYYY-MM-DD
    if not date_str:
        return JsonResponse({"success": False, "error": "Дата не указана"})
    
    try:
        chosen_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"success": False, "error": "Неверный формат даты. Используйте YYYY-MM-DD"})

    # выбранный день
    day = DayClosing.objects.filter(date=chosen_date).first()
    is_closed = bool(day and getattr(day, "is_closed", True))

    # предыдущий день
    prev_date = chosen_date - timedelta(days=1)
    prev_day = DayClosing.objects.filter(date=prev_date).first()
    prev_closed = bool(prev_day and getattr(prev_day, "is_closed", True))
    
    # если предыдущий день не закрыт → сразу вернуть False
    
    if DayClosing.objects.all().exists():
        return JsonResponse({"success": True, "is_closed": is_closed, "last_day_not_closed": not prev_closed})
    else:
        return JsonResponse({"success": True, "is_closed": False, "last_day_not_closed": False})

    

   
    

    
    

    
# # rabotaet no ne polnopaya otchetnost 
# @csrf_exempt
# def close_day(request):
#     if request.method != "POST":
#         return JsonResponse({"success": False, "error": "Только POST запросы разрешены"})

#     data = json.loads(request.body)
#     close_date = data.get("date")
#     reason = data.get("reason", "")
#     user_id = data.get("user_id")

#     if not close_date:
#         return JsonResponse({"success": False, "error": "choose close date"})
    
#     if not user_id:
#         return JsonResponse({"success": False, "error": "youDidntAuthenticated"})
    
#     if not User.objects.filter(id=user_id).exists():
#         return JsonResponse({"success": False, "error": "youDidntAuthenticated"})
    
#     user = User.objects.get(id=user_id)
    
#     #############################################################################################################################################################
#     #############################################################################################################################################################
#     ####### start towar oborot otchet
    
#     close_date_format = datetime.strptime(close_date, "%Y-%m-%d").date()

#     day_start = close_date_format
#     day_end = close_date_format + timedelta(days=1)

#     product_units = (
#         ProductUnit.objects
#         .filter(is_default_for_sale=True)
#         .select_related("unit")
#     )

#     unit_map = {pu.product_id: pu for pu in product_units}

#     turnover_product = defaultdict(lambda: defaultdict(dict))

#     # ---------------------------------------------
#     # 1) НАЧАЛЬНЫЕ ОСТАТКИ
#     # ---------------------------------------------

#     for w in Warehouse.objects.all():
#         start_items = (
#             InvoiceItem.objects
#             .filter(
#                 invoice__entry_created_at_handle__lt=day_start,
#                 invoice__canceled_at__isnull=True
#             )
#             .filter(
#                 Q(invoice__warehouse=w) |
#                 Q(invoice__warehouse2=w)
#             )
#             .select_related(
#                 "product", "product__category", "product__base_unit", "invoice"
#             )
#         )

#         for item in start_items:
#             p = item.product
#             inv = item.invoice

#             product_data = turnover_product[w.id].get(p.id)

#             if not product_data:
#                 pu = unit_map.get(p.id)

#                 if pu:
#                     unit = pu.unit.name
#                     conversion_factor = Decimal(pu.conversion_factor)
#                 else:
#                     unit = p.base_unit.name if p.base_unit else ""
#                     conversion_factor = Decimal("1")

#                 product_data = {
#                     "id": p.id,
#                     "category": p.category.name if p.category else "",
#                     "name": p.name,
#                     "unit": unit,
#                     "price": p.wholesale_price or Decimal("0"),

#                     "start_qty": Decimal("0"),

#                     "oborot_prihod_qty": Decimal("0"),
#                     "oborot_wozwrat_qty": Decimal("0"),
#                     "oborot_rashod_qty": Decimal("0"),

#                     "oborot_prihod_price": Decimal("0"),
#                     "oborot_wozwrat_price": Decimal("0"),
#                     "oborot_rashod_price": Decimal("0"),

#                     "end_qty": Decimal("0"),
#                     "conversion_factor": conversion_factor,
#                 }

#                 turnover_product[w.id][p.id] = product_data

#             cf = product_data["conversion_factor"]
#             qty = item.selected_quantity / cf

#             if inv.wozwrat_or_prihod == "prihod":
#                 if inv.warehouse_id == w.id:
#                     product_data["start_qty"] += qty

#             elif inv.wozwrat_or_prihod == "rashod":
#                 if inv.warehouse_id == w.id:
#                     product_data["start_qty"] -= qty

#             elif inv.wozwrat_or_prihod == "wozwrat":
#                 if inv.warehouse_id == w.id:
#                     product_data["start_qty"] += qty

#             elif inv.wozwrat_or_prihod == "transfer":
#                 if inv.warehouse_id == w.id:
#                     product_data["start_qty"] -= qty
#                 elif inv.warehouse2_id == w.id:
#                     product_data["start_qty"] += qty

#         # ---------------------------------------------
#         # 2) ОБОРОТ ЗА ДЕНЬ
#         # ---------------------------------------------

#         turnover_items = (
#             InvoiceItem.objects
#             .filter(
#                 invoice__entry_created_at_handle__gte=day_start,
#                 invoice__entry_created_at_handle__lt=day_end,
#                 invoice__canceled_at__isnull=True
#             )
#             .filter(
#                 Q(invoice__warehouse=w) |
#                 Q(invoice__warehouse2=w)
#             )
#             .select_related(
#                 "product", "product__category", "product__base_unit", "invoice"
#             )
#         )

#         for item in turnover_items:
#             p = item.product
#             inv = item.invoice

#             product_data = turnover_product[w.id].get(p.id)

#             if not product_data:
#                 pu = unit_map.get(p.id)

#                 if pu:
#                     unit = pu.unit.name
#                     conversion_factor = Decimal(pu.conversion_factor)
#                 else:
#                     unit = p.base_unit.name if p.base_unit else ""
#                     conversion_factor = Decimal("1")

#                 product_data = {
#                     "id": p.id,
#                     "category": p.category.name if p.category else "",
#                     "name": p.name,
#                     "unit": unit,
#                     "price": p.wholesale_price or Decimal("0"),

#                     "start_qty": Decimal("0"),

#                     "oborot_prihod_qty": Decimal("0"),
#                     "oborot_wozwrat_qty": Decimal("0"),
#                     "oborot_rashod_qty": Decimal("0"),

#                     "oborot_prihod_price": Decimal("0"),
#                     "oborot_wozwrat_price": Decimal("0"),
#                     "oborot_rashod_price": Decimal("0"),

#                     "end_qty": Decimal("0"),
#                     "conversion_factor": conversion_factor,
#                 }

#                 turnover_product[w.id][p.id] = product_data

#             cf = product_data["conversion_factor"]
#             qty = item.selected_quantity / cf
#             price = qty * item.selected_price

#             if inv.wozwrat_or_prihod == "prihod":
#                 if inv.warehouse_id == w.id:
#                     product_data["oborot_prihod_qty"] += qty
#                     product_data["oborot_prihod_price"] += price

#             elif inv.wozwrat_or_prihod == "rashod":
#                 if inv.warehouse_id == w.id:
#                     product_data["oborot_rashod_qty"] += qty
#                     product_data["oborot_rashod_price"] += price

#             elif inv.wozwrat_or_prihod == "wozwrat":
#                 if inv.warehouse_id == w.id:
#                     product_data["oborot_wozwrat_qty"] += qty
#                     product_data["oborot_wozwrat_price"] += price

#             elif inv.wozwrat_or_prihod == "transfer":
#                 if inv.warehouse_id == w.id:
#                     product_data["oborot_rashod_qty"] += qty
#                     product_data["oborot_rashod_price"] += price
#                 elif inv.warehouse2_id == w.id:
#                     product_data["oborot_prihod_qty"] += qty
#                     product_data["oborot_prihod_price"] += price

#         # ---------------------------------------------
#         # 3) КОНЕЧНЫЙ ОСТАТОК (END_QTY)
#         # ---------------------------------------------

#         for product_data in turnover_product[w.id].values():
#             product_data["end_qty"] = (
#                 product_data["start_qty"]
#                 + product_data["oborot_prihod_qty"]
#                 + product_data["oborot_wozwrat_qty"]
#                 - product_data["oborot_rashod_qty"]
#             )
            
#         # for product_data in turnover_product[w.id].values():
#         #     if product_data["id"] == 1736:
#         #         ic(product_data)
        
#     # ------------------------------------------------
#     # 6) Excel
#     # ------------------------------------------------
#     # ================== СТИЛИ ==================

#     # Шапка
#     HEADER_FILL = PatternFill(
#         fill_type="solid",
#         fgColor="4472C4"   # синий
#     )
#     HEADER_FONT = Font(
#         bold=True,
#         color="FFFFFF"    # белый текст
#     )
#     CENTER_ALIGN = Alignment(
#         horizontal="center",
#         vertical="center",
#         wrap_text=True
#     )

#     # Категория
#     CATEGORY_FILL = PatternFill(
#         fill_type="solid",
#         fgColor="E7E6E6"
#     )
#     CATEGORY_FONT = Font(bold=True)

#     # Обычные ячейки
#     NORMAL_FONT = Font(color="000000")
#     LEFT_ALIGN = Alignment(vertical="center", horizontal="left")
#     RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

#     # Границы
#     THIN_BORDER = Border(
#         left=Side(style="thin"),
#         right=Side(style="thin"),
#         top=Side(style="thin"),
#         bottom=Side(style="thin"),
#     )

#     # Форматы чисел
#     PRICE_FMT = '#,##0.00'
#     QTY_FMT = '#,##0.###'

#     # Ширина колонок
#     COLUMN_WIDTHS = {
#         "A": 25,
#         "B": 40,
#         "C": 10,
#         "D": 12,
#         "E": 8, "F": 8,
#         "G": 8, "H": 8,
#         "I": 8, "J": 8,
#         "K": 8, "L": 8,
#         "M": 8, "N": 8,
#     }
    
#     TOTAL_FILL = PatternFill(
#         fill_type="solid",
#         fgColor="D9D9D9"   # светло-серый
#     )

#     TOTAL_FONT = Font(
#         bold=True
#     )

#     # ================== EXCEL ==================

#     wb = Workbook()
#     wb.remove(wb.active)

#     for warehouse_id, products in turnover_product.items():
#         warehouse = Warehouse.objects.get(id=warehouse_id)
#         ws = wb.create_sheet(title=warehouse.name[:31])

#         # ширина колонок
#         for col, width in COLUMN_WIDTHS.items():
#             ws.column_dimensions[col].width = width

#         # ---------- ШАПКА ----------
#         ws["A1"] = "Категория"
#         ws["B1"] = "Товар"
#         ws["C1"] = "Ед."
#         ws["D1"] = "Цена"

#         ws.merge_cells("E1:F1"); ws["E1"] = "Остаток на начало"
#         ws.merge_cells("G1:H1"); ws["G1"] = "Приход"
#         ws.merge_cells("I1:J1"); ws["I1"] = "Возврат"
#         ws.merge_cells("K1:L1"); ws["K1"] = "Расход"
#         ws.merge_cells("M1:N1"); ws["M1"] = "Конечный остаток"

#         ws["E2"] = "Кол-во"; ws["F2"] = "Сумма"
#         ws["G2"] = "Кол-во"; ws["H2"] = "Сумма"
#         ws["I2"] = "Кол-во"; ws["J2"] = "Сумма"
#         ws["K2"] = "Кол-во"; ws["L2"] = "Сумма"
#         ws["M2"] = "Кол-во"; ws["N2"] = "Сумма"

#         ws.freeze_panes = "A3"

#         # стиль шапки
#         for row in (1, 2):
#             for col in range(1, 15):
#                 cell = ws.cell(row=row, column=col)
#                 cell.fill = HEADER_FILL
#                 cell.font = HEADER_FONT
#                 cell.alignment = CENTER_ALIGN
#                 cell.border = THIN_BORDER

#         # ---------- ДАННЫЕ ----------
#         row_num = 3
#         current_category = None
#         category_start_row = None
#         warehouse_start_row = row_num

#         for data in sorted(products.values(), key=lambda x: (x["category"], x["name"])):

#             # ===== НОВАЯ КАТЕГОРИЯ =====
#             if data["category"] != current_category:

#                 # ---- Итог по предыдущей категории ----
#                 if current_category is not None:
#                     ws[f"B{row_num}"] = f"ИТОГО по категории: {current_category}"

#                     for col in ["E","F","G","H","I","J","K","L","M","N"]:
#                         ws[f"{col}{row_num}"] = f"=SUM({col}{category_start_row}:{col}{row_num-1})"

#                     for col in range(1, 15):
#                         cell = ws.cell(row=row_num, column=col)
#                         cell.fill = TOTAL_FILL
#                         cell.font = TOTAL_FONT
#                         cell.border = THIN_BORDER
#                         cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN

#                     row_num += 2  # пустая строка после итога

#                 # ---- Строка категории ----
#                 ws.merge_cells(f"A{row_num}:N{row_num}")
#                 cell = ws[f"A{row_num}"]
#                 cell.value = data["category"]
#                 cell.fill = CATEGORY_FILL
#                 cell.font = CATEGORY_FONT
#                 cell.alignment = LEFT_ALIGN
#                 cell.border = THIN_BORDER

#                 for col in range(2, 15):
#                     ws.cell(row=row_num, column=col).border = THIN_BORDER

#                 current_category = data["category"]
#                 category_start_row = row_num + 1
#                 row_num += 1

#             # ===== СТРОКА ТОВАРА =====
#             ws[f"B{row_num}"] = data["name"]
#             ws[f"C{row_num}"] = data["unit"]
#             ws[f"D{row_num}"] = float(data["price"])

#             ws[f"E{row_num}"] = float(data["start_qty"])
#             ws[f"F{row_num}"] = float(data["start_qty"] * data["price"])

#             ws[f"G{row_num}"] = float(data["oborot_prihod_qty"])
#             ws[f"H{row_num}"] = float(data["oborot_prihod_price"])

#             ws[f"I{row_num}"] = float(data["oborot_wozwrat_qty"])
#             ws[f"J{row_num}"] = float(data["oborot_wozwrat_price"])

#             ws[f"K{row_num}"] = float(data["oborot_rashod_qty"])
#             ws[f"L{row_num}"] = float(data["oborot_rashod_price"])

#             ws[f"M{row_num}"] = float(data["end_qty"])
#             ws[f"N{row_num}"] = float(data["end_qty"] * data["price"])

#             for col in range(1, 15):
#                 cell = ws.cell(row=row_num, column=col)
#                 cell.font = NORMAL_FONT
#                 cell.border = THIN_BORDER
#                 cell.alignment = LEFT_ALIGN if col <= 3 else RIGHT_ALIGN

#             ws[f"D{row_num}"].number_format = PRICE_FMT
#             for col in ["F", "H", "J", "L", "N"]:
#                 ws[f"{col}{row_num}"].number_format = PRICE_FMT
#             for col in ["E", "G", "I", "K", "M"]:
#                 ws[f"{col}{row_num}"].number_format = QTY_FMT

#             row_num += 1

#         # ===== ИТОГ ПО ПОСЛЕДНЕЙ КАТЕГОРИИ =====
#         ws[f"B{row_num}"] = f"ИТОГО по категории: {current_category}"

#         for col in ["E","F","G","H","I","J","K","L","M","N"]:
#             ws[f"{col}{row_num}"] = f"=SUM({col}{category_start_row}:{col}{row_num-1})"

#         for col in range(1, 15):
#             cell = ws.cell(row=row_num, column=col)
#             cell.fill = TOTAL_FILL
#             cell.font = TOTAL_FONT
#             cell.border = THIN_BORDER
#             cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN

#         row_num += 2

#         # ===== ИТОГ ПО СКЛАДУ =====
#         ws[f"B{row_num}"] = "ИТОГО ПО СКЛАДУ"

#         for col in ["E","F","G","H","I","J","K","L","M","N"]:
#             ws[f"{col}{row_num}"] = f"=SUM({col}{warehouse_start_row}:{col}{row_num-1})"

#         for col in range(1, 15):
#             cell = ws.cell(row=row_num, column=col)
#             cell.fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
#             cell.font = Font(bold=True)
#             cell.border = THIN_BORDER
#             cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN

#     # ================== СОХРАНЕНИЕ ==================

#     excel_buffer = BytesIO()
#     wb.save(excel_buffer)
#     excel_buffer.seek(0)
    
#     report, _ = DayReport.objects.get_or_create(
#         date=datetime.strptime(close_date, "%Y-%m-%d").date(),
#         report_type="turnover",
#         defaults={
#             "created_by": user,
#             "comment": reason
#         }
#     )

#     report.file.save(
#         "turnover.xlsx",   # имя не важно, upload_to решает
#         ContentFile(excel_buffer.read()),
#         save=True
#     )

#     # report = CloseDayAllReportExcel.objects.create(
#     #     date=close_date_format,
#     #     created_by=user,
#     #     comment=reason
#     # )

#     # report.file.save(
#     #     "tovar_oborot.xlsx",
#     #     ContentFile(excel_buffer.read()),
#     #     save=True
#     # )

            
            
#     # ic(day_start)        
#     # ic(day_end)        
#     # ic(turnover_product)
#     # for warehouse_id, values in turnover_product.items():
#     #     if warehouse_id == 1:
#     #         for product_id, value in values.items():
#     #             if product_id == 606:
#     #                 ic(value)
#     ####### start towar oborot otchet    
#     #############################################################################################################################################################
#     #############################################################################################################################################################
    

     

#     try:
#         with transaction.atomic():
#             1/0
#             # если уже закрыт
#             if DayClosing.objects.filter(date=close_date).exists():
#                 transaction.set_rollback(True)
#                 return JsonResponse({"success": False, "error": "День уже закрыт"})
            
#             day_closing = DayClosing.objects.create(date=close_date)

#             # обновляем статус
#             day_closing.closed_at = timezone.now()
#             day_closing.closed_by = user
#             day_closing.note = reason
#             day_closing.save()

#             # логируем
#             DayClosingLog.objects.create(
#                 day_closing=day_closing,
#                 action="close",
#                 performed_by=user,
#                 reason=reason
#             )

#             # ФУНКЦИИ ДЛЯ ВЫЧИСЛЕНИЯ БАЛАНСОВ ПО КАЖДОМУ СЧЕТУ
#             def calculate_balance_by_account(partner, target_date, account_number):
#                 """Вычисляет баланс по конкретному счету"""
#                 entries = Entry.objects.filter(
#                     transaction__partner=partner,
#                     transaction__date__lte=target_date,
#                     account__number=account_number
#                 )
#                 debit = entries.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
#                 credit = entries.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
#                 return debit - credit

#             # снимки балансов партнёров ОТДЕЛЬНО ПО КАЖДОМУ СЧЕТУ
#             for partner in Partner.objects.all():
#                 balance_60_usd = calculate_balance_by_account(partner, close_date, '60')
#                 balance_62_tmt = calculate_balance_by_account(partner, close_date, '62')
#                 balance_75_usd = calculate_balance_by_account(partner, close_date, '75')
#                 balance_76_tmt = calculate_balance_by_account(partner, close_date, '76')
                
#                 # Итоговые балансы по валютам (для обратной совместимости)
#                 total_usd = balance_60_usd + balance_75_usd
#                 total_tmt = balance_62_tmt + balance_76_tmt
                
#                 PartnerBalanceSnapshot.objects.create(
#                     closing=day_closing,
#                     partner=partner,
#                     balance_60_usd=balance_60_usd,
#                     balance_62_tmt=balance_62_tmt,
#                     balance_75_usd=balance_75_usd,
#                     balance_76_tmt=balance_76_tmt,
#                     balance_usd=total_usd,  # для обратной совместимости
#                     balance_tmt=total_tmt,  # для обратной совместимости
#                     balance=Decimal('0.000')  # старое поле
#                 )

#             # снимки складов (остается без изменений)
#             for wp in WarehouseProduct.objects.all():
#                 StockSnapshot.objects.create(
#                     closing=day_closing,
#                     warehouse=wp.warehouse,
#                     product=wp.product,
#                     purchase_price=wp.product.purchase_price,
#                     retail_price=wp.product.retail_price,
#                     wholesale_price=wp.product.wholesale_price,
#                     discount_price=wp.product.discount_price,
#                     firma_price=wp.product.firma_price,
#                     quantity=wp.quantity
#                 )

#     except Exception as e:
#         # ic(str(e))
#         return JsonResponse({"success": False, "error": str(e)})
    
#     return JsonResponse({"success": True, "message": "day success is closed", "date": close_date})

# Удаляем товары которые НЕ участвовали в обороте за день, если ВСЕ три поля = 0: oborot_prihod_qty == 0and oborot_wozwrat_qty == 0 and oborot_rashod_qty == 0
def filter_turnover_products(turnover_product):
    """
    Оставляет только товары, у которых был оборот за день
    """
    result = {}

    for warehouse_id, products in turnover_product.items():
        filtered_products = {}

        for product_id, data in products.items():
            if (
                data["oborot_prihod_qty"] != Decimal("0")
                or data["oborot_wozwrat_qty"] != Decimal("0")
                or data["oborot_rashod_qty"] != Decimal("0")
            ):
                filtered_products[product_id] = data

        # если в складе остались товары — добавляем склад
        if filtered_products:
            result[warehouse_id] = filtered_products

    return result


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def close_day(request):
    user = request.user          # ✅ настоящий пользователь
    data = request.data

    close_date = data.get("date")
    reason = data.get("reason", "")
    

    if not close_date:
        return Response({"success": False, "error": "choose close date"})
  
    
    # если уже закрыт
    if DayClosing.objects.filter(date=close_date).exists():
        return JsonResponse({"success": False, "error": "День уже закрыт"})
    
    #############################################################################################################################################################
    #############################################################################################################################################################
    ####### start towar oborot otchet i detail (2 otchyota)
    
    close_date_format = datetime.strptime(close_date, "%Y-%m-%d").date()
    convert_close_date = close_date_format.strftime("%d.%m.%Y") # dlya date w models.py
    
    day_start = close_date_format
    day_end = close_date_format + timedelta(days=1)

    product_units = (
        ProductUnit.objects
        .filter(is_default_for_sale=True)
        .select_related("unit")
    )

    unit_map = {pu.product_id: pu for pu in product_units}

    turnover_product = {}
    turnover_product_detail = {}

    # ---------------------------------------------
    # 1) НАЧАЛЬНЫЕ ОСТАТКИ
    # ---------------------------------------------
    
    last_closing = DayClosing.objects.filter(
        date__lt=close_date
    ).order_by("-date").first()
    
    snap_map = {}
    #  and close_date != "2025-12-27"
    if last_closing:
        snapshots = StockSnapshot.objects.filter(closing=last_closing)
        snap_map = {
            (s.warehouse_id, s.product_id): s.quantity
            for s in snapshots
        }

  
    for w in Warehouse.objects.all():
        turnover_product[w.id] = {}
        
        # Получаем ВСЕ товары (все продукты из базы)
        # products = Product.objects.all().select_related('category', 'base_unit')
        products = WarehouseProduct.objects.filter(warehouse__id=w.id).select_related('product')
        
        
        for pr in products:
            p = pr.product
            
            
            pu = unit_map.get(p.id)
            if pu:
                unit = pu.unit.name
                conversion_factor = Decimal(pu.conversion_factor)
            else:
                unit = p.base_unit.name if p.base_unit else ""
                conversion_factor = Decimal("1")

            turnover_product[w.id][p.id] = {
                "id": p.id,
                "category": p.category.name if p.category else "",
                "name": p.name,
                "unit": unit,
                "price": p.wholesale_price or Decimal("0"),

                # 👇 ВАЖНО: начинаем с НУЛЯ, будем считать из истории
                "start_qty": Decimal("0"),

                "oborot_prihod_qty": Decimal("0"),
                "oborot_wozwrat_qty": Decimal("0"),
                "oborot_rashod_qty": Decimal("0"),

                "oborot_prihod_price": Decimal("0"),
                "oborot_wozwrat_price": Decimal("0"),
                "oborot_rashod_price": Decimal("0"),

                "end_qty": Decimal("0"),
                "conversion_factor": conversion_factor,
            }
            
        #  and close_date != "2025-12-27"
        if last_closing:
            for p_id, product_data in turnover_product[w.id].items():
                
                product_data["start_qty"] = snap_map.get(
                    (w.id, p_id),
                    Decimal("0")
                )
                
        else:
            
            # Считаем начальные остатки из всех операций ДО начала дня
            start_items = (
                InvoiceItem.objects
                .filter(
                    invoice__entry_created_at_handle__lt=day_start,
                    invoice__canceled_at__isnull=True
                )
                .filter(
                    Q(invoice__warehouse=w) |
                    Q(invoice__warehouse2=w)
                )
                .select_related(
                    "product", "product__category", "product__base_unit", "invoice"
                )
            )
            # test = 0
            for item in start_items:
                p = item.product
                inv = item.invoice

                product_data = turnover_product[w.id].get(p.id)

                if not product_data:
                    continue

                cf = product_data["conversion_factor"]
                qty = item.selected_quantity / cf

                if inv.wozwrat_or_prihod == "prihod":
                    if inv.warehouse_id == w.id:
                        product_data["start_qty"] += qty

                elif inv.wozwrat_or_prihod == "rashod":
                    if inv.warehouse_id == w.id:
                        product_data["start_qty"] -= qty

                elif inv.wozwrat_or_prihod == "wozwrat":
                    if inv.warehouse_id == w.id:
                        product_data["start_qty"] += qty

                elif inv.wozwrat_or_prihod == "transfer":
                    if inv.warehouse_id == w.id:
                        product_data["start_qty"] -= qty
                    elif inv.warehouse2_id == w.id:
                        product_data["start_qty"] += qty

        # ---------------------------------------------
        # 2) ОБОРОТ ЗА ДЕНЬ
        # ---------------------------------------------

        turnover_items = (
            InvoiceItem.objects
            .filter(
                invoice__entry_created_at_handle__gte=day_start,
                invoice__entry_created_at_handle__lt=day_end,
                invoice__canceled_at__isnull=True
            )
            .filter(
                Q(invoice__warehouse=w) |
                Q(invoice__warehouse2=w)
            )
            .select_related(
                "product", "product__category", "product__base_unit", "invoice"
            )
        )

        products_map = {}
        for item in turnover_items:
            p = item.product
            
            if p.id not in products_map:
                products_map[p.id] = {
                    "product_id": p.id,
                    "product_name": p.name,
                    "wholesale_price": p.wholesale_price,
                    "start_qty": turnover_product[w.id][p.id]["start_qty"],
                    "prihod": Decimal("0"),
                    "rashod": Decimal("0"),
                    "wozwrat": Decimal("0"),
                    "end_qty": Decimal("0"),
                    "_ops_index": {},   # ← для invoice
                }
            product = products_map[p.id]
    
            inv = item.invoice
            product_data = turnover_product[w.id].get(p.id)

            if not product_data:
                continue

            cf = product_data["conversion_factor"]
            qty = item.selected_quantity / cf
            
            op_type = inv.wozwrat_or_prihod
            direction = None
            if op_type == "transfer":
                if inv.warehouse_id == w.id:
                    # Текущий склад - источник (расход)
                    direction = "out"
                    op_type_for_detail = "rashod"
                elif inv.warehouse2_id == w.id:
                    # Текущий склад - получатель (приход)
                    direction = "in"
                    op_type_for_detail = "prihod"
                else:
                    # Не должно происходить, но на всякий случай
                    continue
            else:
                # Для не-transfer операций направление не нужно
                op_type_for_detail = op_type
            if op_type_for_detail  == "prihod":
                product["prihod"] += qty
            elif op_type_for_detail  == "rashod":
                product["rashod"] += qty
            elif op_type_for_detail  == "wozwrat":
                product["wozwrat"] += qty
            # row_key = f"{inv.id}_{transfer_status}"
            # row_key = (inv.id, p.id, w.id, direction)
            # row_key = item.id
            if direction:
                row_key = (item.id, direction)
            else:
                row_key = item.id
            ops = product["_ops_index"]
            if row_key not in ops:
                ops[row_key] = {
                    "invoice_id": inv.id,
                    "date": inv.entry_created_at_handle,
                    "partner": inv.partner.name if inv.partner else "",
                    "comment": inv.comment,
                    "type": op_type_for_detail,
                    "qty": Decimal("0"),
                    "price": item.selected_price,
                    "sum": Decimal("0"),
                }
            ops[row_key]["qty"] += qty
            ops[row_key]["sum"] += qty * item.selected_price
            
            
        
            price = qty * item.selected_price

            if inv.wozwrat_or_prihod == "prihod":
                if inv.warehouse_id == w.id:
                    product_data["oborot_prihod_qty"] += qty
                    product_data["oborot_prihod_price"] += price

            elif inv.wozwrat_or_prihod == "rashod":
                if inv.warehouse_id == w.id:
                    product_data["oborot_rashod_qty"] += qty
                    product_data["oborot_rashod_price"] += price

            elif inv.wozwrat_or_prihod == "wozwrat":
                if inv.warehouse_id == w.id:
                    product_data["oborot_wozwrat_qty"] += qty
                    product_data["oborot_wozwrat_price"] += price

            elif inv.wozwrat_or_prihod == "transfer":
                if inv.warehouse_id == w.id:
                    product_data["oborot_rashod_qty"] += qty
                    product_data["oborot_rashod_price"] += price
                elif inv.warehouse2_id == w.id:
                    product_data["oborot_prihod_qty"] += qty
                    product_data["oborot_prihod_price"] += price
                    
        result_products = []
        for product in products_map.values():
            product["operations"] = sorted(
                product["_ops_index"].values(),
                key=lambda x: x["date"]
            )
            product["end_qty"] = (
                product["start_qty"]
                + product["prihod"]
                + product["wozwrat"]
                - product["rashod"]
            )
            del product["_ops_index"]
            result_products.append(product)
            
        turnover_product_detail[w.id] = result_products
                      
      
             

        # ---------------------------------------------
        # 3) КОНЕЧНЫЙ ОСТАТОК (END_QTY)
        # ---------------------------------------------
        
        for product_data in turnover_product[w.id].values():
            product_data["end_qty"] = (
                product_data["start_qty"]
                + product_data["oborot_prihod_qty"]
                + product_data["oborot_wozwrat_qty"]
                - product_data["oborot_rashod_qty"]
            )
            
    # for warehouse_id, products in turnover_product_detail.items():
    #     for product in products:
    #         for op in product.get("operations", []):
    #             comment = op.get("comment", "")
    #             if comment and "test" in comment.lower():
    #                 # ic(product)
    #                 pass
    #     if warehouse_id == 2:
    #         ic(products)
                
    # ------------------------------------------------
    # 6) Excel
    # ------------------------------------------------
    # ================== СТИЛИ ==================

    # Шапка
    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="4472C4"   # синий
    )
    HEADER_FONT = Font(
        bold=True,
        color="FFFFFF"    # белый текст
    )
    CENTER_ALIGN = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    # Категория
    CATEGORY_FILL = PatternFill(
        fill_type="solid",
        fgColor="E7E6E6"
    )
    CATEGORY_FONT = Font(bold=True)

    # Обычные ячейки
    NORMAL_FONT = Font(color="000000")
    LEFT_ALIGN = Alignment(vertical="center", horizontal="left")
    RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

    # Границы
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Форматы чисел
    PRICE_FMT = '#,##0.00'
    QTY_FMT = '#,##0.###'

    # Ширина колонок
    COLUMN_WIDTHS = {
        "A": 25,
        "B": 40,
        "C": 10,
        "D": 12,
        "E": 8, "F": 8,
        "G": 8, "H": 8,
        "I": 8, "J": 8,
        "K": 8, "L": 8,
        "M": 8, "N": 8,
    }
    
    TOTAL_FILL = PatternFill(
        fill_type="solid",
        fgColor="D9D9D9"   # светло-серый
    )

    TOTAL_FONT = Font(
        bold=True
    )

    GRAY_FILL_0 = PatternFill(
        fill_type="solid",
        fgColor="F5F5F5"
    )
    
    GRAY_FILL_1 = PatternFill(
        fill_type="solid",
        fgColor="EDEDED"
    )
    
    GRAY_FILL_2 = PatternFill(
        fill_type="solid",
        fgColor="DCDCDC"
    )
    
    GRAY_FILL_3 = PatternFill(
        fill_type="solid",
        fgColor="C8C8C8"
    )
    
    GRAY_FILL_4 = PatternFill(
        fill_type="solid",
        fgColor="B0B0B0"
    )
    
    GREEN_FILL_0 = PatternFill(
        fill_type="solid",
        fgColor="E2F0D9"
    )
    
    
    GREEN_FILL_1 = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )
    
    GREEN_FILL_2 = PatternFill(
        fill_type="solid",
        fgColor="92D050"
    )
    
    GREEN_FILL_3 = PatternFill(
        fill_type="solid",
        fgColor="006100"
    )
    
    RED_FONT = Font(color="FF0000")
    GREEN_FONT = Font(color="006400")
    BLUE_FONT = Font(color="0000FF")
    # ================== EXCEL ==================

    wb_product_oborot = Workbook()
    wb_product_oborot.remove(wb_product_oborot.active)
    
    turnover_product_only_turnover = filter_turnover_products(turnover_product)

    # wse towary
    for warehouse_id, products in turnover_product.items():
        warehouse = Warehouse.objects.get(id=warehouse_id)
        ws = wb_product_oborot.create_sheet(title=f"{warehouse.name[:31]}_all")

        # ширина колонок
        for col, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

        # ---------- ШАПКА ----------
        ws["A1"] = "Категория"
        ws["B1"] = "Товар"
        ws["C1"] = "Ед."
        ws["D1"] = "Цена"

        ws.merge_cells("E1:F1"); ws["E1"] = "Остаток на начало"
        ws.merge_cells("G1:H1"); ws["G1"] = "Приход"
        ws.merge_cells("I1:J1"); ws["I1"] = "Возврат"
        ws.merge_cells("K1:L1"); ws["K1"] = "Расход"
        ws.merge_cells("M1:N1"); ws["M1"] = "Конечный остаток"

        ws["E2"] = "Кол-во"; ws["F2"] = "Сумма"
        ws["G2"] = "Кол-во"; ws["H2"] = "Сумма"
        ws["I2"] = "Кол-во"; ws["J2"] = "Сумма"
        ws["K2"] = "Кол-во"; ws["L2"] = "Сумма"
        ws["M2"] = "Кол-во"; ws["N2"] = "Сумма"

        ws.freeze_panes = "A3"

        # стиль шапки
        for row in (1, 2):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER

        # ---------- ДАННЫЕ ----------
        row_num = 3
        current_category = None
        category_start_row = None
        warehouse_start_row = row_num

        # Фильтруем товары - ПОКАЗЫВАЕМ ВСЕ ТОВАРЫ
        # Убрана фильтрация по нулевым оборотам
        filtered_products = products
        warehouse_product_rows = []
        for data in sorted(filtered_products.values(), key=lambda x: (x["category"], x["name"])):
            

            # ===== НОВАЯ КАТЕГОРИЯ =====
            if data["category"] != current_category:

                # ---- Итог по предыдущей категории ----
                if current_category is not None:
                    ws[f"B{row_num}"] = f"ИТОГО по категории: {current_category}"

                    for col in ["E","F","G","H","I","J","K","L","M","N"]:
                        ws[f"{col}{row_num}"] = f"=SUM({col}{category_start_row}:{col}{row_num-1})"

                    for col in range(1, 15):
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill = TOTAL_FILL
                        cell.font = TOTAL_FONT
                        cell.border = THIN_BORDER
                        cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN

                    row_num += 2  # пустая строка после итога

                # ---- Строка категории ----
                ws.merge_cells(f"A{row_num}:N{row_num}")
                cell = ws[f"A{row_num}"]
                cell.value = data["category"]
                cell.fill = CATEGORY_FILL
                cell.font = CATEGORY_FONT
                cell.alignment = LEFT_ALIGN
                cell.border = THIN_BORDER

                for col in range(2, 15):
                    ws.cell(row=row_num, column=col).border = THIN_BORDER

                current_category = data["category"]
                category_start_row = row_num + 1
                row_num += 1

            # ===== СТРОКА ТОВАРА =====
            
            
            ws[f"A{row_num}"] = ""  # Категория (пустая в строке товара)
            ws[f"B{row_num}"] = data["name"]
            ws[f"C{row_num}"] = data["unit"]
            ws[f"D{row_num}"] = float(data["price"])

            ws[f"E{row_num}"] = float(data["start_qty"])
            ws[f"F{row_num}"] = float(data["start_qty"] * data["price"])

            ws[f"G{row_num}"] = float(data["oborot_prihod_qty"])
            ws[f"H{row_num}"] = float(data["oborot_prihod_price"])

            ws[f"I{row_num}"] = float(data["oborot_wozwrat_qty"])
            ws[f"J{row_num}"] = float(data["oborot_wozwrat_price"])

            ws[f"K{row_num}"] = float(data["oborot_rashod_qty"])
            ws[f"L{row_num}"] = float(data["oborot_rashod_price"])

            ws[f"M{row_num}"] = float(data["end_qty"])
            ws[f"N{row_num}"] = float(data["end_qty"] * data["price"])

            for col in range(1, 15):
                cell = ws.cell(row=row_num, column=col)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT_ALIGN if col <= 3 else RIGHT_ALIGN

            ws[f"D{row_num}"].number_format = PRICE_FMT
            for col in ["F", "H", "J", "L", "N"]:
                ws[f"{col}{row_num}"].number_format = PRICE_FMT
            for col in ["E", "G", "I", "K", "M"]:
                ws[f"{col}{row_num}"].number_format = QTY_FMT
            
            warehouse_product_rows.append(row_num)
            row_num += 1

        # ===== ИТОГ ПО ПОСЛЕДНЕЙ КАТЕГОРИИ =====
        if current_category is not None:
            ws[f"B{row_num}"] = f"ИТОГО по категории: {current_category}"

            for col in ["E","F","G","H","I","J","K","L","M","N"]:
                ws[f"{col}{row_num}"] = f"=SUM({col}{category_start_row}:{col}{row_num-1})"

            for col in range(1, 15):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN

            row_num += 2

        # ===== ИТОГ ПО СКЛАДУ =====
        ws[f"B{row_num}"] = "ИТОГО ПО СКЛАДУ"

        for col in ["E","F","G","H","I","J","K","L","M","N"]:

            if not warehouse_product_rows:
                ws[f"{col}{row_num}"] = 0
                continue

            ranges = []
            start = prev = warehouse_product_rows[0]

            for r in warehouse_product_rows[1:]:
                if r == prev + 1:
                    prev = r
                else:
                    ranges.append(f"{col}{start}:{col}{prev}")
                    start = prev = r

            ranges.append(f"{col}{start}:{col}{prev}")

            ws[f"{col}{row_num}"] = f"=SUM({','.join(ranges)})"

        for col in range(1, 15):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN

    # towary tolko s oborotami za den
    for warehouse_id, products in turnover_product_only_turnover.items():
        warehouse = Warehouse.objects.get(id=warehouse_id)
        ws = wb_product_oborot.create_sheet(title=warehouse.name[:31])

        # ширина колонок
        for col, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

        # ---------- ШАПКА ----------
        ws["A1"] = "Категория"
        ws["B1"] = "Товар"
        ws["C1"] = "Ед."
        ws["D1"] = "Цена"

        ws.merge_cells("E1:F1"); ws["E1"] = "Остаток на начало"
        ws.merge_cells("G1:H1"); ws["G1"] = "Приход"
        ws.merge_cells("I1:J1"); ws["I1"] = "Возврат"
        ws.merge_cells("K1:L1"); ws["K1"] = "Расход"
        ws.merge_cells("M1:N1"); ws["M1"] = "Конечный остаток"

        ws["E2"] = "Кол-во"; ws["F2"] = "Сумма"
        ws["G2"] = "Кол-во"; ws["H2"] = "Сумма"
        ws["I2"] = "Кол-во"; ws["J2"] = "Сумма"
        ws["K2"] = "Кол-во"; ws["L2"] = "Сумма"
        ws["M2"] = "Кол-во"; ws["N2"] = "Сумма"

        ws.freeze_panes = "A3"

        # стиль шапки
        for row in (1, 2):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = CENTER_ALIGN
                cell.border = THIN_BORDER

        # ---------- ДАННЫЕ ----------
        row_num = 3
        current_category = None
        category_start_row = None
        warehouse_start_row = row_num

        # Фильтруем товары - ПОКАЗЫВАЕМ ВСЕ ТОВАРЫ
        # Убрана фильтрация по нулевым оборотам
        filtered_products = products
        warehouse_product_rows = []
        for data in sorted(filtered_products.values(), key=lambda x: (x["category"], x["name"])):
            

            # ===== НОВАЯ КАТЕГОРИЯ =====
            if data["category"] != current_category:

                # ---- Итог по предыдущей категории ----
                if current_category is not None:
                    ws[f"B{row_num}"] = f"ИТОГО по категории: {current_category}"

                    for col in ["E","F","G","H","I","J","K","L","M","N"]:
                        ws[f"{col}{row_num}"] = f"=SUM({col}{category_start_row}:{col}{row_num-1})"

                    for col in range(1, 15):
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill = TOTAL_FILL
                        cell.font = TOTAL_FONT
                        cell.border = THIN_BORDER
                        cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN

                    row_num += 2  # пустая строка после итога

                # ---- Строка категории ----
                ws.merge_cells(f"A{row_num}:N{row_num}")
                cell = ws[f"A{row_num}"]
                cell.value = data["category"]
                cell.fill = CATEGORY_FILL
                cell.font = CATEGORY_FONT
                cell.alignment = LEFT_ALIGN
                cell.border = THIN_BORDER

                for col in range(2, 15):
                    ws.cell(row=row_num, column=col).border = THIN_BORDER

                current_category = data["category"]
                category_start_row = row_num + 1
                row_num += 1

            # ===== СТРОКА ТОВАРА =====
            
            
            ws[f"A{row_num}"] = ""  # Категория (пустая в строке товара)
            ws[f"B{row_num}"] = data["name"]
            ws[f"C{row_num}"] = data["unit"]
            ws[f"D{row_num}"] = float(data["price"])

            ws[f"E{row_num}"] = float(data["start_qty"])
            ws[f"F{row_num}"] = float(data["start_qty"] * data["price"])

            ws[f"G{row_num}"] = float(data["oborot_prihod_qty"])
            ws[f"H{row_num}"] = float(data["oborot_prihod_price"])

            ws[f"I{row_num}"] = float(data["oborot_wozwrat_qty"])
            ws[f"J{row_num}"] = float(data["oborot_wozwrat_price"])

            ws[f"K{row_num}"] = float(data["oborot_rashod_qty"])
            ws[f"L{row_num}"] = float(data["oborot_rashod_price"])

            ws[f"M{row_num}"] = float(data["end_qty"])
            ws[f"N{row_num}"] = float(data["end_qty"] * data["price"])

            for col in range(1, 15):
                cell = ws.cell(row=row_num, column=col)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = LEFT_ALIGN if col <= 3 else RIGHT_ALIGN

            ws[f"D{row_num}"].number_format = PRICE_FMT
            for col in ["F", "H", "J", "L", "N"]:
                ws[f"{col}{row_num}"].number_format = PRICE_FMT
            for col in ["E", "G", "I", "K", "M"]:
                ws[f"{col}{row_num}"].number_format = QTY_FMT
            
            warehouse_product_rows.append(row_num)
            row_num += 1

        # ===== ИТОГ ПО ПОСЛЕДНЕЙ КАТЕГОРИИ =====
        if current_category is not None:
            ws[f"B{row_num}"] = f"ИТОГО по категории: {current_category}"

            for col in ["E","F","G","H","I","J","K","L","M","N"]:
                ws[f"{col}{row_num}"] = f"=SUM({col}{category_start_row}:{col}{row_num-1})"

            for col in range(1, 15):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN

            row_num += 2

        # ===== ИТОГ ПО СКЛАДУ =====
        ws[f"B{row_num}"] = "ИТОГО ПО СКЛАДУ"

        for col in ["E","F","G","H","I","J","K","L","M","N"]:

            if not warehouse_product_rows:
                ws[f"{col}{row_num}"] = 0
                continue

            ranges = []
            start = prev = warehouse_product_rows[0]

            for r in warehouse_product_rows[1:]:
                if r == prev + 1:
                    prev = r
                else:
                    ranges.append(f"{col}{start}:{col}{prev}")
                    start = prev = r

            ranges.append(f"{col}{start}:{col}{prev}")

            ws[f"{col}{row_num}"] = f"=SUM({','.join(ranges)})"

        for col in range(1, 15):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            cell.alignment = RIGHT_ALIGN if col >= 5 else LEFT_ALIGN
             
    # ================== СОХРАНЕНИЕ ==================
    
    

    excel_buffer = BytesIO()
    wb_product_oborot.save(excel_buffer)
    excel_buffer.seek(0)
    product_oborot_content = excel_buffer.read()
    
    

    ####### end towar oborot otchet i detail (2 otchyota)
    #############################################################################################################################################################
    #############################################################################################################################################################
    
    
        
    #############################################################################################################################################################
    #############################################################################################################################################################
    ####### start, skidka, Отклонение от оптовой цены
    
    # ====== СКЛАДЫ ======
    warehouses = Warehouse.objects.all()

    # ====== WORKBOOK ======
    wb_skidka = Workbook()
    wb_skidka.remove(wb_skidka.active)

    # =========================================================
    # 📊 ЛИСТ: ПО СКЛАДАМ (СВОДКА)
    # =========================================================
    ws_total = wb_skidka.create_sheet(title="ПО СКЛАДАМ")

    ws_total.merge_cells("A1:E1")
    ws_total["A1"] = f"Отклонение от оптовой цены за {convert_close_date}"
    ws_total["A1"].fill = HEADER_FILL
    ws_total["A1"].font = HEADER_FONT
    ws_total["A1"].alignment = CENTER_ALIGN

    ws_total["A3"] = "Склад"
    ws_total["B3"] = "Выручка"
    ws_total["C3"] = "Скидки"
    ws_total["D3"] = "Наценки"
    ws_total["E3"] = "Итоговое отклонение"

    ws_total.column_dimensions["A"].width = 18
    ws_total.column_dimensions["E"].width = 28

    for c in ["A3", "B3", "C3", "D3", "E3"]:
        ws_total[c].fill = CATEGORY_FILL
        ws_total[c].border = THIN_BORDER
        ws_total[c].alignment = CENTER_ALIGN

    row_total = 4

    # =========================================================
    # 🔁 ОДИН ПРОХОД ПО СКЛАДАМ
    # =========================================================
    for w in warehouses:

        total_all_price = Decimal('0')
        otkloneniy_wsego = Decimal('0')
        skidki = Decimal('0')
        nasenki = Decimal('0')
        opt_sum = Decimal('0')

        table = []

        turnover_items = (
            InvoiceItem.objects
            .filter(
                invoice__entry_created_at_handle__gte=day_start,
                invoice__entry_created_at_handle__lt=day_end,
                invoice__canceled_at__isnull=True,
                invoice__wozwrat_or_prihod="rashod",
                invoice__warehouse=w,
            )
            .select_related(
                "invoice",
                "invoice__partner",
                "product"
            )
        )

        # ================== ПОДСЧЁТ ==================
        for t in turnover_items:
            opt_sum += t.wholesale_price * t.selected_quantity
            total_all_price += t.selected_quantity * t.selected_price

            otkloneniye = (
                t.selected_price - t.wholesale_price
            ) * t.selected_quantity

            otkloneniy_wsego += otkloneniye

            if otkloneniye < 0:
                skidki += abs(otkloneniye)
            elif otkloneniye > 0:
                nasenki += otkloneniye

            if t.selected_price != t.product.wholesale_price:
                table.append({
                    "partner_name": t.invoice.partner.name,
                    "invoice": t.invoice,
                    "product_name": t.product.name,
                    "unit": t.unit_name_on_selected_warehouses,
                    "wholesale_price": t.product.wholesale_price,
                    "selected_price": t.selected_price,
                    "selected_quantity": t.selected_quantity,
                    "total_selected_price": t.selected_price * t.selected_quantity,
                    "difference": (
                        t.selected_price * t.selected_quantity
                        - t.product.wholesale_price * t.selected_quantity
                    ),
                })

        percent = Decimal('0')
        if opt_sum > 0:
            percent = (otkloneniy_wsego / opt_sum * 100).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP
            )

        # =====================================================
        # 🧾 ЗАПИСЬ В СВОДНЫЙ ЛИСТ
        # =====================================================
        ws_total[f"A{row_total}"] = w.name
        ws_total[f"B{row_total}"] = total_all_price
        ws_total[f"C{row_total}"] = -skidki if skidki > 0 else Decimal('0')
        ws_total[f"D{row_total}"] = nasenki
        ws_total[f"E{row_total}"] = otkloneniy_wsego

        for col in ["B", "C", "D", "E"]:
            ws_total[f"{col}{row_total}"].number_format = '#,##0.00'
            ws_total[f"{col}{row_total}"].border = THIN_BORDER

        ws_total[f"A{row_total}"].border = THIN_BORDER
        row_total += 1

        # =====================================================
        # 📄 ЛИСТ КОНКРЕТНОГО СКЛАДА
        # =====================================================
        ws = wb_skidka.create_sheet(title=w.name[:31])

        ws.merge_cells("A1:J1")
        ws["A1"] = f"Отклонение от оптовой цены за {convert_close_date} — {w.name}"
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = HEADER_FONT
        ws["A1"].alignment = CENTER_ALIGN

        headers = [
            "№", "Партнер", "Комментарий", "Товар", "Ед.",
            "Оптовая цена", "Цена продажи", "Кол-во",
            "Всего цена продажи", "Разница"
        ]

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = CATEGORY_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        row = 3
        for i, value in enumerate(table, start=1):
            ws[f"A{row}"] = i
            ws[f"B{row}"] = value["partner_name"]
            ws[f"C{row}"] = f"№{value['invoice'].pk} {value['invoice'].comment}"
            ws[f"D{row}"] = value["product_name"]
            ws[f"E{row}"] = value["unit"]
            ws[f"F{row}"] = value["wholesale_price"]
            ws[f"G{row}"] = value["selected_price"]
            ws[f"H{row}"] = value["selected_quantity"]
            ws[f"I{row}"] = value["total_selected_price"]
            ws[f"J{row}"] = value["difference"]

            for c in ["F", "G", "I", "J"]:
                ws[f"{c}{row}"].number_format = '#,##0.00'
            ws[f"H{row}"].number_format = '#,##0.###'

            for c in ["A","B","C","D","E","F","G","H","I","J"]:
                ws[f"{c}{row}"].border = THIN_BORDER

            row += 1

        # ===== ИТОГИ =====
        row += 1
        ws[f"B{row}"] = "Показатель"
        ws[f"C{row}"] = "Значение"
        ws[f"B{row}"].fill = CATEGORY_FILL
        ws[f"C{row}"].fill = CATEGORY_FILL
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"C{row}"].border = THIN_BORDER

        labels = [
            ("Общая выручка", total_all_price),
            ("Отклонение всего", otkloneniy_wsego),
            ("Скидки", -skidki if skidki > 0 else Decimal('0')),
            ("Наценки", nasenki),
            ("% отклонения", percent),
        ]

        for i, (label, val) in enumerate(labels, start=1):
            ws[f"B{row+i}"] = label
            ws[f"C{row+i}"] = val
            ws[f"B{row+i}"].border = THIN_BORDER
            ws[f"C{row+i}"].border = THIN_BORDER

        ws[f"C{row+1}"].number_format = '#,##0.00'
        ws[f"C{row+2}"].number_format = '#,##0.00'
        ws[f"C{row+3}"].number_format = '#,##0.00'
        ws[f"C{row+4}"].number_format = '#,##0.00'
        ws[f"C{row+5}"].number_format = '0.00"%"'

    # =========================================================
    # 💾 СОХРАНЕНИЕ
    # =========================================================
    excel_buffer = BytesIO()
    wb_skidka.save(excel_buffer)
    excel_buffer.seek(0)
    skidka_content = excel_buffer.read()
    
    ####### end, skidka, Отклонение от оптовой цены
    #############################################################################################################################################################
    #############################################################################################################################################################
    
    
    #############################################################################################################################################################
    #############################################################################################################################################################
    ####### start exel, buh oborot towarow detail
    wb_product_oborot_detail = Workbook()
    wb_product_oborot_detail.remove(wb_product_oborot_detail.active)
    
    
    for warehouse_id, products in turnover_product_detail.items():
        warehouse = Warehouse.objects.get(id=warehouse_id)
        ws = wb_product_oborot_detail.create_sheet(title=warehouse.name[:31])
        
        ws.merge_cells("A1:L1")
        ws.merge_cells("A2:L2")
        ws["A1"] = "ДЕТАЛЬНЫЙ ОТЧЁТ ПО ТОВАРАМ"
        ws["A2"] = str(convert_close_date)
        ws["A1"].alignment = CENTER_ALIGN
        ws["A2"].alignment = CENTER_ALIGN
        ws["A1"].font = CATEGORY_FONT
        ws["A2"].font = CATEGORY_FONT
        
        row = 4
        
        for product in products:
            wholesale_price = product.get("wholesale_price", "")
            ws.merge_cells(f"A{row}:L{row}")
            ws[f"A{row}"] = f'{product["product_name"]} (Цена:{wholesale_price})'
            ws[f"A{row}"].alignment = CENTER_ALIGN
            ws[f"A{row}"].font = CATEGORY_FONT
            
            row += 1
            ws[f"A{row}"] = "Дата"
            ws[f"B{row}"] = "Партнёр"
            ws[f"C{row}"] = "Комментарий"
            ws[f"D{row}"] = "Цена"
            
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 35
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"G{row}", f"I{row}", f"K{row}"]:
                ws[i].fill = GRAY_FILL_1
                ws[i].border = THIN_BORDER
                ws[i].alignment = CENTER_ALIGN
                ws[i].font = CATEGORY_FONT
                
            
            
            ws.merge_cells(f"E{row}:F{row}")
            ws[f"E{row}"] = "Приход"
            
            ws.merge_cells(f"G{row}:H{row}")
            ws[f"G{row}"] = "Возврат"
            
            ws.merge_cells(f"I{row}:J{row}")
            ws[f"I{row}"] = "Расход"
            
            ws.merge_cells(f"K{row}:L{row}")
            ws[f"K{row}"] = "Остаток"
            
            row += 1
            ws[f"E{row}"] = "Кол-во"
            ws[f"F{row}"] = "Всего"
            ws[f"G{row}"] = "Кол-во"
            ws[f"H{row}"] = "Всего"
            ws[f"I{row}"] = "Кол-во"
            ws[f"J{row}"] = "Всего"
            ws[f"K{row}"] = "Кол-во"
            ws[f"L{row}"] = "Всего"
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}"]:
                ws[i].fill = GRAY_FILL_0
                ws[i].border = THIN_BORDER
                ws[i].alignment = CENTER_ALIGN
                ws[i].font = CATEGORY_FONT
                
            
            row += 1
            ws[f"A{row}"] = "Остаток на начало"
            ws[f"K{row}"] = product["start_qty"]
            ws[f"L{row}"] = product["start_qty"] * product["wholesale_price"]
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}"]:
                ws[i].border = THIN_BORDER
                ws[i].font = CATEGORY_FONT
            
            row += 1
            running_qty = product["start_qty"]
            rashod_total_qty = 0
            rashod_total_price = 0
            prihod_total_qty = 0
            prihod_total_price = 0
            wozwrat_total_qty = 0
            wozwrat_total_price = 0
            
            ops = product.get("operations", [])
            if not ops:
                continue
            
            for op in product.get("operations", []):
                ws[f"D{row}"].number_format = PRICE_FMT
                ws[f"F{row}"].number_format = PRICE_FMT
                ws[f"H{row}"].number_format = PRICE_FMT
                ws[f"J{row}"].number_format = PRICE_FMT
                ws[f"L{row}"].number_format = PRICE_FMT

                ws[f"E{row}"].number_format = QTY_FMT
                ws[f"G{row}"].number_format = QTY_FMT
                ws[f"I{row}"].number_format = QTY_FMT
                ws[f"K{row}"].number_format = QTY_FMT
                
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}"]:
                    ws[i].border = THIN_BORDER

                ws[f"A{row}"] = op["date"].strftime("%d.%m.%Y")
                ws[f"B{row}"] = op["partner"]
                ws[f"C{row}"] = f'№{op["invoice_id"]} {op["comment"]}'
                
                ws[f"D{row}"] = op["price"]
       
                if op["type"] == "prihod":
                    # end_qty += op["qty"]
                    ws[f"E{row}"] = op["qty"]
                    ws[f"F{row}"] = op["qty"] * op["price"]
                    ws[f"E{row}"].font = GREEN_FONT
                    ws[f"F{row}"].font = GREEN_FONT
                    running_qty += op["qty"]
                    prihod_total_qty += op["qty"]
                    prihod_total_price += op["qty"] * op["price"]
                    
                if op["type"] == "wozwrat":
                    # end_qty += op["qty"]
                    ws[f"G{row}"] = op["qty"]
                    ws[f"H{row}"] = op["qty"] * op["price"]
                    ws[f"G{row}"].font = RED_FONT
                    ws[f"H{row}"].font = RED_FONT
                    running_qty += op["qty"]
                    wozwrat_total_qty += op["qty"]
                    wozwrat_total_price += op["qty"] * op["price"]
                    
                    # ws[f"K{row}"] = product["start_qty"] + op["qty"]
                    # ws[f"L{row}"] = (product["start_qty"] + op["qty"]) * op["price"]
                    
                if op["type"] == "rashod":
                    # end_qty -= op["qty"]
                    ws[f"I{row}"] = op["qty"]
                    ws[f"J{row}"] = op["qty"] * op["price"]
                    ws[f"I{row}"].font = BLUE_FONT
                    ws[f"J{row}"].font = BLUE_FONT
                    running_qty -= op["qty"]
                    rashod_total_qty += op["qty"]
                    rashod_total_price += op["qty"] * op["price"]
                    
                    # ws[f"K{row}"] = product["start_qty"] - op["qty"]
                    # ws[f"L{row}"] = (product["start_qty"] - op["qty"]) * op["price"]
                ws[f"K{row}"] = running_qty
                ws[f"L{row}"] = running_qty * product["wholesale_price"]
                    
              
                row += 1
                
            ws[f"A{row}"] = "Итого"
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}"]:
                ws[i].border = THIN_BORDER
                ws[i].font = CATEGORY_FONT
            
            ws[f"D{row}"].number_format = PRICE_FMT
            ws[f"F{row}"].number_format = PRICE_FMT
            ws[f"H{row}"].number_format = PRICE_FMT
            ws[f"J{row}"].number_format = PRICE_FMT
            ws[f"L{row}"].number_format = PRICE_FMT

            ws[f"E{row}"].number_format = QTY_FMT
            ws[f"G{row}"].number_format = QTY_FMT
            ws[f"I{row}"].number_format = QTY_FMT
            ws[f"K{row}"].number_format = QTY_FMT
            
            ws[f"E{row}"] = prihod_total_qty if prihod_total_qty != 0 else ""
            ws[f"F{row}"] = prihod_total_price if prihod_total_price != 0 else ""
            
            ws[f"G{row}"] = wozwrat_total_qty if wozwrat_total_qty != 0 else ""
            ws[f"H{row}"] = wozwrat_total_price if wozwrat_total_price != 0 else ""
            
            ws[f"I{row}"] = rashod_total_qty if rashod_total_qty != 0 else ""
            ws[f"J{row}"] = rashod_total_price if rashod_total_price != 0 else ""
            
            ws[f"K{row}"] = running_qty
            ws[f"L{row}"] = running_qty * wholesale_price
            
            row += 1
            row += 1
     
                

    # ===== СОХРАНЕНИЕ =====
    excel_buffer = BytesIO()
    wb_product_oborot_detail.save(excel_buffer)
    excel_buffer.seek(0)
    product_oborot_detail_content = excel_buffer.read()
    

    
    
    ####### end exel, buh oborot towarow detail
    #############################################################################################################################################################
    #############################################################################################################################################################
    
    #############################################################################################################################################################
    #############################################################################################################################################################
    ####### start card prowodok
    
    # =========================================================
    # ОСНОВНОЙ ОСВ
    # =========================================================
    
    
    
    accounts_OSW = {}
    
    for account in Account.objects.all():
        accounts_OSW[account.id] = {
            'id': account.id,
            'number': account.number,
            'name': account.name,
            'parent_id': account.parent_id,
            'debit_start': Decimal('0'),
            'credit_start': Decimal('0'),
            'opening_balance': Decimal('0'),
            'debit_turnover': Decimal('0'),
            'credit_turnover': Decimal('0'),
            'closing_balance': Decimal('0'),
        }
        
    
        


    opening_OSW = (
        Entry.objects
        .filter(transaction__date__lt=day_start)
        .values('account')
        .annotate(
            debit_sum=Sum('debit'),
            credit_sum=Sum('credit')
        )
    )
    
    

    for o in opening_OSW:
        debit = o['debit_sum'] or Decimal('0')
        credit = o['credit_sum'] or Decimal('0')
        
        
        

        acc = accounts_OSW[o['account']]
        acc['debit_start'] = debit
        acc['credit_start'] = credit
        acc['opening_balance'] = debit - credit
        
        
    turnover_OSW = (
        Entry.objects
        .filter(transaction__date__gte=day_start, transaction__date__lt=day_end)
        .values('account')
        .annotate(
            debit_sum=Sum('debit'),
            credit_sum=Sum('credit')
        )
    )
    
    for t in turnover_OSW:
        debit = t['debit_sum'] or Decimal('0')
        credit = t['credit_sum'] or Decimal('0')

        acc = accounts_OSW[t['account']]
        acc['debit_turnover'] = debit
        acc['credit_turnover'] = credit
        
    for acc in accounts_OSW.values():
        acc['closing_balance'] = (
            acc['opening_balance']
            + acc['debit_turnover']
            - acc['credit_turnover']
        )
        
    total_debit = sum(acc['debit_turnover'] for acc in accounts_OSW.values())
    total_credit = sum(acc['credit_turnover'] for acc in accounts_OSW.values())

    if total_debit != total_credit:
        raise ValueError("НАРУШЕНА ДВОЙНАЯ ЗАПИСЬ")
        
    # сортируем счета так, чтобы дети шли первыми
    accounts_sorted = sorted(
        accounts_OSW.values(),
        key=lambda x: x['number'].count('.'),
        reverse=True
    )

    for acc in accounts_sorted:
        parent_id = acc['parent_id']
        if parent_id:
            parent = accounts_OSW[parent_id]

            parent['debit_start'] += acc['debit_start']
            parent['credit_start'] += acc['credit_start']
            parent['opening_balance'] += acc['opening_balance']

            parent['debit_turnover'] += acc['debit_turnover']
            parent['credit_turnover'] += acc['credit_turnover']
            parent['closing_balance'] += acc['closing_balance']
     
     
     
    wb_OSW = Workbook()
    wb_OSW.remove(wb_OSW.active)
    
    ws = wb_OSW.create_sheet(title="ОСВ")
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws["A1"] = "Оборотно-сальдовая ведомость"
    ws["A2"] = str(convert_close_date)
    
    ws["A1"].alignment = CENTER_ALIGN
    ws["A2"].alignment = CENTER_ALIGN
    
    ws["A1"].font = CATEGORY_FONT
    ws["A2"].font = CATEGORY_FONT
    
    
    
    # row += 1
    
    ws["A4"] = "Счёт"
    ws["B4"] = "Название счёта"
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    
    ws.merge_cells("C4:D4")
    ws["C4"] = "Сальдо на начало"
    
    ws.merge_cells("E4:F4")
    ws["E4"] = "Обороты за период"
    
    ws.merge_cells("G4:H4")
    ws["G4"] = "Сальдо на конец"
    
    ws["C5"] = "Дебет"
    ws["D5"] = "Кредит"
    
    ws["E5"] = "Дебет"
    ws["F5"] = "Кредит"
    
    ws["G5"] = "Дебет"
    ws["H5"] = "Кредит"
    
    for i in ["A4", "B4", "C4", "D4", "E4", "F4", "G4", "H4"]:
        ws[i].alignment = CENTER_ALIGN
        ws[i].font = CATEGORY_FONT
        ws[i].fill = GRAY_FILL_1
        ws[i].border = THIN_BORDER
        
    for i in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5"]:
        ws[i].alignment = CENTER_ALIGN
        ws[i].font = CATEGORY_FONT
        ws[i].fill = GRAY_FILL_1
        ws[i].border = THIN_BORDER
        
        
        

    row = 6
    
    # сначала найдём все parent_id
    parent_ids = {
        acc['parent_id']
        for acc in accounts_OSW.values()
        if acc['parent_id']
    }

    # листовые счета = те, кто не родитель
    leaf_accounts = [
        acc for acc_id, acc in accounts_OSW.items()
        if acc_id not in parent_ids
    ]
    
    detail_sheets = {}

    for acc in leaf_accounts:
        title=f"{acc['number']} {acc['name'][:20]}"
        ws_detail = wb_OSW.create_sheet(title=title)
        detail_sheets[acc["id"]] = ws_detail

        ws_detail["A1"] = "← ОСВ"
        ws_detail["A1"].font = Font(bold=True, color="0000FF")
        ws_detail["A1"].hyperlink = f"#'ОСВ'!A1"
        
        if acc['number'].startswith(("40", "42")):
            ws_detail.merge_cells("A2:N2")
            ws_detail.merge_cells("A3:N3")
        else:
            ws_detail.merge_cells("A2:H2")
        ws_detail["A2"] = f"Счёт {acc['number']} — {acc['name']}"
        ws_detail["A2"].alignment = CENTER_ALIGN
        ws_detail["A2"].font = TOTAL_FONT
        
       
        ws_detail["A3"] = convert_close_date
        ws_detail["A3"].alignment = CENTER_ALIGN
        ws_detail["A3"].font = TOTAL_FONT
        
    
        
    for account_id, data in accounts_OSW.items():
        cell = ws[f"B{row}"]
        cell.value = data["number"]
        
        title=f"{data['number']} {data['name'][:20]}"
        if data["id"] in detail_sheets:
            cell.hyperlink = f"#'{title}'!A1"
            cell.style = "Hyperlink"
        
        for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}"]:
            ws[i].border = THIN_BORDER
            
        ws[f"C{row}"].number_format = PRICE_FMT
        ws[f"D{row}"].number_format = PRICE_FMT
        ws[f"E{row}"].number_format = PRICE_FMT
        ws[f"F{row}"].number_format = PRICE_FMT
        ws[f"G{row}"].number_format = PRICE_FMT
        ws[f"H{row}"].number_format = PRICE_FMT

     
            
        ws[f"A{row}"] = data["number"]
        ws[f"B{row}"] = data["name"]
        
        ws[f"C{row}"] = data["debit_start"]
        ws[f"D{row}"] = data["credit_start"]
        
        ws[f"E{row}"] = data["debit_turnover"]
        ws[f"F{row}"] = data["credit_turnover"]
        ws[f"E{row}"].font = GREEN_FONT
        ws[f"F{row}"].font = RED_FONT
        
        
        if data["closing_balance"] >= 0:
            ws[f"G{row}"] = data["closing_balance"]   # дебет
        else:
            ws[f"H{row}"] = abs(data["closing_balance"])  # кредит
            
        row += 1
        
    grand_total = {
        'debit_start': Decimal('0'),
        'credit_start': Decimal('0'),
        'debit_turnover': Decimal('0'),
        'credit_turnover': Decimal('0'),
        'closing_debit': Decimal('0'),
        'closing_credit': Decimal('0'),
    }
    
    for acc in accounts_OSW.values():
        grand_total['debit_start'] += acc['debit_start']
        grand_total['credit_start'] += acc['credit_start']
        grand_total['debit_turnover'] += acc['debit_turnover']
        grand_total['credit_turnover'] += acc['credit_turnover']

        if acc['closing_balance'] >= 0:
            grand_total['closing_debit'] += acc['closing_balance']
        else:
            grand_total['closing_credit'] += abs(acc['closing_balance'])
            
            
    

    grand_total = {
        'debit_start': sum(a['debit_start'] for a in leaf_accounts),
        'credit_start': sum(a['credit_start'] for a in leaf_accounts),
        'debit_turnover': sum(a['debit_turnover'] for a in leaf_accounts),
        'credit_turnover': sum(a['credit_turnover'] for a in leaf_accounts),
        'closing_debit': sum(
            a['closing_balance'] for a in leaf_accounts if a['closing_balance'] >= 0
        ),
        'closing_credit': sum(
            abs(a['closing_balance']) for a in leaf_accounts if a['closing_balance'] < 0
        ),
    }
            
            
    ws[f"A{row}"] = "ИТОГО"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    ws[f"C{row}"] = grand_total['debit_start']
    ws[f"D{row}"] = grand_total['credit_start']
    ws[f"E{row}"] = grand_total['debit_turnover']
    ws[f"F{row}"] = grand_total['credit_turnover']
    ws[f"G{row}"] = grand_total['closing_debit']
    ws[f"H{row}"] = grand_total['closing_credit']

    for col in "ABCDEFGH":
        ws[f"{col}{row}"].font = CATEGORY_FONT
        ws[f"{col}{row}"].border = THIN_BORDER
        ws[f"{col}{row}"].fill = GRAY_FILL_1
        
    for col in "CDEFGH":
        ws[f"{col}{row}"].number_format = PRICE_FMT
        
    row += 1
    
    ws[f"A{row}"] = "САЛЬДО"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    saldo_start = grand_total['debit_start'] - grand_total['credit_start']
    saldo_turnover = grand_total['debit_turnover'] - grand_total['credit_turnover']
    saldo_end = grand_total['closing_debit'] - grand_total['closing_credit']
    if saldo_start > 0:
        ws[f"C{row}"] = saldo_start
    else:
        ws[f"D{row}"] = abs(saldo_start)
    if saldo_turnover > 0:
        ws[f"E{row}"] = saldo_turnover
    else:
        ws[f"F{row}"] = abs(saldo_turnover)
    if saldo_end > 0:    
        ws[f"G{row}"] = saldo_end
    else:
        ws[f"H{row}"] = abs(saldo_end)

    for col in "ABCDEFGH":
        ws[f"{col}{row}"].font = CATEGORY_FONT
        ws[f"{col}{row}"].border = THIN_BORDER
        ws[f"{col}{row}"].fill = GRAY_FILL_1
        
    for col in "CDEFGH":
        ws[f"{col}{row}"].number_format = PRICE_FMT
        
        
    # =========================================================
    # SUB ОТЧЁТЫ
    # =========================================================
    # ic| account_id: 19, account_number: '25'
    # ic| account_id: 20, account_number: '26'
    # ic| account_id: 21, account_number: '27'
    # ic| account_id: 22, account_number: '28'
    # ic| account_id: 12, account_number: '40.1'
    # ic| account_id: 13, account_number: '40.2'
    # ic| account_id: 15, account_number: '42.1'
    # ic| account_id: 6, account_number: '46.1'
    # ic| account_id: 7, account_number: '46.2'
    # ic| account_id: 9, account_number: '47.1'
    # ic| account_id: 10, account_number: '47.2'
    # ic| account_id: 16, account_number: '50'
    # ic| account_id: 18, account_number: '52'
    # ic| account_id: 1, account_number: '60'
    # ic| account_id: 2, account_number: '62'
    # ic| account_id: 3, account_number: '75'
    # ic| account_id: 4, account_number: '76'
    # ic| account_id: 17, account_number: '80'
    

    osw_detail_level_1 = {}
    partner_snap_map = {}
    if last_closing:
        partner_balance_snapshots = PartnerBalanceSnapshot.objects.filter(closing=last_closing).select_related("partner")
        partner_snap_map = {
            (s.partner.id): {
                    "balance_60_usd": s.balance_60_usd, 
                    "balance_62_tmt": s.balance_62_tmt, 
                    "balance_75_usd": s.balance_75_usd, 
                    "balance_76_tmt": s.balance_76_tmt,
                    
                    "balance_60_usd_credit": s.balance_60_usd_credit,
                    "balance_60_usd_debit": s.balance_60_usd_debit,
                    
                    "balance_62_tmt_credit": s.balance_62_tmt_credit,
                    "balance_62_tmt_debit": s.balance_62_tmt_debit,
                    
                    "balance_75_usd_credit": s.balance_75_usd_credit,
                    "balance_75_usd_debit": s.balance_75_usd_debit,
                    
                    "balance_76_tmt_credit": s.balance_76_tmt_credit,
                    "balance_76_tmt_debit": s.balance_76_tmt_debit,

                }
            for s in partner_balance_snapshots
        }
    
    for acc in leaf_accounts:
        account_id = acc["id"]
        account_number = acc["number"]
        

        if account_number.startswith("60") or account_number.startswith("62"):
            ic(account_id, account_number)
            partners = Partner.objects.all().select_related("agent")
            
            account_60_62 = {}
            for p in partners:
                partner = {
                    "id": p.id,
                    "name": p.name,
                    "type": p.type,
                }
                
                if p.agent:
                    agent = {
                        "id": p.agent.id,
                        "name": p.agent.name,
                    }
                else:
                    agent = {}
                    
                account_60_62[p.id] = {
                    "partner": partner,
                    "agent": agent,
                    "debit_start": Decimal("0.00"),
                    "credit_start": Decimal("0.00"),
                    "debit_turnover": Decimal("0.00"),
                    "credit_turnover": Decimal("0.00"),
                    "debit_end": Decimal("0.00"),
                    "credit_end": Decimal("0.00"),
                }
            
            if last_closing:
                for p_id, value in partner_snap_map.items():
                    if p_id not in account_60_62:
                        continue
                    if account_number == "60":
                        account_60_62[p_id]["debit_start"] += value["balance_60_usd_debit"]
                        account_60_62[p_id]["credit_start"] += value["balance_60_usd_credit"]
                    else:
                        account_60_62[p_id]["debit_start"] += value["balance_62_tmt_debit"]
                        account_60_62[p_id]["credit_start"] += value["balance_62_tmt_credit"]
                    
            else:
                entries_start = (
                    Entry.objects
                    .filter(
                        account_id=account_id,
                        transaction__date__lt=close_date_format
                    )
                    .select_related(
                        'transaction',
                        'partner__agent',
                        'product',
                        'warehouse',
                        'transaction__invoice',
                    )
                    .order_by('transaction__date', 'id')
                )
                
                for e in entries_start:
                    if not e.partner:
                        continue
                    partner_id = e.partner.id
                    
                    
                    account_60_62[partner_id]["debit_start"] += money(e.debit)
                    account_60_62[partner_id]["credit_start"] += money(e.credit)
                    
            entries_turnover = (
                Entry.objects
                .filter(
                    account_id=account_id,
                    transaction__date__gte=day_start,
                    transaction__date__lt=day_end,
                )
                .select_related(
                    'transaction',
                    'partner__agent',
                    'product',
                    'warehouse',
                    'transaction__invoice',
                )
                .order_by('transaction__date', 'id')
            )
            
            for e in entries_turnover:
                    if not e.partner:
                        continue
                    partner_id = e.partner.id
                    
                    account_60_62[partner_id]["debit_turnover"] += money(e.debit)
                    account_60_62[partner_id]["credit_turnover"] += money(e.credit)
            
            
            
            total_credit_start = Decimal("0.00")        
            total_debit_start = Decimal("0.00")        
            total_credit_turnover = Decimal("0.00")        
            total_debit_turnover = Decimal("0.00")        
            total_credit_end = Decimal("0.00")        
            total_debit_end = Decimal("0.00")
                    
            for p_id, row in account_60_62.items():
                debit_end_raw = row["debit_start"] + row["debit_turnover"]
                credit_end_raw = row["credit_start"] + row["credit_turnover"]

                saldo = debit_end_raw - credit_end_raw

                if saldo >= 0:
                    row["debit_end"] = saldo
                    row["credit_end"] = Decimal("0.00")
                else:
                    row["debit_end"] = Decimal("0.00")
                    row["credit_end"] = -saldo
                    
                # 👉 ИТОГИ СЧИТАЕМ ЗДЕСЬ
                total_debit_start += row["debit_start"]
                total_credit_start += row["credit_start"]

                total_debit_turnover += row["debit_turnover"]
                total_credit_turnover += row["credit_turnover"]

                total_debit_end += row["debit_end"]
                total_credit_end += row["credit_end"]
                    
            grand_total_60_62 =  {
                "total_credit_start": total_credit_start,
                "total_debit_start": total_debit_start,
                "total_credit_turnover": total_credit_turnover,
                "total_debit_turnover": total_debit_turnover,
                "total_credit_end": total_credit_end,
                "total_debit_end": total_debit_end,
            }
            
            
                    
            ws_detail = detail_sheets[account_id]
            ws_detail.freeze_panes = "A7"
            ws_detail["A5"] = "№"
            ws_detail.column_dimensions["A"].width = 5
            ws_detail["B5"] = "Agent"
            ws_detail.column_dimensions["B"].width = 20
            ws_detail["C5"] = "Субконто"
            ws_detail.column_dimensions["C"].width = 45
            
            ws_detail.merge_cells("D5:E5")
            ws_detail["D5"] = "Сальдо на начало"

            
            ws_detail.merge_cells("F5:G5")
            ws_detail["F5"] = "Обороты за период"
         
            
            ws_detail.merge_cells("H5:I5")
            ws_detail["H5"] = "Сальдо на конец"
       
            
            
            for i in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", 
                      "A6", "B6", "C6", "D6", "E6", "F6", "G6", "H6", "I6"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = CENTER_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
                
                
            for i in ["D", "E", "F", "G", "H", "I"]:
                ws_detail.column_dimensions[i].width = 15
       
                
     
            ws_detail["D6"] = "Дебит"
            ws_detail["E6"] = "Кредит"
            
            ws_detail["F6"] = "Дебит"
            ws_detail["G6"] = "Кредит"
            
            ws_detail["H6"] = "Дебит"
            ws_detail["I6"] = "Кредит"
            
            
            row = 7
            count = 1
            for p_id, v in account_60_62.items():
                
                ws_detail[f"D{row}"].number_format = PRICE_FMT
                ws_detail[f"E{row}"].number_format = PRICE_FMT
                ws_detail[f"F{row}"].number_format = PRICE_FMT
                ws_detail[f"G{row}"].number_format = PRICE_FMT
                ws_detail[f"H{row}"].number_format = PRICE_FMT
                ws_detail[f"I{row}"].number_format = PRICE_FMT
                
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                    ws_detail[i].border = THIN_BORDER
                
        
        
                ws_detail[f"A{row}"] = count
                ws_detail[f"B{row}"] = v["agent"]["name"] if v["agent"] else ""
                ws_detail[f"C{row}"] = v["partner"]["name"] if v["partner"] else ""
                
                ws_detail[f"D{row}"] = v["debit_start"]
                ws_detail[f"E{row}"] = v["credit_start"]
                
                ws_detail[f"F{row}"] = v["debit_turnover"]
                ws_detail[f"G{row}"] = v["credit_turnover"]
                
                ws_detail[f"H{row}"] = v["debit_end"]
                ws_detail[f"I{row}"] = v["credit_end"]
                
                count += 1
                row += 1
                
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = RIGHT_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
              
            
            ws_detail.merge_cells(f"A{row}:C{row}")
            ws_detail[f"A{row}"].alignment = LEFT_ALIGN
            ws_detail[f"A{row}"] = "Итого развернутое:"
            
            ws_detail[f"D{row}"].number_format = PRICE_FMT
            ws_detail[f"E{row}"].number_format = PRICE_FMT
            ws_detail[f"F{row}"].number_format = PRICE_FMT
            ws_detail[f"G{row}"].number_format = PRICE_FMT
            ws_detail[f"H{row}"].number_format = PRICE_FMT
            ws_detail[f"I{row}"].number_format = PRICE_FMT
            
            ws_detail[f"D{row}"] = grand_total_60_62["total_debit_start"]
            ws_detail[f"E{row}"] = grand_total_60_62["total_credit_start"]
            
            ws_detail[f"F{row}"] = grand_total_60_62["total_debit_turnover"]
            ws_detail[f"G{row}"] = grand_total_60_62["total_credit_turnover"]
            
            ws_detail[f"H{row}"] = grand_total_60_62["total_debit_end"]
            ws_detail[f"I{row}"] = grand_total_60_62["total_credit_end"]
            
            row += 1
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = RIGHT_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
              
            
            ws_detail.merge_cells(f"A{row}:C{row}")
            ws_detail[f"A{row}"].alignment = LEFT_ALIGN
            ws_detail[f"A{row}"] = "Всего:"
            
            ws_detail[f"D{row}"].number_format = PRICE_FMT
            ws_detail[f"E{row}"].number_format = PRICE_FMT
            ws_detail[f"F{row}"].number_format = PRICE_FMT
            ws_detail[f"G{row}"].number_format = PRICE_FMT
            ws_detail[f"H{row}"].number_format = PRICE_FMT
            ws_detail[f"I{row}"].number_format = PRICE_FMT
            
            saldo_start = grand_total_60_62["total_debit_start"] - grand_total_60_62["total_credit_start"]
            ws_detail[f"D{row}"] = saldo_start if saldo_start > 0 else 0
            ws_detail[f"E{row}"] = abs(saldo_start) if saldo_start < 0 else 0
            
            saldo_turnover = grand_total_60_62["total_debit_turnover"] - grand_total_60_62["total_credit_turnover"]
            ws_detail[f"F{row}"] = saldo_turnover if saldo_turnover > 0 else 0
            ws_detail[f"G{row}"] = abs(saldo_turnover) if saldo_turnover < 0 else 0
            
            saldo_end = grand_total_60_62["total_debit_end"] - grand_total_60_62["total_credit_end"]
            ws_detail[f"H{row}"] = saldo_end if saldo_end > 0 else 0
            ws_detail[f"I{row}"] = abs(saldo_end) if saldo_end < 0 else 0
            
 
        
        elif account_number.startswith("40") or account_number.startswith("42"):
            # w_acc = WarehouseAccount.objects.get(account_id=account_id)
            account_40_42 = {}
            w_acc = (
                WarehouseAccount.objects
                .select_related("warehouse")
                .get(account_id=account_id)
            )

            products = Product.objects.filter(
                warehouse_products__warehouse_id=w_acc.warehouse_id
            ).distinct()

            
            unit_map = get_unit_map()
            
            for p in products:
                if not p.category:
                    continue
                unit, cf = get_unit_and_cf(unit_map, p)
                category_id = p.category.id
                category_name = p.category.name
                # if p.id == 601:
                #     ic(p.name, unit, cf)
                if category_id not in account_40_42:
                    account_40_42[category_id] = {
                        "category": {
                            "id": category_id,
                            "name": category_name,
                        },
                        "totals": {
                            "start_qty": Decimal("0.00"),
                            "start_price": Decimal("0.00"),

                            "prihod_qty": Decimal("0.00"),
                            "prihod_price": Decimal("0.00"),

                            "wozwrat_qty": Decimal("0.00"),
                            "wozwrat_price": Decimal("0.00"),

                            "rashod_qty": Decimal("0.00"),
                            "rashod_price": Decimal("0.00"),

                            "end_qty": Decimal("0.00"),
                            "end_price": Decimal("0.00"),
                        },
                        "products": {}
                    }

                account_40_42[category_id]["products"][p.id] = {
                    "product": {
                        "id": p.id,
                        "name": p.name,
                        "unit": unit,
                        "cf": Decimal(cf),
                        "wholsale_price": p.wholesale_price,
                    },

                    "start_qty": Decimal("0.00"),
                    "start_price": Decimal("0.00"),

                    "prihod_qty": Decimal("0.00"),
                    "prihod_price": Decimal("0.00"),

                    "wozwrat_qty": Decimal("0.00"),
                    "wozwrat_price": Decimal("0.00"),

                    "rashod_qty": Decimal("0.00"),
                    "rashod_price": Decimal("0.00"),

                    "end_qty": Decimal("0.00"),
                    "end_price": Decimal("0.00"),
                }
                
            start_items = InvoiceItem.objects.filter(
                invoice__entry_created_at_handle__lt=close_date_format,
                invoice__canceled_at__isnull=True
            ).filter(
                Q(invoice__warehouse_id=w_acc.warehouse_id) |
                Q(invoice__warehouse2_id=w_acc.warehouse_id)
            ).select_related(
                "product", "product__category", "product__base_unit", "invoice"
            )
            
            grand_total = {
                "start_qty": Decimal("0.00"),
                "start_price": Decimal("0.00"),

                "prihod_qty": Decimal("0.00"),
                "prihod_price": Decimal("0.00"),

                "wozwrat_qty": Decimal("0.00"),
                "wozwrat_price": Decimal("0.00"),

                "rashod_qty": Decimal("0.00"),
                "rashod_price": Decimal("0.00"),

                "end_qty": Decimal("0.00"),
                "end_price": Decimal("0.00"),
            }
            
            for item in start_items:
                p = item.product
                inv = item.invoice
                
                # Товар уже инициализирован выше
                if not p.category:
                    continue
                if (p.category.id not in account_40_42 or p.id not in account_40_42[p.category.id]["products"]):
                    continue
                cat_totals = account_40_42[p.category.id]["totals"]
                prod = account_40_42[p.category.id]["products"][p.id]
                cf = prod["product"]["cf"]
                qty = Decimal(item.selected_quantity) / cf
                calculated_price = qty * Decimal(p.wholesale_price)
                if p.name == 'UYP-231, Srup 6.3*70 "PAiiA" (2kg/5guty)':
                    ic(qty)
                    ic(item.wholesale_price)
                    ic(calculated_price)
                # Проверяем принадлежность к выбранным складам
                if inv.wozwrat_or_prihod == "prihod":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["start_qty"] += qty
                        prod["start_price"] += calculated_price
                        
                        cat_totals["start_qty"] += qty
                        cat_totals["start_price"] += calculated_price
                        
                        grand_total["start_qty"] += qty
                        grand_total["start_price"] += calculated_price

                elif inv.wozwrat_or_prihod == "rashod":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["start_qty"] -= qty
                        prod["start_price"] -= calculated_price
                        
                        cat_totals["start_qty"] -= qty
                        cat_totals["start_price"] -= calculated_price
                        
                        grand_total["start_qty"] -= qty
                        grand_total["start_price"] -= calculated_price

                elif inv.wozwrat_or_prihod == "wozwrat":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["start_qty"] += qty
                        prod["start_price"] += calculated_price
                        
                        cat_totals["start_qty"] += qty
                        cat_totals["start_price"] += calculated_price
                        
                        grand_total["start_qty"] += qty
                        grand_total["start_price"] += calculated_price
                    
                elif inv.wozwrat_or_prihod == "transfer":
                    # если со склада (и склад в выбранных)
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["start_qty"] -= qty
                        prod["start_price"] -= calculated_price
                        
                        cat_totals["start_qty"] -= qty
                        cat_totals["start_price"] -= calculated_price
                        
                        grand_total["start_qty"] -= qty
                        grand_total["start_price"] -= calculated_price
                        
                    # если на склад (и склад в выбранных)
                    elif inv.warehouse2_id == w_acc.warehouse_id:
                        prod["start_qty"] += qty
                        prod["start_price"] += calculated_price
                        
                        cat_totals["start_qty"] += qty
                        cat_totals["start_price"] += calculated_price
                        
                        grand_total["start_qty"] += qty
                        grand_total["start_price"] += calculated_price
                        
            turnover_items = InvoiceItem.objects.filter(
                invoice__entry_created_at_handle__gte=day_start,
                invoice__entry_created_at_handle__lt=day_end,
                invoice__canceled_at__isnull=True
            ).filter(
                Q(invoice__warehouse_id=w_acc.warehouse_id) |
                Q(invoice__warehouse2_id=w_acc.warehouse_id)
            ).select_related(
                "product", "product__category", "product__base_unit", "invoice"
            ).order_by(
                "product__category__name", "product__name"
            )
            
            
            for item in turnover_items:
                p = item.product
                inv = item.invoice
                
                # Товар уже инициализирован выше
                if not p.category:
                    continue
                if (p.category.id not in account_40_42 or p.id not in account_40_42[p.category.id]["products"]):
                    continue
                cat_totals = account_40_42[p.category.id]["totals"]
                prod = account_40_42[p.category.id]["products"][p.id]
                cf = prod["product"]["cf"]
                qty = Decimal(item.selected_quantity) / cf
                calculated_price = qty * Decimal(item.selected_price)
                # Проверяем принадлежность к выбранным складам
                if inv.wozwrat_or_prihod == "prihod":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["prihod_qty"] += qty
                        prod["prihod_price"] += calculated_price
                        
                        cat_totals["prihod_qty"] += qty
                        cat_totals["prihod_price"] += calculated_price
                        
                        grand_total["prihod_qty"] += qty
                        grand_total["prihod_price"] += calculated_price

                elif inv.wozwrat_or_prihod == "rashod":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["rashod_qty"] += qty
                        prod["rashod_price"] += calculated_price
                        
                        cat_totals["rashod_qty"] += qty
                        cat_totals["rashod_price"] += calculated_price
                        
                        grand_total["rashod_qty"] += qty
                        grand_total["rashod_price"] += calculated_price

                elif inv.wozwrat_or_prihod == "wozwrat":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["wozwrat_qty"] += qty
                        prod["wozwrat_price"] += calculated_price
                        
                        cat_totals["wozwrat_qty"] += qty
                        cat_totals["wozwrat_price"] += calculated_price
                        
                        grand_total["wozwrat_qty"] += qty
                        grand_total["wozwrat_price"] += calculated_price
                    
                elif inv.wozwrat_or_prihod == "transfer":
                    # если со склада (и склад в выбранных)
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["rashod_qty"] += qty
                        prod["rashod_price"] += calculated_price
                        
                        cat_totals["rashod_qty"] += qty
                        cat_totals["rashod_price"] += calculated_price
                        
                        grand_total["rashod_qty"] += qty
                        grand_total["rashod_price"] += calculated_price
                        
                    # если на склад (и склад в выбранных)
                    elif inv.warehouse2_id == w_acc.warehouse_id:
                        prod["prihod_qty"] += qty
                        prod["prihod_price"] += calculated_price
                        
                        cat_totals["prihod_qty"] += qty
                        cat_totals["prihod_price"] += calculated_price
                        
                        grand_total["prihod_qty"] += qty
                        grand_total["prihod_price"] += calculated_price
                        
            for cat in account_40_42.values():
                for prod in cat["products"].values():
                    prod["end_qty"] = (
                        prod["start_qty"]
                        + prod["prihod_qty"]
                        - prod["rashod_qty"]
                        + prod["wozwrat_qty"]
                    )

                    prod["end_price"] = (
                        prod["start_price"]
                        + prod["prihod_price"]
                        - prod["rashod_price"]
                        + prod["wozwrat_price"]
                    )
                    
            for cat in account_40_42.values():
                totals = cat["totals"]
                totals["end_qty"] = (
                    totals["start_qty"]
                    + totals["prihod_qty"]
                    - totals["rashod_qty"]
                    + totals["wozwrat_qty"]
                )
                totals["end_price"] = (
                    totals["start_price"]
                    + totals["prihod_price"]
                    - totals["rashod_price"]
                    + totals["wozwrat_price"]
                )

            grand_total["end_qty"] = (
                grand_total["start_qty"]
                + grand_total["prihod_qty"]
                - grand_total["rashod_qty"]
                + grand_total["wozwrat_qty"]
            )
            grand_total["end_price"] = (
                grand_total["start_price"]
                + grand_total["prihod_price"]
                - grand_total["rashod_price"]
                + grand_total["wozwrat_price"]
            )

            
            ws_detail = detail_sheets[account_id]
            ws_detail.freeze_panes = "A7"
            
      
            ws_detail["A5"] = "№"
            ws_detail.column_dimensions["A"].width = 8
            ws_detail["B5"] = "Наименование товара"
            ws_detail.column_dimensions["B"].width = 55
            ws_detail["C5"] = "Ед."
            ws_detail.column_dimensions["C"].width = 8
            
            ws_detail["D5"] = "Цена"
            ws_detail.column_dimensions["D"].width = 8
            ws_detail[f"D5"].number_format = PRICE_FMT
            
            ws_detail.merge_cells("E5:F5")
            ws_detail["E5"] = "Остаток на начало"
   

            
            ws_detail.merge_cells("G5:H5")
            ws_detail["G5"] = "Приход"
         
            
            ws_detail.merge_cells("I5:J5")
            ws_detail["I5"] = "Возврат"
            
            ws_detail.merge_cells("K5:L5")
            ws_detail["K5"] = "Расход"
            
            ws_detail.merge_cells("M5:N5")
            ws_detail["M5"] = "Остаток на конец"
       
            
            
            for i in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", 
                      "A6", "B6", "C6", "D6", "E6", "F6", "G6", "H6", "I6", "J6", "K6", "L6", "M6", "N6"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = CENTER_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
                
                
            for i in ["E", "F",  "G", "H", "I", "J", "K", "L", "M", "N"]:
                ws_detail.column_dimensions[i].width = 11
       
                
     
            ws_detail["E6"] = "Кол-во"
            ws_detail["F6"] = "Всего"
            
            ws_detail["G6"] = "Кол-во"
            ws_detail["H6"] = "Всего"
            
            ws_detail["I6"] = "Кол-во"
            ws_detail["J6"] = "Всего"
            
            ws_detail["K6"] = "Кол-во"
            ws_detail["L6"] = "Всего"
            
            ws_detail["M6"] = "Кол-во"
            ws_detail["N6"] = "Всего"
            
            
            
     
            
            
            row = 7
            
            
                
            count = 1
            # for cat_id, values in account_40_42.items():
            for cat_id, values in sorted(
                account_40_42.items(),
                key=lambda item: item[1]["category"]["name"].lower()
            ):
                cat_name = values["category"]["name"]
                totals = values["totals"]
                ws_detail.merge_cells(f"A{row}:N{row}")
                ws_detail[f"A{row}"].fill = GREEN_FILL_0
                ws_detail[f"A{row}"].font = TOTAL_FONT
                ws_detail[f"A{row}"].alignment = LEFT_ALIGN
                ws_detail[f"A{row}"] = cat_name
                
                
                
                for col in "FHJLN":
                    ws_detail[f"{col}{row}"].number_format = PRICE_FMT
                    
                for col in "EGIKM":
                    ws_detail[f"{col}{row}"].number_format = QTY_FMT
                
                row += 1
                
                for products, value in values["products"].items():
                    product = value["product"]
                    unit = product["unit"]
                    wholesale_price = product["wholsale_price"]
                    
                    start_qty = value["start_qty"]
                    start_price = value["start_price"]
                    
                    prihod_qty = value["prihod_qty"]
                    prihod_price = value["prihod_price"]
                    
                    
                    wozwrat_qty = value["wozwrat_qty"]
                    wozwrat_price = value["wozwrat_price"]
                    
                    rashod_qty = value["rashod_qty"]
                    rashod_price = value["rashod_price"]
                    
                    end_qty = value["end_qty"]
                    end_price = value["end_price"] 
                    
                    ws_detail[f"G{row}"].font = GREEN_FONT
                    ws_detail[f"H{row}"].font = GREEN_FONT
                    ws_detail[f"I{row}"].font = RED_FONT
                    ws_detail[f"J{row}"].font = RED_FONT
                    ws_detail[f"K{row}"].font = BLUE_FONT
                    ws_detail[f"L{row}"].font = BLUE_FONT
                    
                    for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                        ws_detail[i].border = THIN_BORDER
                        
                    for col in "FHJLN":
                        ws_detail[f"{col}{row}"].number_format = PRICE_FMT
                        
                    for col in "EGIKM":
                        ws_detail[f"{col}{row}"].number_format = QTY_FMT
                
                    ws_detail[f"A{row}"] = count
                    ws_detail[f"B{row}"] = product["name"]
                    ws_detail[f"C{row}"] = unit
                    ws_detail[f"D{row}"] = wholesale_price
                    
                    ws_detail[f"E{row}"] = start_qty
                    ws_detail[f"F{row}"] = start_price
                    
                    ws_detail[f"G{row}"] = prihod_qty
                    ws_detail[f"H{row}"] = prihod_price
                    
                    ws_detail[f"I{row}"] = wozwrat_qty
                    ws_detail[f"J{row}"] = wozwrat_price
                    
                    ws_detail[f"K{row}"] = rashod_qty
                    ws_detail[f"L{row}"] = rashod_price
                    
                    ws_detail[f"M{row}"] = end_qty
                    ws_detail[f"N{row}"] = end_price
                    
                
                    count += 1
                    row += 1
            
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                        ws_detail[i].border = THIN_BORDER
                        
                ws_detail.merge_cells(f"A{row}:D{row}")
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                        ws_detail[i].fill = GRAY_FILL_0
                ws_detail[f"A{row}"].font = TOTAL_FONT
                ws_detail[f"A{row}"].alignment = RIGHT_ALIGN
                ws_detail[f"A{row}"] = f"Итого по категории: {cat_name}"
                
              
                ws_detail[f"E{row}"] = totals["start_qty"]
                ws_detail[f"F{row}"] = totals["start_price"]
                ws_detail[f"G{row}"] = totals["prihod_qty"]
                ws_detail[f"H{row}"] = totals["prihod_price"]
                ws_detail[f"I{row}"] = totals["wozwrat_qty"]
                ws_detail[f"J{row}"] = totals["wozwrat_price"]
                ws_detail[f"K{row}"] = totals["rashod_qty"]
                ws_detail[f"L{row}"] = totals["rashod_price"]
                ws_detail[f"M{row}"] = totals["end_qty"]
                ws_detail[f"N{row}"] = totals["end_price"]
                                
                row += 1
                
            ws_detail[f"E{row}"] = grand_total["start_qty"]
            ws_detail[f"F{row}"] = grand_total["start_price"]
            ws_detail[f"G{row}"] = grand_total["prihod_qty"]
            ws_detail[f"H{row}"] = grand_total["prihod_price"]
            ws_detail[f"I{row}"] = grand_total["wozwrat_qty"]
            ws_detail[f"J{row}"] = grand_total["wozwrat_price"]
            ws_detail[f"K{row}"] = grand_total["rashod_qty"]
            ws_detail[f"L{row}"] = grand_total["rashod_price"]
            ws_detail[f"M{row}"] = grand_total["end_qty"]
            ws_detail[f"N{row}"] = grand_total["end_price"]
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws_detail[i].border = THIN_BORDER
                
            for col in "FHJLN":
                ws_detail[f"{col}{row}"].number_format = PRICE_FMT
                
            for col in "EGIKM":
                ws_detail[f"{col}{row}"].number_format = QTY_FMT
                        
            ws_detail.merge_cells(f"A{row}:D{row}")
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws_detail[i].fill = GREEN_FILL_1
       
            ws_detail[f"A{row}"].font = TOTAL_FONT
            ws_detail[f"A{row}"].alignment = RIGHT_ALIGN
            ws_detail[f"A{row}"] = "ВСЕГО:"
        
            
       
     
                    
           
        else:
            # poka drugie account continue potom dodelaem
            continue
    # for p_id, value in account_60_62.items():
    #     if p_id == 79:
    #         ic(value)
    
    # # ic(partner_balance_snapshots)
            
    # for p_id, value in partner_snap_map.items():
    #     if p_id == 79:
    #         ic(value)
            
    # entries = Entry.objects.filter(
    #     partner_id=79,
    #     transaction__date__lt=day_start,
    #     account__number="60"
    # )
    # debit = entries.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
    # credit = entries.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
    # ic(debit)
    # ic(credit)
    # ic(debit - credit)
    
            
        
    excel_buffer = BytesIO()
    wb_OSW.save(excel_buffer)
    excel_buffer.seek(0)
    OSW_content = excel_buffer.read()
        
        

    
    
    ####### end card prowodok
    #############################################################################################################################################################
    #############################################################################################################################################################
    
    


    try:
        with transaction.atomic():
            # 1/0
            

            day_closing = DayClosing.objects.create(date=close_date_format)

            # обновляем статус
            day_closing.closed_at = timezone.now()
            day_closing.closed_by = user
            day_closing.note = reason
            day_closing.save()

            # логируем
            DayClosingLog.objects.create(
                day_closing=day_closing,
                action="close",
                performed_by=user,
                reason=reason
            )

            # ФУНКЦИИ ДЛЯ ВЫЧИСЛЕНИЯ БАЛАНСОВ ПО КАЖДОМУ СЧЕТУ
            def calculate_balance_by_account(partner, target_date, account_number):
                entries = Entry.objects.filter(
                    partner=partner,
                    transaction__date__lte=target_date,
                    account__number=account_number
                )
                debit = entries.aggregate(total=Sum('debit'))['total'] or Decimal('0.00')
                credit = entries.aggregate(total=Sum('credit'))['total'] or Decimal('0.00')
                balance = debit - credit
                return balance, debit, credit

            # снимки балансов партнёров ОТДЕЛЬНО ПО КАЖДОМУ СЧЕТУ
            for partner in Partner.objects.all():
                usd_60 = calculate_balance_by_account(partner, close_date, '60')
                balance_60_usd = usd_60[0]
                balance_60_usd_debit = usd_60[1]
                balance_60_usd_credit = usd_60[2]
                
                tmt_62 = calculate_balance_by_account(partner, close_date, '62')
                balance_62_tmt = tmt_62[0]
                balance_62_tmt_debit = tmt_62[1]
                balance_62_tmt_credit = tmt_62[2]
                
                usd_75 = calculate_balance_by_account(partner, close_date, '75')
                balance_75_usd = usd_75[0]
                balance_75_usd_debit = usd_75[1]
                balance_75_usd_credit = usd_75[2]
                
                tmt_76 = calculate_balance_by_account(partner, close_date, '76')
                balance_76_tmt = tmt_76[0]
                balance_76_tmt_debit = tmt_76[1]
                balance_76_tmt_credit = tmt_76[2]
                
                
                # Итоговые балансы по валютам (для обратной совместимости)
                total_usd = balance_60_usd + balance_75_usd
                total_tmt = balance_62_tmt + balance_76_tmt
                
                PartnerBalanceSnapshot.objects.create(
                    closing=day_closing,
                    partner=partner,
                    balance_60_usd=balance_60_usd,
                    balance_62_tmt=balance_62_tmt,
                    balance_75_usd=balance_75_usd,
                    balance_76_tmt=balance_76_tmt,
                    balance_usd=total_usd,
                    balance_tmt=total_tmt,
                    balance=Decimal('0.000'),
                    
                    balance_60_usd_credit=balance_60_usd_credit,
                    balance_60_usd_debit=balance_60_usd_debit,
                    
                    balance_62_tmt_credit=balance_62_tmt_credit,
                    balance_62_tmt_debit=balance_62_tmt_debit,
                    
                    balance_75_usd_credit=balance_75_usd_credit,
                    balance_75_usd_debit=balance_75_usd_debit,
                    
                    balance_76_tmt_credit=balance_76_tmt_credit,
                    balance_76_tmt_debit=balance_76_tmt_debit,
                )

            # снимки складов
            for wp in WarehouseProduct.objects.all():
                StockSnapshot.objects.create(
                    closing=day_closing,
                    warehouse=wp.warehouse,
                    product=wp.product,
                    purchase_price=wp.product.purchase_price,
                    retail_price=wp.product.retail_price,
                    wholesale_price=wp.product.wholesale_price,
                    discount_price=wp.product.discount_price,
                    firma_price=wp.product.firma_price,
                    # quantity=wp.quantity
                    quantity = turnover_product[wp.warehouse_id][wp.product_id]["end_qty"]
                )
                
            # oborot product excel
            report_product_oborot, _ = DayReport.objects.get_or_create(
                date=close_date_format,
                report_type=f"OBOROT_TOWAR",
                defaults={
                    "created_by": user,
                    "comment": reason
                }
            )

            filename = f"OBOROT_TOWAR_{convert_close_date}.xlsx"
            report_product_oborot.file.save(
                filename,
                ContentFile(product_oborot_content),
                save=True
            )
            
            # skidka excel
            report_skidka, _ = DayReport.objects.get_or_create(
                date=close_date_format,
                report_type=f"SKIDKA",
                defaults={
                    "created_by": user,
                    "comment": reason
                }
            )

            filename = f"SKIDKA_{convert_close_date}.xlsx"

            report_skidka.file.save(
                filename,
                ContentFile(skidka_content),
                save=True
            )
            
            # oborot product detail
            report_product_oborot_detail, _ = DayReport.objects.get_or_create(
                date=close_date_format,
                report_type="OBOROT_TOWAR_DETAIL",
                defaults={
                    "created_by": user,
                    "comment": reason
                }
            )

            filename = f"OBOROT_TOWAR_DETAIL_{convert_close_date}.xlsx"
            report_product_oborot_detail.file.save(
                filename,
                ContentFile(product_oborot_detail_content),
                save=True
            )
            
            
            # OSW
            report_OSW, _ = DayReport.objects.get_or_create(
                date=close_date_format,
                report_type="OSW",
                defaults={
                    "created_by": user,
                    "comment": reason
                }
            )

            filename = f"OSW_{convert_close_date}.xlsx"
            report_OSW.file.save(
                filename,
                ContentFile(OSW_content),
                save=True
            )

    except Exception as e:
        return Response({"success": False, "error": str(e)})
    
    return Response({"success": True, "message": "day success is closed", "date": close_date})


# close day
########################################################################################################################################################################## END

@csrf_exempt
def universal_entries(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST request required"}, status=400)
    
    try:
        data = json.loads(request.body)
        date = parse_date(data.get("date"))
        debit_number = data.get("debit")
        credit_number = data.get("credit")
        amount = Decimal(data.get("amount"))
        description = data.get("description", "")
        
        # Найти счета
        debit_account = Account.objects.get(number=debit_number)
        credit_account = Account.objects.get(number=credit_number)
        
        try:
            with transaction.atomic():
        
                # Создать транзакцию
                transaction_obj = Transaction.objects.create(
                    date=date,
                    description=description,
                    created_by=request.user
                )
                
                # Создать проводки
                Entry.objects.create(
                    transaction=transaction_obj,
                    account=debit_account,
                    debit=amount,
                    credit=Decimal('0.00')
                )
                Entry.objects.create(
                    transaction=transaction_obj,
                    account=credit_account,
                    debit=Decimal('0.00'),
                    credit=amount
                )
                
                return JsonResponse({"success": True, "transaction_id": transaction_obj.id})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    
    except Account.DoesNotExist:
        return JsonResponse({"error": "Дебет или кредит счет не найден"}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
    
    
    
@require_GET
def get_transaction_journal(request):
    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")

    if not date_from or not date_to:
        return JsonResponse({"error": "Не указаны даты"}, status=400)

    allowed_accounts = ['50', '52', '60', '62', '75', '76']

    transactions = Transaction.objects.filter(
        date__range=[date_from, date_to],
        partner__isnull=False
    ).prefetch_related('entries', 'partner')

    journal = []

    for tr in transactions:
        entries = tr.entries.all()
        account_numbers = [e.account.number for e in entries]

        # Проверяем, что хотя бы один счет в списке allowed_accounts
        if not any(acc in allowed_accounts for acc in account_numbers):
            continue

        # Проверяем, что все счета транзакции входят в allowed_accounts
        if not all(acc in allowed_accounts for acc in account_numbers):
            continue

        # Фильтруем проводки только по нужным счетам (для дебета/кредита)
        filtered_entries = [e for e in entries if e.account.number in allowed_accounts]

        debit_accounts = ', '.join(str(e.account.number) for e in filtered_entries if e.debit > 0)
        credit_accounts = ', '.join(str(e.account.number) for e in filtered_entries if e.credit > 0)
        amount = sum(e.debit for e in filtered_entries)

        journal.append({
            "date": tr.date.strftime("%d.%m.%Y"),
            "operation": f"{tr.description}, {tr.partner.name if tr.partner else ''}",
            "debit": debit_accounts,
            "credit": credit_accounts,
            "amount": float(amount)
        })

    total_amount = sum(item['amount'] for item in journal)
    period_str = f"Period from {date_from} to {date_to}"

    return JsonResponse({
        "journal": journal,
        "period": period_str,
        "totalAmount": total_amount
    }, safe=False)