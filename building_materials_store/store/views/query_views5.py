from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime
from rest_framework.decorators import api_view, permission_classes
from ..models import Partner, Invoice, Transaction, WarehouseProduct, Entry, Account, Warehouse, FreeItemForInvoiceItem, UnitForInvoiceItem, Product, UnitOfMeasurement, Employee, ProductUnit, ProductImage, InvoiceItem, WarehouseAccount
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.views.decorators.csrf import csrf_exempt
from collections import defaultdict
from django.db.models import Sum, F, Q
from django.http import JsonResponse
from icecream import ic
from django.db import transaction as db_transaction
import pandas as pd
from datetime import date
from ..my_func.get_unit_map import get_unit_map 
from ..my_func.get_unit_and_cf import get_unit_and_cf 
from ..my_func.date_convert_for_excel import format_date_ru 

from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


GRAY_FILL_0 = PatternFill(
    fill_type="solid",
    fgColor="F5F5F5"
)

GRAY_FILL_1 = PatternFill(
    fill_type="solid",
    fgColor="EDEDED"
)
    
CENTER_ALIGN = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

# Категория
CATEGORY_FILL = PatternFill(
    fill_type="solid",
    fgColor="E7E6E6"
)
CATEGORY_FONT = Font(bold=True)

LEFT_ALIGN = Alignment(vertical="center", horizontal="left")
RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

    # Границы
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Форматы чисел
PRICE_FMT = '#,##0.00'
QTY_FMT = '#,##0.###'

RED_FONT = Font(color="FF0000")
GREEN_FONT = Font(color="006400")
BLUE_FONT = Font(color="0000FF")

GREEN_FILL_0 = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9"
    )


GREEN_FILL_1 = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)
    

################################################################################################################################################################
################################################################################################################################################################
# BuhOborotTowarow (3 warianta moy, chatGpt i DeepSeek) START



# rabochiyy no claude ai dal isprawlwnnuyu wersiyu
# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def BuhOborotTowarow(request):
#     dateFrom = request.GET.get("dateFrom")
#     dateTo = request.GET.get("dateTo")
    
#     # Поддержка нового параметра warehouses и старого warehouse
#     warehouses_param = request.GET.get("warehouses", "")
#     warehouse_param = request.GET.get("warehouse", "")
    
#     categories = request.GET.get("categories")
#     products = request.GET.get("products")
#     emptyTurnovers = True if request.GET.get("emptyTurnovers") == "1" else False
    
#     # Обработка параметра складов
#     warehouse_ids = []
    
#     if warehouses_param:
#         try:
#             warehouse_ids = [int(id.strip()) for id in warehouses_param.split(",") if id.strip().isdigit()]
#         except ValueError:
#             return JsonResponse({"status": "error", "message": "Invalid warehouse IDs"}, status=400)
#     elif warehouse_param and warehouse_param.isdigit():
#         warehouse_ids = [int(warehouse_param)]
    
#     if not warehouse_ids:
#         return JsonResponse({"status": "error", "message": "Warehouse(s) not selected"}, status=400)
    
#     # ic(warehouse_ids)
#     product_units = (
#         ProductUnit.objects
#         .filter(is_default_for_sale=True)
#         .select_related("unit")
#     )
    
#     unit_map = {
#         pu.product_id: pu
#         for pu in product_units
#     }
    
#     if categories:
#         category_ids = [int(c) for c in categories.split(",") if c.isdigit()]
#     else:
#         category_ids = []
        
#     if products:
#         product_ids = [int(c) for c in products.split(",") if c.isdigit()]
#     else:
#         product_ids = []
        

#     # Баланс по товарам
#     balance = {}

#     # ---------------------------------------------
#     # 1) SQL №1 — начальные остатки (суммируем по всем складам сразу)
#     # ---------------------------------------------
#     start_items = InvoiceItem.objects.filter(
#         invoice__entry_created_at_handle__lt=dateFrom
#     ).filter(
#         Q(invoice__warehouse_id__in=warehouse_ids) |
#         Q(invoice__warehouse2_id__in=warehouse_ids)
#     ).select_related(
#         "product", "product__category", "product__base_unit", "invoice"
#     )
    
#     if category_ids:
#         start_items = start_items.filter(product__category_id__in=category_ids)
        
#     if product_ids:
#         start_items = start_items.filter(product__id__in=product_ids)

#     for item in start_items:
#         p = item.product
#         inv = item.invoice
#         if inv.canceled_at != None:
#             continue
        
#         if p.id not in balance:
#             pu = unit_map.get(p.id)
#             if pu:
#                 unit = pu.unit.name
#                 conversion_factor = pu.conversion_factor
#             else:
#                 unit = p.base_unit.name if p.base_unit else ""
#                 conversion_factor = 1
                
#             balance[p.id] = {
#                 "id": p.id,
#                 "category": p.category.name,
#                 "name": p.name,
#                 "unit": unit,
#                 "price": p.wholesale_price,
#                 "selected_quantity": Decimal("0"),
#                 "oborot_selected_quantity_girdeji": Decimal("0"),
#                 "oborot_selected_quantity_chykdajy": Decimal("0"),
#                 "oborot_selected_quantity_wozwrat": Decimal("0"),
#                 "conversion_factor": conversion_factor,
#             }
        
#         cf = Decimal(balance[p.id]["conversion_factor"])
        
        
#         # Проверяем принадлежность к выбранным складам
#         if inv.wozwrat_or_prihod == "prihod":
#             if inv.warehouse_id in warehouse_ids:
#                 balance[p.id]["selected_quantity"] += (item.selected_quantity / cf)

#         elif inv.wozwrat_or_prihod == "rashod":
#             if inv.warehouse_id in warehouse_ids:
#                 balance[p.id]["selected_quantity"] -= (item.selected_quantity / cf)

#         elif inv.wozwrat_or_prihod == "wozwrat":
#             if inv.warehouse_id in warehouse_ids:
#                 balance[p.id]["selected_quantity"] += (item.selected_quantity / cf)
            
#         elif inv.wozwrat_or_prihod == "transfer":
#             # если со склада (и склад в выбранных)
#             if inv.warehouse_id in warehouse_ids:
#                 balance[p.id]["selected_quantity"] -= (item.selected_quantity / cf)
#             # если на склад (и склад в выбранных)
#             elif inv.warehouse2_id in warehouse_ids:
#                 balance[p.id]["selected_quantity"] += (item.selected_quantity / cf)

#     # ---------------------------------------------
#     # 2) SQL №2 — обороты в периоде (суммируем по всем складам)
#     # ---------------------------------------------
#     period_items = InvoiceItem.objects.filter(
#         invoice__entry_created_at_handle__gte=dateFrom,
#         invoice__entry_created_at_handle__lte=dateTo
#     ).filter(
#         Q(invoice__warehouse_id__in=warehouse_ids) |
#         Q(invoice__warehouse2_id__in=warehouse_ids)
#     ).select_related(
#         "product", "product__category", "product__base_unit", "invoice"
#     ).order_by(
#         "product__category__name", "product__name"
#     )
    
#     # ic(period_items)
#     # ic(warehouse_ids)
    

    
#     if category_ids:
#         period_items = period_items.filter(product__category_id__in=category_ids)
        
#     if product_ids:
#         period_items = period_items.filter(product__id__in=product_ids)

#     for item in period_items:
#         p = item.product
#         # if p.id == 606:
#         #     ic(p)
#         inv = item.invoice
        
#         if inv.canceled_at != None:
#             continue
        
#         if p.id not in balance:
#             pu = unit_map.get(p.id)
#             if pu:
#                 unit = pu.unit.name
#                 conversion_factor = pu.conversion_factor
#             else:
#                 unit = p.base_unit.name if p.base_unit else ""
#                 conversion_factor = 1

#             balance[p.id] = {
#                 "id": p.id,
#                 "category": p.category.name,
#                 "name": p.name,
#                 "unit": unit,
#                 "price": p.wholesale_price,
#                 "selected_quantity": Decimal("0"),
#                 "oborot_selected_quantity_girdeji": Decimal("0"),
#                 "oborot_selected_quantity_chykdajy": Decimal("0"),
#                 "oborot_selected_quantity_wozwrat": Decimal("0"),
#                 "conversion_factor": conversion_factor,
#             }
        
#         cf = Decimal(balance[p.id]["conversion_factor"])
        
#         if inv.wozwrat_or_prihod == "prihod":
#             if inv.warehouse_id in warehouse_ids:
#                 balance[p.id]["oborot_selected_quantity_girdeji"] += (item.selected_quantity / cf)

#         elif inv.wozwrat_or_prihod == "rashod":
#             if inv.warehouse_id in warehouse_ids:
#                 balance[p.id]["oborot_selected_quantity_chykdajy"] += (item.selected_quantity / cf)

#         elif inv.wozwrat_or_prihod == "wozwrat":
#             if inv.warehouse_id in warehouse_ids:
#                     balance[p.id]["oborot_selected_quantity_wozwrat"] += (item.selected_quantity / cf)
#         elif inv.wozwrat_or_prihod == "transfer":
#             # уход со склада (и склад в выбранных)
#             if inv.warehouse_id in warehouse_ids:
#                 balance[p.id]["oborot_selected_quantity_chykdajy"] += (item.selected_quantity / cf)
#             # приход на склад (и склад в выбранных)
#             elif inv.warehouse2_id in warehouse_ids:
#                 balance[p.id]["oborot_selected_quantity_girdeji"] += (item.selected_quantity / cf)

#     # ---------------------------------------------
#     # 3) Итог
#     # ---------------------------------------------
#     for p_id, d in balance.items():
#         d["end_selected_quantity"] = (
#             d["selected_quantity"]
#             + d["oborot_selected_quantity_girdeji"]
#             - d["oborot_selected_quantity_chykdajy"]
#             + d["oborot_selected_quantity_wozwrat"]
#         )
        
#     if not emptyTurnovers:
#         balance = {
#             p_id: d
#             for p_id, d in balance.items()
#             if d["oborot_selected_quantity_girdeji"] != Decimal("0")
#             or d["oborot_selected_quantity_chykdajy"] != Decimal("0")
#             or d["oborot_selected_quantity_wozwrat"] != Decimal("0")
#         }

#     return JsonResponse({"data": list(balance.values())})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def BuhOborotTowarow(request):
    dateFrom = request.GET.get("dateFrom")
    dateTo = request.GET.get("dateTo")
    
    # Поддержка нового параметра warehouses и старого warehouse
    warehouses_param = request.GET.get("warehouses", "")
    warehouse_param = request.GET.get("warehouse", "")
    
    categories = request.GET.get("categories")
    products = request.GET.get("products")
    emptyTurnovers = True if request.GET.get("emptyTurnovers") == "1" else False
    
    # Обработка параметра складов
    warehouse_ids = []
    
    if warehouses_param:
        try:
            warehouse_ids = [int(id.strip()) for id in warehouses_param.split(",") if id.strip().isdigit()]
        except ValueError:
            return JsonResponse({"status": "error", "message": "Invalid warehouse IDs"}, status=400)
    elif warehouse_param and warehouse_param.isdigit():
        warehouse_ids = [int(warehouse_param)]
    
    if not warehouse_ids:
        return JsonResponse({"status": "error", "message": "Warehouse(s) not selected"}, status=400)
    
    product_units = (
        ProductUnit.objects
        .filter(is_default_for_sale=True)
        .select_related("unit")
    )
    
    unit_map = {
        pu.product_id: pu
        for pu in product_units
    }
    
    if categories:
        category_ids = [int(c) for c in categories.split(",") if c.isdigit()]
    else:
        category_ids = []
        
    if products:
        product_ids = [int(c) for c in products.split(",") if c.isdigit()]
    else:
        product_ids = []

    # Баланс по товарам
    balance = {}

    # ---------------------------------------------
    # 0) ИНИЦИАЛИЗАЦИЯ ВСЕХ ТОВАРОВ НА ВЫБРАННЫХ СКЛАДАХ
    # ---------------------------------------------
    
    # Получаем все product_id которые есть на выбранных складах
    warehouse_products = WarehouseProduct.objects.filter(
        warehouse_id__in=warehouse_ids
    ).select_related(
        "product",
        "product__category",
        "product__base_unit"
    )
    
    # Применяем фильтры по категориям и товарам
    if category_ids:
        warehouse_products = warehouse_products.filter(product__category_id__in=category_ids)
    
    if product_ids:
        warehouse_products = warehouse_products.filter(product__id__in=product_ids)
    
    # Собираем все product_id
    all_product_ids = set()
    for wp in warehouse_products:
        all_product_ids.add(wp.product.id)
    
    # Также добавляем товары из истории операций (до и в периоде)
    historical_product_ids = InvoiceItem.objects.filter(
        Q(invoice__entry_created_at_handle__lte=dateTo),
        Q(invoice__warehouse_id__in=warehouse_ids) | Q(invoice__warehouse2_id__in=warehouse_ids),
        invoice__canceled_at__isnull=True
    )
    
    if category_ids:
        historical_product_ids = historical_product_ids.filter(product__category_id__in=category_ids)
    
    if product_ids:
        historical_product_ids = historical_product_ids.filter(product__id__in=product_ids)
    
    historical_product_ids = historical_product_ids.values_list('product_id', flat=True).distinct()
    
    for pid in historical_product_ids:
        all_product_ids.add(pid)
    
    # Инициализируем все товары
    products_to_init = Product.objects.filter(
        id__in=all_product_ids
    ).select_related('category', 'base_unit')
    
    for p in products_to_init:
        pu = unit_map.get(p.id)
        if pu:
            unit = pu.unit.name
            conversion_factor = pu.conversion_factor
        else:
            unit = p.base_unit.name if p.base_unit else ""
            conversion_factor = 1
        
        balance[p.id] = {
            "id": p.id,
            "category": p.category.name if p.category else "",
            "name": p.name,
            "unit": unit,
            "price": p.wholesale_price or Decimal("0"),
            "selected_quantity": Decimal("0"),
            "oborot_selected_quantity_girdeji": Decimal("0"),
            "oborot_selected_quantity_chykdajy": Decimal("0"),
            "oborot_selected_quantity_wozwrat": Decimal("0"),
            # Добавляем фактические суммы оборотов
            "oborot_girdeji_price": Decimal("0"),
            "oborot_chykdajy_price": Decimal("0"),
            "oborot_wozwrat_price": Decimal("0"),
            "conversion_factor": conversion_factor,
        }

    # ---------------------------------------------
    # 1) SQL №1 — начальные остатки (суммируем по всем складам сразу)
    # ---------------------------------------------
    start_items = InvoiceItem.objects.filter(
        invoice__entry_created_at_handle__lt=dateFrom,
        invoice__canceled_at__isnull=True
    ).filter(
        Q(invoice__warehouse_id__in=warehouse_ids) |
        Q(invoice__warehouse2_id__in=warehouse_ids)
    ).select_related(
        "product", "product__category", "product__base_unit", "invoice"
    )
    
    if category_ids:
        start_items = start_items.filter(product__category_id__in=category_ids)
        
    if product_ids:
        start_items = start_items.filter(product__id__in=product_ids)

    for item in start_items:
        p = item.product
        inv = item.invoice
        
        # Товар уже инициализирован выше
        if p.id not in balance:
            continue
        
        cf = Decimal(balance[p.id]["conversion_factor"])
        
        # Проверяем принадлежность к выбранным складам
        if inv.wozwrat_or_prihod == "prihod":
            if inv.warehouse_id in warehouse_ids:
                balance[p.id]["selected_quantity"] += (item.selected_quantity / cf)

        elif inv.wozwrat_or_prihod == "rashod":
            if inv.warehouse_id in warehouse_ids:
                balance[p.id]["selected_quantity"] -= (item.selected_quantity / cf)

        elif inv.wozwrat_or_prihod == "wozwrat":
            if inv.warehouse_id in warehouse_ids:
                balance[p.id]["selected_quantity"] += (item.selected_quantity / cf)
            
        elif inv.wozwrat_or_prihod == "transfer":
            # если со склада (и склад в выбранных)
            if inv.warehouse_id in warehouse_ids:
                balance[p.id]["selected_quantity"] -= (item.selected_quantity / cf)
            # если на склад (и склад в выбранных)
            elif inv.warehouse2_id in warehouse_ids:
                balance[p.id]["selected_quantity"] += (item.selected_quantity / cf)

    # ---------------------------------------------
    # 2) SQL №2 — обороты в периоде (суммируем по всем складам)
    # ---------------------------------------------
    period_items = InvoiceItem.objects.filter(
        invoice__entry_created_at_handle__gte=dateFrom,
        invoice__entry_created_at_handle__lte=dateTo,
        invoice__canceled_at__isnull=True
    ).filter(
        Q(invoice__warehouse_id__in=warehouse_ids) |
        Q(invoice__warehouse2_id__in=warehouse_ids)
    ).select_related(
        "product", "product__category", "product__base_unit", "invoice"
    ).order_by(
        "product__category__name", "product__name"
    )
    
    if category_ids:
        period_items = period_items.filter(product__category_id__in=category_ids)
        
    if product_ids:
        period_items = period_items.filter(product__id__in=product_ids)

    for item in period_items:
        p = item.product
        inv = item.invoice
        
        # Товар уже инициализирован выше
        if p.id not in balance:
            continue
        
        cf = Decimal(balance[p.id]["conversion_factor"])
        qty = item.selected_quantity / cf
        price = qty * item.selected_price  # ФАКТИЧЕСКАЯ ЦЕНА ИЗ НАКЛАДНОЙ
        
        if inv.wozwrat_or_prihod == "prihod":
            if inv.warehouse_id in warehouse_ids:
                balance[p.id]["oborot_selected_quantity_girdeji"] += qty
                balance[p.id]["oborot_girdeji_price"] += price

        elif inv.wozwrat_or_prihod == "rashod":
            if inv.warehouse_id in warehouse_ids:
                balance[p.id]["oborot_selected_quantity_chykdajy"] += qty
                balance[p.id]["oborot_chykdajy_price"] += price

        elif inv.wozwrat_or_prihod == "wozwrat":
            if inv.warehouse_id in warehouse_ids:
                balance[p.id]["oborot_selected_quantity_wozwrat"] += qty
                balance[p.id]["oborot_wozwrat_price"] += price
                
        elif inv.wozwrat_or_prihod == "transfer":
            # уход со склада (и склад в выбранных)
            if inv.warehouse_id in warehouse_ids:
                balance[p.id]["oborot_selected_quantity_chykdajy"] += qty
                balance[p.id]["oborot_chykdajy_price"] += price
            # приход на склад (и склад в выбранных)
            elif inv.warehouse2_id in warehouse_ids:
                balance[p.id]["oborot_selected_quantity_girdeji"] += qty
                balance[p.id]["oborot_girdeji_price"] += price

    # ---------------------------------------------
    # 3) Итог
    # ---------------------------------------------
    for p_id, d in balance.items():
        d["end_selected_quantity"] = (
            d["selected_quantity"]
            + d["oborot_selected_quantity_girdeji"]
            - d["oborot_selected_quantity_chykdajy"]
            + d["oborot_selected_quantity_wozwrat"]
        )
        
    if not emptyTurnovers:
        balance = {
            p_id: d
            for p_id, d in balance.items()
            if d["selected_quantity"] != Decimal("0")  # начальный остаток
            or d["oborot_selected_quantity_girdeji"] != Decimal("0")
            or d["oborot_selected_quantity_chykdajy"] != Decimal("0")
            or d["oborot_selected_quantity_wozwrat"] != Decimal("0")
            or d["end_selected_quantity"] != Decimal("0")  # конечный остаток
        }

    return JsonResponse({"data": list(balance.values())})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def BuhOborotTowarowExcel(request):
    dateFrom = request.GET.get("dateFrom")
    dateTo = request.GET.get("dateTo")
    
    
    day_start = datetime.strptime(dateFrom, "%Y-%m-%d").date()
    day_end = datetime.strptime(dateTo, "%Y-%m-%d").date()
    
    # Поддержка нового параметра warehouses и старого warehouse
    warehouses_param = request.GET.get("warehouses", "")
    warehouse_param = request.GET.get("warehouse", "")
    
    categories = request.GET.get("categories")
    products = request.GET.get("products")
    emptyTurnovers = True if request.GET.get("emptyTurnovers") == "1" else False
    
    # Обработка параметра складов
    warehouse_ids = []
    
    if warehouses_param:
        try:
            warehouse_ids = [int(id.strip()) for id in warehouses_param.split(",") if id.strip().isdigit()]
        except ValueError:
            return JsonResponse({"status": "error", "message": "Invalid warehouse IDs"}, status=400)
    elif warehouse_param and warehouse_param.isdigit():
        warehouse_ids = [int(warehouse_param)]
    
    if not warehouse_ids:
        return JsonResponse({"status": "error", "message": "Warehouse(s) not selected"}, status=400)
    

    products = Product.objects.filter(
        Q(warehouse_products__warehouse_id__in=warehouse_ids) |
        Q(invoiceitem__invoice__warehouse_id__in=warehouse_ids) |
        Q(invoiceitem__invoice__warehouse2_id__in=warehouse_ids)
    ).distinct()
        
    unit_map = get_unit_map()
    account_40_42 = {}
    for p in products:
        if not p.category:
            continue
        unit, cf = get_unit_and_cf(unit_map, p)
        category_id = p.category.id
        category_name = p.category.name
        # if p.id == 601:
        #     ic(p.name, unit, cf)
        if category_id not in account_40_42:
            account_40_42[category_id] = {
                "category": {
                    "id": category_id,
                    "name": category_name,
                },
                "totals": {
                    "start_qty": Decimal("0.00"),
                    "start_price": Decimal("0.00"),

                    "prihod_qty": Decimal("0.00"),
                    "prihod_price": Decimal("0.00"),

                    "wozwrat_qty": Decimal("0.00"),
                    "wozwrat_price": Decimal("0.00"),

                    "rashod_qty": Decimal("0.00"),
                    "rashod_price": Decimal("0.00"),

                    "end_qty": Decimal("0.00"),
                    "end_price": Decimal("0.00"),
                },
                "products": {}
            }

        account_40_42[category_id]["products"][p.id] = {
            "product": {
                "id": p.id,
                "name": p.name,
                "unit": unit,
                "cf": Decimal(cf),
                "wholsale_price": p.wholesale_price,
            },

            "start_qty": Decimal("0.00"),
            "start_price": Decimal("0.00"),

            "prihod_qty": Decimal("0.00"),
            "prihod_price": Decimal("0.00"),

            "wozwrat_qty": Decimal("0.00"),
            "wozwrat_price": Decimal("0.00"),

            "rashod_qty": Decimal("0.00"),
            "rashod_price": Decimal("0.00"),

            "end_qty": Decimal("0.00"),
            "end_price": Decimal("0.00"),
        }
        
    start_items = InvoiceItem.objects.filter(
        invoice__entry_created_at_handle__lt=day_start,
        invoice__canceled_at__isnull=True
    ).filter(
        Q(invoice__warehouse_id__in=warehouse_ids) |
        Q(invoice__warehouse2_id__in=warehouse_ids)
    ).select_related(
        "product", "product__category", "product__base_unit", "invoice"
    )
    
    grand_total = {
        "start_qty": Decimal("0.00"),
        "start_price": Decimal("0.00"),

        "prihod_qty": Decimal("0.00"),
        "prihod_price": Decimal("0.00"),

        "wozwrat_qty": Decimal("0.00"),
        "wozwrat_price": Decimal("0.00"),

        "rashod_qty": Decimal("0.00"),
        "rashod_price": Decimal("0.00"),

        "end_qty": Decimal("0.00"),
        "end_price": Decimal("0.00"),
    }
            
        
    for item in start_items:
        p = item.product
        inv = item.invoice
        
        # Товар уже инициализирован выше
        if not p.category:
            continue
        if (p.category.id not in account_40_42 or p.id not in account_40_42[p.category.id]["products"]):
            continue
        cat_totals = account_40_42[p.category.id]["totals"]
        prod = account_40_42[p.category.id]["products"][p.id]
        cf = prod["product"]["cf"]
        qty = Decimal(item.selected_quantity) / cf
        calculated_price = qty * Decimal(item.selected_price)
        # Проверяем принадлежность к выбранным складам
        if inv.wozwrat_or_prihod == "prihod":
            if inv.warehouse_id in warehouse_ids:
                prod["start_qty"] += qty
                prod["start_price"] += calculated_price
                
                cat_totals["start_qty"] += qty
                cat_totals["start_price"] += calculated_price
                
                grand_total["start_qty"] += qty
                grand_total["start_price"] += calculated_price

        elif inv.wozwrat_or_prihod == "rashod":
            if inv.warehouse_id in warehouse_ids:
                prod["start_qty"] -= qty
                prod["start_price"] -= calculated_price
                
                cat_totals["start_qty"] -= qty
                cat_totals["start_price"] -= calculated_price
                
                grand_total["start_qty"] -= qty
                grand_total["start_price"] -= calculated_price

        elif inv.wozwrat_or_prihod == "wozwrat":
            if inv.warehouse_id in warehouse_ids:
                prod["start_qty"] += qty
                prod["start_price"] += calculated_price
                
                cat_totals["start_qty"] += qty
                cat_totals["start_price"] += calculated_price
                
                grand_total["start_qty"] += qty
                grand_total["start_price"] += calculated_price
            
        elif inv.wozwrat_or_prihod == "transfer":
            # если со склада (и склад в выбранных)
            if inv.warehouse_id in warehouse_ids:
                prod["start_qty"] -= qty
                prod["start_price"] -= calculated_price
                
                cat_totals["start_qty"] -= qty
                cat_totals["start_price"] -= calculated_price
                
                grand_total["start_qty"] -= qty
                grand_total["start_price"] -= calculated_price
                
            # если на склад (и склад в выбранных)
            elif inv.warehouse2_id in warehouse_ids:
                prod["start_qty"] += qty
                prod["start_price"] += calculated_price
                
                cat_totals["start_qty"] += qty
                cat_totals["start_price"] += calculated_price
                
                grand_total["start_qty"] += qty
                grand_total["start_price"] += calculated_price
                
    turnover_items = InvoiceItem.objects.filter(
        invoice__entry_created_at_handle__gte=day_start,
        invoice__entry_created_at_handle__lte=day_end,
        invoice__canceled_at__isnull=True
    ).filter(
        Q(invoice__warehouse_id__in=warehouse_ids) |
        Q(invoice__warehouse2_id__in=warehouse_ids)
    ).select_related(
        "product", "product__category", "product__base_unit", "invoice"
    ).order_by(
        "product__category__name", "product__name"
    )
    
    for item in turnover_items:
        p = item.product
        inv = item.invoice
        
        # Товар уже инициализирован выше
        if not p.category:
            continue
        if (p.category.id not in account_40_42 or p.id not in account_40_42[p.category.id]["products"]):
            continue
        cat_totals = account_40_42[p.category.id]["totals"]
        prod = account_40_42[p.category.id]["products"][p.id]
        cf = prod["product"]["cf"]
        qty = Decimal(item.selected_quantity) / cf
        calculated_price = qty * Decimal(item.selected_price)
        # Проверяем принадлежность к выбранным складам
        if inv.wozwrat_or_prihod == "prihod":
            if inv.warehouse_id in warehouse_ids:
                prod["prihod_qty"] += qty
                prod["prihod_price"] += calculated_price
                
                cat_totals["prihod_qty"] += qty
                cat_totals["prihod_price"] += calculated_price
                
                grand_total["prihod_qty"] += qty
                grand_total["prihod_price"] += calculated_price

        elif inv.wozwrat_or_prihod == "rashod":
            if inv.warehouse_id in warehouse_ids:
                prod["rashod_qty"] += qty
                prod["rashod_price"] += calculated_price
                
                cat_totals["rashod_qty"] += qty
                cat_totals["rashod_price"] += calculated_price
                
                grand_total["rashod_qty"] += qty
                grand_total["rashod_price"] += calculated_price

        elif inv.wozwrat_or_prihod == "wozwrat":
            if inv.warehouse_id in warehouse_ids:
                prod["wozwrat_qty"] += qty
                prod["wozwrat_price"] += calculated_price
                
                cat_totals["wozwrat_qty"] += qty
                cat_totals["wozwrat_price"] += calculated_price
                
                grand_total["wozwrat_qty"] += qty
                grand_total["wozwrat_price"] += calculated_price
            
        elif inv.wozwrat_or_prihod == "transfer":
            # если со склада (и склад в выбранных)
            if inv.warehouse_id in warehouse_ids:
                prod["rashod_qty"] += qty
                prod["rashod_price"] += calculated_price
                
                cat_totals["rashod_qty"] += qty
                cat_totals["rashod_price"] += calculated_price
                
                grand_total["rashod_qty"] += qty
                grand_total["rashod_price"] += calculated_price
                
            # если на склад (и склад в выбранных)
            elif inv.warehouse2_id in warehouse_ids:
                prod["prihod_qty"] += qty
                prod["prihod_price"] += calculated_price
                
                cat_totals["prihod_qty"] += qty
                cat_totals["prihod_price"] += calculated_price
                
                grand_total["prihod_qty"] += qty
                grand_total["prihod_price"] += calculated_price
                
    for cat in account_40_42.values():
        for prod in cat["products"].values():
            prod["end_qty"] = (
                prod["start_qty"]
                + prod["prihod_qty"]
                - prod["rashod_qty"]
                + prod["wozwrat_qty"]
            )

            prod["end_price"] = (
                prod["start_price"]
                + prod["prihod_price"]
                - prod["rashod_price"]
                + prod["wozwrat_price"]
            )
            
    for cat in account_40_42.values():
        totals = cat["totals"]
        totals["end_qty"] = (
            totals["start_qty"]
            + totals["prihod_qty"]
            - totals["rashod_qty"]
            + totals["wozwrat_qty"]
        )
        totals["end_price"] = (
            totals["start_price"]
            + totals["prihod_price"]
            - totals["rashod_price"]
            + totals["wozwrat_price"]
        )
        
    grand_total["end_qty"] = (
        grand_total["start_qty"]
        + grand_total["prihod_qty"]
        - grand_total["rashod_qty"]
        + grand_total["wozwrat_qty"]
    )
    grand_total["end_price"] = (
        grand_total["start_price"]
        + grand_total["prihod_price"]
        - grand_total["rashod_price"]
        + grand_total["wozwrat_price"]
    )
    
    w_acc = (
        WarehouseAccount.objects
        .select_related("warehouse", "account")
        .filter(warehouse_id__in=warehouse_ids)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Бух. Оборот товаров"

    account_warehouse = ", ".join(
        sorted({f"{w.account.number} — {w.warehouse.name}" for w in w_acc})
    )

    dateFrom_ru = format_date_ru(dateFrom)
    dateTo_ru = format_date_ru(dateTo)

    # === Заголовок ===
    ws.merge_cells("A2:N2")
    ws.merge_cells("A3:N3")
    ws.merge_cells("A4:N4")

    ws["A2"] = "Бухгалтерский оборот товаров"
    ws["A3"] = f"Склад(ы): {account_warehouse}"
    ws["A4"] = f"Период: {dateFrom_ru} — {dateTo_ru}"

    for cell in ["A2", "A3", "A4"]:
        ws[cell].font = CATEGORY_FONT
        ws[cell].alignment = CENTER_ALIGN
        
    ws.freeze_panes = "A7"
    
    ws["A5"] = "№"
    ws.column_dimensions["A"].width = 8
    ws["B5"] = "Наименование товара"
    ws.column_dimensions["B"].width = 55
    ws["C5"] = "Ед."
    ws.column_dimensions["C"].width = 8
    
    ws["D5"] = "Цена"
    ws.column_dimensions["D"].width = 8
    ws[f"D5"].number_format = PRICE_FMT
    
    ws.merge_cells("E5:F5")
    ws["E5"] = "Остаток на начало"


    
    ws.merge_cells("G5:H5")
    ws["G5"] = "Приход"
    
    
    ws.merge_cells("I5:J5")
    ws["I5"] = "Возврат"
    
    ws.merge_cells("K5:L5")
    ws["K5"] = "Расход"
    
    ws.merge_cells("M5:N5")
    ws["M5"] = "Остаток на конец"

    
    
    for i in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", 
                "A6", "B6", "C6", "D6", "E6", "F6", "G6", "H6", "I6", "J6", "K6", "L6", "M6", "N6"]:
        ws[i].font = CATEGORY_FONT
        ws[i].alignment = CENTER_ALIGN
        ws[i].fill = GRAY_FILL_1
        ws[i].border = THIN_BORDER
        
        
    for i in ["E", "F",  "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws.column_dimensions[i].width = 11

        

    ws["E6"] = "Кол-во"
    ws["F6"] = "Всего"
    
    ws["G6"] = "Кол-во"
    ws["H6"] = "Всего"
    
    ws["I6"] = "Кол-во"
    ws["J6"] = "Всего"
    
    ws["K6"] = "Кол-во"
    ws["L6"] = "Всего"
    
    ws["M6"] = "Кол-во"
    ws["N6"] = "Всего"
    
    
    row = 7
    count = 1
    for cat_id, values in sorted(
        account_40_42.items(),
        key=lambda item: item[1]["category"]["name"].lower()
    ):
        cat_name = values["category"]["name"]
        totals = values["totals"]
        ws.merge_cells(f"A{row}:N{row}")
        ws[f"A{row}"].fill = GREEN_FILL_0
        ws[f"A{row}"].font = CATEGORY_FONT
        ws[f"A{row}"].alignment = LEFT_ALIGN
        ws[f"A{row}"] = cat_name
        
        
        
        for col in "FHJLN":
            ws[f"{col}{row}"].number_format = PRICE_FMT
            
        for col in "EGIKM":
            ws[f"{col}{row}"].number_format = QTY_FMT
        
        row += 1
        
        for products, value in values["products"].items():
            product = value["product"]
            unit = product["unit"]
            wholesale_price = product["wholsale_price"]
            
            start_qty = value["start_qty"]
            start_price = value["start_price"]
            
            prihod_qty = value["prihod_qty"]
            prihod_price = value["prihod_price"]
            
            
            wozwrat_qty = value["wozwrat_qty"]
            wozwrat_price = value["wozwrat_price"]
            
            rashod_qty = value["rashod_qty"]
            rashod_price = value["rashod_price"]
            
            end_qty = value["end_qty"]
            end_price = value["end_price"] 
            
            ws[f"G{row}"].font = GREEN_FONT
            ws[f"H{row}"].font = GREEN_FONT
            ws[f"I{row}"].font = RED_FONT
            ws[f"J{row}"].font = RED_FONT
            ws[f"K{row}"].font = BLUE_FONT
            ws[f"L{row}"].font = BLUE_FONT
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws[i].border = THIN_BORDER
                
            for col in "FHJLN":
                ws[f"{col}{row}"].number_format = PRICE_FMT
                
            for col in "EGIKM":
                ws[f"{col}{row}"].number_format = QTY_FMT
        
            ws[f"A{row}"] = count
            ws[f"B{row}"] = product["name"]
            ws[f"C{row}"] = unit
            ws[f"D{row}"] = wholesale_price
            
            ws[f"E{row}"] = start_qty
            ws[f"F{row}"] = start_price
            
            ws[f"G{row}"] = prihod_qty
            ws[f"H{row}"] = prihod_price
            
            ws[f"I{row}"] = wozwrat_qty
            ws[f"J{row}"] = wozwrat_price
            
            ws[f"K{row}"] = rashod_qty
            ws[f"L{row}"] = rashod_price
            
            ws[f"M{row}"] = end_qty
            ws[f"N{row}"] = end_price
            
        
            count += 1
            row += 1
    
        for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws[i].border = THIN_BORDER
                
        ws.merge_cells(f"A{row}:D{row}")
        for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws[i].fill = GRAY_FILL_0
        ws[f"A{row}"].font = CATEGORY_FONT
        ws[f"A{row}"].alignment = RIGHT_ALIGN
        ws[f"A{row}"] = f"Итого по категории: {cat_name}"
        
        
        ws[f"E{row}"] = totals["start_qty"]
        ws[f"F{row}"] = totals["start_price"]
        ws[f"G{row}"] = totals["prihod_qty"]
        ws[f"H{row}"] = totals["prihod_price"]
        ws[f"I{row}"] = totals["wozwrat_qty"]
        ws[f"J{row}"] = totals["wozwrat_price"]
        ws[f"K{row}"] = totals["rashod_qty"]
        ws[f"L{row}"] = totals["rashod_price"]
        ws[f"M{row}"] = totals["end_qty"]
        ws[f"N{row}"] = totals["end_price"]
                        
        row += 1
        
    ws[f"E{row}"] = grand_total["start_qty"]
    ws[f"F{row}"] = grand_total["start_price"]
    ws[f"G{row}"] = grand_total["prihod_qty"]
    ws[f"H{row}"] = grand_total["prihod_price"]
    ws[f"I{row}"] = grand_total["wozwrat_qty"]
    ws[f"J{row}"] = grand_total["wozwrat_price"]
    ws[f"K{row}"] = grand_total["rashod_qty"]
    ws[f"L{row}"] = grand_total["rashod_price"]
    ws[f"M{row}"] = grand_total["end_qty"]
    ws[f"N{row}"] = grand_total["end_price"]
    
    for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
        ws[i].border = THIN_BORDER
        
    for col in "FHJLN":
        ws[f"{col}{row}"].number_format = PRICE_FMT
        
    for col in "EGIKM":
        ws[f"{col}{row}"].number_format = QTY_FMT
                
    ws.merge_cells(f"A{row}:D{row}")
    for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
        ws[i].fill = GREEN_FILL_1

    ws[f"A{row}"].font = CATEGORY_FONT
    ws[f"A{row}"].alignment = RIGHT_ALIGN
    ws[f"A{row}"] = "ВСЕГО:"
            
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"buh_oborot_tovarov_{dateFrom}_{dateTo}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    print("EXCEL RESPONSE SENT")
    return response

    

                
    
    # data = []
    # ic(dateFrom)
    # ic(dateTo)
    # ic(warehouses_param)
    # ic(categories)
    # ic(products)
    # ic(emptyTurnovers)
    # ic(warehouse_ids)

    # return Response(data)
    
    
# BuhOborotTowarow (3 warianta moy, chatGpt i DeepSeek) END
################################################################################################################################################################
################################################################################################################################################################



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_active_warehouses(request):
    ic("get_active_warehouses")
    warehouses = Warehouse.objects.all()
    data = []
    if warehouses:
        for w in warehouses:
            d = {
                "id": w.id,
                "name": w.name,
                "is_active": w.is_active,
                "location": w.location,
                "currency_name": w.currency.name if w.currency else "",
                "currency_code": w.currency.code if w.currency else "",
            }
            data.append(d)
    # ic(data)
    return JsonResponse({
        "data": data,
    })
    
    
@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_initial_stock(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=400)

    if "file" not in request.FILES:
        return JsonResponse({"error": "No file provided"}, status=400)

    excel_file = request.FILES["file"]
    
    warehouse = Warehouse.objects.get(name="Sklad 1 USD")
    # ic(warehouse)

    try:
        # Читаем Excel
        df = pd.read_excel(excel_file)
        # for _, row in df.iterrows():
        #     # ic(row)
        #     name = str(row["haryt"]).strip()
        #     qty = Decimal(row["sany"])
        #     price = Decimal(row["baha"])
        #     unit = row["ol"]
        #     unit_obj = UnitOfMeasurement.objects.get(name=unit)
            
        #     p = Product.objects.get(name=name)
        #     ic(name, qty, price, unit, unit_obj.name)

        # Ожидаемые колонки:
        # # name, quantity, price (или любые твои)
        # required_columns = ["name", "quantity"]
        # for col in required_columns:
        #     if col not in df.columns:
        #         return JsonResponse({"error": f"Column '{col}' not found in Excel"}, status=400)
        
        dt = date(2025, 8, 31)

        

        count = 0
        try:
            with db_transaction.atomic():
                # Создаём системный invoice
                invoice = Invoice.objects.create(
                    wozwrat_or_prihod="prihod",
                    comment="Первоначальный приход из Excel",
                    invoice_date=dt,
                    entry_created_at=dt,
                    created_at_handle=dt,
                    updated_at_handle=dt,
                    entry_created_at_handle=dt,
                    created_by=request.user,
                    warehouse=warehouse,
                    is_entry=True,
                )
                
                
                transaction = Transaction.objects.create(
                    description="Первоначальный приход из Excel",
                    date=dt,
                    invoice=invoice,
                    created_by=request.user
                )
                
                warehouse_account = Account.objects.get(number="40.1")
                fond_account = Account.objects.get(number="80")
                 
            
                for _, row in df.iterrows():
                    name = str(row["haryt"]).strip()
                    # qty = Decimal(row["sany"])
                    # price = Decimal(row["baha"])
                    qty = Decimal(str(row["sany"]))
                    price = Decimal(str(row["baha"]))
                    unit = row["ol"]
                    unit_obj = UnitOfMeasurement.objects.get(name=unit)
                    
                    # amount = (qty * price).quantize(Decimal("0.01"))
                    amount = (qty * price).quantize(
                        Decimal("0.01"),
                        rounding=ROUND_HALF_UP
                    )       

                    product = Product.objects.get(name=name)

                    # Создаём InvoiceItem
                    InvoiceItem.objects.create(
                        invoice=invoice,
                        product=product,
                        selected_quantity=qty,
                        selected_price=price,
                        unit_name_on_selected_warehouses=unit,
                        base_unit_obj=unit_obj,
                    )
                    
                    Entry.objects.create(
                        transaction=transaction,
                        account=warehouse_account,
                        product=product,
                        warehouse=warehouse,
                        debit=amount
                    )

                    Entry.objects.create(
                        transaction=transaction,
                        account=fond_account,
                        credit=amount
                    )

                    count += 1
        except Exception as e:
            ic(e)
            # create error
            return JsonResponse({"status": "error", "message": "transactionChange", "reason_for_the_error": str(e)}, status=400)

        return JsonResponse({
            "status": "success",
            "imported_rows": count,
            # "invoice_id": invoice.id
            "invoice_id": None
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)    
    
    





@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_products_id_and_name(request):
    warehouse_id = request.GET.get("warehouseId")
    data = []
    if warehouse_id:
        products = Product.objects.filter(
            warehouse_products__warehouse_id=warehouse_id,
        )
        if products:
            for p in products:
                d = {
                    "id": p.id,
                    "name": p.name
                }
                data.append(d)
        
    return JsonResponse({
        "data": data,
    })
    
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def product_buh_oborot_detail(request, product_id):
    
    dateFrom = request.GET.get("dateFrom")
    dateTo = request.GET.get("dateTo")
    
    # Поддержка нового параметра warehouses и старого warehouseId
    warehouses_param = request.GET.get("warehouses", "")
    warehouse_id_param = request.GET.get("warehouseId", "")
    
    if warehouses_param:
        # Новый формат: несколько складов
        try:
            warehouse_ids = [int(id.strip()) for id in warehouses_param.split(",") if id.strip().isdigit()]
        except ValueError:
            return JsonResponse({"status": "error", "message": "Invalid warehouse IDs"}, status=400)
    elif warehouse_id_param and warehouse_id_param.isdigit():
        # Старый формат: один склад
        warehouse_ids = [int(warehouse_id_param)]
    else:
        return JsonResponse(
            {"status": "error", "message": "choose correct warehouse(s)"},
            status=400
        )
    
    if not product_id:
        return JsonResponse(
            {"status": "error", "message": "choose correct product"},
            status=400
        )

    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse(
            {"status": "error", "message": "product not found"},
            status=404
        )

    # базовая единица по умолчанию
    unit_name = product.base_unit.name
    conversion_factor = 1

    default_unit = ProductUnit.objects.filter(
        product=product,
        is_default_for_sale=True
    ).select_related('unit').first()

    if default_unit:
        unit_name = default_unit.unit.name
        conversion_factor = float(default_unit.conversion_factor)

    # Данные для каждого склада
    warehouses_data = {}
    all_rows = []
    
    for warehouse_id in warehouse_ids:
        try:
            warehouse_obj = Warehouse.objects.get(id=warehouse_id)
        except Warehouse.DoesNotExist:
            continue
        
        items = (
            InvoiceItem.objects
            .filter(
                product_id=product.id,
                invoice__entry_created_at_handle__lt=dateFrom
            )
            .filter(
                Q(invoice__warehouse_id=warehouse_id) |
                Q(invoice__warehouse2_id=warehouse_id)
            )
            .select_related("invoice")
        )
        
        start_quantity = Decimal("0")
        cf = Decimal(str(conversion_factor))
        
        for item in items:
            inv = item.invoice
            if inv.canceled_at != None:
                continue
            qty = Decimal(item.selected_quantity) / cf

            if inv.wozwrat_or_prihod == "prihod":
                if inv.warehouse_id == warehouse_id:
                    start_quantity += qty

            elif inv.wozwrat_or_prihod == "rashod":
                if inv.warehouse_id == warehouse_id:
                    start_quantity -= qty

            elif inv.wozwrat_or_prihod == "wozwrat":
                if inv.warehouse_id == warehouse_id:
                    start_quantity += qty

            elif inv.wozwrat_or_prihod == "transfer":
                if inv.warehouse_id == warehouse_id:
                    start_quantity -= qty
                elif inv.warehouse2_id == warehouse_id:
                    start_quantity += qty
        
        period_items = (
            InvoiceItem.objects
            .filter(
                product_id=product.id,
                invoice__entry_created_at_handle__gte=dateFrom,
                invoice__entry_created_at_handle__lte=dateTo
            )
            .filter(
                Q(invoice__warehouse_id=warehouse_id) |
                Q(invoice__warehouse2_id=warehouse_id)
            )
            .select_related("invoice", "invoice__partner")
            .order_by("invoice__entry_created_at_handle", "invoice_id")
        )
        
        rows = []
        balance_qty = start_quantity
        cf = Decimal(str(conversion_factor))

        for item in period_items:
            inv = item.invoice
            if inv.canceled_at != None:
                continue
            qty = Decimal(item.selected_quantity) / cf
            price = Decimal(item.selected_price)
            income_qty = Decimal("0")
            outcome_qty = Decimal("0")
            return_qty = Decimal("0")

            # ПРИХОД
            if inv.wozwrat_or_prihod == "prihod" and inv.warehouse_id == warehouse_id:
                income_qty = qty
                balance_qty += qty

            # РАСХОД
            elif inv.wozwrat_or_prihod == "rashod" and inv.warehouse_id == warehouse_id:
                outcome_qty = qty
                balance_qty -= qty

            # ВОЗВРАТ
            elif inv.wozwrat_or_prihod == "wozwrat" and inv.warehouse_id == warehouse_id:
                return_qty = qty
                balance_qty += qty

            # ПЕРЕМЕЩЕНИЕ
            elif inv.wozwrat_or_prihod == "transfer":
                if inv.warehouse_id == warehouse_id:
                    outcome_qty = qty
                    balance_qty -= qty
                elif inv.warehouse2_id == warehouse_id:
                    income_qty = qty
                    balance_qty += qty
            
            row_data = {
                "date": inv.entry_created_at_handle,
                "invoice_id": inv.id,
                "partner": inv.partner.name if inv.partner else "",
                "operation": inv.wozwrat_or_prihod,
                "text": f"{inv.get_wozwrat_or_prihod_display()} №{inv.id} ({inv.comment})",

                "price": price,

                "income_qty": income_qty,
                "income_sum": income_qty * price,

                "outcome_qty": outcome_qty,
                "outcome_sum": outcome_qty * price,

                "return_qty": return_qty,
                "return_sum": return_qty * price,

                "balance_qty": balance_qty,
                "balance_sum": balance_qty * price,

                "warehouse_id": warehouse_id,
            }
            rows.append(row_data)
            all_rows.append(row_data)
        
        total_income_qty = Decimal("0")
        total_income_sum = Decimal("0")
        total_outcome_qty = Decimal("0")
        total_outcome_sum = Decimal("0")
        total_return_qty = Decimal("0")
        total_return_sum = Decimal("0")

        for r in rows:
            total_income_qty += r["income_qty"]
            total_income_sum += r["income_sum"]

            total_outcome_qty += r["outcome_qty"]
            total_outcome_sum += r["outcome_sum"]

            total_return_qty += r["return_qty"]
            total_return_sum += r["return_sum"]

        end_quantity = (
            start_quantity
            + total_income_qty
            + total_return_qty
            - total_outcome_qty
        )
        end_sum = end_quantity * product.wholesale_price
        
        warehouses_data[str(warehouse_id)] = {
            "id": warehouse_id,
            "name": warehouse_obj.name,
            "start_quantity": start_quantity,
            "income_qty": total_income_qty,
            "income_sum": total_income_sum,
            "outcome_qty": total_outcome_qty,
            "outcome_sum": total_outcome_sum,
            "return_qty": total_return_qty,
            "return_sum": total_return_sum,
            "end_quantity": end_quantity,
            "end_sum": end_sum,
        }
    
    # Сортируем общие строки по дате
    all_rows.sort(key=lambda x: x["date"])
    
    # Агрегируем общие данные
    total_start_quantity = sum(wh["start_quantity"] for wh in warehouses_data.values())
    total_income_qty = sum(wh["income_qty"] for wh in warehouses_data.values())
    total_income_sum = sum(wh["income_sum"] for wh in warehouses_data.values())
    total_outcome_qty = sum(wh["outcome_qty"] for wh in warehouses_data.values())
    total_outcome_sum = sum(wh["outcome_sum"] for wh in warehouses_data.values())
    total_return_qty = sum(wh.get("return_qty", 0) for wh in warehouses_data.values())
    total_return_sum = sum(wh.get("return_sum", 0) for wh in warehouses_data.values())
    total_end_quantity = sum(wh["end_quantity"] for wh in warehouses_data.values())
    total_end_sum = total_end_quantity * product.wholesale_price
    
    product_image = (
        ProductImage.objects
        .filter(product=product)
        .only("image")
        .first()
    )

    image_url = product_image.image.url if product_image else None
                    
    data = {
        "product_id": product.id,
        "product_name": product.name,
        "product_unit": unit_name,
        "product_wholesale_price": product.wholesale_price,
        "product_retail_price": product.retail_price,
        "warehouses": warehouses_data,
        "start_quantity": total_start_quantity,
        "image": image_url,
        "turnover": {
            "income_qty": total_income_qty,
            "income_sum": total_income_sum,
            "outcome_qty": total_outcome_qty,
            "outcome_sum": total_outcome_sum,
            "return_qty": total_return_qty,
            "return_sum": total_return_sum,
        },
        "end": {
            "quantity": total_end_quantity,
            "sum": total_end_sum
        },
        "rows": all_rows,
    }
    
    return JsonResponse({"data": data})
                
    
    
    
    
    
    
    