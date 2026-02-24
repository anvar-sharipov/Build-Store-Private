from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime
from rest_framework.decorators import api_view, permission_classes
from ..models import Partner, Invoice, Transaction, DayClosing, StockSnapshot, WarehouseProduct, Entry, Account, Warehouse, FreeItemForInvoiceItem, UnitForInvoiceItem, Product, UnitOfMeasurement, Employee, ProductUnit, ProductImage, InvoiceItem, WarehouseAccount
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.views.decorators.csrf import csrf_exempt
from collections import defaultdict
from django.db.models import Sum, F, Q
from django.http import JsonResponse
from icecream import ic
from django.db import transaction as db_transaction
import pandas as pd
from datetime import date
from ..my_func.get_unit_map import get_unit_map 
from ..my_func.get_unit_and_cf import get_unit_and_cf 
from ..my_func.date_convert_for_excel import format_date_ru 
from ..my_func.date_str_to_dateFormat import date_str_to_dateFormat

from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_detail_entry(request):
    entry_id = request.GET.get("entry_id")

    if not entry_id:
        return HttpResponse("entry_id required", status=400)

    transaction_ = Transaction.objects.get(pk=entry_id)

    debit_acc = None
    debit_partner = None
    credit_acc = None
    credit_partner = None
    price = 0

    for i in transaction_.entries.all():
        if i.debit != 0:
            debit_acc = i.account.number
            price = i.debit
            if i.partner:
                debit_partner = i.partner.name

        if i.credit != 0:
            credit_acc = i.account.number
            price = i.credit
            if i.partner:
                credit_partner = i.partner.name

    # ===== EXCEL =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Детали операции"
    


    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    formatted_date = transaction_.date.strftime("%d.%m.%Y")

    data = [
        ("Дата:", formatted_date),
        ("Номер операции:", transaction_.id),
        ("Дебет счета:", debit_acc),
        ("Дебет субконто:", debit_partner),
        ("Кредит счет:", credit_acc),
        ("Кредит субконто:", credit_partner),
        ("Комментарий:", transaction_.description),
        ("Сумма:", float(price)),
    ]

    row = 1
    for label, value in data:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = bold
        ws[f"A{row}"].border = thin

        ws[f"B{row}"] = value
        ws[f"B{row}"].alignment = center
        ws[f"B{row}"].border = thin

        if label == "Сумма:":
            ws[f"B{row}"].number_format = '#,##0.00'

        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="entry_{transaction_.id}.xlsx"'

    wb.save(response)
    return response
