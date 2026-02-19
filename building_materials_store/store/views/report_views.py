
# from django.shortcuts import render
# import time
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
# from icecream import ic
# from rest_framework.decorators import action
# from rest_framework.filters import OrderingFilter
# from collections import OrderedDict
# from collections import defaultdict
from django.db.models import Value




# from rest_framework import viewsets, status, filters
from .. models import *
# from .serializers import *

from .. serializers.base_serializers import *
from .. serializers.entry_serializers import *
from .. serializers.partner_serializers import *
from .. serializers.product_serializers import *
from .. serializers.register_serializers import *
from .. serializers.sale_invoice_serializers import *

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
# from rest_framework.decorators import api_view, action

# from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
# from django.shortcuts import render, get_object_or_404
# from rest_framework.generics import CreateAPIView
# from django_filters.rest_framework import DjangoFilterBackend
# from .. filters import ProductFilter
# from django.views.decorators.http import require_GET
# from django.http import JsonResponse
from django.contrib.postgres.search import TrigramSimilarity
from django.db.models.functions import Greatest
from django.db.models import Q
# from django.utils.dateparse import parse_datetime, parse_date
# from django.db.models import Sum, F, Count
# from openpyxl.styles import Font
# from rest_framework.exceptions import PermissionDenied
# from django.db import transaction
# from datetime import datetime

# from rest_framework.pagination import PageNumberPagination

# from . base_views import IsInAdminOrWarehouseGroup, CustomPageNumberPagination



class ProductExportExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        ic("gg")
        try:
            # Получаем фильтры или product_ids
            filters = request.data.get("filters", {})
            product_ids = request.data.get("product_ids", [])
            
            warehouse_ids = []
            if product_ids:
                # Старый способ - по ID товаров (ограниченный пагинацией)
                queryset = Product.objects.filter(id__in=product_ids)
            elif filters:
                # Новый способ - применяем все фильтры для получения всех товаров
                queryset = Product.objects.all()
                print('filters', filters)
                # print('filters', filters)
                
                # Применяем поиск
                if 'search' in filters and filters['search']:
                    search_query = filters['search']
                    # queryset = queryset.filter(
                    #     Q(name__icontains=search_query) |
                    #     Q(sku__icontains=search_query) |
                    #     Q(category__name__icontains=search_query) |
                    #     Q(brand__name__icontains=search_query)
                    # )
                    queryset = queryset.annotate(
                        similarity=Greatest(
                            TrigramSimilarity('name', search_query),
                            TrigramSimilarity('sku', search_query),
                            TrigramSimilarity('category__name', search_query),
                            TrigramSimilarity('brand__name', search_query),
                        )
                    ).filter(similarity__gt=0.3).order_by('-similarity')
                
                # Применяем фильтр по категории
                if 'categories' in filters and filters['categories']:
                    queryset = queryset.filter(category_id=filters['categories'])
                
                # Применяем фильтр по бренду
                if 'brands' in filters and filters['brands']:
                    queryset = queryset.filter(brand_id=filters['brands'])
                
                # Применяем фильтр по модели
                if 'models' in filters and filters['models']:
                    queryset = queryset.filter(model_id=filters['models'])

                if 'tags' in filters and filters['tags']:
                    queryset = queryset.filter(model_id=filters['tags'])

                
                if 'warehouse' in filters and filters['warehouse']:
                    warehouses = filters.get('warehouse')
                    if warehouses:
                        warehouse_ids = [int(w_id) for w_id in warehouses.split(',') if w_id.isdigit()]
                        queryset = queryset.filter(warehouse_products__warehouse_id__in=warehouse_ids)

                


                if 'wholesale_price_max' in filters or "wholesale_price_min" in filters:
                    wholesale_max = filters.get('wholesale_price_max', None)
                    wholesale_min = filters.get('wholesale_price_min', None)

                    if wholesale_min is not None:
                        queryset = queryset.filter(wholesale_price__gte=wholesale_min)
                    if wholesale_max is not None:
                        queryset = queryset.filter(wholesale_price__lte=wholesale_max)


                if 'retail_price_max' in filters or "retail_price_min" in filters:
                    retail_max = filters.get('retail_price_max', None)
                    retail_min = filters.get('retail_price_min', None)

                    if retail_min is not None:
                        queryset = queryset.filter(retail_price__gte=retail_min)
                    if retail_max is not None:
                        queryset = queryset.filter(retail_price__lte=retail_max)


                if 'quantity_max' in filters or "quantity_min" in filters:
                    quantity_max = filters.get('quantity_max', None)
                    quantity_min = filters.get('quantity_min', None)

                    if quantity_min is not None:
                        queryset = queryset.filter(quantity__gte=quantity_min)
                    if quantity_max is not None:
                        queryset = queryset.filter(quantity__lte=quantity_max)


                if 'is_active' in filters:
                    if filters['is_active'] == 'true':
                        queryset = queryset.filter(is_active=True)
                    else:
                        queryset = queryset.filter(is_active=False)



                
                
                # Применяем другие фильтры по необходимости
                # Добавьте свои фильтры здесь
                
            else:
                # Если ничего не передано, экспортируем все товары
                queryset = Product.objects.all()

            # Оптимизируем запрос
            queryset = queryset.select_related('base_unit', 'category', 'brand', 'model').order_by('id')
            
            total_count = queryset.count()
            if total_count == 0:
                return Response({"error": "Товары не найдены"}, status=404)
            
            # Ограничиваем количество товаров для экспорта (защита от перегрузки)
            MAX_EXPORT_LIMIT = 10000  # Максимум 10,000 товаров
            if total_count > MAX_EXPORT_LIMIT:
                return Response({
                    "error": f"Слишком много товаров для экспорта. Максимум: {MAX_EXPORT_LIMIT}, найдено: {total_count}"
                }, status=400)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Товары"

            headers = [
                "№","ID", "Наименование", "Категория", "Ед. изм.", "Артикул", "Количество",
                "Цена закупки", "Розничная цена", "Оптовая цена", "Цена со скидкой",
                "Бренд", "Модель", "Вес (кг)", "Объём (м³)", "Длина (см)", "Ширина (см)", "Высота (см)", "Активен"
            ]

            if warehouse_ids:
                warehouse_names = list(
                    Warehouse.objects.filter(id__in=warehouse_ids).values_list('name', flat=True)
                )
                ws.append(["Склады:"] + warehouse_names)  # Первая строка Excel
                ws.append([])  # Пустая строка для отступа
            else:
                ws.append(["Склады: Все"])  # Первая строка Excel
                ws.append([])  # Пустая строка для отступа


            ws.append(headers)

            # Обрабатываем товары партиями для больших объемов
            batch_size = 1000
            row_number = 1
            for i in range(0, total_count, batch_size):
                batch_queryset = queryset[i:i + batch_size]
                
                for product in batch_queryset:
                    if warehouse_ids:
                        quantity = product.warehouse_products.filter(
                            warehouse_id__in=warehouse_ids
                        ).aggregate(total=models.Sum('quantity'))['total'] or 0
                    else:
                        quantity = product.get_total_quantity()

                   
                    unit_name = product.base_unit.name if product.base_unit else ""
                    for unit in product.units.all():
                        if unit.is_default_for_sale and unit.conversion_factor:
                            quantity = quantity / unit.conversion_factor
                            unit_name = unit.unit.name
                            break

                    is_active = "+" if product.is_active else ""
                    ws.append([
                        row_number,
                        product.id,
                        product.name,
                        product.category.name if product.category else "",
                        # product.base_unit.name if product.base_unit else "",
                        unit_name,
                        product.sku or "",
                        # float(product.quantity) if product.quantity else 0,
                        # float(product.total_quantity) if hasattr(product, 'total_quantity') and product.total_quantity else 0,
                        float(quantity),
                        float(product.purchase_price) if product.purchase_price else 0,
                        float(product.retail_price) if product.retail_price else 0,
                        float(product.wholesale_price) if product.wholesale_price else 0,
                        float(product.discount_price) if product.discount_price else 0,
                        product.brand.name if product.brand else "",
                        product.model.name if product.model else "",
                        float(product.weight) if product.weight else 0,
                        float(product.volume) if product.volume else 0,
                        product.length,
                        product.width,
                        product.height,
                        is_active,
                    ])
                    row_number += 1

            # Автоматическая ширина колонок
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="products_export_{total_count}_items.xlsx"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({"error": f"Внутренняя ошибка сервера: {str(e)}"}, status=500)