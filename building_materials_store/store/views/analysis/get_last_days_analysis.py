from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta
from rest_framework.decorators import api_view, permission_classes
from ...models import Invoice, InvoiceItem, ProductUnit, Product, StockSnapshot, ProductImage, Warehouse, FreeProduct, WarehouseProduct, Brand, Category
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.views.decorators.csrf import csrf_exempt
from collections import defaultdict
from django.http import JsonResponse, HttpResponse
from icecream import ic
from django.db import transaction as db_transaction
import pandas as pd
from datetime import date
from django.shortcuts import get_object_or_404
import time
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Q, Case, When, Value, IntegerField, Count, Prefetch, Exists, OuterRef, Min
from rest_framework import status
from django.db.models.functions import Coalesce
from django.utils import timezone









from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from django.http import FileResponse
# import os
# from io import BytesIO
# from django.core.files.base import ContentFile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import CellIsRule
# from openpyxl.worksheet.page import PageMargins

from rest_framework import status
from django.db.models.functions import Coalesce
from openpyxl.worksheet.page import PageMargins
from collections import defaultdict
from django.contrib.postgres.aggregates import ArrayAgg
import statistics



def apply_abcxyz_priority(products):
    ABCXYZ_PRIORITY = {
        "AX": 1,
        "AY": 2,
        "AZ": 3,
        "BX": 4,
        "BY": 5,
        "BZ": 6,
        "CX": 7,
        "CY": 8,
        "CZ": 9,
    }
    for p in products:
        cat = p.get("abcxyz_category")
        p["abcxyz_priority"] = ABCXYZ_PRIORITY.get(cat, 99)

def apply_xyz_analysis(products, product_daily_sales):
    for p in products:
        product_id = p["product_id"]
        sales = product_daily_sales.get(product_id, [])
        
        # берём только дни с продажами
        sales_nonzero = [s for s in sales if s > 0]

        if not sales_nonzero:
            p["xyz_category"] = "Z"
            continue
        
       

        # но учитываем частоту продаж (сколько дней из всех было продажа)
        total_days = len(sales)
        selling_days = len(sales_nonzero)
        frequency = selling_days / total_days  # от 0 до 1

        avg = sum(sales_nonzero) / selling_days
        std = statistics.pstdev(sales_nonzero) if selling_days > 1 else 0
        variation = std / avg if avg > 0 else 999

        # X: стабильный объём И часто продаётся
        if variation <= 1 and frequency >= 0.10:
            xyz = "X"
        # Y: средняя стабильность ИЛИ редко продаётся
        elif variation <= 2 and frequency >= 0.03:
            xyz = "Y"
        else:
            xyz = "Z"

        p["xyz_category"] = xyz
        


def apply_abc_analysis(products):
    
    # total_sales = sum(p["total_sold_price"] for p in products)
    total_sales = sum((p["total_sold_price"] for p in products), Decimal("0"))

    products_sorted = sorted(
        products,
        key=lambda x: x["total_sold_price"],
        reverse=True
    )

    cumulative = Decimal("0")

    for p in products_sorted:

        # share = p["total_sold_price"] / total_sales if total_sales else Decimal("0")
        if total_sales > 0:
            share = p["total_sold_price"] / total_sales
        else:
            share = Decimal("0")
            
        cumulative += share

        if cumulative <= Decimal("0.80"):
            abc = "A"
        elif cumulative <= Decimal("0.95"):
            abc = "B"
        else:
            abc = "C"

        p["abc_category"] = abc

    return products_sorted


def apply_abcxyz(products):
    for p in products:
        abc = p.get("abc_category", "")
        xyz = p.get("xyz_category", "")
        p["abcxyz_category"] = f"{abc}{xyz}"

def get_result(date_from, date_to, warehouse_ids, last_days, sort_brend, brand_ids, category_ids, dont_show_zero, new_product_days_count, full_list, sort_field=None, sort_direction="desc"):
    
    test_ab28 = WarehouseProduct.objects.filter(product__name = "AB-60, BOLOR-646 RASTWORITEL 0.5L (20)")
    
    # ic(test_ab28)
    for i in test_ab28:
        ic(i.product.name)
        ic(i.quantity)
    
    if sort_field == "abcxyz_category":
        sort_field = "abcxyz_priority"
    
    # ic(sort_brend)
    products_qs = (
        InvoiceItem.objects.filter(
            invoice__wozwrat_or_prihod="rashod",
            invoice__is_entry=True,
            invoice__canceled_at__isnull=True,
            invoice__invoice_date__range=(date_from, date_to)
        )
    )
    
    

    # data = (
    #     InvoiceItem.objects.filter(
    #         invoice__wozwrat_or_prihod="rashod",
    #         invoice__is_entry=True,
    #         invoice__canceled_at__isnull=True,
    #         invoice__invoice_date__range=(date_from, date_to)
    #     )
    #     .values("invoice__partner__name")
    #     .annotate(
    #         faktura_count=Count("invoice", distinct=True),
    #         selected_price=Sum("selected_price"),
    #         selected_quantity=Sum("selected_quantity"),
    #         agents=ArrayAgg("invoice__partner__agent__name", distinct=True),
    #     )
    # )
    
    sold_product_ids = set(
        products_qs.values_list("product_id", flat=True)
    )
    unsold_products = Product.objects.exclude(id__in=sold_product_ids)
    
    # ic("tut3")
    
    
    
    
    if warehouse_ids:
        products_qs = products_qs.filter(
            invoice__warehouse_id__in=warehouse_ids
        )
        
        # data = data.filter(
        #     invoice__warehouse_id__in=warehouse_ids
        # )
        
    if brand_ids:
        products_qs = products_qs.filter(
            product__brand_id__in=brand_ids
        )
        unsold_products = unsold_products.filter(brand_id__in=brand_ids)
        
        # data = data.filter(
        #     product__brand_id__in=brand_ids
        # )
        
    if category_ids:
        products_qs = products_qs.filter(
            product__category_id__in=category_ids
        )
        unsold_products = unsold_products.filter(category_id__in=category_ids)
        
        # data = data.filter(
        #     product__category_id__in=category_ids
        # )
        
    # partners_dict = {
    #     row["invoice__partner__name"]: {
    #         "faktura_count": row["faktura_count"],
    #         "agents": row["agents"] or [],
    #         "selected_price": row["selected_price"] or 0,
    #         "selected_quantity": row["selected_quantity"] or 0,
    #     }
    #     for row in data
    # }

    # ic(partners_dict)
    # ic("tut1")
    # --- продажи по дням для XYZ анализа START ---
    daily_sales_qs = (
        InvoiceItem.objects.filter(
            invoice__wozwrat_or_prihod="rashod",
            invoice__is_entry=True,
            invoice__canceled_at__isnull=True,
            invoice__invoice_date__range=(date_from, date_to)
        )
    )
    
    # ic("tut2")

    if warehouse_ids:
        daily_sales_qs = daily_sales_qs.filter(invoice__warehouse_id__in=warehouse_ids)

    if brand_ids:
        daily_sales_qs = daily_sales_qs.filter(product__brand_id__in=brand_ids)

    if category_ids:
        daily_sales_qs = daily_sales_qs.filter(product__category_id__in=category_ids)


    daily_sales_qs = (
        daily_sales_qs
        .values("product_id", "invoice__invoice_date")
        .annotate(qty=Sum("selected_quantity"))
    )
    
    # product_daily_sales = defaultdict(list)
    # for row in daily_sales_qs:
    #     product_daily_sales[row["product_id"]].append(
    #         float(row["qty"])
    #     )
    
    product_daily_sales = defaultdict(lambda: defaultdict(float))
    
    for row in daily_sales_qs:
        product_id = row["product_id"]
        date = row["invoice__invoice_date"].date()

        product_daily_sales[product_id][date] += float(row["qty"])
        
    all_dates = [
        (date_from + timedelta(days=i)).date()
        for i in range(last_days)
    ]
    
    # ic("tut4")
        
    product_daily_sales_list = {}

    for product_id, sales_by_date in product_daily_sales.items():

        sales_list = []

        for d in all_dates:
            sales_list.append(sales_by_date.get(d, 0))

        product_daily_sales_list[product_id] = sales_list
    
    # --- продажи по дням для XYZ анализа END ---
        
    products_qs = (
        products_qs
        .values("product_id",
                "product__category_id",
                "product__category__name",
                "product__brand__name",
                )
        .annotate(
            product_name=F("product__name"),
            category_name=F("product__category__name"),
            brand_name=F("product__brand__name"),
            purchase_price=F("product__purchase_price"),
            wholesale_price=F("product__wholesale_price"),
            unit=F("product__base_unit__name"),
            total_sold=Sum("selected_quantity"),
            # invoices_count=Count("invoice", distinct=True),
            partners_count=Count("invoice__partner", distinct=True),
            # partners_names=ArrayAgg("invoice__partner__name", distinct=True, default=[]),
        )
    )
    
    
    
    
    
    result = defaultdict(list)
    
    all_products = []
    
    # --- Получаем остатки всех товаров одним запросом ---
    stock_qs = WarehouseProduct.objects.all()
    
    if warehouse_ids:
        stock_qs = stock_qs.filter(warehouse_id__in=warehouse_ids)
        
    stock_data = (
        stock_qs
        .values("product_id")
        .annotate(
            total=Coalesce(
                Sum("quantity"),
                Decimal("0.00"),
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )
    )
    
    stock_map = {row["product_id"]: row["total"] for row in stock_data}
    
    # result = []
    
    partners_data = (
        products_qs
        .values(
            "product_id",
            "invoice__partner__name",
        )
        .annotate(
            faktura_count=Count("invoice", distinct=True),
            # total_sum=Sum("selected_price"),
            total_sum=Sum(F("selected_price") * F("selected_quantity")),
            # total_sum=Sum(F("selected_price") * F("product__wholesale_price")),
            wholesale_price=F("product__wholesale_price"),
            selected_quantity=Sum("selected_quantity"),
            agents=ArrayAgg("invoice__partner__agent__name", distinct=True),
        )
    )
    
    partners_by_product = defaultdict(dict)

    for row in partners_data:
        product_id = row["product_id"]
        partner = row["invoice__partner__name"]

        partners_by_product[product_id][partner] = {
            "faktura_count": row["faktura_count"],
            "agents": row["agents"] or [],
            "wholesale_price": row["wholesale_price"] or Decimal("0"),
            "total_sum": row["total_sum"] or Decimal("0"),
            "selected_quantity": row["selected_quantity"] or Decimal("0"),
        }
    
    # ic(partners_by_product)
        
    product_ids = [p["product_id"] for p in products_qs]
    
    first_prihod_qs = (
        InvoiceItem.objects.filter(
            invoice__wozwrat_or_prihod="prihod",
            invoice__is_entry=True,
            invoice__canceled_at__isnull=True,
            product_id__in=product_ids
        )
        .values("product_id")
        .annotate(first_prihod_date=Min("invoice__invoice_date__date"))
    )
    first_prihod_map = {
        row["product_id"]: row["first_prihod_date"]
        for row in first_prihod_qs
    }
    
    today = timezone.now().date()



    images_qs = (
        ProductImage.objects
        .filter(product_id__in=product_ids)
        .values("product_id", "image")
    )

    images_by_product = defaultdict(list)

    for img in images_qs:
        images_by_product[img["product_id"]].append(img["image"])
    
    for p in products_qs:
        
        first_prihod_date = first_prihod_map.get(p["product_id"])
        
        product_age_days = None

        if first_prihod_date:
            product_age_days = (today - first_prihod_date).days
            
        # ic(p["product_name"])
        # ic(product_age_days)
        
        total_sold = p["total_sold"]
        avg_per_day = total_sold / Decimal(last_days) if last_days else Decimal("0")
        stock_qty = stock_map.get(p["product_id"], Decimal("0.00"))
        
        # --- Прогноз ---
        forecast = avg_per_day * Decimal(last_days)
        
        need_to_buy = forecast - stock_qty
        # if need_to_buy < 0:
        #     need_to_buy = Decimal(0)
            
        turnover = (
            total_sold / stock_qty
            if stock_qty and stock_qty > 0
            else 0
        )
        
        # category = p["category_name"] or "Без категории"
        if sort_brend:
            category = p["brand_name"] or "Без бренда"
        else:
            category = p["category_name"] or "Без категории"
            
        # if avg_per_day > 0:
        #     days_of_stock_remaining = stock_qty / avg_per_day
        # else:
        #     days_of_stock_remaining = Decimal("0")
        
        if avg_per_day > 0:
            # days_of_stock_remaining = round(stock_qty / avg_per_day, 2)
            days_of_stock_remaining = (
                (stock_qty / avg_per_day)
                .quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            )
        else:
            # days_of_stock_remaining = Decimal("999")
            days_of_stock_remaining = Decimal("0")
            
        product_data = {
            "product_id": p["product_id"],
            "product_name": p["product_name"],
            "purchase_price": p["purchase_price"],
            "wholesale_price": p["wholesale_price"],
            "unit": p["unit"],
            "total_sold": total_sold,
            "total_sold_price": total_sold * p["purchase_price"],
            "avg_per_day": avg_per_day,
            "avg_per_day_price": avg_per_day * p["purchase_price"],
            "partners_count": p["partners_count"],
            "partners": partners_by_product.get(p["product_id"], {}),
            "stock_qty": stock_qty,
            "stock_qty_price": stock_qty * p["purchase_price"],
            "days_of_stock_remaining": days_of_stock_remaining,
            "turnover": turnover,
            "turnover_price": turnover * p["purchase_price"],
            "forecast": forecast,
            "forecast_price": forecast * p["purchase_price"],
            "need_to_buy": need_to_buy,
            "need_to_buy_price": need_to_buy * p["purchase_price"],
            "images": images_by_product.get(p["product_id"], []),
            "product_age_days": product_age_days,
            "category_name": p["category_name"],
            "brand_name": p["brand_name"],
        }
        
        result[category].append(product_data)
        all_products.append(product_data)
        
        # result[category].append({
        #     "product_id": p["product_id"],
        #     "product_name": p["product_name"],
        #     "purchase_price": p["purchase_price"],
        #     "wholesale_price": p["wholesale_price"],
        #     "unit": p["unit"],
        #     "total_sold": total_sold,
        #     "total_sold_price": total_sold * p["purchase_price"],
        #     "avg_per_day": avg_per_day,
        #     "avg_per_day_price": avg_per_day * p["purchase_price"],
        #     # "invoices_count": p["invoices_count"],
        #     "partners_count": p["partners_count"],
        #     "partners": partners_by_product.get(p["product_id"], {}),
        #     # "partners_names": p["partners_names"],
        #     "stock_qty": stock_qty,
        #     "stock_qty_price": stock_qty * p["purchase_price"],
        #     "days_of_stock_remaining": days_of_stock_remaining,
        #     # "days_of_stock_remaining_price": days_of_stock_remaining * p["purchase_price"],
        #     "turnover": turnover,
        #     "turnover_price": turnover * p["purchase_price"],
        #     "forecast": forecast,
        #     "forecast_price": forecast * p["purchase_price"],
        #     "need_to_buy": need_to_buy,
        #     "need_to_buy_price": need_to_buy * p["purchase_price"],
        #     "images": images_by_product.get(p["product_id"], []),
        # })
   
    if dont_show_zero:
        for p in unsold_products.values(
            "id",
            "name",
            "purchase_price",
            "wholesale_price",
            "base_unit__name",
            "category__name",
            "brand__name"
        ):

            product_id = p["id"]

            stock_qty = stock_map.get(product_id, Decimal("0.00"))

            if sort_brend:
                category = p["brand__name"] or "Без бренда"
            else:
                category = p["category__name"] or "Без категории"
                
            first_prihod_date = first_prihod_map.get(product_id)

            product_age_days = None
            if first_prihod_date:
                product_age_days = (today - first_prihod_date).days

            product_data = {
                "product_id": product_id,
                "product_name": p["name"],
                "purchase_price": p["purchase_price"],
                "wholesale_price": p["wholesale_price"],
                "unit": p["base_unit__name"],
                "total_sold": Decimal("0"),
                "total_sold_price": Decimal("0"),
                "avg_per_day": Decimal("0"),
                "avg_per_day_price": Decimal("0"),
                "partners_count": 0,
                "partners": {},
                "stock_qty": stock_qty,
                "stock_qty_price": stock_qty * p["purchase_price"],
                # "days_of_stock_remaining": Decimal("999"),
                "days_of_stock_remaining": Decimal("0"),
                "turnover": Decimal("0"),
                "turnover_price": Decimal("0"),
                "forecast": Decimal("0"),
                "forecast_price": Decimal("0"),
                "need_to_buy": Decimal("0"),
                "need_to_buy_price": Decimal("0"),
                "images": images_by_product.get(product_id, []),
                "product_age_days": product_age_days,
                "category_name": p["category__name"],
                "brand_name": p["brand__name"],
            }

            result[category].append(product_data)
            all_products.append(product_data)
    
    # for category in result:
    #     result[category] = sorted(result[category], key=lambda x: x["need_to_buy"], reverse=True)
    
    apply_abc_analysis(all_products)
    
    # apply_xyz_analysis(all_products, product_daily_sales)
    apply_xyz_analysis(all_products, product_daily_sales_list)
    
    apply_abcxyz(all_products)
    apply_abcxyz_priority(all_products)
    
    grand_totals = {
            "total_sold": Decimal("0"),
            "total_sold_price": Decimal("0"),
            "avg_per_day": Decimal("0"),
            "avg_per_day_price": Decimal("0"),
            "stock_qty": Decimal("0"),
            "stock_qty_price": Decimal("0"),
            "turnover": Decimal("0"),
            "turnover_price": Decimal("0"),
            "forecast": Decimal("0"),
            "forecast_price": Decimal("0"),
            "need_to_buy": Decimal("0"),
            "need_to_buy_price": Decimal("0"),
        }
    
    if full_list:
        

        
            
        for p in all_products:
            grand_totals["total_sold"] += p["total_sold"]
            grand_totals["total_sold_price"] += p["total_sold_price"]
            grand_totals["avg_per_day"] += p["avg_per_day"]
            grand_totals["avg_per_day_price"] += p["avg_per_day_price"]
            grand_totals["stock_qty"] += p["stock_qty"]
            grand_totals["stock_qty_price"] += p["stock_qty_price"]
            grand_totals["forecast"] += p["forecast"]
            grand_totals["forecast_price"] += p["forecast_price"]
            
            if p["need_to_buy"] > 0:
                grand_totals["need_to_buy"] += p["need_to_buy"]
                grand_totals["need_to_buy_price"] += p["need_to_buy_price"]
                
        total_sold_all = grand_totals["total_sold_price"]
        
        # udelnyy wes
        for p in all_products:
            
            if total_sold_all > 0:
                p["share"] = (
                    p["total_sold_price"] / total_sold_all * Decimal("100")
                ).quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP)
            else:
                p["share"] = Decimal("0")
                
        if sort_field in [
            "total_sold", "total_sold_price", "stock_qty", "stock_qty_price",
            "avg_per_day", "avg_per_day_price", "abcxyz_priority",
            "product_age_days", "days_of_stock_remaining",
            "forecast", "forecast_price", "need_to_buy",
            "need_to_buy_price", "partners_count", "purchase_price", "share"
        ]:

            reverse = sort_direction == "desc"

            all_products = sorted(
                all_products,
                key=lambda x: (x.get(sort_field) is None, x.get(sort_field) or 0),
                reverse=reverse
            )
                
            
        # ic(all_products[:10])
        return {
            "result_list": all_products,
            "grand_totals": grand_totals,
            "is_full_list": full_list
        }
    else:
    
    
        for category in result:

            if sort_field in ["total_sold", "total_sold_price", "stock_qty", "stock_qty_price", "avg_per_day", "avg_per_day_price", "abcxyz_priority", "product_age_days",
                            "days_of_stock_remaining", "forecast", "forecast_price", "need_to_buy", "need_to_buy_price", "partners_count","purchase_price"]:
                reverse = sort_direction == "desc"

                result[category] = sorted(
                    result[category],
                    # key=lambda x: x.get(sort_field, 0),
                    key=lambda x: (x.get(sort_field) is None, x.get(sort_field) or 0),
                    reverse=reverse
                )
                

            else:
                # дефолтная сортировка
                result[category] = sorted(
                    result[category],
                    key=lambda x: x["need_to_buy"],
                    reverse=True
                )
            
        # return result
        # return dict(result)
        result_list = []
        
        total_product_count = 0
        
        
        
        for category, products in result.items():
            category_totals = {
                "total_sold": Decimal("0"),
                "total_sold_price": Decimal("0"),
                "avg_per_day": Decimal("0"),
                "avg_per_day_price": Decimal("0"),
                "stock_qty": Decimal("0"),
                "stock_qty_price": Decimal("0"),
                "turnover": Decimal("0"),
                "turnover_price": Decimal("0"),
                "forecast": Decimal("0"),
                "forecast_price": Decimal("0"),
                "need_to_buy": Decimal("0"),
                "need_to_buy_price": Decimal("0"),
            }
            
            for p in products:
                total_product_count += 1
                
                category_totals["total_sold"] += p["total_sold"]
                category_totals["total_sold_price"] += p["total_sold_price"]
                # category_totals["avg_per_day"] += p["avg_per_day"]
                category_totals["avg_per_day_price"] += p["avg_per_day_price"]
                category_totals["stock_qty"] += p["stock_qty"]
                category_totals["stock_qty_price"] += p["stock_qty_price"]
                # category_totals["turnover"] += p["turnover"]
                # category_totals["turnover_price"] += p["turnover_price"]
                category_totals["forecast"] += p["forecast"]
                category_totals["forecast_price"] += p["forecast_price"]
                
                if p["need_to_buy"] > 0:
                    category_totals["need_to_buy"] += p["need_to_buy"]
                    category_totals["need_to_buy_price"] += p["need_to_buy_price"]
                
            if category_totals["stock_qty"] > 0:
                category_totals["turnover"] = (
                    category_totals["total_sold"] /
                    category_totals["stock_qty"]
                )
            else:
                category_totals["turnover"] = Decimal("0")

            category_totals["turnover_price"] = (
                category_totals["turnover"] *
                category_totals["total_sold_price"]
            )
            
            category_totals["avg_per_day"] = (
                category_totals["total_sold"] / Decimal(last_days)
                if last_days
                else Decimal("0")
            )
            
            # ic(products)
                
            result_list.append({
                "category": category,
                "products": products,
                "category_totals": category_totals
            })
            
            grand_totals["total_sold"] += category_totals["total_sold"]
            grand_totals["total_sold_price"] += category_totals["total_sold_price"]
            grand_totals["avg_per_day_price"] += category_totals["avg_per_day_price"]
            grand_totals["stock_qty"] += category_totals["stock_qty"]
            grand_totals["stock_qty_price"] += category_totals["stock_qty_price"]
            grand_totals["forecast"] += category_totals["forecast"]
            grand_totals["forecast_price"] += category_totals["forecast_price"]
            grand_totals["need_to_buy"] += category_totals["need_to_buy"]
            grand_totals["need_to_buy_price"] += category_totals["need_to_buy_price"]
            
        if grand_totals["stock_qty"] > 0:
            grand_totals["turnover"] = (
                grand_totals["total_sold"] /
                grand_totals["stock_qty"]
            )
        else:
            grand_totals["turnover"] = Decimal("0")

        grand_totals["turnover_price"] = (
            grand_totals["turnover"] *
            grand_totals["total_sold_price"]
        )

        grand_totals["avg_per_day"] = (
            grand_totals["total_sold"] / Decimal(last_days)
            if last_days
            else Decimal("0")
        )
        
        total_sold_all = grand_totals["total_sold_price"]
        
        for category_data in result_list:
            for p in category_data["products"]:
                if total_sold_all > 0:
                    p["share"] = (
                        p["total_sold_price"] / total_sold_all * Decimal("100")
                    ).quantize(Decimal("0.00001"), rounding=ROUND_HALF_UP)
                else:
                    p["share"] = Decimal("0")
                    
        # сортировка по share
        if sort_field == "share":
            reverse = sort_direction == "desc"

            for category_data in result_list:
                category_data["products"] = sorted(
                    category_data["products"],
                    key=lambda x: x.get("share", 0),
                    reverse=reverse
                )
            

        # --- перенос "Без категории" / "Без бренда" вниз ---
        if sort_brend:
            empty_name = "Без бренда"
        else:
            empty_name = "Без категории"

        result_list = sorted(
            result_list,
            key=lambda x: (x["category"] == empty_name, x["category"])
        )
        grand_totals["total_product_count"] = total_product_count
        # return result_list
        return {
            "result_list": result_list,
            "grand_totals": grand_totals,
            "is_full_list": full_list
        }
        

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_last_days_analysis(request):
    ic("get_last_days_analysis")
    

    last_days = int(request.GET.get("lastDaysCount", 0)) 
    new_product_days_count = int(request.GET.get("newProductDaysCount", 0))
    warehouses = request.GET.get("warehouses", "")
    sort_brend = request.GET.get("sortBrend", "") == "true" 
    full_list = request.GET.get("fullList", "") == "true" 
    

    
    brands = request.GET.get("brands", "")
    categories = request.GET.get("categories", "") 
    
    dont_show_zero = request.GET.get("dontShowZero", "") == "true"

              
    sort_field = request.GET.get("sortField", "")
    sort_direction = request.GET.get("sortDirection", "")
    
    
    if last_days <= 0:
        return Response({"error": "Invalid period"}, status=400)
    
    warehouse_ids = []
    if warehouses:
        warehouse_ids = [int(w) for w in warehouses.split(",")]
        
    brand_ids = []
    if brands:
        brand_ids = [int(b) for b in brands.split(",")]
        
    category_ids = []
    if categories:
        category_ids = [int(c) for c in categories.split(",")]
        
    date_to = timezone.now()
    date_from = date_to - timedelta(days=last_days)
    
    # ic(date_from, date_to)
    
    result = get_result(date_from, date_to, warehouse_ids, last_days, sort_brend, brand_ids, category_ids, dont_show_zero, new_product_days_count, full_list, sort_field, sort_direction)
    
    return Response(result)




CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_VERTICAL = Alignment(vertical="center")

RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="CFE2F3")

CATEGORY_FILL = PatternFill("solid", fgColor="E9ECEF")

RED_FILL = PatternFill("solid", fgColor="F8D7DA")     # скидка
GREEN_FILL = PatternFill("solid", fgColor="D1E7DD")   # наценка

GRAY_FILL_0 = PatternFill(fill_type="solid", fgColor="F5F5F5")
GRAY_FILL_1 = PatternFill(fill_type="solid", fgColor="EDEDED")
GRAY_FILL_2 = PatternFill( fill_type="solid", fgColor="DCDCDC")
GRAY_FILL_3 = PatternFill(fill_type="solid", fgColor="C8C8C8")
GRAY_FILL_4 = PatternFill(fill_type="solid", fgColor="B0B0B0")

GREEN_FILL_0 = PatternFill(fill_type="solid", fgColor="E2F0D9")
GREEN_FILL_1 = PatternFill(fill_type="solid", fgColor="C6EFCE")
GREEN_FILL_2 = PatternFill(fill_type="solid", fgColor="92D050")
GREEN_FILL_3 = PatternFill(fill_type="solid", fgColor="006100")

ABCXYZ_COLORS = {
    "AX": GREEN_FILL_3,
    "AY": GREEN_FILL_2,
    "AZ": GREEN_FILL_1,

    "BX": GRAY_FILL_2,
    "BY": GRAY_FILL_1,
    "BZ": GRAY_FILL_0,

    "CX": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "CY": PatternFill(fill_type="solid", fgColor="FCE5CD"),
    "CZ": PatternFill(fill_type="solid", fgColor="F8D7DA"),
}

RED_FONT = Font(color="FF0000")
GREEN_FONT = Font(color="006400")
BLUE_FONT = Font(color="0000FF")
YELLOW_FONT = Font(color="B8860B")

PRICE_FMT = '#,##0.00'
QTY_FMT = '#,##0.###'


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_last_days_analysis_excel(request):
    ic("get_last_days_analysis_excel")
    
    def get_abcxyz_recommendation(code):

        recommendations = {
            "AX": "🔥 Держать большой запас",
            "AY": "✔ Держать запас",
            "AZ": "📦 Средний запас",

            "BX": "📊 Контроль продаж",
            "BY": "⚠ Уменьшить запас",
            "BZ": "⬇ Сокращать закупки",

            "CX": "📦 Минимальный запас",
            "CY": "📉 Под заказ",
            "CZ": "❌ Убрать / распродать",
        }

        return recommendations.get(code, "")
    

    last_days = int(request.GET.get("lastDaysCount", 0))
    new_product_days_count = int(request.GET.get("newProductDaysCount", 0))
    warehouses = request.GET.get("warehouses", "")
    
    sort_brend = request.GET.get("sortBrend", "") == "true"
    
    full_list = request.GET.get("fullList", "") == "true" 
    
    brands = request.GET.get("brands", "")
    categories = request.GET.get("categories", "")
    
    dont_show_zero = request.GET.get("dontShowZero", "") == "true"
    
    sort_field = request.GET.get("sortField", "")
    sort_direction = request.GET.get("sortDirection", "")
    
    if last_days <= 0:
        return Response({"error": "Invalid period"}, status=400)
    
    warehouse_ids = []
    if warehouses:
        warehouse_ids = [int(w) for w in warehouses.split(",")]
        
        
    brand_ids = []
    if brands:
        brand_ids = [int(b) for b in brands.split(",")]
        
 
    category_ids = []
    if categories:
        category_ids = [int(c) for c in categories.split(",")]
        

    date_to = timezone.now()
    date_from = date_to - timedelta(days=last_days)
    
    date_from_str = date_from.strftime("%d.%m.%Y")
    date_to_str = date_to.strftime("%d.%m.%Y")
    
    # ic(date_from, date_to)
    
    result = get_result(date_from, date_to, warehouse_ids, last_days, sort_brend, brand_ids, category_ids, dont_show_zero, new_product_days_count, full_list, sort_field, sort_direction)

    
    wb = Workbook()
    ws = wb.active
    ws.title = "Aнализ продаж"
    
    
    if full_list:
        ws.merge_cells("A1:W1")
    else:
        ws.merge_cells("A1:U1")
        
    ws["A1"] = f"Aнализ продаж"
    ws["A1"].alignment = CENTER
    ws["A1"].font = Font(size=14, bold=True)
    ws.row_dimensions[1].height = 28
    
    if full_list:
        ws.merge_cells("A2:W2")
    else:
        ws.merge_cells("A2:U2")
    ws["A2"] = f"За последние {str(last_days)} дней с {date_from_str} по {date_to_str}"
    ws["A2"].alignment = CENTER
    ws["A2"].font = Font(size=12)
    
    if full_list:
        ws.merge_cells("A3:W3")
    else:
        ws.merge_cells("A3:U3")
        
    if warehouse_ids:
        # warehouses_name = ""
        # for w in warehouse_ids:
        #     ware = Warehouse.objects.get(pk=w)
        
        #     warehouses_name += f"{ware.name}, "
        warehouses_name = ", ".join(
            Warehouse.objects.filter(id__in=warehouse_ids)
            .values_list("name", flat=True)
        )
    else:
        warehouses_name = "Все"
    
    ws["A3"] = f"Склад: {warehouses_name}"
    
    row = 4
    
    
    if brand_ids:
        if full_list:
            ws.merge_cells(f"A{row}:W{row}")
        else:
            ws.merge_cells(f"A{row}:U{row}")
            
        brands_name = ", ".join(
            Brand.objects.filter(id__in=brand_ids)
            .values_list("name", flat=True)
        )
  
        ws[f"A{row}"] = f"Бренд: {brands_name}"
        
        row += 1
        
    if category_ids:
        if full_list:
            ws.merge_cells(f"A{row}:W{row}")
        else:
            ws.merge_cells(f"A{row}:U{row}")

        categories_name = ", ".join(
            Category.objects.filter(id__in=category_ids)
            .values_list("name", flat=True)
        )
  
        ws[f"A{row}"] = f"Категория: {categories_name}"
        
        row += 1
        
        
    ws.freeze_panes = f"C{row+2}"
    
    
    # ===== Шапка =====

    
    
    ws.row_dimensions[row].height = 50
    
    
    ws[f"A{row}"] = f"№"
    
    ws[f"B{row}"] = f"Наименования товара" 
    
    
    ws[f"C{row}"] = f"цена закупки за штуку"
    ws[f"D{row}"] = f"ед. изм."
    ws[f"E{row}"] = f"Возраст товара (дни)"
    ws[f"F{row}"] = f"ABCXYZ"
    
    if full_list:
        ws[f"G{row}"] = f"Категория"
        ws[f"H{row}"] = f"Бренд"
        ws[f"I{row}"] = f"Клиент"
        ws.merge_cells(f"J{row}:K{row}")
        ws[f"J{row}"] = f"Продано"
        ws.merge_cells(f"L{row}:M{row}")
        ws[f"L{row}"] = f"Средняя продажа в день"
        ws.merge_cells(f"N{row}:O{row}") 
        ws[f"N{row}"] = f"Остаток на складе"
        ws[f"P{row}"] = f"Остаток дней запаса"
        ws.merge_cells(f"Q{row}:R{row}")
        ws[f"Q{row}"] = f"Прогноз на следующие {str(last_days)} дней"
        ws.merge_cells(f"S{row}:T{row}")
        ws[f"S{row}"] = f"Нужно закупить"
        ws[f"U{row}"] = "Удельный вес %"
        ws[f"V{row}"] = "Накопительный %"
        ws[f"W{row}"] = "Рекомендация"
        
        
        for i in "ABCDEFGHIJKLMNOPQRSTUVW":
            ws[f"{i}{row}"].font = Font(bold=True)
            ws[f"{i}{row}"].alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            ws[f"{i}{row}"].fill = GRAY_FILL_1
            ws[f"{i}{row}"].border = THIN_BORDER
        row += 1
        headers2 = [
        "",              
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "кол-во",
        "цена",
        "кол-во",
        "цена",
        "кол-во",
        "цена",
        "",
        "кол-во",
        "цена",
        "кол-во",
        "цена",
        "",    
        "",    
        "",    
    ]
    else:
        ws[f"G{row}"] = f"Клиент"
        ws.merge_cells(f"H{row}:I{row}")
        ws[f"H{row}"] = f"Продано"
        ws.merge_cells(f"J{row}:K{row}")
        ws[f"J{row}"] = f"Средняя продажа в день"
        ws.merge_cells(f"L{row}:M{row}") 
        ws[f"L{row}"] = f"Остаток на складе"
        ws[f"N{row}"] = f"Остаток дней запаса"
        ws.merge_cells(f"O{row}:P{row}")
        ws[f"O{row}"] = f"Прогноз на следующие {str(last_days)} дней"
        ws.merge_cells(f"Q{row}:R{row}")
        ws[f"Q{row}"] = f"Нужно закупить"
        ws[f"S{row}"] = "Удельный вес %"
        ws[f"T{row}"] = "Накопительный %"
        ws[f"U{row}"] = "Рекомендация"
    
    
        for i in "ABCDEFGHIJKLMNOPQRSTU":
            ws[f"{i}{row}"].font = Font(bold=True)
            ws[f"{i}{row}"].alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            ws[f"{i}{row}"].fill = GRAY_FILL_1
            ws[f"{i}{row}"].border = THIN_BORDER
        

        row += 1
    
        headers2 = [
            "",              
            "",
            "",
            "",
            "",
            "",
            "",
            "кол-во",
            "цена",
            "кол-во",
            "цена",
            "кол-во",
            "цена",
            "",
            "кол-во",
            "цена",
            "кол-во",
            "цена",
            "",    
            "",    
            "",    
        ]
        
    for col, header in enumerate(headers2, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.fill = GRAY_FILL_1
        cell.border = THIN_BORDER
        

    # ширина колонок
    
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 13
    if full_list:
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 20
    else:
        ws.column_dimensions["G"].width = 13
        ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 13
    ws.column_dimensions["K"].width = 13
    ws.column_dimensions["L"].width = 13
    ws.column_dimensions["M"].width = 13
    ws.column_dimensions["N"].width = 13
    ws.column_dimensions["O"].width = 13
    ws.column_dimensions["P"].width = 13
    ws.column_dimensions["Q"].width = 13
    ws.column_dimensions["R"].width = 13
    if full_list:
        ws.column_dimensions["S"].width = 28
        ws.column_dimensions["T"].width = 28
        ws.column_dimensions["U"].width = 13
        ws.column_dimensions["V"].width = 13
        ws.column_dimensions["W"].width = 28
    else:
        ws.column_dimensions["S"].width = 13
        ws.column_dimensions["T"].width = 13
        ws.column_dimensions["U"].width = 28
    
    row += 1

    result_list = result["result_list"]
    grand_totals = result["grand_totals"]
    
    cumulative = Decimal("0")
    
    first_data_row = row
    
    if full_list:
        count = 0
        
        for p in result_list:
            # products = i["products"]
            
            
         
            count += 1
            ws[f"A{row}"] = count   
            ws[f"A{row}"].alignment = CENTER   
            
            ws[f"B{row}"] = p["product_name"]   
            ws[f"B{row}"].alignment = Alignment(wrap_text=True)
            
            ws[f"C{row}"] = p["purchase_price"]  
            
            ws[f"D{row}"] = p["unit"]  
            ws[f"D{row}"].alignment = CENTER_ALIGN  

            abcxyz = p.get("abcxyz_category", "")
            ws[f"E{row}"] = p["product_age_days"]
            ws[f"E{row}"].alignment = CENTER_ALIGN
            
            ws[f"F{row}"] = abcxyz
            ws[f"F{row}"].alignment = CENTER_ALIGN
            ws[f"F{row}"].font = Font(bold=True)
            
            
            
            recommendation = get_abcxyz_recommendation(p.get("abcxyz_category"))

            ws[f"W{row}"] = recommendation
            ws[f"W{row}"].alignment = Alignment(vertical="center", wrap_text=True)

            if abcxyz in ABCXYZ_COLORS:
                ws[f"F{row}"].fill = ABCXYZ_COLORS[abcxyz]
                
            ws[f"G{row}"] = p["category_name"]  
            # ws[f"G{row}"].alignment = CENTER  
            
            ws[f"H{row}"] = p["brand_name"]  
            # ws[f"H{row}"].alignment = CENTER 
                        
            ws[f"I{row}"] = p["partners_count"]  
            ws[f"I{row}"].alignment = CENTER  
            
            ws[f"J{row}"] = p["total_sold"]  
            ws[f"K{row}"] = p["total_sold_price"]  
            ws[f"J{row}"].font = Font(bold=True)
            ws[f"K{row}"].font = Font(bold=True) 
            
            ws[f"L{row}"] = p["avg_per_day"]  
            ws[f"M{row}"] = p["avg_per_day_price"]  
            ws[f"N{row}"] = p["stock_qty"]  
            ws[f"O{row}"] = p["stock_qty_price"] 
        
            
            ws[f"P{row}"] = p["days_of_stock_remaining"]  
            if p["days_of_stock_remaining"] < 7:
                ws[f"P{row}"].font = RED_FONT
            elif p["days_of_stock_remaining"] < 30:
                ws[f"P{row}"].font = YELLOW_FONT
            elif p["days_of_stock_remaining"] >= 30:
                ws[f"P{row}"].font = GREEN_FONT
                
            ws[f"Q{row}"] = p["forecast"]  
            ws[f"R{row}"] = p["forecast_price"] 
            
            
            ws[f"S{row}"] = p["need_to_buy"]  
            ws[f"T{row}"] = p["need_to_buy_price"] 
            
            ws[f"U{row}"] = p["share"] 
            
            ws[f"V{row}"] = f"=SUM($U${first_data_row}:U{row})"
            
            if p["need_to_buy"] > 0:
                ws[f"S{row}"].font = RED_FONT
                ws[f"T{row}"].font = RED_FONT
            else:
                ws[f"S{row}"].font = GREEN_FONT
                ws[f"T{row}"].font = GREEN_FONT
            
            for i in "ABCDEFGHIJKLMNOPQRSTUVW":
                ws[f"{i}{row}"].border = THIN_BORDER 
                
            for i in "CGHIJKLMNOPQRSTUVW":
                ws[f"{i}{row}"].alignment = CENTER_VERTICAL 
                
            for i in "EIJLNPQSU":
                ws[f"{i}{row}"].number_format = QTY_FMT 
                
            for i in "CKMORTW":
                ws[f"{i}{row}"].number_format = PRICE_FMT 
            
            row += 1
    else:
        cumulative = Decimal("0")
        
        for i in result_list:
            category = i["category"]
            category_totals = i["category_totals"]
            products = i["products"]
            
            ws.merge_cells(f"A{row}:U{row}")
            ws[f"A{row}"] = category
            ws[f"A{row}"].font = Font(size=12, bold=True)
            ws[f"A{row}"].fill = GRAY_FILL_1
            ws[f"A{row}"].border = THIN_BORDER
            
            row += 1
            count = 0
            
            for p in products:
                count += 1
                ws[f"A{row}"] = count   
                ws[f"A{row}"].alignment = CENTER   
                
                ws[f"B{row}"] = p["product_name"]   
                ws[f"B{row}"].alignment = Alignment(wrap_text=True)
                
                ws[f"C{row}"] = p["purchase_price"]  
                
                ws[f"D{row}"] = p["unit"]  
                ws[f"D{row}"].alignment = CENTER_ALIGN  
                
                # ws[f"E{row}"] = p["abcxyz_category"]
                # ws[f"E{row}"].alignment = CENTER
                # ws[f"E{row}"].font = Font(bold=True)
                
                abcxyz = p.get("abcxyz_category", "")
                ws[f"E{row}"] = p["product_age_days"]
                ws[f"E{row}"].alignment = CENTER_ALIGN
                
                ws[f"F{row}"] = abcxyz
                ws[f"F{row}"].alignment = CENTER_ALIGN
                ws[f"F{row}"].font = Font(bold=True)
                
                recommendation = get_abcxyz_recommendation(p.get("abcxyz_category"))

                ws[f"U{row}"] = recommendation
                ws[f"U{row}"].alignment = Alignment(vertical="center", wrap_text=True)

                if abcxyz in ABCXYZ_COLORS:
                    ws[f"F{row}"].fill = ABCXYZ_COLORS[abcxyz]
                            
                ws[f"G{row}"] = p["partners_count"]  
                ws[f"G{row}"].alignment = CENTER  
                
                ws[f"H{row}"] = p["total_sold"]  
                ws[f"I{row}"] = p["total_sold_price"]  
                ws[f"H{row}"].font = Font(bold=True)
                ws[f"I{row}"].font = Font(bold=True) 
                
                ws[f"J{row}"] = p["avg_per_day"]  
                ws[f"K{row}"] = p["avg_per_day_price"]  
                ws[f"L{row}"] = p["stock_qty"]  
                ws[f"M{row}"] = p["stock_qty_price"] 
            
                
                ws[f"N{row}"] = p["days_of_stock_remaining"]  
                if p["days_of_stock_remaining"] < 7:
                    ws[f"N{row}"].font = RED_FONT
                elif p["days_of_stock_remaining"] < 30:
                    ws[f"N{row}"].font = YELLOW_FONT
                elif p["days_of_stock_remaining"] >= 30:
                    ws[f"N{row}"].font = GREEN_FONT
                    
                ws[f"O{row}"] = p["forecast"]  
                ws[f"P{row}"] = p["forecast_price"] 
                
                
                ws[f"Q{row}"] = p["need_to_buy"]  
                ws[f"R{row}"] = p["need_to_buy_price"] 
                
                ws[f"S{row}"] = p["share"] 
                
                cumulative += p["share"]   # ← накопление
                ws[f"T{row}"] = f"=SUM($S${first_data_row}:S{row})"
                
           
                
                if p["need_to_buy"] > 0:
                    ws[f"Q{row}"].font = RED_FONT
                    ws[f"R{row}"].font = RED_FONT
                else:
                    ws[f"Q{row}"].font = GREEN_FONT
                    ws[f"R{row}"].font = GREEN_FONT
                
                for i in "ABCDEFGHIJKLMNOPQRSTU":
                    ws[f"{i}{row}"].border = THIN_BORDER 
                    
                for i in "CGHIJKLMNOPQRSU":
                    ws[f"{i}{row}"].alignment = CENTER_VERTICAL 
                    
                for i in "EGHJLNOQS":
                    ws[f"{i}{row}"].number_format = QTY_FMT 
                    
                for i in "DIKMPRU":
                    ws[f"{i}{row}"].number_format = PRICE_FMT 
                    
                # ws[f"C{row}"].number_format = PRICE_FMT
                
                row += 1
                
            # category_totals
            # ws.merge_cells(f"A{row}:E{row}")
            ws[f"B{row}"] = f"Итого {category} "
            ws[f"B{row}"].alignment = RIGHT_ALIGN
            
            for i in "ABCDEFGHIJKLMNOPQRSTU":
                ws[f"{i}{row}"].border = THIN_BORDER 
                ws[f"{i}{row}"].font = HEADER_FONT 
                ws[f"{i}{row}"].fill = GRAY_FILL_1
                
        
                
            for i in "GHJLNOQS":
                ws[f"{i}{row}"].number_format = QTY_FMT 
                
            for i in "IKMPR":
                ws[f"{i}{row}"].number_format = PRICE_FMT 
                
            ws[f"H{row}"] = category_totals["total_sold"]  
            ws[f"I{row}"] = category_totals["total_sold_price"]  
        
            
            ws[f"J{row}"] = category_totals["avg_per_day"]  
            ws[f"K{row}"] = category_totals["avg_per_day_price"]  
            ws[f"L{row}"] = category_totals["stock_qty"]  
            ws[f"M{row}"] = category_totals["stock_qty_price"]  
            
            # ws[f"L{row}"] = category_totals["stock_qty_price"]  
            
            
            ws[f"O{row}"] = category_totals["forecast"]  
            ws[f"P{row}"] = category_totals["forecast_price"] 
            ws[f"Q{row}"] = category_totals["need_to_buy"]  
            ws[f"R{row}"] = category_totals["need_to_buy_price"] 
            
            row += 1
            
    ws[f"B{row}"] = f"ИТОГО ВСЕГО "
    ws[f"B{row}"].alignment = RIGHT_ALIGN
    
    for i in "ABCDEFGHIJKLMNOPQRSTUVW":
        ws[f"{i}{row}"].border = THIN_BORDER 
        ws[f"{i}{row}"].font = HEADER_FONT 
        ws[f"{i}{row}"].fill = GRAY_FILL_1
        
    
        
    for i in "JLNPQS":
        ws[f"{i}{row}"].number_format = QTY_FMT 
        
    for i in "KMPRT":
        ws[f"{i}{row}"].number_format = PRICE_FMT 
        
    ws[f"J{row}"] = grand_totals["total_sold"]  
    ws[f"K{row}"] = grand_totals["total_sold_price"]  
    
    
    ws[f"L{row}"] = grand_totals["avg_per_day"]  
    ws[f"M{row}"] = grand_totals["avg_per_day_price"]  
    ws[f"N{row}"] = grand_totals["stock_qty"]  
    ws[f"O{row}"] = grand_totals["stock_qty_price"]  
    
    # ws[f"K{row}"] = grand_totals["stock_qty_price"]  
    
    ws[f"Q{row}"] = grand_totals["forecast"]  
    ws[f"R{row}"] = grand_totals["forecast_price"] 
    ws[f"S{row}"] = grand_totals["need_to_buy"]  
    ws[f"T{row}"] = grand_totals["need_to_buy_price"] 
    
    row += 1
            
        
        
        
        
 
    
    # ===== RESPONSE =====
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="sale_analiz_{str(date_from_str)}_{str(date_to_str)}.xlsx"'
    )
    
    # ===== НАСТРОЙКА ПЕЧАТИ =====

    # Поля страницы (в дюймах!)
    ws.page_margins = PageMargins(
        left=0.3,        # Левый
        right=0.2,     # Правый
        top=0.2,       # Верхний
        bottom=0.2,    # Нижний
    )

    # Альбомная ориентация (если нужно — можно убрать)
    # ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # Разместить не более чем на 1 странице по ширине
    # ws.page_setup.fitToWidth = 1
    # ws.page_setup.fitToHeight = False
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = без ограничения по высоте

    # Центрирование по горизонтали при печати
    ws.print_options.horizontalCentered = True

    wb.save(response)
    return response


