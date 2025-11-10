from ..models import *
from django.http import JsonResponse
from icecream import ic
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
from django.utils.dateparse import parse_date
from django.db.models import Sum, F, Q
import json
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from datetime import datetime
from io import BytesIO
from openpyxl.styles import PatternFill
from django.db import transaction
from datetime import datetime, date
from collections import defaultdict

from decimal import Decimal
from django.db import transaction as db_transaction






# поиск продукта по id и складу
def get_product_by_id_and_warehouse(request):
    product_id = request.GET.get('product_id')
    main_product_id = request.GET.get('main_product_id')
    warehouse_id = request.GET.get('warehouse_id')
    warehouse = Warehouse.objects.get(id=warehouse_id)
    
    main_product_obj = Product.objects.get(id=main_product_id)

    if not product_id or not warehouse_id:
        return JsonResponse({'error': 'Missing parameters'}, status=400)

    try:
        # получаем запись на складе
        wp = WarehouseProduct.objects.select_related('product').get(
            product_id=product_id,
            warehouse_id=warehouse_id
        )
        product = wp.product
        
   
        base_unit_obj = {"id":product.base_unit.id, "name":product.base_unit.name}
        
        images = []
        product_images = product.images.all()
        if product_images.exists():
            for i in product_images:
                images.append({
                    "alt_text": i.alt_text,
                    "id": i.id,
                    "image": i.image.url,
                    "product": product.id,
                })
                
        auqntity_for_per_unit_main_product = FreeProduct.objects.get(main_product=main_product_obj, gift_product=product).quantity_per_unit
                
        if warehouse:
            quantity = product.warehouse_products.filter(
                warehouse_id=warehouse
            ).aggregate(total=models.Sum('quantity'))['total'] or 0
            base_quantity_in_stock = quantity
        else:
            quantity = product.get_total_quantity()

        unit_name = product.base_unit.name if product.base_unit else ""
        for unit in product.units.all():
            if unit.is_default_for_sale and unit.conversion_factor:
                # print('quantity', quantity)
                # print('unit.conversion_factor', unit.conversion_factor)
                quantity = float(quantity) / float(unit.conversion_factor)
                unit_name = unit.unit.name
                break
            
            
        units_obj = ProductUnit.objects.filter(product=product, )
        units = []
        if units_obj.exists():
            units.append({
                "base_unit_name": base_unit_obj.name,
                "conversion_factor": units_obj.conversion_factor,
                "is_default_for_sale": units_obj.is_default_for_sale,
                "unit_name": units_obj.unit.name
                
            })
            
        warehouses_data = [{
            "quantity": float(wp.quantity),
            "warehouse_id": warehouse_id,
            "warehouse_name": warehouse.name
        }]
            
 
        data = {
            'base_quantity_in_stock': float(wp.quantity),
            "base_unit_obj": base_unit_obj,
            "discount_price": product.discount_price,
            "firma_price": product.firma_price,
            "height": product.height,
            "id": product.id,
            "images": images,
            "is_active": product.is_active,
            "is_custom_price": True,
            "is_gift": True,
            "length": product.length,
            "name": product.name,
            "purchase_price": product.purchase_price,
            "qr_code": product.qr_code,
            "quantity_on_selected_warehouses": quantity,
            "retail_price": product.retail_price,
            "selected_price": 0,
            "selected_quantity": auqntity_for_per_unit_main_product,
            "sku": product.sku,
            # "total_quantity": product.firma_price,
            'unit_name_on_selected_warehouses': unit_name,
            "units": units,
            "volume": product.volume,
            "warehouses_data": product.firma_price,
            "warehouses_data": warehouses_data,
            "weight": product.weight,
            "wholesale_price": product.wholesale_price,
            "width": product.width,
            # "firma_price": product.firma_price,
            # "firma_price": product.firma_price,
            # "firma_price": product.firma_price,
            
            
            # 'id': product.id,
            # 'name': product.name,
            # 'price': product.retail_price,   # или любая цена
            
            # 'free_items': [
            #     {
            #         'id': f.gift_product.id,
            #         'name': f.gift_product.name,
            #         'quantity_per_unit': float(f.quantity_per_unit)
            #     } for f in product.free_items.all()
            # ]
        }
        return JsonResponse(data)
    except WarehouseProduct.DoesNotExist:
        return JsonResponse({'error': 'Product not found in this warehouse'}, status=404)
    
    
    
    
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transaction_detail(request, id):
    transaction_obj = Transaction.objects.get(pk=id)

    data = {
        "transaction_id": transaction_obj.id,
    }

    if transaction_obj.invoice:
        invoice = transaction_obj.invoice
        data["invoice_id"] = invoice.id,
        


    return Response(data, status=status.HTTP_200_OK)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_product_for_print_qr(request):
    products = Product.objects.prefetch_related('images').all()
    
    data = []
    for product in products:
        item = {
            "id": product.id,
            "name": product.name,
            "qr_code": product.qr_code,
            "images": []
        }
        
        # Добавляем все изображения товара
        for product_image in product.images.all():
            image_data = {
                "url": request.build_absolute_uri(product_image.image.url),
                "alt_text": product_image.alt_text or product.name
            }
            item["images"].append(image_data)
        
        data.append(item)

    return Response(data, status=status.HTTP_200_OK)




@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_margin_date(request):
    try:
        date_focus = DateFocus.objects.latest('id')
        data = {"date_focus": date_focus.dateFocus}
    except DateFocus.DoesNotExist:
        data = {}

    return Response(data, status=200)


@csrf_exempt
def set_date_focus(request):
    if request.method == "POST":
        today = datetime.now().date()
        date_str = today.strftime("%Y-%m-%d")
        DateFocus.objects.all().delete()
        DateFocus.objects.create(dateFocus=today)

        return JsonResponse({"date_focus": str(today)})
    return JsonResponse({"error": "Method not allowed"}, status=405)




##################################################################################################################################################################
##################################################################################################################################################################
##################################################################################################################################################################
# osw2

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_account_for_osw2(request):
    # Получаем даты
    date_from = request.GET.get('dateFrom')
    date_to = request.GET.get('dateTo')

    if date_from:
        date_from = parse_date(date_from)
    if date_to:
        date_to = parse_date(date_to)

    # Активные счета
    accounts = Account.objects.filter(is_active=True)
    parent_accounts = []

    data = []
    data_dict = {}  # ключ = номер счета, значение = агрегированные данные
    
    # Функция для добавления или суммирования счета в data_dict
    def add_or_update(acc_number, acc_name):
        if acc_number not in data_dict:
            data_dict[acc_number] = {
                "number": acc_number,
                "name": acc_name,
                "initial_debit": float(initial_debit),
                "initial_credit": float(initial_credit),
                "initial_balance": float(initial_balance),
                "debit": float(period_debit),
                "credit": float(period_credit),
                "period_debit": float(period_debit),
                "period_credit": float(period_credit),
                
                "final_debit": float(initial_debit) + float(period_debit),
                "final_credit": float(initial_credit) + float(period_credit),
                "final_balance": float(final_balance),
            }
        else:
            d = data_dict[acc_number]
            d['initial_debit'] += float(initial_debit)
            d['initial_credit'] += float(initial_credit)
            d['initial_balance'] += float(initial_balance)
            d['debit'] += float(period_debit)
            d['credit'] += float(period_credit)
            d['period_debit'] += float(period_debit)
            d['period_credit'] += float(period_credit)
            
            d["final_debit"] += float(initial_debit) + float(period_debit)
            d["final_credit"] += float(initial_credit) + float(period_credit)
            d['final_balance'] += float(final_balance)

    for a in accounts:
        # Начальный остаток
        entries_before = Entry.objects.filter(account=a)
        if date_from:
            entries_before = entries_before.filter(transaction__date__lt=date_from)

        initial_debit = entries_before.aggregate(total=Sum('debit'))['total'] or 0
        initial_credit = entries_before.aggregate(total=Sum('credit'))['total'] or 0
        initial_balance = initial_debit - initial_credit

        # Проводки в период
        entries_period = Entry.objects.filter(account=a)
        if date_from and date_to:
            entries_period = entries_period.filter(transaction__date__range=[date_from, date_to])
        elif date_from:
            entries_period = entries_period.filter(transaction__date__gte=date_from)
        elif date_to:
            entries_period = entries_period.filter(transaction__date__lte=date_to)

        period_debit = entries_period.aggregate(total=Sum('debit'))['total'] or 0
        period_credit = entries_period.aggregate(total=Sum('credit'))['total'] or 0

        final_balance = initial_balance + period_debit - period_credit

        

        # Добавляем ребёнка
        add_or_update(a.number, a.name)

        # Добавляем родителя, если есть
        if a.parent:
      
            if a.parent.number not in parent_accounts:
                parent_accounts.append(a.parent.number)
            add_or_update(a.parent.number, a.parent.name)

    # Преобразуем словарь в список для ответа

    
    for acc, values in data_dict.items():
        if acc in parent_accounts:
 
            values["is_parent"] = True
        else:
            values["is_parent"] = False
    data = list(data_dict.values())


    return Response(data, status=200)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_detail_account(request):
    account_number = request.GET.get('account')
    date_from = request.GET.get('dateFrom')
    date_to = request.GET.get('dateTo')

    if date_from:
        date_from = parse_date(date_from)
    if date_to:
        date_to = parse_date(date_to)

    try:
        account = Account.objects.get(number=account_number)
    except Account.DoesNotExist:
        return Response({"error": "Account not found"}, status=404)

    all_accounts = [account] + list(account.children.all())

    # Начальное сальдо до date_from
    if date_from:
        initial_entries = Entry.objects.filter(
            account__in=all_accounts,
            transaction__date__lt=date_from
        )
        initial_debit = sum([float(e.debit) for e in initial_entries])
        initial_credit = sum([float(e.credit) for e in initial_entries])
    else:
        initial_debit = initial_credit = 0

    # Проводки за выбранный период
    entries = Entry.objects.filter(account__in=all_accounts)
    if date_from and date_to:
        entries = entries.filter(transaction__date__range=[date_from, date_to])
    elif date_from:
        entries = entries.filter(transaction__date__gte=date_from)
    elif date_to:
        entries = entries.filter(transaction__date__lte=date_to)

    transactions = []
    for e in entries.select_related('transaction'):
        transactions.append({
            "transaction_id": e.transaction.id,
            "date": e.transaction.date.strftime("%Y-%m-%d"),
            "description": e.transaction.description,
            "debit": float(e.debit),
            "credit": float(e.credit),
            "created_by": e.transaction.created_by.username if e.transaction.created_by else None,
        })

    # Итоги по оборотам и конечное сальдо
    period_debit = sum([t['debit'] for t in transactions])
    period_credit = sum([t['credit'] for t in transactions])
    final_debit = initial_debit + period_debit
    final_credit = initial_credit + period_credit

    osw = [{
        "initial_debit": initial_debit,
        "initial_credit": initial_credit,
        "debit": period_debit,
        "credit": period_credit,
        "final_debit": final_debit,
        "final_credit": final_credit,
    }]

    return Response({
        "account": {
            "number": account.number,
            "name": account.name,
        },
        "transactions": transactions,
        "osw": osw,
    })

# osw2
##################################################################################################################################################################
##################################################################################################################################################################
##################################################################################################################################################################



# rabotayushiy create entry no deepseek predlojil chutka izmenit chtoby rabotali impoert export entries (ego kod snizu)
# @csrf_exempt
# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def create_entry(request):
#     if request.method == "POST":

#         try:
#             with transaction.atomic():
#                 data = json.loads(request.body)

#                 date = parse_date(data.get("date"))
#                 if not date:
#                     return JsonResponse({"status": "error", "message": "Неверная дата"}, status=400)

#                 debit_acc_number = data.get("debitAccount")
#                 credit_acc_number = data.get("creditAccount")
#                 amount = Decimal(data.get("amount") or 0)
#                 if amount <= 0:
#                     return JsonResponse({"status": "error", "message": "Сумма должна быть больше 0"}, status=400)

#                 comment = data.get("comment", "")
#                 partner_id = data.get("partnerId")
#                 partner = Partner.objects.get(id=partner_id) if partner_id else None
                
#                 # USD
#                 if partner:
#                     rule_usd = CustomePostingRule.objects.filter(operation__code="sale", directory_type=partner.type, amount_type="revenue", currency__code="USD").first()
#                     rule_tmt = CustomePostingRule.objects.filter(operation__code="sale", directory_type=partner.type, amount_type="revenue", currency__code="TMT").first()
                    
#                     if not rule_usd and not rule_tmt:
#                         return JsonResponse({"status": "error", "message": "No posting rule for this partner type and currency"}, status=400)
                    
               
                
                
#                 # Весь процесс в атомарной транзакции
#                 with db_transaction.atomic():
#                     transaction_obj = Transaction.objects.create(
#                         date=date,
#                         description=comment,
#                         partner=partner,
#                         created_by=request.user
#                     )

#                     debit_account = Account.objects.get(number=debit_acc_number)
#                     credit_account = Account.objects.get(number=credit_acc_number)
                    
              
                    
#                     if partner:
#                 
#                         if credit_account == rule_tmt.debit_account:
#                       
#                             partner.balance_tmt += amount
#                         elif credit_account == rule_usd.debit_account:
#                             partner.balance_usd += amount
#                      
                            
#                         if debit_account == rule_tmt.debit_account:
#                             partner.balance_tmt -= amount
#               
#                         elif debit_account == rule_usd.debit_account:
#                             partner.balance_usd -= amount
#                       
                            
#                         partner.save()
                        
                        
                    
                    
                    
                    

#                     Entry.objects.create(
#                         transaction=transaction_obj,
#                         account=debit_account,
#                         debit=amount,
#                         credit=Decimal("0.00")
#                     )

#                     Entry.objects.create(
#                         transaction=transaction_obj,
#                         account=credit_account,
#                         debit=Decimal("0.00"),
#                         credit=amount
#                     )
#                     # 1/0
#                 return JsonResponse({"message": "success entry", "transaction_id": transaction_obj.id})

#         except Account.DoesNotExist:
#             return JsonResponse({"status": "error", "message": "Счёт не найден"}, status=400)
#         except Partner.DoesNotExist:
#             return JsonResponse({"status": "error", "message": "Партнёр не найден"}, status=400)
#         except Exception as e:
#             ic(e)
#             # create error
#             return JsonResponse({"status": "error", "message": "transactionChange", "reason_for_the_error": str(e)}, status=400)





@csrf_exempt
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_entry(request):
    """Создание проводки с улучшенной обработкой для импорта"""
    if request.method == "POST":
        ic(request.user)
        try:
            with transaction.atomic():
                data = json.loads(request.body)

                # Поддержка разных форматов даты
                date_str = data.get("date")
                if not date_str:
                    return JsonResponse({"status": "error", "message": "Дата обязательна"}, status=400)
                
                # Парсим дату в разных форматах
                try:
                    if isinstance(date_str, str):
                        # Сначала пробуем стандартный парсинг
                        date_obj = parse_date(date_str)
                        if not date_obj:
                            # Пробуем другие форматы
                            try:
                                date_obj = datetime.strptime(date_str, '%d.%m.%Y').date()
                            except ValueError:
                                try:
                                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                                except ValueError:
                                    return JsonResponse({"status": "error", "message": "Неверный формат даты"}, status=400)
                        # Конвертируем в datetime для модели
                        datetime_obj = datetime.combine(date_obj, datetime.min.time())
                    else:
                        return JsonResponse({"status": "error", "message": "Неверный формат даты"}, status=400)
                except Exception as e:
                    ic(f"Date parsing error: {e}")
                    return JsonResponse({"status": "error", "message": "Ошибка парсинга даты"}, status=400)

                # Проверка закрытых дней
                if DayClosing.objects.filter(date=date_obj).exists():
                    return JsonResponse({"status": "error", "message": f"День {date_obj.strftime('%d.%m.%Y')} закрыт"}, status=400)

                # Получаем данные счетов
                debit_acc_number = data.get("debitAccount") or data.get("debit_account_number")
                credit_acc_number = data.get("creditAccount") or data.get("credit_account_number")
                
                if not debit_acc_number or not credit_acc_number:
                    return JsonResponse({"status": "error", "message": "Не указаны счета дебета или кредита"}, status=400)

                amount = Decimal(data.get("amount") or 0)
                if amount <= 0:
                    return JsonResponse({"status": "error", "message": "Сумма должна быть больше 0"}, status=400)

                comment = data.get("comment", "") or data.get("description", "")
                partner_id = data.get("partnerId") or data.get("partner_id")
                partner = Partner.objects.get(id=partner_id) if partner_id else None
                
                # Дополнительные поля
                product_id = data.get("product_id") or data.get("product", {}).get("id")
                product = Product.objects.get(id=product_id) if product_id else None
                
                warehouse_id = data.get("warehouse_id") or data.get("warehouse", {}).get("id")
                warehouse = Warehouse.objects.get(id=warehouse_id) if warehouse_id else None

                # Валидация счетов
                try:
                    debit_account = Account.objects.get(number=debit_acc_number, is_active=True)
                    credit_account = Account.objects.get(number=credit_acc_number, is_active=True)
                except Account.DoesNotExist as e:
                    return JsonResponse({"status": "error", "message": f"Счет не найден или неактивен: {str(e)}"}, status=400)

                # USD/TMT логика для партнеров
                if partner and warehouse:
                    # Получаем валюту склада
                    warehouse_currency_code = warehouse.currency.code if warehouse.currency else None
                    
                    if warehouse_currency_code:
                        rule = CustomePostingRule.objects.filter(
                            operation__code="sale", 
                            directory_type=partner.type, 
                            amount_type="revenue", 
                            currency__code=warehouse_currency_code
                        ).first()
                        
                        if rule:
                            ic(f"Found rule for {warehouse_currency_code}:", rule.debit_account, rule.credit_account)

                # Весь процесс в атомарной транзакции
                with db_transaction.atomic():
                    transaction_obj = Transaction.objects.create(
                        date=datetime_obj,  # Используем datetime объект
                        description=comment,
                        partner=partner,
                        created_by=request.user
                    )

                    # Логика обновления баланса партнера (если нужно)
                    if partner and warehouse and warehouse.currency:
                        currency_code = warehouse.currency.code
                        # Простая логика - можно доработать под конкретные правила
                        if currency_code == "TMT":
                            # Логика для TMT
                            pass
                        elif currency_code == "USD":
                            # Логика для USD
                            pass

                    # Создаем проводки
                    Entry.objects.create(
                        transaction=transaction_obj,
                        account=debit_account,
                        debit=amount,
                        credit=Decimal("0.00"),
                        product=product,
                        warehouse=warehouse
                    )

                    Entry.objects.create(
                        transaction=transaction_obj,
                        account=credit_account,
                        debit=Decimal("0.00"),
                        credit=amount,
                        product=product,
                        warehouse=warehouse
                    )

                return JsonResponse({
                    "message": "success entry", 
                    "transaction_id": transaction_obj.id
                })

        except Account.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Счёт не найден"}, status=400)
        except Partner.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Партнёр не найден"}, status=400)
        except Product.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Продукт не найден"}, status=400)
        except Warehouse.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Склад не найден"}, status=400)
        except Exception as e:
            ic(e)
            return JsonResponse({"status": "error", "message": "transactionChange", "reason_for_the_error": str(e)}, status=400)




@csrf_exempt
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_entry(request, id):
    """Редактирование существующей проводки"""
    try:
        data = json.loads(request.body)

        with transaction.atomic():

            # Находим транзакцию
            try:
                transaction_obj = Transaction.objects.get(id=id)
            except Transaction.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Транзакция не найдена"}, status=404)

            # Парсинг даты
            date_str = data.get("date")
            if not date_str:
                return JsonResponse({"status": "error", "message": "Дата обязательна"}, status=400)

            date_obj = parse_date(date_str)
            if not date_obj:
                try:
                    date_obj = datetime.strptime(date_str, '%d.%m.%Y').date()
                except ValueError:
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                    except ValueError:
                        return JsonResponse({"status": "error", "message": "Неверный формат даты"}, status=400)

            # Проверка закрытого дня
            if DayClosing.objects.filter(date=date_obj).exists():
                return JsonResponse({"status": "error", "message": f"День {date_obj} закрыт"}, status=400)

            # Счета
            debit_acc_number = data.get("debitAccount")
            credit_acc_number = data.get("creditAccount")

            try:
                debit_account = Account.objects.get(number=debit_acc_number)
                credit_account = Account.objects.get(number=credit_acc_number)
            except Account.DoesNotExist:
                return JsonResponse({"status": "error", "message": "Указан неверный счёт"}, status=400)

            # Сумма
            amount = Decimal(data.get("amount") or 0)
            if amount <= 0:
                return JsonResponse({"status": "error", "message": "Сумма должна быть > 0"}, status=400)

            # Partner / Product / Warehouse
            partner_id = data.get("partnerId")
            product_id = data.get("product_id")
            warehouse_id = data.get("warehouse_id")

            partner = Partner.objects.get(id=partner_id) if partner_id else None
            product = Product.objects.get(id=product_id) if product_id else None
            warehouse = Warehouse.objects.get(id=warehouse_id) if warehouse_id else None

            # Обновляем Transaction
            transaction_obj.date = datetime.combine(date_obj, datetime.min.time())
            transaction_obj.description = data.get("comment", "")
            transaction_obj.partner = partner
            transaction_obj.save()

            # Удаляем старые проводки
            Entry.objects.filter(transaction=transaction_obj).delete()

            # Создаём новые (дебет)
            Entry.objects.create(
                transaction=transaction_obj,
                account=debit_account,
                debit=amount,
                credit=Decimal("0"),
                product=product,
                warehouse=warehouse
            )

            # Кредит
            Entry.objects.create(
                transaction=transaction_obj,
                account=credit_account,
                debit=Decimal("0"),
                credit=amount,
                product=product,
                warehouse=warehouse
            )

            return JsonResponse({"status": "success", "message": "Проводка обновлена"})

    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=400)
    
    
    
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_entry(request, entry_id):
    try:
        transaction = Transaction.objects.get(id=entry_id)
        
        # Дополнительные проверки перед удалением
        if transaction.invoice:
            return JsonResponse({
                "error": "Нельзя удалить проводку, связанную с фактурой. Сначала удалите фактуру."
            }, status=400)
        
        # Проверяем права пользователя
        if transaction.created_by and transaction.created_by != request.user:
            return JsonResponse({
                "error": "Вы можете удалять только свои проводки"
            }, status=403)
        
        # Логируем удаление
        print(f"Пользователь {request.user} удалил проводку {entry_id}")
        
        transaction_id = transaction.id
        transaction.delete()
        
        return JsonResponse({
            "message": "Проводка успешно удалена",
            "deleted_id": transaction_id
        })
        
    except Transaction.DoesNotExist:
        return JsonResponse({"error": "Проводка не найдена"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"Ошибка при удалении: {str(e)}"}, status=500)




# # samyy perwyy kod rabotaet no ochen skudnyy po kolonkam i ne uchitywaet wozwrat
# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def upload_sales_excel_for_analis(request):
#     file = request.FILES.get("file")
#     if not file:
#         return HttpResponse("Файл не получен", status=400)

#     try:
#         wb = load_workbook(filename=file, data_only=True)
#         ws = wb.active

#         products = {}

#         # собираем данные
#         for row in ws.iter_rows(min_row=6, values_only=True):
#             if row[0] and "Hemmesi" in str(row[0]):
#                 break

#             product_name = row[9]  # Harydyň ady
#             document = row[3]      # Dokument
#             date_str = row[1]      # Senesi
#             quantity = float(row[12] or 0)  # Mukdary
#             price = float(row[11] or 0)     # Bahasy
#             client = row[4]       # Kime

#             try:
#                 date_ = datetime.strptime(date_str, "%d.%m.%y").date()
#             except Exception:
#                 continue

#             if product_name not in products:
#                 products[product_name] = {
#                     "price": price,
#                     "sale_quantity": 0,
#                     "total_sale": 0,
#                     "purchase_date": None,
#                     "last_sale_date": None,
#                     "clients": set()
#                 }

#             if "Girdeji faktura" in document:
#                 products[product_name]["purchase_date"] = date_
#             elif "Söwda" in document:
#                 products[product_name]["sale_quantity"] += quantity
#                 products[product_name]["total_sale"] += price * quantity
#                 products[product_name]["last_sale_date"] = (
#                     date_ if not products[product_name]["last_sale_date"]
#                     else max(products[product_name]["last_sale_date"], date_)
#                 )
#                 if client:
#                     products[product_name]["clients"].add(client)

#         # вычисляем диапазон для градиента
#         all_days = [
#             (v["last_sale_date"] - v["purchase_date"]).days
#             for v in products.values()
#             if v["purchase_date"] and v["last_sale_date"]
#         ]
#         min_days = min(all_days) if all_days else 0
#         max_days = max(all_days) if all_days else 1

#         def get_fill(days):
#             if days is None:
#                 return PatternFill(fill_type=None)
#             ratio = (days - min_days) / (max_days - min_days + 1e-5)
#             if ratio < 0.5:
#                 green = 255
#                 red = int(255 * ratio * 2)
#                 blue = 0
#             else:
#                 red = 255
#                 green = int(255 * (1 - (ratio-0.5)*2))
#                 blue = 0
#             hex_color = f"{red:02X}{green:02X}{blue:02X}"
#             return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

#         # создаём новый Excel
#         out_wb = Workbook()
#         out_ws = out_wb.active
#         out_ws.title = "Sales Analysis"

#         headers = [
#             "Наименование товара",
#             "Цена (Сумма/Количество)",
#             "Количество продажи",
#             "Сумма продажи",
#             "Дата Прихода",
#             "Дата Продажи последней штуки",
#             "Количество дней",
#             "Количество клиента"
#         ]
#         out_ws.append(headers)

#         # заполняем данные и красим строки
#         for name, info in products.items():
#             purchase_date = info["purchase_date"]
#             last_sale_date = info["last_sale_date"]
#             days_diff = (last_sale_date - purchase_date).days if purchase_date and last_sale_date else None

#             out_ws.append([
#                 name,
#                 info["price"],
#                 info["sale_quantity"],
#                 info["total_sale"],
#                 purchase_date.isoformat() if purchase_date else None,
#                 last_sale_date.isoformat() if last_sale_date else None,
#                 days_diff,
#                 len(info["clients"])
#             ])

#             row_idx = out_ws.max_row
#             fill = get_fill(days_diff)
#             for col_idx in range(1, len(headers)+1):
#                 out_ws.cell(row=row_idx, column=col_idx).fill = fill

#         # сохраняем и возвращаем
#         output = BytesIO()
#         out_wb.save(output)
#         output.seek(0)

#         response = HttpResponse(
#             output,
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = 'attachment; filename="sales_analysis.xlsx"'
#         return response

#     except Exception as e:
#         import traceback
#         print(traceback.format_exc())
#         return HttpResponse(f"Ошибка при обработке файла: {e}", status=500)
    
    
    
    

# kod rabochiy i uchitywaet wozwrat no kolonki wse rawno skudnye
# def _parse_date(cell):
#     """Поддерживаем datetime, 'dd.mm.YYYY', 'dd.mm.yy' и пустые значения."""
#     if cell is None:
#         return None
#     if isinstance(cell, datetime):
#         return cell.date()
#     s = str(cell).strip()
#     if not s:
#         return None
#     for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
#         try:
#             return datetime.strptime(s, fmt).date()
#         except Exception:
#             pass
#     return None


# def _to_float(cell):
#     try:
#         return float(cell) if cell is not None else 0.0
#     except Exception:
#         return 0.0
# # rabotaet no nado dobawit neskolko kolonok poprosil azamat
# @api_view(['POST'])
# @permission_classes([IsAuthenticated])
# def upload_sales_excel_for_analis_with_return(request):
#     file = request.FILES.get("file")
#     if not file:
#         return HttpResponse("Файл не получен", status=400)

#     try:
#         wb = load_workbook(filename=file, data_only=True)
#         ws = wb.active

#         # products: ключ = название товара, значение = список партий (FIFO)
#         products = {}

#         for row in ws.iter_rows(min_row=6, values_only=True):
#             # стоп-условие
#             if row[0] and "Hemmesi" in str(row[0]):
#                 break

#             remark = str(row[2]).strip().lower() if row[2] else ""  # C - Resmi-nama belgisi
#             document = str(row[3] or "")                            # D - Dokument
#             date_cell = row[1]                                      # B - дата
#             product_name = row[9]                                   # J - Harydyň ady
#             qty_in_row = _to_float(row[12])                         # M - количество
#             price = _to_float(row[11])                              # L - цена
#             client = row[4]                                         # E - клиент

#             if not product_name or date_cell is None:
#                 continue

#             date_ = _parse_date(date_cell)
#             if date_ is None:
#                 continue

#             # Инициализация
#             if product_name not in products:
#                 products[product_name] = []

#             # === Приход ===
#             if "Girdeji faktura" in document:
#                 if "wozwrat" in remark:
#                     continue

#                 purchase_qty = qty_in_row
#                 party = {
#                     "purchase_date": date_,
#                     "purchase_qty": purchase_qty,
#                     "remaining": purchase_qty,
#                     "sale_qty": 0.0,
#                     "total_sale": 0.0,
#                     "price": price,              # добавили цену за единицу
#                     "last_sale_touch": None,
#                     "sold_out_date": None,
#                     "clients_net": {}
#                 }
#                 products[product_name].append(party)

#             # === Продажа / Возврат ===
#             elif "Söwda" in document:
#                 is_return = "wozwrat" in remark
#                 qty_to_process = qty_in_row

#                 if not products.get(product_name):
#                     continue

#                 for party in products[product_name]:
#                     if qty_to_process <= 0:
#                         break

#                     if is_return:
#                         can_take_back = min(party["sale_qty"], qty_to_process)
#                         if can_take_back <= 0:
#                             continue
#                         party["sale_qty"] -= can_take_back
#                         party["total_sale"] -= can_take_back * price
#                         party["remaining"] += can_take_back
#                         if client:
#                             party["clients_net"][client] = party["clients_net"].get(client, 0.0) - can_take_back
#                         if party["remaining"] > 0:
#                             party["sold_out_date"] = None
#                         qty_to_process -= can_take_back

#                     else:
#                         if party["remaining"] <= 0:
#                             continue
#                         take = min(party["remaining"], qty_to_process)
#                         party["remaining"] -= take
#                         party["sale_qty"] += take
#                         party["total_sale"] += take * price
#                         party["last_sale_touch"] = date_ if not party["last_sale_touch"] else max(party["last_sale_touch"], date_)
#                         if client:
#                             party["clients_net"][client] = party["clients_net"].get(client, 0.0) + take
#                         if party["remaining"] == 0:
#                             party["sold_out_date"] = date_
#                         qty_to_process -= take

#                 if qty_to_process > 0:
#                     pass

#             else:
#                 continue

#         # === Формируем Excel ===
#         out_wb = Workbook()
#         out_ws = out_wb.active
#         out_ws.title = "Sales Analysis"

#         headers = [
#             "Наименование товара",
#             "Цена за штуку",
#             "Дата прихода партии",
#             "Дата продажи последней банки этой партии",
#             "Кол-во в партии (приход)",
#             "Продано из этой партии (шт)",
#             "Остаток партии (шт)",
#             "Сумма продаж по партии",
#             "Количество клиентов",
#             "Количество дней (до последней продажи)"
#         ]
#         out_ws.append(headers)
#         out_ws.freeze_panes = out_ws['A2']

#         # Заполняем строки
#         for product_name, parties in products.items():
#             for p in parties:
#                 last_unit_date = p["sold_out_date"] if p["sold_out_date"] else p["last_sale_touch"]
#                 clients_count = sum(1 for q in p["clients_net"].values() if q > 0)
#                 days_diff = (last_unit_date - p["purchase_date"]).days if last_unit_date and p["purchase_date"] else None

#                 out_ws.append([
#                     product_name,
#                     p["price"],
#                     p["purchase_date"].isoformat(),
#                     last_unit_date.isoformat() if last_unit_date else None,
#                     p["purchase_qty"],
#                     p["sale_qty"],
#                     p["remaining"],
#                     p["total_sale"],
#                     clients_count,
#                     days_diff
#                 ])

#         # === Заливка ячеек по количеству дней ===
#         all_days = []
#         for parties in products.values():
#             for p in parties:
#                 d = p["sold_out_date"] if p["sold_out_date"] else p["last_sale_touch"]
#                 if d and p["purchase_date"]:
#                     all_days.append((d - p["purchase_date"]).days)

#         min_days = min(all_days) if all_days else 0
#         max_days = max(all_days) if all_days else 1

#         def get_fill(days):
#             if days is None:
#                 return PatternFill(fill_type=None)
#             ratio = (days - min_days) / (max_days - min_days + 1e-5)
#             red = int(255 * ratio)
#             green = int(255 * (1 - ratio))
#             hex_color = f"{red:02X}{green:02X}00"
#             return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

#         for row_idx in range(2, out_ws.max_row + 1):
#             purchase_cell = out_ws.cell(row=row_idx, column=3).value  # C
#             last_unit_cell = out_ws.cell(row=row_idx, column=4).value  # D
#             try:
#                 pd = datetime.fromisoformat(purchase_cell).date() if purchase_cell else None
#                 ld = datetime.fromisoformat(last_unit_cell).date() if last_unit_cell else None
#                 days = (ld - pd).days if (pd and ld) else None
#             except Exception:
#                 days = None
#             fill = get_fill(days)
#             for col in range(1, len(headers) + 1):
#                 out_ws.cell(row=row_idx, column=col).fill = fill

#         output = BytesIO()
#         out_wb.save(output)
#         output.seek(0)

#         response = HttpResponse(
#             output,
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = 'attachment; filename="sales_analysis_by_party.xlsx"'
#         return response

#     except Exception as e:
#         import traceback
#         print(traceback.format_exc())
#         return HttpResponse(f"Ошибка при обработке файла: {e}", status=500)







# Gotowyy kod sdelannoe s chatGPT
def _parse_date1(cell):
    """Поддерживаем datetime, 'dd.mm.YYYY', 'dd.mm.yy' и пустые значения."""
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return cell.date()
    s = str(cell).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def _to_float(cell):
    try:
        return float(cell) if cell is not None else 0.0
    except Exception:
        return 0.0

def parse_excel_date(cell_value):
    if isinstance(cell_value, datetime):
        return cell_value.date()
    elif isinstance(cell_value, str):
        try:
            return datetime.strptime(cell_value, "%d.%m.%Y").date()
        except Exception:
            return None
    return None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_sales_excel_for_analis(request):
    file = request.FILES.get("file")
    if not file:
        return HttpResponse("Файл не получен", status=400)

    try:
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.active

        products = {}

        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] and "Hemmesi" in str(row[0]):
                break

            remark = str(row[2]).strip().lower() if row[2] else ""
            document = str(row[3] or "")
            date_cell = row[1]
            product_name = row[9]
            qty_in_row = _to_float(row[12])
            price = _to_float(row[11])
            client = row[4]

            if not product_name or date_cell is None:
                continue

            date_ = _parse_date1(date_cell)
            if date_ is None:
                continue

            if product_name not in products:
                products[product_name] = []

            # Приход
            if "Girdeji faktura" in document:
                if "wozwrat" in remark:
                    continue
                purchase_qty = qty_in_row
                party = {
                    "purchase_date": date_,
                    "purchase_qty": purchase_qty,
                    "remaining": purchase_qty,
                    "sale_qty": 0.0,
                    "total_sale": 0.0,
                    "price": price,
                    "last_sale_touch": None,
                    "sold_out_date": None,
                    "clients_net": {}
                }
                products[product_name].append(party)

            # Продажа / возврат
            elif "Söwda" in document:
                is_return = "wozwrat" in remark
                qty_to_process = qty_in_row

                if not products.get(product_name):
                    continue

                for party in products[product_name]:
                    if qty_to_process <= 0:
                        break

                    if is_return:
                        can_take_back = min(party["sale_qty"], qty_to_process)
                        if can_take_back <= 0:
                            continue
                        party["sale_qty"] -= can_take_back
                        party["total_sale"] -= can_take_back * price
                        party["remaining"] += can_take_back
                        if client:
                            party["clients_net"][client] = party["clients_net"].get(client, 0.0) - can_take_back
                        if party["remaining"] > 0:
                            party["sold_out_date"] = None
                        qty_to_process -= can_take_back
                    else:
                        if party["remaining"] <= 0:
                            continue
                        take = min(party["remaining"], qty_to_process)
                        party["remaining"] -= take
                        party["sale_qty"] += take
                        party["total_sale"] += take * price
                        party["last_sale_touch"] = date_ if not party["last_sale_touch"] else max(party["last_sale_touch"], date_)
                        if client:
                            party["clients_net"][client] = party["clients_net"].get(client, 0.0) + take
                        if party["remaining"] == 0:
                            party["sold_out_date"] = date_
                        qty_to_process -= take

        # Формируем Excel
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Sales Analysis"

        headers = [
                "Название товара",
                "Цена за единицу",
                "Дата когда партия пришла на склад",
                "Дата когда продали последнюю единицу из партии",
                "Сколько штук было в партии при поступлении",
                "Сколько штук уже продали из этой партии",
                "Сколько штук осталось непроданными в партии",
                "На какую сумму продали товар из этой партии",
                "На какую сумму остался непроданный товар в партии",
                "На какую сумму поступила вся партия изначально",
                "Скольким разным клиентам продали товар из этой партии",
                "Сколько дней прошло от прихода партии до последней продажи",
                "Сколько дней товар отсутствовал после распродажи партии (пока не пришла новая)",
                "Сколько дней товар лежит без продаж (от последней продажи до сегодня или до новой партии)",
                "Общее количество пассивных дней (без остатка + без движения)",
                "На сколько дней хватит текущего остатка при текущей скорости продаж",
                "Сколько штук товара мы упустили из-за отсутствия на складе (могли бы продать)",
                "На какую сумму мы потеряли продажи из-за отсутствия товара"
            ]

        out_ws.append(headers)
        out_ws.freeze_panes = out_ws['A2']
        today = date.today()

        # План по упущенной продаже
        # Считаем среднюю продажу за день для товара по всем партиям
        for product_name, parties in products.items():
            # Общая продажа и дни для расчета средней продажи в день
            total_sold_all = sum(p["sale_qty"] for p in parties)
            total_days_all = sum(
                ((p["sold_out_date"] or p["last_sale_touch"]) - p["purchase_date"]).days
                for p in parties
                if (p["sold_out_date"] or p["last_sale_touch"])
            )
            avg_sale_per_day_product = total_sold_all / total_days_all if total_days_all else 0

            for i, p in enumerate(parties):
                last_unit_date = p["sold_out_date"] if p["sold_out_date"] else p["last_sale_touch"]
                clients_count = sum(1 for q in p["clients_net"].values() if q > 0)
                days_diff = (last_unit_date - p["purchase_date"]).days if last_unit_date and p["purchase_date"] else None

                sum_remaining = p["price"] * p["remaining"]
                sum_purchase = p["price"] * p["purchase_qty"]

                # Дни без остатка
                if p["sold_out_date"]:
                    next_party = parties[i + 1] if i + 1 < len(parties) else None
                    if next_party:
                        days_no_stock = (next_party["purchase_date"] - p["sold_out_date"]).days
                    else:
                        days_no_stock = (today - p["sold_out_date"]).days
                else:
                    days_no_stock = 0

                # Дни без движения
                if not p["last_sale_touch"]:
                    days_no_movement = (today - p["purchase_date"]).days
                else:
                    next_party = parties[i + 1] if i + 1 < len(parties) else None
                    if next_party:
                        days_no_movement = (next_party["purchase_date"] - p["last_sale_touch"]).days
                    else:
                        days_no_movement = (today - p["last_sale_touch"]).days

                total_passive_days = (days_no_stock or 0) + (days_no_movement or 0)

                # План дней при текущем остатке (для текущей партии)
                if p["sale_qty"] > 0 and days_diff:
                    avg_sale_per_day_party = p["sale_qty"] / days_diff
                    plan_days = p["remaining"] / avg_sale_per_day_party
                else:
                    plan_days = None

                # === Новые колонки: упущенные продажи ===
                days_no_stock = max(days_no_stock, 0)
                missed_qty = round(avg_sale_per_day_product * days_no_stock, 2) if avg_sale_per_day_product > 0 else 0
                missed_sum = round(missed_qty * p["price"], 2)

                out_ws.append([
                    product_name,
                    p["price"],
                    p["purchase_date"].strftime("%d.%m.%Y") if p["purchase_date"] else None,
                    last_unit_date.strftime("%d.%m.%Y") if last_unit_date else None,
                    p["purchase_qty"],
                    p["sale_qty"],
                    p["remaining"],
                    p["total_sale"],
                    sum_remaining,
                    sum_purchase,
                    clients_count,
                    days_diff,
                    days_no_stock,
                    days_no_movement,
                    total_passive_days,
                    round(plan_days, 2) if plan_days is not None else None,
                    missed_qty,    # упущенные продажи количество
                    missed_sum     # упущенные продажи сумма
                ])

        # Заливка ячеек по количеству дней
        all_days = []
        for parties in products.values():
            for p in parties:
                d = p["sold_out_date"] if p["sold_out_date"] else p["last_sale_touch"]
                if d and p["purchase_date"]:
                    all_days.append((d - p["purchase_date"]).days)

        min_days = min(all_days) if all_days else 0
        max_days = max(all_days) if all_days else 1

        def get_fill(days):
            if days is None:
                return PatternFill(fill_type=None)
            ratio = (days - min_days) / (max_days - min_days + 1e-5)
            red = int(255 * ratio)
            green = int(255 * (1 - ratio))
            hex_color = f"{red:02X}{green:02X}00"
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        for row_idx in range(2, out_ws.max_row + 1):
            purchase_cell = out_ws.cell(row=row_idx, column=3).value
            last_unit_cell = out_ws.cell(row=row_idx, column=4).value

            pd = _parse_date1(purchase_cell)
            ld = _parse_date1(last_unit_cell)
            days = (ld - pd).days if (pd and ld) else None

            fill = get_fill(days)
            for col in range(1, len(headers) + 1):
                out_ws.cell(row=row_idx, column=col).fill = fill

        output = BytesIO()
        out_wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sales_analysis_by_party.xlsx"'
        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return HttpResponse(f"Ошибка при обработке файла: {e}", status=500)








# polnyyk kod kotoryy dal mne claude.ai
def _parse_date2(cell):
    """Поддерживаем datetime, 'dd.mm.YYYY', 'dd.mm.yy' и пустые значения."""
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return cell.date()
    s = str(cell).strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def _to_float(cell):
    try:
        return float(cell) if cell is not None else 0.0
    except Exception:
        return 0.0




@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_sales_excel_for_analis_with_return(request):
    file = request.FILES.get("file")
    if not file:
        return HttpResponse("Файл не получен", status=400)

    try:
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.active

        products = {}

        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] and "Hemmesi" in str(row[0]):
                break

            remark = str(row[2]).strip().lower() if row[2] else ""
            document = str(row[3] or "")
            date_cell = row[1]
            product_name = row[9]
            qty_in_row = _to_float(row[12])
            price = _to_float(row[11])
            client = row[4]

            if not product_name or date_cell is None:
                continue

            date_ = _parse_date2(date_cell)
            if date_ is None:
                continue

            if product_name not in products:
                products[product_name] = []

            # Приход
            if "Girdeji faktura" in document:
                if "wozwrat" in remark:
                    continue
                purchase_qty = qty_in_row
                party = {
                    "purchase_date": date_,
                    "purchase_qty": purchase_qty,
                    "remaining": purchase_qty,
                    "sale_qty": 0.0,
                    "total_sale": 0.0,
                    "price": price,
                    "last_sale_touch": None,
                    "sold_out_date": None,
                    "clients_net": {}
                }
                products[product_name].append(party)

            # Продажа / возврат
            elif "Söwda" in document:
                is_return = "wozwrat" in remark
                qty_to_process = qty_in_row

                if not products.get(product_name):
                    continue

                for party in products[product_name]:
                    if qty_to_process <= 0:
                        break

                    if is_return:
                        can_take_back = min(party["sale_qty"], qty_to_process)
                        if can_take_back <= 0:
                            continue
                        party["sale_qty"] -= can_take_back
                        party["total_sale"] -= can_take_back * price
                        party["remaining"] += can_take_back
                        if client:
                            party["clients_net"][client] = party["clients_net"].get(client, 0.0) - can_take_back
                        if party["remaining"] > 0:
                            party["sold_out_date"] = None
                        qty_to_process -= can_take_back
                    else:
                        if party["remaining"] <= 0:
                            continue
                        take = min(party["remaining"], qty_to_process)
                        party["remaining"] -= take
                        party["sale_qty"] += take
                        party["total_sale"] += take * price
                        party["last_sale_touch"] = date_ if not party["last_sale_touch"] else max(party["last_sale_touch"], date_)
                        if client:
                            party["clients_net"][client] = party["clients_net"].get(client, 0.0) + take
                        if party["remaining"] == 0:
                            party["sold_out_date"] = date_
                        qty_to_process -= take

        # Расчет упущенных продаж ПО ТОВАРУ
        today = date.today()
        product_missed_sales = {}

        for product_name, parties in products.items():
            if not parties:
                continue
            
            # Сортируем партии по дате прихода
            sorted_parties = sorted(parties, key=lambda x: x["purchase_date"])
            
            # 1. Находим периоды когда товар был полностью распродан
            total_days_out_of_stock = 0
            
            for i, party in enumerate(sorted_parties):
                if party["sold_out_date"]:
                    next_party = sorted_parties[i + 1] if i + 1 < len(sorted_parties) else None
                    if next_party:
                        # Период между распродажей и следующим приходом
                        gap_days = (next_party["purchase_date"] - party["sold_out_date"]).days
                        total_days_out_of_stock += max(0, gap_days)
                    else:
                        # Последняя партия распродана, товара нет до сегодня
                        gap_days = (today - party["sold_out_date"]).days
                        total_days_out_of_stock += max(0, gap_days)
            
            # 2. Считаем среднюю продажу в день (только когда товар БЫЛ)
            total_sold = sum(p["sale_qty"] for p in sorted_parties)
            
            # Дни когда товар был в продаже
            if sorted_parties[0]["purchase_date"] and sorted_parties[-1].get("sold_out_date"):
                # Если последняя партия распродана
                total_period = (sorted_parties[-1]["sold_out_date"] - sorted_parties[0]["purchase_date"]).days
            elif sorted_parties[0]["purchase_date"] and sorted_parties[-1].get("last_sale_touch"):
                # Если последняя партия еще есть
                total_period = (sorted_parties[-1]["last_sale_touch"] - sorted_parties[0]["purchase_date"]).days
            else:
                total_period = 0
            
            # Вычитаем дни отсутствия товара
            days_with_stock = max(1, total_period - total_days_out_of_stock)
            
            # Средняя продажа в день
            avg_sale_per_day = total_sold / days_with_stock if days_with_stock > 0 else 0
            
            # 3. Упущенные продажи = средняя * дни без товара
            missed_qty = round(avg_sale_per_day * total_days_out_of_stock, 2)
            missed_sum = round(missed_qty * (sorted_parties[0]["price"] if sorted_parties else 0), 2)
            
            product_missed_sales[product_name] = {
                "missed_qty": missed_qty,
                "missed_sum": missed_sum,
                "days_out_of_stock": total_days_out_of_stock
            }

        # Формируем Excel
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Sales Analysis"

        headers = [
                "Название товара",
                "Цена за единицу",
                "Дата когда партия пришла на склад",
                "Дата когда продали последнюю единицу из партии",
                "Сколько штук было в партии при поступлении",
                "Сколько штук уже продали из этой партии",
                "Сколько штук осталось непроданными в партии",
                "На какую сумму продали товар из этой партии",
                "На какую сумму остался непроданный товар в партии",
                "На какую сумму поступила вся партия изначально",
                "Скольким разным клиентам продали товар из этой партии",
                "Сколько дней прошло от прихода партии до последней продажи",
                "Сколько дней товар отсутствовал после распродажи партии (пока не пришла новая)",
                "Сколько дней товар лежит без продаж (от последней продажи до сегодня или до новой партии)",
                "Общее количество пассивных дней (без остатка + без движения)",
                "На сколько дней хватит текущего остатка при текущей скорости продаж",
                "Сколько штук товара мы упустили из-за отсутствия на складе (могли бы продать)",
                "На какую сумму мы потеряли продажи из-за отсутствия товара"
            ]

        out_ws.append(headers)
        out_ws.freeze_panes = out_ws['A2']

        # План по упущенной продаже
        for product_name, parties in products.items():
            missed_data = product_missed_sales.get(product_name, {"missed_qty": 0, "missed_sum": 0})
            
            for i, p in enumerate(parties):
                last_unit_date = p["sold_out_date"] if p["sold_out_date"] else p["last_sale_touch"]
                clients_count = sum(1 for q in p["clients_net"].values() if q > 0)
                days_diff = (last_unit_date - p["purchase_date"]).days if last_unit_date and p["purchase_date"] else None

                sum_remaining = p["price"] * p["remaining"]
                sum_purchase = p["price"] * p["purchase_qty"]

                # Дни без остатка
                if p["sold_out_date"]:
                    next_party = parties[i + 1] if i + 1 < len(parties) else None
                    if next_party:
                        days_no_stock = (next_party["purchase_date"] - p["sold_out_date"]).days
                    else:
                        days_no_stock = (today - p["sold_out_date"]).days
                else:
                    days_no_stock = 0

                # Дни без движения
                if not p["last_sale_touch"]:
                    days_no_movement = (today - p["purchase_date"]).days
                else:
                    next_party = parties[i + 1] if i + 1 < len(parties) else None
                    if next_party:
                        days_no_movement = (next_party["purchase_date"] - p["last_sale_touch"]).days
                    else:
                        days_no_movement = (today - p["last_sale_touch"]).days

                total_passive_days = (days_no_stock or 0) + (days_no_movement or 0)

                # План дней при текущем остатке (для текущей партии)
                if p["sale_qty"] > 0 and days_diff:
                    avg_sale_per_day_party = p["sale_qty"] / days_diff
                    plan_days = p["remaining"] / avg_sale_per_day_party
                else:
                    plan_days = None

                # Упущенные продажи показываем только для ПЕРВОЙ партии товара
                if i == 0:
                    missed_qty = missed_data["missed_qty"]
                    missed_sum = missed_data["missed_sum"]
                else:
                    missed_qty = 0
                    missed_sum = 0

                out_ws.append([
                    product_name,
                    p["price"],
                    p["purchase_date"].strftime("%d.%m.%Y") if p["purchase_date"] else None,
                    last_unit_date.strftime("%d.%m.%Y") if last_unit_date else None,
                    p["purchase_qty"],
                    p["sale_qty"],
                    p["remaining"],
                    p["total_sale"],
                    sum_remaining,
                    sum_purchase,
                    clients_count,
                    days_diff,
                    days_no_stock,
                    days_no_movement,
                    total_passive_days,
                    round(plan_days, 2) if plan_days is not None else None,
                    missed_qty,
                    missed_sum
                ])

        # Заливка ячеек по количеству дней
        all_days = []
        for parties in products.values():
            for p in parties:
                d = p["sold_out_date"] if p["sold_out_date"] else p["last_sale_touch"]
                if d and p["purchase_date"]:
                    all_days.append((d - p["purchase_date"]).days)

        min_days = min(all_days) if all_days else 0
        max_days = max(all_days) if all_days else 1

        def get_fill(days):
            if days is None:
                return PatternFill(fill_type=None)
            ratio = (days - min_days) / (max_days - min_days + 1e-5)
            red = int(255 * ratio)
            green = int(255 * (1 - ratio))
            hex_color = f"{red:02X}{green:02X}00"
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        for row_idx in range(2, out_ws.max_row + 1):
            purchase_cell = out_ws.cell(row=row_idx, column=3).value
            last_unit_cell = out_ws.cell(row=row_idx, column=4).value
            pd = parse_excel_date(purchase_cell)
            ld = parse_excel_date(last_unit_cell)
            days = (ld - pd).days if (pd and ld) else None
            fill = get_fill(days)
            for col in range(1, len(headers) + 1):
                out_ws.cell(row=row_idx, column=col).fill = fill

        output = BytesIO()
        out_wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sales_analysis_by_party.xlsx"'
        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return HttpResponse(f"Ошибка при обработке файла: {e}", status=500)



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def get_cards(request):
    data = request.data
    date_from = data.get('date_from')
    date_to = data.get('date_to')

    if not date_from or not date_to:
        return JsonResponse({"status": "error", "message": "choose date"}, status=400)

    try:
        date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
        date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"status": "error", "message": "invalid date format"}, status=400)

    if date_to_obj < date_from_obj:
        return JsonResponse({"status": "error", "message": "To date must be after From date"}, status=400)

    result = []

    # --- 🔹 Перебираем все активные счета
    for acc in Account.objects.filter(is_active=True).order_by("number"):

        # --- 🔹 Находим всех потомков (включая сам счёт)
        all_accounts = Account.objects.filter(
            Q(id=acc.id) | Q(parent=acc) | Q(parent__parent=acc)
        )

        # ---- 1️⃣ Сальдо на начало ----
        before_entries = Entry.objects.filter(
            account__in=all_accounts,
            transaction__date__lt=date_from_obj
        ).aggregate(
            debit=Sum('debit', default=Decimal('0.00')),
            credit=Sum('credit', default=Decimal('0.00'))
        )
        saldo_start = before_entries['debit'] - before_entries['credit']

        # ---- 2️⃣ Проводки за период ----
        period_entries = Entry.objects.filter(
            account__in=all_accounts,
            transaction__date__range=[date_from_obj, date_to_obj]
        ).select_related('transaction').order_by('transaction__date')

        movements_dict = {}
        for e in period_entries:
            t_id = e.transaction.id
            if t_id not in movements_dict:
                movements_dict[t_id] = {
                    "date": e.transaction.date.strftime("%d.%m.%Y"),
                    "description": e.transaction.description,
                    "debit": float(e.debit),
                    "credit": float(e.credit),
                }
            else:
                movements_dict[t_id]['debit'] += float(e.debit)
                movements_dict[t_id]['credit'] += float(e.credit)

        movements = list(movements_dict.values())

        # ---- 3️⃣ Обороты за период ----
        period_sum = period_entries.aggregate(
            debit=Sum('debit', default=Decimal('0.00')),
            credit=Sum('credit', default=Decimal('0.00'))
        )

        # ---- 4️⃣ Сальдо на конец ----
        saldo_end = saldo_start + period_sum['debit'] - period_sum['credit']

        # ---- ✅ Пропускаем счета с нулевым сальдо на начало И конец И без движений ----
        if saldo_start == 0 and saldo_end == 0 and not movements:
            continue

        # ---- ✅ Добавляем в результат
        result.append({
            "id": acc.id,
            "account": f"{acc.number} {acc.name}",
            "saldo_start": float(saldo_start),
            "movements": movements,
            "debit_turnover": float(period_sum['debit']),
            "credit_turnover": float(period_sum['credit']),
            "saldo_end": float(saldo_end),
        })

    return Response({
        "date_from": date_from,
        "date_to": date_to,
        "accounts": result,
    })


@api_view(['GET'])
def get_account_cards(request, id):
    date_from_str = request.GET.get('dateFrom')
    date_to_str = request.GET.get('dateTo')
    date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date() if date_from_str else None
    date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date() if date_to_str else None

    account = Account.objects.get(id=id)
    
    
    
    # my_transactions = Transaction.objects.filter(date__range=[date_from, date_to],)
    
    
    
    
    

    # Собираем все дочерние счета рекурсивно
    def get_all_children(acc):
        children = list(acc.children.all())
        all_accs = [acc]
        for c in children:
            all_accs.extend(get_all_children(c))
        return all_accs
    


    all_accounts = get_all_children(account)
    account_ids = [a.id for a in all_accounts]
    

    # выбираем все транзакции за период, где есть проводки по account_ids
    transactions = (
        Transaction.objects.filter(
            date__range=[date_from, date_to],
            entries__account__in=account_ids
        )
        .distinct()
        .select_related('partner')
        .prefetch_related('entries')
        .order_by('date', 'id')
    )
    

    # группируем по партнёрам
    partner_groups = defaultdict(list)
    for tr in transactions:
        partner_id = tr.partner_id or 0
        partner_groups[partner_id].append(tr)
        


    cards = []
    for partner_id, trans in partner_groups.items():
        partner_name = trans[0].partner.name if partner_id else ""
        saldo_start = Decimal(0)
        debit_turnover = Decimal(0)
        credit_turnover = Decimal(0)
        current_saldo = Decimal(saldo_start)
        
        saldo_start_data = Entry.objects.filter(
            account__in=account_ids,
            transaction__partner_id=partner_id,
            transaction__date__lt=date_from
        ).aggregate(
            debit_sum=Sum('debit', default=Decimal(0)),
            credit_sum=Sum('credit', default=Decimal(0))
        )
        saldo_start = saldo_start_data['debit_sum'] - saldo_start_data['credit_sum']
        current_saldo = saldo_start
        



        movements = []
        seen = set()
        for tr in trans:
            if tr.id in seen:
                continue
            seen.add(tr.id)

            # суммируем дебет/кредит по всем выбранным счетам
            sums = tr.entries.filter(account__in=account_ids).aggregate(
                debit_sum=Sum('debit', default=Decimal(0)),
                credit_sum=Sum('credit', default=Decimal(0))
            )
            debit = sums['debit_sum'] or Decimal(0)
            credit = sums['credit_sum'] or Decimal(0)

            current_saldo += debit - credit
            debit_turnover += debit
            credit_turnover += credit

            movements.append({
                "date": tr.date.strftime("%d.%m.%Y"),
                "description": tr.description or "",
                "debit": f"{debit:.2f}" if debit else "",
                "credit": f"{credit:.2f}" if credit else "",
                "saldo": f"{current_saldo:.2f}",
            })

        card = {
            "account": account.name,
            "partner": partner_name,
            "date_from": date_from.strftime("%d.%m.%Y"),
            "date_to": date_to.strftime("%d.%m.%Y"),
            "saldo_start": f"{saldo_start:.2f}",
            "saldo_end": f"{current_saldo:.2f}",
            "debit_turnover": f"{debit_turnover:.2f}",
            "credit_turnover": f"{credit_turnover:.2f}",
            "movements": movements,
        }
        cards.append(card)

    return Response(cards)

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_account_cards(request, id):
#     date_from = request.GET.get('dateFrom')
#     date_to = request.GET.get('dateTo')

#     # Родительский счёт
#     parent_account = Account.objects.get(id=id)

#     # Все дочерние счета
#     child_accounts = parent_account.children.all()
#     all_accounts = [parent_account] + list(child_accounts)

#     # Сначала собираем все движения за период
#     movements_by_partner = {}

#     for acc in all_accounts:
#         entries = Entry.objects.filter(
#             account=acc,
#             transaction__date__gte=date_from,
#             transaction__date__lte=date_to
#         ).select_related('transaction', 'transaction__invoice', 'transaction__partner')

#         for e in entries:
#             partner_name = e.transaction.partner.name if e.transaction.partner else "Без партнёра"
#             if partner_name not in movements_by_partner:
#                 movements_by_partner[partner_name] = {
#                     "id": parent_account.id,
#                     "account": str(parent_account),
#                     "partner": partner_name,
#                     "date_from": date_from,
#                     "date_to": date_to,
#                     "saldo_start": Decimal('0.00'),
#                     "debit_turnover": Decimal('0.00'),
#                     "credit_turnover": Decimal('0.00'),
#                     "saldo_end": Decimal('0.00'),
#                     "movements": []
#                 }

#             debit = e.debit or Decimal('0.00')
#             credit = e.credit or Decimal('0.00')
#             description = ""
#             if e.transaction.invoice:
#                 description = f"Faktura #{e.transaction.invoice.id}"
#                 if e.transaction.invoice.comment:
#                     description += f" {e.transaction.invoice.comment}"
#             else:
#                 description = e.transaction.description

#             movements_by_partner[partner_name]["movements"].append({
#                 "date": e.transaction.date.strftime('%d.%m.%Y'),
#                 "description": description,
#                 "debit": debit,
#                 "credit": credit
#             })

#             # Считаем обороты
#             movements_by_partner[partner_name]["debit_turnover"] += debit
#             movements_by_partner[partner_name]["credit_turnover"] += credit

#     # Считаем saldo_start и saldo_end для каждого партнёра
#     for partner_name, data in movements_by_partner.items():
#         saldo_start = Decimal('0.00')
#         # Начальный остаток на date_from
#         for acc in all_accounts:
#             entries_before = Entry.objects.filter(
#                 account=acc,
#                 transaction__date__lt=date_from,
#                 transaction__partner__name=partner_name if partner_name != "Без партнёра" else None
#             )
#             sums = entries_before.aggregate(
#                 debit_sum=Sum('debit'), credit_sum=Sum('credit')
#             )
#             debit_sum = sums['debit_sum'] or Decimal('0.00')
#             credit_sum = sums['credit_sum'] or Decimal('0.00')
#             saldo_start += debit_sum - credit_sum
#         data["saldo_start"] = saldo_start

#         # Текущий остаток по движениям
#         current_saldo = saldo_start
#         for m in sorted(data["movements"], key=lambda x: x['date']):
#             current_saldo += m['debit'] - m['credit']
#             m['saldo'] = current_saldo

#         data["saldo_end"] = current_saldo

#     # Возвращаем список карточек по партнёрам
#     return Response(list(movements_by_partner.values()))


