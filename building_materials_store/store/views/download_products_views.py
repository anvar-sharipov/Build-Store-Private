
from django.http import JsonResponse, HttpResponse
# from django.views.decorators.http import require_GET, require_POST
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Q, Sum, Count, F, DecimalField, ExpressionWrapper, Prefetch
from datetime import datetime, timedelta
# from django.utils.dateparse import parse_date

# from django.db.models import F
# from django.contrib.postgres.search import TrigramSimilarity
from ..models import Account, Entry, DayClosing, PartnerBalanceSnapshot, Partner, WarehouseAccount, Product, InvoiceItem
from icecream import ic
# import json
# from collections import defaultdict

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from django.http import FileResponse
# import os
# from io import BytesIO
# from django.core.files.base import ContentFile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import CellIsRule
# from openpyxl.worksheet.page import PageMargins

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from io import BytesIO




from ..my_func.get_unit_map import get_unit_map 
from ..my_func.get_unit_and_cf import get_unit_and_cf 
from ..my_func.str_to_int_list import str_to_int_list
from ..my_func.date_convert_for_excel import format_date_ru 




# ================== СТИЛИ ==================

# Шапка
HEADER_FILL = PatternFill(fill_type="solid",fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN_WRAP = Alignment(vertical="center",horizontal="left",wrap_text=True)

# Категория
CATEGORY_FILL = PatternFill(fill_type="solid", fgColor="E7E6E6")
CATEGORY_FONT = Font(bold=True)

# Обычные ячейки
NORMAL_FONT = Font(color="000000")
LEFT_ALIGN = Alignment(vertical="center", horizontal="left")
RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

# Границы
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

LEFT_BORDER = Border(left=Side(style="thin"))
RIGHT_BORDER = Border(right=Side(style="thin"))
TOP_BORDER = Border(top=Side(style="thin"))
BOTTOM_BORDER = Border(bottom=Side(style="thin"))


# Форматы чисел
PRICE_FMT = '#,##0.00'
QTY_FMT = '#,##0.###'

# Ширина колонок
COLUMN_WIDTHS = {
    "A": 25,
    "B": 40,
    "C": 10,
    "D": 12,
    "E": 8, "F": 8,
    "G": 8, "H": 8,
    "I": 8, "J": 8,
    "K": 8, "L": 8,
    "M": 8, "N": 8,
}

TOTAL_FILL = PatternFill(fill_type="solid",fgColor="D9D9D9")

TOTAL_FONT = Font(bold=True)

GRAY_FILL_0 = PatternFill(fill_type="solid", fgColor="F5F5F5")
GRAY_FILL_1 = PatternFill(fill_type="solid", fgColor="EDEDED")
GRAY_FILL_2 = PatternFill( fill_type="solid", fgColor="DCDCDC")
GRAY_FILL_3 = PatternFill(fill_type="solid", fgColor="C8C8C8")
GRAY_FILL_4 = PatternFill(fill_type="solid", fgColor="B0B0B0")

GREEN_FILL_0 = PatternFill(fill_type="solid", fgColor="E2F0D9")
GREEN_FILL_1 = PatternFill(fill_type="solid", fgColor="C6EFCE")
GREEN_FILL_2 = PatternFill(fill_type="solid", fgColor="92D050")
GREEN_FILL_3 = PatternFill(fill_type="solid", fgColor="006100")

RED_FONT = Font(color="FF0000")
GREEN_FONT = Font(color="006400")
BLUE_FONT = Font(color="0000FF")

def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)




@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_excel_products_diapazon(request):
    
    
    
    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")
    
    warehouse = request.GET.get("warehouse", "")
    categories = request.GET.get("categories", "")
    search = request.GET.get("search", "")
    is_active = request.GET.get("is_active", "")
    
    warehouse_list = str_to_int_list(warehouse)
    categories_list = str_to_int_list(categories)
  
    
    
    # ic(warehouse_list)
    # ic(categories_list)
    # ic(search)
    # ic(is_active)
    # ic(request)
    

    if not date_from or not date_to:
        return Response(
            {"error": "choose diapazon date"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        day_start = datetime.strptime(date_from, "%Y-%m-%d").date()
        day_end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return Response(
            {"error": "invalid diapazon date format"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if day_start > day_end:
        return Response(
            {"error": "date start must be <= date end"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    day_start_str = day_start.strftime("%d.%m.%Y")
    day_end_str = day_end.strftime("%d.%m.%Y")
    
    if not warehouse_list:
        # ic('tut')
        return Response(
            {"error": "select_warehouse"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    
    # ic(day_start)
    # ic(day_end)
    
    products = Product.objects.filter(
        Q(warehouse_products__warehouse_id__in=warehouse_list) |
        Q(invoiceitem__invoice__warehouse_id__in=warehouse_list) |
        Q(invoiceitem__invoice__warehouse2_id__in=warehouse_list)
    ).distinct()
    
    if search:
        products = products.filter(name__icontains=search)
        
    if categories_list:
        products = products.filter(category__id__in=categories_list)
        
    unit_map = get_unit_map()
    account_40_42 = {}
    for p in products:
        if not p.category:
            continue
        unit, cf = get_unit_and_cf(unit_map, p)
        category_id = p.category.id
        category_name = p.category.name
        # if p.id == 601:
        #     ic(p.name, unit, cf)
        if category_id not in account_40_42:
            account_40_42[category_id] = {
                "category": {
                    "id": category_id,
                    "name": category_name,
                },
                "totals": {
                    "start_qty": Decimal("0.00"),
                    "start_price": Decimal("0.00"),

                    "prihod_qty": Decimal("0.00"),
                    "prihod_price": Decimal("0.00"),

                    "wozwrat_qty": Decimal("0.00"),
                    "wozwrat_price": Decimal("0.00"),

                    "rashod_qty": Decimal("0.00"),
                    "rashod_price": Decimal("0.00"),

                    "end_qty": Decimal("0.00"),
                    "end_price": Decimal("0.00"),
                },
                "products": {}
            }

        account_40_42[category_id]["products"][p.id] = {
            "product": {
                "id": p.id,
                "name": p.name,
                "unit": unit,
                "cf": Decimal(cf),
                "wholsale_price": p.wholesale_price,
            },

            "start_qty": Decimal("0.00"),
            "start_price": Decimal("0.00"),

            "prihod_qty": Decimal("0.00"),
            "prihod_price": Decimal("0.00"),

            "wozwrat_qty": Decimal("0.00"),
            "wozwrat_price": Decimal("0.00"),

            "rashod_qty": Decimal("0.00"),
            "rashod_price": Decimal("0.00"),

            "end_qty": Decimal("0.00"),
            "end_price": Decimal("0.00"),
        }
        
    start_items = InvoiceItem.objects.filter(
        invoice__entry_created_at_handle__lt=day_start,
        invoice__canceled_at__isnull=True
    ).filter(
        Q(invoice__warehouse_id__in=warehouse_list) |
        Q(invoice__warehouse2_id__in=warehouse_list)
    ).select_related(
        "product", "product__category", "product__base_unit", "invoice"
    )
    
    grand_total = {
        "start_qty": Decimal("0.00"),
        "start_price": Decimal("0.00"),

        "prihod_qty": Decimal("0.00"),
        "prihod_price": Decimal("0.00"),

        "wozwrat_qty": Decimal("0.00"),
        "wozwrat_price": Decimal("0.00"),

        "rashod_qty": Decimal("0.00"),
        "rashod_price": Decimal("0.00"),

        "end_qty": Decimal("0.00"),
        "end_price": Decimal("0.00"),
    }
            
        
    for item in start_items:
        p = item.product
        inv = item.invoice
        
        # Товар уже инициализирован выше
        if not p.category:
            continue
        if (p.category.id not in account_40_42 or p.id not in account_40_42[p.category.id]["products"]):
            continue
        cat_totals = account_40_42[p.category.id]["totals"]
        prod = account_40_42[p.category.id]["products"][p.id]
        cf = prod["product"]["cf"]
        qty = Decimal(item.selected_quantity) / cf
        calculated_price = qty * Decimal(p.wholesale_price)
        # Проверяем принадлежность к выбранным складам
        if inv.wozwrat_or_prihod == "prihod":
            if inv.warehouse_id in warehouse_list:
                prod["start_qty"] += qty
                prod["start_price"] += calculated_price
                
                cat_totals["start_qty"] += qty
                cat_totals["start_price"] += calculated_price
                
                grand_total["start_qty"] += qty
                grand_total["start_price"] += calculated_price

        elif inv.wozwrat_or_prihod == "rashod":
            if inv.warehouse_id in warehouse_list:
                prod["start_qty"] -= qty
                prod["start_price"] -= calculated_price
                
                cat_totals["start_qty"] -= qty
                cat_totals["start_price"] -= calculated_price
                
                grand_total["start_qty"] -= qty
                grand_total["start_price"] -= calculated_price

        elif inv.wozwrat_or_prihod == "wozwrat":
            if inv.warehouse_id in warehouse_list:
                prod["start_qty"] += qty
                prod["start_price"] += calculated_price
                
                cat_totals["start_qty"] += qty
                cat_totals["start_price"] += calculated_price
                
                grand_total["start_qty"] += qty
                grand_total["start_price"] += calculated_price
            
        elif inv.wozwrat_or_prihod == "transfer":
            # если со склада (и склад в выбранных)
            if inv.warehouse_id in warehouse_list:
                prod["start_qty"] -= qty
                prod["start_price"] -= calculated_price
                
                cat_totals["start_qty"] -= qty
                cat_totals["start_price"] -= calculated_price
                
                grand_total["start_qty"] -= qty
                grand_total["start_price"] -= calculated_price
                
            # если на склад (и склад в выбранных)
            elif inv.warehouse2_id in warehouse_list:
                prod["start_qty"] += qty
                prod["start_price"] += calculated_price
                
                cat_totals["start_qty"] += qty
                cat_totals["start_price"] += calculated_price
                
                grand_total["start_qty"] += qty
                grand_total["start_price"] += calculated_price
                
    turnover_items = InvoiceItem.objects.filter(
        invoice__entry_created_at_handle__gte=day_start,
        invoice__entry_created_at_handle__lte=day_end,
        invoice__canceled_at__isnull=True
    ).filter(
        Q(invoice__warehouse_id__in=warehouse_list) |
        Q(invoice__warehouse2_id__in=warehouse_list)
    ).select_related(
        "product", "product__category", "product__base_unit", "invoice"
    ).order_by(
        "product__category__name", "product__name"
    )
    
    for item in turnover_items:
        p = item.product
        inv = item.invoice
        
        # Товар уже инициализирован выше
        if not p.category:
            continue
        if (p.category.id not in account_40_42 or p.id not in account_40_42[p.category.id]["products"]):
            continue
        cat_totals = account_40_42[p.category.id]["totals"]
        prod = account_40_42[p.category.id]["products"][p.id]
        cf = prod["product"]["cf"]
        qty = Decimal(item.selected_quantity) / cf
        calculated_price = qty * Decimal(item.selected_price)
        # Проверяем принадлежность к выбранным складам
        if inv.wozwrat_or_prihod == "prihod":
            if inv.warehouse_id in warehouse_list:
                prod["prihod_qty"] += qty
                prod["prihod_price"] += calculated_price
                
                cat_totals["prihod_qty"] += qty
                cat_totals["prihod_price"] += calculated_price
                
                grand_total["prihod_qty"] += qty
                grand_total["prihod_price"] += calculated_price

        elif inv.wozwrat_or_prihod == "rashod":
            if inv.warehouse_id in warehouse_list:
                prod["rashod_qty"] += qty
                prod["rashod_price"] += calculated_price
                
                cat_totals["rashod_qty"] += qty
                cat_totals["rashod_price"] += calculated_price
                
                grand_total["rashod_qty"] += qty
                grand_total["rashod_price"] += calculated_price

        elif inv.wozwrat_or_prihod == "wozwrat":
            if inv.warehouse_id in warehouse_list:
                prod["wozwrat_qty"] += qty
                prod["wozwrat_price"] += calculated_price
                
                cat_totals["wozwrat_qty"] += qty
                cat_totals["wozwrat_price"] += calculated_price
                
                grand_total["wozwrat_qty"] += qty
                grand_total["wozwrat_price"] += calculated_price
            
        elif inv.wozwrat_or_prihod == "transfer":
            # если со склада (и склад в выбранных)
            if inv.warehouse_id in warehouse_list:
                prod["rashod_qty"] += qty
                prod["rashod_price"] += calculated_price
                
                cat_totals["rashod_qty"] += qty
                cat_totals["rashod_price"] += calculated_price
                
                grand_total["rashod_qty"] += qty
                grand_total["rashod_price"] += calculated_price
                
            # если на склад (и склад в выбранных)
            elif inv.warehouse2_id in warehouse_list:
                prod["prihod_qty"] += qty
                prod["prihod_price"] += calculated_price
                
                cat_totals["prihod_qty"] += qty
                cat_totals["prihod_price"] += calculated_price
                
                grand_total["prihod_qty"] += qty
                grand_total["prihod_price"] += calculated_price
                
    for cat in account_40_42.values():
        for prod in cat["products"].values():
            prod["end_qty"] = (
                prod["start_qty"]
                + prod["prihod_qty"]
                - prod["rashod_qty"]
                + prod["wozwrat_qty"]
            )

            prod["end_price"] = (
                prod["start_price"]
                + prod["prihod_price"]
                - prod["rashod_price"]
                + prod["wozwrat_price"]
            )
            
    for cat in account_40_42.values():
        totals = cat["totals"]
        totals["end_qty"] = (
            totals["start_qty"]
            + totals["prihod_qty"]
            - totals["rashod_qty"]
            + totals["wozwrat_qty"]
        )
        totals["end_price"] = (
            totals["start_price"]
            + totals["prihod_price"]
            - totals["rashod_price"]
            + totals["wozwrat_price"]
        )
        
    grand_total["end_qty"] = (
        grand_total["start_qty"]
        + grand_total["prihod_qty"]
        - grand_total["rashod_qty"]
        + grand_total["wozwrat_qty"]
    )
    grand_total["end_price"] = (
        grand_total["start_price"]
        + grand_total["prihod_price"]
        - grand_total["rashod_price"]
        + grand_total["wozwrat_price"]
    )
    
    w_acc = (
        WarehouseAccount.objects
        .select_related("warehouse", "account")
        .filter(warehouse_id__in=warehouse_list)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Бух. Оборот товаров"

    account_warehouse = ", ".join(
        sorted({f"{w.account.number} — {w.warehouse.name}" for w in w_acc})
    )

    dateFrom_ru = format_date_ru(date_from)
    dateTo_ru = format_date_ru(date_to)

    # === Заголовок ===
    ws.merge_cells("A2:N2")
    ws.merge_cells("A3:N3")
    ws.merge_cells("A4:N4")

    ws["A2"] = "Бухгалтерский оборот товаров"
    ws["A3"] = f"Склад(ы): {account_warehouse}"
    ws["A4"] = f"Период: {dateFrom_ru} — {dateTo_ru}"

    for cell in ["A2", "A3", "A4"]:
        ws[cell].font = CATEGORY_FONT
        ws[cell].alignment = CENTER_ALIGN
        
    ws.freeze_panes = "A7"
    
    ws["A5"] = "№"
    ws.column_dimensions["A"].width = 8
    ws["B5"] = "Наименование товара"
    ws.column_dimensions["B"].width = 55
    ws["C5"] = "Ед."
    ws.column_dimensions["C"].width = 8
    
    ws["D5"] = "Цена"
    ws.column_dimensions["D"].width = 8
    ws[f"D5"].number_format = PRICE_FMT
    
    ws.merge_cells("E5:F5")
    ws["E5"] = "Остаток на начало"


    
    ws.merge_cells("G5:H5")
    ws["G5"] = "Приход"
    
    
    ws.merge_cells("I5:J5")
    ws["I5"] = "Возврат"
    
    ws.merge_cells("K5:L5")
    ws["K5"] = "Расход"
    
    ws.merge_cells("M5:N5")
    ws["M5"] = "Остаток на конец"

    
    
    for i in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", 
                "A6", "B6", "C6", "D6", "E6", "F6", "G6", "H6", "I6", "J6", "K6", "L6", "M6", "N6"]:
        ws[i].font = CATEGORY_FONT
        ws[i].alignment = CENTER_ALIGN
        ws[i].fill = GRAY_FILL_1
        ws[i].border = THIN_BORDER
        
        
    for i in ["E", "F",  "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws.column_dimensions[i].width = 11

        

    ws["E6"] = "Кол-во"
    ws["F6"] = "Всего"
    
    ws["G6"] = "Кол-во"
    ws["H6"] = "Всего"
    
    ws["I6"] = "Кол-во"
    ws["J6"] = "Всего"
    
    ws["K6"] = "Кол-во"
    ws["L6"] = "Всего"
    
    ws["M6"] = "Кол-во"
    ws["N6"] = "Всего"
    
    
    row = 7
    count = 1
    for cat_id, values in sorted(
        account_40_42.items(),
        key=lambda item: item[1]["category"]["name"].lower()
    ):
        cat_name = values["category"]["name"]
        totals = values["totals"]
        ws.merge_cells(f"A{row}:N{row}")
        ws[f"A{row}"].fill = GREEN_FILL_0
        ws[f"A{row}"].font = CATEGORY_FONT
        ws[f"A{row}"].alignment = LEFT_ALIGN
        ws[f"A{row}"] = cat_name
        
        
        
        for col in "FHJLN":
            ws[f"{col}{row}"].number_format = PRICE_FMT
            
        for col in "EGIKM":
            ws[f"{col}{row}"].number_format = QTY_FMT
        
        row += 1
        
        for products, value in values["products"].items():
            product = value["product"]
            unit = product["unit"]
            wholesale_price = product["wholsale_price"]
            
            start_qty = value["start_qty"]
            start_price = value["start_price"]
            
            prihod_qty = value["prihod_qty"]
            prihod_price = value["prihod_price"]
            
            
            wozwrat_qty = value["wozwrat_qty"]
            wozwrat_price = value["wozwrat_price"]
            
            rashod_qty = value["rashod_qty"]
            rashod_price = value["rashod_price"]
            
            end_qty = value["end_qty"]
            end_price = value["end_price"] 
            
            ws[f"G{row}"].font = GREEN_FONT
            ws[f"H{row}"].font = GREEN_FONT
            ws[f"I{row}"].font = RED_FONT
            ws[f"J{row}"].font = RED_FONT
            ws[f"K{row}"].font = BLUE_FONT
            ws[f"L{row}"].font = BLUE_FONT
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws[i].border = THIN_BORDER
                
            for col in "FHJLN":
                ws[f"{col}{row}"].number_format = PRICE_FMT
                
            for col in "EGIKM":
                ws[f"{col}{row}"].number_format = QTY_FMT
        
            ws[f"A{row}"] = count
            ws[f"B{row}"] = product["name"]
            ws[f"C{row}"] = unit
            ws[f"D{row}"] = wholesale_price
            
            ws[f"E{row}"] = start_qty
            ws[f"F{row}"] = start_price
            
            ws[f"G{row}"] = prihod_qty
            ws[f"H{row}"] = prihod_price
            
            ws[f"I{row}"] = wozwrat_qty
            ws[f"J{row}"] = wozwrat_price
            
            ws[f"K{row}"] = rashod_qty
            ws[f"L{row}"] = rashod_price
            
            ws[f"M{row}"] = end_qty
            ws[f"N{row}"] = end_price
            
        
            count += 1
            row += 1
    
        for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws[i].border = THIN_BORDER
                
        ws.merge_cells(f"A{row}:D{row}")
        for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws[i].fill = GRAY_FILL_0
        ws[f"A{row}"].font = CATEGORY_FONT
        ws[f"A{row}"].alignment = RIGHT_ALIGN
        ws[f"A{row}"] = f"Итого по категории: {cat_name}"
        
        
        ws[f"E{row}"] = totals["start_qty"]
        ws[f"F{row}"] = totals["start_price"]
        ws[f"G{row}"] = totals["prihod_qty"]
        ws[f"H{row}"] = totals["prihod_price"]
        ws[f"I{row}"] = totals["wozwrat_qty"]
        ws[f"J{row}"] = totals["wozwrat_price"]
        ws[f"K{row}"] = totals["rashod_qty"]
        ws[f"L{row}"] = totals["rashod_price"]
        ws[f"M{row}"] = totals["end_qty"]
        ws[f"N{row}"] = totals["end_price"]
                        
        row += 1
        
    ws[f"E{row}"] = grand_total["start_qty"]
    ws[f"F{row}"] = grand_total["start_price"]
    ws[f"G{row}"] = grand_total["prihod_qty"]
    ws[f"H{row}"] = grand_total["prihod_price"]
    ws[f"I{row}"] = grand_total["wozwrat_qty"]
    ws[f"J{row}"] = grand_total["wozwrat_price"]
    ws[f"K{row}"] = grand_total["rashod_qty"]
    ws[f"L{row}"] = grand_total["rashod_price"]
    ws[f"M{row}"] = grand_total["end_qty"]
    ws[f"N{row}"] = grand_total["end_price"]
    
    for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
        ws[i].border = THIN_BORDER
        
    for col in "FHJLN":
        ws[f"{col}{row}"].number_format = PRICE_FMT
        
    for col in "EGIKM":
        ws[f"{col}{row}"].number_format = QTY_FMT
                
    ws.merge_cells(f"A{row}:D{row}")
    for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
        ws[i].fill = GREEN_FILL_1

    ws[f"A{row}"].font = CATEGORY_FONT
    ws[f"A{row}"].alignment = RIGHT_ALIGN
    ws[f"A{row}"] = "ВСЕГО:"
            
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"buh_oborot_tovarov_{day_start_str}_{day_end_str}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    print("EXCEL RESPONSE SENT")
    return response

    
    
    
   
   
    
        
    data = []
    
    return Response(data)


