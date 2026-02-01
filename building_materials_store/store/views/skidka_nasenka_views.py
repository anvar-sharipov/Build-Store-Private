# from django.http import JsonResponse, HttpResponse
# from django.views.decorators.http import require_GET, require_POST
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
# from decimal import Decimal, ROUND_HALF_UP
# from django.db.models import Q, Sum
# from datetime import datetime, timedelta
# from django.utils.dateparse import parse_date

# from django.db.models import F
# from django.contrib.postgres.search import TrigramSimilarity
# from ..models import *

# import json
# from collections import defaultdict

# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from django.http import FileResponse
# import os
# from io import BytesIO
# from django.core.files.base import ContentFile
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import CellIsRule
# from openpyxl.worksheet.page import PageMargins

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from icecream import ic
import time
from datetime import datetime
from ..models import Warehouse, InvoiceItem
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import F, ExpressionWrapper, DecimalField

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from django.http import FileResponse
from io import BytesIO


# ===== СТИЛИ =====
THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

HEADER_FONT = Font(bold=True, size=13)
HEADER_FILL = PatternFill("solid", fgColor="CFE2F3")

CATEGORY_FILL = PatternFill("solid", fgColor="E9ECEF")

RED_FILL = PatternFill("solid", fgColor="F8D7DA")     # скидка
GREEN_FILL = PatternFill("solid", fgColor="D1E7DD")   # наценка

def build_excel_response(result, date_from, date_to):
    wb = Workbook()
    wb.remove(wb.active)

    date_str = datetime.now().strftime("%d.%m.%Y")

   
    # =====================================================
    # 📄 ЛИСТЫ СКЛАДОВ
    # =====================================================
    for w in result:
        ws = wb.create_sheet(title=w["name"][:31])

        # было A1:J1 → стало A1:K1
        ws.merge_cells("A1:J1")
        ws["A1"] = f"Отклонение от оптовой цены за {date_from} - {date_to} {w['name']}"
        ws["A1"].font = HEADER_FONT
        ws["A1"].fill = HEADER_FILL
        ws["A1"].alignment = CENTER_ALIGN

        headers = [
            "№",
            "Партнёр",
            "Комментарий",      # ✅ НОВАЯ КОЛОНКА
            "Товар",
            "Ед.",
            "Оптовая цена",
            "Цена продажи",
            "Кол-во",
            "Всего продажа",
            "Разница",
        ]

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = CATEGORY_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            # ws.column_dimensions[get_column_letter(col)].width = 20
            
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 7
        ws.column_dimensions["F"].width = 7
        ws.column_dimensions["G"].width = 7
        ws.column_dimensions["H"].width = 7
        ws.column_dimensions["I"].width = 7
        
        ws.freeze_panes = "A3"

        row = 3
        for i, t in enumerate(w["table"], start=1):
            ws[f"A{row}"] = i
            ws[f"B{row}"] = t["partner_name"]
            ws[f"C{row}"] = f"№{t['invoice_id']} {t.get('invoice_comment', '')}"  # ✅ комментарий
            ws[f"D{row}"] = t["product_name"]
            ws[f"E{row}"] = t["unit"]
            ws[f"F{row}"] = t["wholesale_price"]
            ws[f"G{row}"] = t["selected_price"]
            ws[f"H{row}"] = t["selected_quantity"]
            ws[f"I{row}"] = t["total_selected_price"]
            ws[f"J{row}"] = t["difference"]

            # цвет разницы
            if t["difference"] < 0:
                ws[f"J{row}"].font = Font(color="9C0006")
            else:
                ws[f"J{row}"].font = Font(color="006100")

            for col in ["A","B","C","D","E","F","G","H","I","J"]:
                ws[f"{col}{row}"].border = THIN_BORDER

            for col in ["F", "G", "I", "J"]:
                ws[f"{col}{row}"].number_format = "#,##0.00"

            ws[f"H{row}"].number_format = "#,##0.###"

            row += 1

        # ===== ИТОГО ПО ТАБЛИЦЕ =====
        ws[f"I{row}"] = "ИТОГО"
        ws[f"J{row}"] = w["otkloneniy_wsego"]
        ws[f"I{row}"].alignment = RIGHT_ALIGN

        for col in ["A","B","C","D","E","F","G","H","I","J"]:
            ws[f"{col}{row}"].border = THIN_BORDER
            ws[f"{col}{row}"].fill = CATEGORY_FILL
            
            
        # ===== ИТОГОВЫЕ ПОКАЗАТЕЛИ ПОД ТАБЛИЦЕЙ =====
        row += 2   # одна пустая строка

        ws[f"B{row}"] = "Показатель"
        ws[f"C{row}"] = "Значение"

        ws[f"B{row}"].fill = CATEGORY_FILL
        ws[f"C{row}"].fill = CATEGORY_FILL
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"C{row}"].border = THIN_BORDER
        ws[f"B{row}"].alignment = CENTER_ALIGN
        ws[f"C{row}"].alignment = CENTER_ALIGN

        labels = [
            ("Общая выручка", w["total_all_price"]),
            ("Отклонение всего", w["otkloneniy_wsego"]),
            ("Скидки", -w["skidki"] if w["skidki"] > 0 else Decimal("0")),
            ("Наценки", w["nasenki"]),
            ("% отклонения", w["percent"]),
        ]

        for i, (label, val) in enumerate(labels, start=1):
            ws[f"B{row+i}"] = label
            ws[f"C{row+i}"] = val

            ws[f"B{row+i}"].border = THIN_BORDER
            ws[f"C{row+i}"].border = THIN_BORDER

            if isinstance(val, Decimal):
                ws[f"C{row+i}"].number_format = "#,##0.00"

        # процент
        ws[f"C{row+5}"].number_format = '0.00"%"'
            
        
        
    # =====================================================
    # 📊 СВОДНЫЙ ЛИСТ
    # =====================================================
    ws_total = wb.create_sheet("ПО СКЛАДАМ")

    ws_total.merge_cells("A1:E1")
    ws_total["A1"] = f"Отклонение от оптовой цены за {date_from} - {date_to}"
    ws_total["A1"].font = HEADER_FONT
    ws_total["A1"].fill = HEADER_FILL
    ws_total["A1"].alignment = CENTER_ALIGN

    headers = ["Склад", "Выручка", "Скидки", "Наценки", "Итоговое отклонение"]
    for col, h in enumerate(headers, start=1):
        cell = ws_total.cell(row=3, column=col, value=h)
        cell.fill = CATEGORY_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        ws_total.column_dimensions[get_column_letter(col)].width = 22

    row_total = 4

    for w in result:
        ws_total[f"A{row_total}"] = w["name"]
        ws_total[f"B{row_total}"] = w["total_all_price"]
        ws_total[f"C{row_total}"] = -w["skidki"] if w["skidki"] > 0 else Decimal("0")
        ws_total[f"D{row_total}"] = w["nasenki"]
        ws_total[f"E{row_total}"] = w["otkloneniy_wsego"]

        for col in ["A", "B", "C", "D", "E"]:
            ws_total[f"{col}{row_total}"].border = THIN_BORDER

        for col in ["B", "C", "D", "E"]:
            ws_total[f"{col}{row_total}"].number_format = "#,##0.00"

        row_total += 1


    # =====================================================
    # 💾 ОТДАЧА ФАЙЛА
    # =====================================================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return FileResponse(
        buffer,
        as_attachment=True,
        filename="skidka_nacenka.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def skidka_nasenka(request):
    user = request.user
    
    if not user.groups.filter(name="admin").exists():
        return Response(
            {"message": "onlyAdmin"},
            status=403
        )
    
    date_from_str = request.query_params.get("date_from")
    date_to_str = request.query_params.get("date_to")
    
    if not date_from_str or not date_to_str:
        return Response({"message": "date_required"}, status=400)
    
    
    
    date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
    date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
    
    partners = request.query_params.get("partners")
    warehouses = request.query_params.get("warehouses")
    agents = request.query_params.get("agents")
    products = request.query_params.get("products")
    users = request.query_params.get("users")
    sortPrice = request.query_params.get("sortPrice")
    
    print_excel = request.query_params.get("printExcel") == "true"
    
    # ic(print_excel)
    
    
    
    warehouse_ids = []
    partner_ids = []
    agent_ids = []
    product_ids = []
    user_ids = []
    
    if warehouses:
        warehouse_ids = [
            int(pk) for pk in warehouses.split(",") if pk.isdigit()
        ]
        
    if not warehouse_ids:
        return Response({"message": "select_warehouse"}, status=400)
        
    if partners:
        partner_ids = [
            int(pk) for pk in partners.split(",") if pk.isdigit()
        ]
        
    if agents:
        agent_ids = [
            int(pk) for pk in agents.split(",") if pk.isdigit()
        ]
        
    if products:
        product_ids = [
            int(pk) for pk in products.split(",") if pk.isdigit()
        ]
        
    if users:
        user_ids = [
            int(pk) for pk in users.split(",") if pk.isdigit()
        ]
        
    result = []

    for w in Warehouse.objects.filter(id__in=warehouse_ids):
        total_all_price = Decimal('0')
        otkloneniy_wsego = Decimal('0')
        skidki = Decimal('0')
        nasenki = Decimal('0')
        opt_sum = Decimal('0')

        table = []

        turnover_items = (
            InvoiceItem.objects
            .filter(
                invoice__entry_created_at_handle__gte=date_from,
                invoice__entry_created_at_handle__lte=date_to,
                invoice__canceled_at__isnull=True,
                invoice__wozwrat_or_prihod="rashod",
                invoice__warehouse=w,
            )
            .select_related(
                "invoice",
                "invoice__partner",
                "product"
            ).annotate(
                difference_db=ExpressionWrapper(
                    (F("selected_price") - F("wholesale_price")) * F("selected_quantity"),
                    output_field=DecimalField(max_digits=14, decimal_places=2),
                )
            )
        )
        
        if partner_ids:
            turnover_items = turnover_items.filter(
                invoice__partner__id__in=partner_ids
            )
            
        if agent_ids:
            turnover_items = turnover_items.filter(
                invoice__partner__agent__id__in=agent_ids
            )
            
        if product_ids:
            turnover_items = turnover_items.filter(
                product__id__in=product_ids
            )

        if user_ids:
            turnover_items = turnover_items.filter(
                invoice__entry_created_by__id__in=user_ids
            )
            
        
        if sortPrice == "asc":
            turnover_items = turnover_items.order_by("difference_db")
        elif sortPrice == "desc":
            turnover_items = turnover_items.order_by("-difference_db")
        else:
            turnover_items = turnover_items.order_by("-invoice__pk")


        for t in turnover_items:
            opt_sum += t.wholesale_price * t.selected_quantity
            total_all_price += t.selected_price * t.selected_quantity

            # otkloneniye = (
            #     t.selected_price - t.wholesale_price
            # ) * t.selected_quantity
            otkloneniye = t.difference_db

            otkloneniy_wsego += otkloneniye

            if otkloneniye < 0:
                skidki += abs(otkloneniye)
            elif otkloneniye > 0:
                nasenki += otkloneniye

            if t.selected_price != t.product.wholesale_price:
                table.append({
                    "partner_name": t.invoice.partner.name,
                    "invoice_id": t.invoice.id,
                    "invoice_comment": t.invoice.comment,
                    "product_name": t.product.name,
                    "unit": t.unit_name_on_selected_warehouses,
                    "wholesale_price": t.product.wholesale_price,
                    "selected_price": t.selected_price,
                    "selected_quantity": t.selected_quantity,
                    "total_selected_price": t.selected_price * t.selected_quantity,
                    "difference": t.difference_db,
                })

        percent = Decimal('0')
        if opt_sum > 0:
            percent = (otkloneniy_wsego / opt_sum * 100).quantize(
                Decimal('0.01'),
                rounding=ROUND_HALF_UP
            )
            
       
        result.append({
            "id": w.id,
            "name": w.name,
            "table": table,
            "total_all_price": total_all_price,
            "otkloneniy_wsego": otkloneniy_wsego,
            "skidki": skidki,
            "nasenki": nasenki,
            "opt_sum": opt_sum,
            "percent": percent,
        })
        
    if print_excel:
        return build_excel_response(result, date_from.strftime("%d.%m.%Y"), date_to.strftime("%d.%m.%Y"))
    else:
        return Response({ "warehouses": result })




