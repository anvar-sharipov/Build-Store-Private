
from django.http import JsonResponse, HttpResponse
# from django.views.decorators.http import require_GET, require_POST
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Q, Sum, Count, F, DecimalField, ExpressionWrapper, Prefetch
from datetime import datetime, timedelta
# from django.utils.dateparse import parse_date

# from django.db.models import F
# from django.contrib.postgres.search import TrigramSimilarity
from ..models import Entry, Transaction
from icecream import ic
# import json
# from collections import defaultdict

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from django.http import FileResponse
# import os
# from io import BytesIO
# from django.core.files.base import ContentFile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import CellIsRule
# from openpyxl.worksheet.page import PageMargins

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models.functions import Coalesce
from openpyxl.worksheet.page import PageMargins
from collections import defaultdict



CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)








@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_excel_entries_diapazon(request):

    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")

    if not date_from or not date_to:
        return Response({"error": "choose diapazon date"}, status=400)

    try:
        day_start = datetime.strptime(date_from, "%Y-%m-%d").date()
        day_end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return Response({"error": "invalid date format"}, status=400)

    if day_start > day_end:
        return Response({"error": "date start must be <= date end"}, status=400)
    
    
    
    day_start_str = day_start.strftime("%d.%m.%Y")
    day_end_str = day_end.strftime("%d.%m.%Y")

    # 🔥 ВАЖНО — фильтрация через __date__range
    transactions = (
        Transaction.objects
        .filter(
            date__date__range=[day_start, day_end],
            invoice__isnull=True
        )
        .prefetch_related("entries__account", "entries__partner")
        .order_by("-date")
    )

    data = []
    
    

    for transaction in transactions:
        debit_entry_obj = None
        credit_entry_obj = None

        for entry in transaction.entries.all():
            if entry.debit != 0:
                debit_entry_obj = entry
            elif entry.credit != 0:
                credit_entry_obj = entry

        if not debit_entry_obj or not credit_entry_obj:
            continue
        

        data.append({
            "id": transaction.id,
            "date": transaction.date.strftime("%Y-%m-%d"),
            "comment": transaction.description,
            "debit_account": debit_entry_obj.account.number,
            "credit_account": credit_entry_obj.account.number,
            "amount": float(debit_entry_obj.debit),
        })
        
        
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал операций"
    
    
    
    
    
    
    
    # ===== Заголовок =====
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Журнал операций за {day_start_str} - {day_end_str}"
    ws["A1"].alignment = CENTER
    ws["A1"].font = Font(size=14, bold=True)
    ws.row_dimensions[1].height = 30

    # ===== Шапка =====
    headers = [
        "№",
        "Дата",
        "№ операции",
        "Комментарий",
        "Сумма",
        "Дебет",
        "Дебет субконто",
        "Кредит",
        "Кредит субконто",
    ]
    ws.row_dimensions[3].height = 22

    row = 3
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = THIN

    # ширина колонок
    widths = [5, 10, 10, 25, 11, 7, 25, 7, 45]
    
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    row += 1
    counter = 0
    total_sum = Decimal("0")
    
    # account_totals = defaultdict(lambda: {"debit": Decimal("0"), "credit": Decimal("0")})
    account_totals = defaultdict(
        lambda: {"debit": Decimal("0"), "credit": Decimal("0"), "name": ""}
    )
    total_debit = Decimal("0")
    total_credit = Decimal("0")

    # ===== ДАННЫЕ =====
    
    for transaction in transactions:

        debit_entry = None
        credit_entry = None

        for entry in transaction.entries.all():
            if entry.debit != 0:
                debit_entry = entry
            elif entry.credit != 0:
                credit_entry = entry

        if not debit_entry or not credit_entry:
            continue

        counter += 1
        amount = debit_entry.debit
        total_sum += amount
        
        # накопление оборотов
        # account_totals[debit_entry.account.number]["debit"] += debit_entry.debit
        # account_totals[credit_entry.account.number]["credit"] += credit_entry.credit
        # Дебет
        account_totals[debit_entry.account.number]["debit"] += debit_entry.debit
        account_totals[debit_entry.account.number]["name"] = debit_entry.account.name

        # Кредит
        account_totals[credit_entry.account.number]["credit"] += credit_entry.credit
        account_totals[credit_entry.account.number]["name"] = credit_entry.account.name

        total_debit += debit_entry.debit
        total_credit += credit_entry.credit

        # субконто только для 60 и 75
        debit_partner = ""
        credit_partner = ""

        if debit_entry.account.number.startswith(("60", "75")) and debit_entry.partner:
            debit_partner = debit_entry.partner.name

        if credit_entry.account.number.startswith(("60", "75")) and credit_entry.partner:
            credit_partner = credit_entry.partner.name

        values = [
            counter,
            transaction.date.strftime("%d.%m.%Y"),
            transaction.id,
            transaction.description,
            float(amount),
            debit_entry.account.number,
            debit_partner,
            credit_entry.account.number,
            credit_partner,
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = val
            cell.border = THIN

            # Сумма
            if col == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = RIGHT

            # Центрирование
            if col in [1, 2, 3]:
                cell.alignment = CENTER

            # Дебет — зелёный
            if col == 6:
                cell.font = Font(size=12, color="008000", bold=True)  # зелёный
                cell.alignment = CENTER

            # Кредит — красный
            if col == 8:
                cell.font = Font(size=12, color="FF0000", bold=True)  # красный
                cell.alignment = CENTER

            # Перенос текста для субконто
            if col in [7, 9]:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                
        # ws.row_dimensions[row].height = 20

        row += 1


    # ===== ИТОГО =====
    ws[f"D{row}"] = "ИТОГО:"
    ws[f"D{row}"].font = Font(bold=True)
    ws[f"D{row}"].alignment = RIGHT

    ws[f"E{row}"] = float(total_sum)
    ws[f"E{row}"].font = Font(bold=True)
    ws[f"E{row}"].number_format = '#,##0.00'
    ws[f"E{row}"].alignment = RIGHT

    for col in range(1, 10):
        ws.cell(row=row, column=col).border = THIN
        
        
        
   
        
        
    # ================================
    # 3 SHEET — ОБОРОТЫ ПО СЧЕТАМ
    # ================================

    ws3 = wb.create_sheet(title="Обороты по счетам")
    
    ws3.merge_cells("A1:C1")

    ws3["A1"] = f"Обороты по счетам за {day_start_str} - {day_end_str}"
    ws3["A1"].font = Font(size=14, bold=True)
    ws3["A1"].alignment = CENTER

    headers = ["Счет", "Дебет", "Кредит"]
    row = 3

    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=row, column=col)
        cell.value = header
        # cell.font = Font(bold=True)
        cell.font = Font(size=12, bold=True)
        cell.border = THIN
        cell.alignment = CENTER

    row += 1

    for acc, sums in sorted(account_totals.items()):
        # ws3.cell(row=row, column=1).value = acc
        # ws3.cell(row=row, column=2).value = float(sums["debit"])
        # ws3.cell(row=row, column=3).value = float(sums["credit"])
        
        # cell_acc = ws3.cell(row=row, column=1)
        # cell_acc.value = acc
        # cell_acc.alignment = CENTER   # ← ВОТ ЭТО ДОБАВЛЕНО
        
        account_name = sums["name"]

        cell_acc = ws3.cell(row=row, column=1)
        cell_acc.value = f"{acc} — {account_name}"
        cell_acc.alignment = Alignment(horizontal="left", vertical="center")

        cell_debit = ws3.cell(row=row, column=2)
        cell_debit.value = float(sums["debit"])
        cell_debit.number_format = '#,##0.00'
        cell_debit.alignment = RIGHT

        cell_credit = ws3.cell(row=row, column=3)
        cell_credit.value = float(sums["credit"])
        cell_credit.number_format = '#,##0.00'
        cell_credit.alignment = RIGHT

        ws3.cell(row=row, column=2).number_format = '#,##0.00'
        ws3.cell(row=row, column=3).number_format = '#,##0.00'

        for col in range(1, 4):
            ws3.cell(row=row, column=col).border = THIN

        row += 1
        
    # ======================
    # ИТОГО
    # ======================

    total_row_start = row

    ws3.cell(row=row, column=1).value = "ИТОГО:"
    ws3.cell(row=row, column=1).font = Font(bold=True)
    ws3.cell(row=row, column=1).alignment = RIGHT

    ws3.cell(row=row, column=2).value = float(total_debit)
    ws3.cell(row=row, column=3).value = float(total_credit)

    ws3.cell(row=row, column=2).number_format = '#,##0.00'
    ws3.cell(row=row, column=3).number_format = '#,##0.00'

    ws3.cell(row=row, column=2).font = Font(bold=True)
    ws3.cell(row=row, column=3).font = Font(bold=True)

    ws3.cell(row=row, column=2).alignment = RIGHT
    ws3.cell(row=row, column=3).alignment = RIGHT

    for col in range(1, 4):
        ws3.cell(row=row, column=col).border = THIN

    row += 1


    # ======================
    # САЛЬДО
    # ======================

    saldo = total_debit - total_credit

    ws3.cell(row=row, column=1).value = "САЛЬДО:"
    ws3.cell(row=row, column=1).font = Font(bold=True)
    ws3.cell(row=row, column=1).alignment = RIGHT

    ws3.cell(row=row, column=2).value = float(saldo)
    ws3.cell(row=row, column=2).number_format = '#,##0.00'
    ws3.cell(row=row, column=2).font = Font(bold=True)
    ws3.cell(row=row, column=2).alignment = RIGHT

    # если баланс не 0 — подсветить красным
    if saldo != 0:
        ws3.cell(row=row, column=2).font = Font(bold=True, color="FF0000")

    for col in range(1, 4):
        ws3.cell(row=row, column=col).border = THIN

    # ширина колонок
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20

    # ===== RESPONSE =====
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="journal_{day_start_str}_{day_end_str}.xlsx"'
    )
    
    # ===== НАСТРОЙКА ПЕЧАТИ =====

    # Поля страницы (в дюймах!)
    ws.page_margins = PageMargins(
        left=0.3,        # Левый
        right=0.2,     # Правый
        top=0.2,       # Верхний
        bottom=0.2,    # Нижний
    )

    # Альбомная ориентация (если нужно — можно убрать)
    # ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # Разместить не более чем на 1 странице по ширине
    # ws.page_setup.fitToWidth = 1
    # ws.page_setup.fitToHeight = False
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 = без ограничения по высоте

    # Центрирование по горизонтали при печати
    ws.print_options.horizontalCentered = True

    wb.save(response)
    return response





    # return Response({
    #     "count": len(data),
    #     "data": data
    # })











