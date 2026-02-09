
from django.http import JsonResponse, HttpResponse
# from django.views.decorators.http import require_GET, require_POST
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Q, Sum, Count, F, DecimalField, ExpressionWrapper, Prefetch
from datetime import datetime, timedelta
# from django.utils.dateparse import parse_date

# from django.db.models import F
# from django.contrib.postgres.search import TrigramSimilarity
from ..models import Account, Entry, DayClosing, PartnerBalanceSnapshot, Partner, WarehouseAccount, Product, InvoiceItem
from icecream import ic
# import json
# from collections import defaultdict

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from django.http import FileResponse
# import os
# from io import BytesIO
# from django.core.files.base import ContentFile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import CellIsRule
# from openpyxl.worksheet.page import PageMargins

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from io import BytesIO




from ..my_func.get_unit_map import get_unit_map 
from ..my_func.get_unit_and_cf import get_unit_and_cf 
from ..my_func.str_to_int_list import str_to_int_list
from ..my_func.date_convert_for_excel import format_date_ru 



# ================== СТИЛИ ==================

# Шапка
HEADER_FILL = PatternFill(fill_type="solid",fgColor="4472C4")
HEADER_FONT = Font(bold=True, size=16)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN_WRAP = Alignment(vertical="center",horizontal="left",wrap_text=True)

# Категория
CATEGORY_FILL = PatternFill(fill_type="solid", fgColor="E7E6E6")
CATEGORY_FONT = Font(bold=True)

# Обычные ячейки
NORMAL_FONT = Font(color="000000")
LEFT_ALIGN = Alignment(vertical="center", horizontal="left")
RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

# Границы
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

LEFT_BORDER = Border(left=Side(style="thin"))
RIGHT_BORDER = Border(right=Side(style="thin"))
TOP_BORDER = Border(top=Side(style="thin"))
BOTTOM_BORDER = Border(bottom=Side(style="thin"))


# Форматы чисел
PRICE_FMT = '#,##0.00'
QTY_FMT = '#,##0.###'

# Ширина колонок
COLUMN_WIDTHS = {
    "A": 25,
    "B": 40,
    "C": 10,
    "D": 12,
    "E": 8, "F": 8,
    "G": 8, "H": 8,
    "I": 8, "J": 8,
    "K": 8, "L": 8,
    "M": 8, "N": 8,
}

TOTAL_FILL = PatternFill(fill_type="solid",fgColor="D9D9D9")

TOTAL_FONT = Font(bold=True)

GRAY_FILL_0 = PatternFill(fill_type="solid", fgColor="F5F5F5")
GRAY_FILL_1 = PatternFill(fill_type="solid", fgColor="EDEDED")
GRAY_FILL_2 = PatternFill( fill_type="solid", fgColor="DCDCDC")
GRAY_FILL_3 = PatternFill(fill_type="solid", fgColor="C8C8C8")
GRAY_FILL_4 = PatternFill(fill_type="solid", fgColor="B0B0B0")

GREEN_FILL_0 = PatternFill(fill_type="solid", fgColor="E2F0D9")
GREEN_FILL_1 = PatternFill(fill_type="solid", fgColor="C6EFCE")
GREEN_FILL_2 = PatternFill(fill_type="solid", fgColor="92D050")
GREEN_FILL_3 = PatternFill(fill_type="solid", fgColor="006100")

RED_FONT = Font(color="FF0000")
GREEN_FONT = Font(color="006400")
BLUE_FONT = Font(color="0000FF")

def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)



def download_excel_partners_diapazon(request):
    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")
    
    if not date_from or not date_to:
        return Response(
            {"error": "choose diapazon date"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        day_start = datetime.strptime(date_from, "%Y-%m-%d").date()
        day_end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return Response(
            {"error": "invalid diapazon date format"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if day_start > day_end:
        return Response(
            {"error": "date start must be <= date end"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    day_start_str = day_start.strftime("%d.%m.%Y")
    day_end_str = day_end.strftime("%d.%m.%Y")
    
    
    
    partner_type = request.GET.get("type", "")
    sort = request.GET.get("sort", "")
    search = request.GET.get("search", "")
    is_active = request.GET.get("is_active", "")
    
    ic(partner_type)
    ic(sort)
    ic(search)
    ic(is_active)
    
 
        
    partners = Partner.objects.all().select_related("agent")
    ic("len(partners)", len(partners))
    
    if search:
        partners = partners.filter(name__icontains=search)
        ic("len(partners)2", len(partners))
        
        
    if is_active:
        is_active2 = True if is_active == "true" else False
        partners = partners.filter(is_active=is_active2)
        ic("len(partners)3", len(partners))
        
    if partner_type:
        partners = partners.filter(type=partner_type)
        ic("len(partners)4", len(partners))
        
    partner_ids = partners.values_list("id", flat=True)
        
    last_closing = DayClosing.objects.filter(
        date__lt=day_start
    ).order_by("-date").first()
        
    partner_snap_map = {}
    wb = Workbook()
    
    if last_closing:
        partner_balance_snapshots = PartnerBalanceSnapshot.objects.filter(closing=last_closing).select_related("partner")
        partner_snap_map = {
            (s.partner.id): {
                    "balance_60_usd": s.balance_60_usd, 
                    "balance_62_tmt": s.balance_62_tmt, 
                    "balance_75_usd": s.balance_75_usd, 
                    "balance_76_tmt": s.balance_76_tmt,
                    
                    "balance_60_usd_credit": s.balance_60_usd_credit,
                    "balance_60_usd_debit": s.balance_60_usd_debit,
                    
                    "balance_62_tmt_credit": s.balance_62_tmt_credit,
                    "balance_62_tmt_debit": s.balance_62_tmt_debit,
                    
                    "balance_75_usd_credit": s.balance_75_usd_credit,
                    "balance_75_usd_debit": s.balance_75_usd_debit,
                    
                    "balance_76_tmt_credit": s.balance_76_tmt_credit,
                    "balance_76_tmt_debit": s.balance_76_tmt_debit,

                }
            for s in partner_balance_snapshots
        }
        
    for account_number in ["60", "62"]:    
        account_60_62 = {}
        for p in partners:
            partner = {
                "id": p.id,
                "name": p.name,
                "type": p.type,
            }
            
            if p.agent:
                agent = {
                    "id": p.agent.id,
                    "name": p.agent.name,
                }
            else:
                agent = {}
                
            account_60_62[p.id] = {
                "partner": partner,
                "agent": agent,
                "debit_start": Decimal("0.00"),
                "credit_start": Decimal("0.00"),
                "debit_turnover": Decimal("0.00"),
                "credit_turnover": Decimal("0.00"),
                "debit_end": Decimal("0.00"),
                "credit_end": Decimal("0.00"),
            }
            
        if last_closing:
            for p_id, value in partner_snap_map.items():
                if p_id not in account_60_62:
                    continue
                if account_number == "60":
                    account_60_62[p_id]["debit_start"] += value["balance_60_usd_debit"]
                    account_60_62[p_id]["credit_start"] += value["balance_60_usd_credit"]
                else:
                    account_60_62[p_id]["debit_start"] += value["balance_62_tmt_debit"]
                    account_60_62[p_id]["credit_start"] += value["balance_62_tmt_credit"]
                    
        else:
            entries_start = (
                Entry.objects
                .filter(
                    account__number=account_number,
                    transaction__date__lt=day_end,
                    partner__in = partner_ids
                )
                .select_related(
                    'transaction',
                    'partner__agent',
                    'product',
                    'warehouse',
                    'transaction__invoice',
                )
                .order_by('transaction__date', 'id')
            )
            
            for e in entries_start:
                if not e.partner:
                    continue
                partner_id = e.partner.id
                
                
                account_60_62[partner_id]["debit_start"] += money(e.debit)
                account_60_62[partner_id]["credit_start"] += money(e.credit)
                
        entries_turnover = (
            Entry.objects
            .filter(
                account__number=account_number,
                transaction__date__gte=day_start,
                transaction__date__lte=day_end,
                partner__in = partner_ids
            )
            .select_related(
                'transaction',
                'partner__agent',
                'product',
                'warehouse',
                'transaction__invoice',
            )
            .order_by('transaction__date', 'id')
        )
        
        for e in entries_turnover:
                if not e.partner:
                    continue
                partner_id = e.partner.id
                ic(account_60_62[partner_id])
                
                account_60_62[partner_id]["debit_turnover"] += money(e.debit)
                account_60_62[partner_id]["credit_turnover"] += money(e.credit)
                
        total_credit_start = Decimal("0.00")        
        total_debit_start = Decimal("0.00")        
        total_credit_turnover = Decimal("0.00")        
        total_debit_turnover = Decimal("0.00")        
        total_credit_end = Decimal("0.00")        
        total_debit_end = Decimal("0.00")
        
        
        for p_id, row in account_60_62.items():
            debit_end_raw = row["debit_start"] + row["debit_turnover"]
            credit_end_raw = row["credit_start"] + row["credit_turnover"]

            saldo = debit_end_raw - credit_end_raw

            if saldo >= 0:
                row["debit_end"] = saldo
                row["credit_end"] = Decimal("0.00")
            else:
                row["debit_end"] = Decimal("0.00")
                row["credit_end"] = -saldo
                
            # 👉 ИТОГИ СЧИТАЕМ ЗДЕСЬ
            total_debit_start += row["debit_start"]
            total_credit_start += row["credit_start"]

            total_debit_turnover += row["debit_turnover"]
            total_credit_turnover += row["credit_turnover"]

            total_debit_end += row["debit_end"]
            total_credit_end += row["credit_end"]
            
        grand_total_60_62 =  {
            "total_credit_start": total_credit_start,
            "total_debit_start": total_debit_start,
            "total_credit_turnover": total_credit_turnover,
            "total_debit_turnover": total_debit_turnover,
            "total_credit_end": total_credit_end,
            "total_debit_end": total_debit_end,
        }    
        
        ic(account_60_62)
    
        if account_number == "60":
            ws_detail = wb.active
            ws_detail.title = "Счет 60"
            
            ws_detail.merge_cells("A1:I1")
            ws_detail["A1"] = "Счёт 60 — Клиент USD"
            ws_detail["A1"].font = HEADER_FONT
            ws_detail["A1"].alignment = CENTER_ALIGN
            
            ws_detail.merge_cells("A2:I2")
            ws_detail["A2"] = f"{day_start_str} - {day_end_str}"
            ws_detail["A2"].font = Font(bold=True)
            ws_detail["A2"].alignment = CENTER_ALIGN
            
            ws_detail.freeze_panes = "A5"
            ws_detail["A3"] = "№"
            ws_detail.column_dimensions["A"].width = 5
            ws_detail["B3"] = "Agent"
            ws_detail.column_dimensions["B"].width = 20
            ws_detail["C3"] = "Субконто"
            ws_detail.column_dimensions["C"].width = 45
            
            
            
            
            
            ws_detail.merge_cells("D3:E3")
            ws_detail["D3"] = "Сальдо на начало"

            
            ws_detail.merge_cells("F3:G3")
            ws_detail["F3"] = "Обороты за период"
         
            
            ws_detail.merge_cells("H3:I3")
            ws_detail["H3"] = "Сальдо на конец"
            
            for i in ["A3", "B3", "C3", "D3", "E3", "F3", "G3", "H3", "I3", 
                      "A4", "B4", "C4", "D4", "E4", "F4", "G4", "H4", "I4"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = CENTER_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
                
                
            for i in ["D", "E", "F", "G", "H", "I"]:
                ws_detail.column_dimensions[i].width = 15
                
                
            ws_detail["D4"] = "Дебит"
            ws_detail["E4"] = "Кредит"
            
            ws_detail["F4"] = "Дебит"
            ws_detail["G4"] = "Кредит"
            
            ws_detail["H4"] = "Дебит"
            ws_detail["I4"] = "Кредит"
            
            row = 5
            count = 1
            for p_id, v in account_60_62.items():
                
                ws_detail[f"D{row}"].number_format = PRICE_FMT
                ws_detail[f"E{row}"].number_format = PRICE_FMT
                ws_detail[f"F{row}"].number_format = PRICE_FMT
                ws_detail[f"G{row}"].number_format = PRICE_FMT
                ws_detail[f"H{row}"].number_format = PRICE_FMT
                ws_detail[f"I{row}"].number_format = PRICE_FMT
                
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                    ws_detail[i].border = THIN_BORDER
                
        
        
                ws_detail[f"A{row}"] = count
                ws_detail[f"B{row}"] = v["agent"]["name"] if v["agent"] else ""
                ws_detail[f"C{row}"] = v["partner"]["name"] if v["partner"] else ""
                
                start_saldo_debit = Decimal("0")
                start_saldo_credit = Decimal("0")
                saldo_start_row = v["debit_start"] - v["credit_start"]
                if saldo_start_row > 0:
                    start_saldo_debit = saldo_start_row
                elif saldo_start_row < 0:
                    start_saldo_credit = abs(saldo_start_row)
                ws_detail[f"D{row}"] = start_saldo_debit
                ws_detail[f"E{row}"] = start_saldo_credit
                
                ws_detail[f"F{row}"] = v["debit_turnover"]
                ws_detail[f"G{row}"] = v["credit_turnover"]
                
                ws_detail[f"H{row}"] = v["debit_end"]
                ws_detail[f"I{row}"] = v["credit_end"]
                
                count += 1
                row += 1
                
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = RIGHT_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
            
            
            ws_detail.merge_cells(f"A{row}:C{row}")
            ws_detail[f"A{row}"].alignment = LEFT_ALIGN
            ws_detail[f"A{row}"] = "Итого развернутое:"
            
            ws_detail[f"D{row}"].number_format = PRICE_FMT
            ws_detail[f"E{row}"].number_format = PRICE_FMT
            ws_detail[f"F{row}"].number_format = PRICE_FMT
            ws_detail[f"G{row}"].number_format = PRICE_FMT
            ws_detail[f"H{row}"].number_format = PRICE_FMT
            ws_detail[f"I{row}"].number_format = PRICE_FMT
            
            ws_detail[f"D{row}"] = grand_total_60_62["total_debit_start"]
            ws_detail[f"E{row}"] = grand_total_60_62["total_credit_start"]
            
            ws_detail[f"F{row}"] = grand_total_60_62["total_debit_turnover"]
            ws_detail[f"G{row}"] = grand_total_60_62["total_credit_turnover"]
            
            ws_detail[f"H{row}"] = grand_total_60_62["total_debit_end"]
            ws_detail[f"I{row}"] = grand_total_60_62["total_credit_end"]
            
            
            row += 1
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = RIGHT_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
                
            ws_detail.merge_cells(f"A{row}:C{row}")
            ws_detail[f"A{row}"].alignment = LEFT_ALIGN
            ws_detail[f"A{row}"] = "Всего:"
            
            ws_detail[f"D{row}"].number_format = PRICE_FMT
            ws_detail[f"E{row}"].number_format = PRICE_FMT
            ws_detail[f"F{row}"].number_format = PRICE_FMT
            ws_detail[f"G{row}"].number_format = PRICE_FMT
            ws_detail[f"H{row}"].number_format = PRICE_FMT
            ws_detail[f"I{row}"].number_format = PRICE_FMT
            
            saldo_start = grand_total_60_62["total_debit_start"] - grand_total_60_62["total_credit_start"]
            ws_detail[f"D{row}"] = saldo_start if saldo_start > 0 else 0
            ws_detail[f"E{row}"] = abs(saldo_start) if saldo_start < 0 else 0
            
            saldo_turnover = grand_total_60_62["total_debit_turnover"] - grand_total_60_62["total_credit_turnover"]
            ws_detail[f"F{row}"] = saldo_turnover if saldo_turnover > 0 else 0
            ws_detail[f"G{row}"] = abs(saldo_turnover) if saldo_turnover < 0 else 0
            
            saldo_end = grand_total_60_62["total_debit_end"] - grand_total_60_62["total_credit_end"]
            ws_detail[f"H{row}"] = saldo_end if saldo_end > 0 else 0
            ws_detail[f"I{row}"] = abs(saldo_end) if saldo_end < 0 else 0
        else:
            ws_62 = wb.create_sheet(title="Счет 62")
            
            ws_62.merge_cells("A1:I1")
            ws_62["A1"] = "Счёт 62 — Клиент TMT"
            ws_62["A1"].font = HEADER_FONT
            ws_62["A1"].alignment = CENTER_ALIGN
            
            ws_62.merge_cells("A2:I2")
            ws_62["A2"] = f"{day_start_str} - {day_end_str}"
            ws_62["A2"].font = Font(bold=True)
            ws_62["A2"].alignment = CENTER_ALIGN
            
            ws_62.freeze_panes = "A5"
            ws_62["A3"] = "№"
            ws_62.column_dimensions["A"].width = 5
            ws_62["B3"] = "Agent"
            ws_62.column_dimensions["B"].width = 20
            ws_62["C3"] = "Субконто"
            ws_62.column_dimensions["C"].width = 45
            
            
            
            
            
            ws_62.merge_cells("D3:E3")
            ws_62["D3"] = "Сальдо на начало"

            
            ws_62.merge_cells("F3:G3")
            ws_62["F3"] = "Обороты за период"
         
            
            ws_62.merge_cells("H3:I3")
            ws_62["H3"] = "Сальдо на конец"
            
            for i in ["A3", "B3", "C3", "D3", "E3", "F3", "G3", "H3", "I3", 
                      "A4", "B4", "C4", "D4", "E4", "F4", "G4", "H4", "I4"]:
                ws_62[i].font = TOTAL_FONT
                ws_62[i].alignment = CENTER_ALIGN
                ws_62[i].fill = GRAY_FILL_1
                ws_62[i].border = THIN_BORDER
                
                
            for i in ["D", "E", "F", "G", "H", "I"]:
                ws_62.column_dimensions[i].width = 15
                
                
            ws_62["D4"] = "Дебит"
            ws_62["E4"] = "Кредит"
            
            ws_62["F4"] = "Дебит"
            ws_62["G4"] = "Кредит"
            
            ws_62["H4"] = "Дебит"
            ws_62["I4"] = "Кредит"
            
            row = 5
            count = 1
            for p_id, v in account_60_62.items():
                
                ws_62[f"D{row}"].number_format = PRICE_FMT
                ws_62[f"E{row}"].number_format = PRICE_FMT
                ws_62[f"F{row}"].number_format = PRICE_FMT
                ws_62[f"G{row}"].number_format = PRICE_FMT
                ws_62[f"H{row}"].number_format = PRICE_FMT
                ws_62[f"I{row}"].number_format = PRICE_FMT
                
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                    ws_62[i].border = THIN_BORDER
                
        
        
                ws_62[f"A{row}"] = count
                ws_62[f"B{row}"] = v["agent"]["name"] if v["agent"] else ""
                ws_62[f"C{row}"] = v["partner"]["name"] if v["partner"] else ""
                
                start_saldo_debit = Decimal("0")
                start_saldo_credit = Decimal("0")
                saldo_start_row = v["debit_start"] - v["credit_start"]
                if saldo_start_row > 0:
                    start_saldo_debit = saldo_start_row
                elif saldo_start_row < 0:
                    start_saldo_credit = abs(saldo_start_row)
                ws_62[f"D{row}"] = start_saldo_debit
                ws_62[f"E{row}"] = start_saldo_credit
                
                ws_62[f"F{row}"] = v["debit_turnover"]
                ws_62[f"G{row}"] = v["credit_turnover"]
                
                ws_62[f"H{row}"] = v["debit_end"]
                ws_62[f"I{row}"] = v["credit_end"]
                
                count += 1
                row += 1
                
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                ws_62[i].font = TOTAL_FONT
                ws_62[i].alignment = RIGHT_ALIGN
                ws_62[i].fill = GRAY_FILL_1
                ws_62[i].border = THIN_BORDER
            
            
            ws_62.merge_cells(f"A{row}:C{row}")
            ws_62[f"A{row}"].alignment = LEFT_ALIGN
            ws_62[f"A{row}"] = "Итого развернутое:"
            
            ws_62[f"D{row}"].number_format = PRICE_FMT
            ws_62[f"E{row}"].number_format = PRICE_FMT
            ws_62[f"F{row}"].number_format = PRICE_FMT
            ws_62[f"G{row}"].number_format = PRICE_FMT
            ws_62[f"H{row}"].number_format = PRICE_FMT
            ws_62[f"I{row}"].number_format = PRICE_FMT
            
            ws_62[f"D{row}"] = grand_total_60_62["total_debit_start"]
            ws_62[f"E{row}"] = grand_total_60_62["total_credit_start"]
            
            ws_62[f"F{row}"] = grand_total_60_62["total_debit_turnover"]
            ws_62[f"G{row}"] = grand_total_60_62["total_credit_turnover"]
            
            ws_62[f"H{row}"] = grand_total_60_62["total_debit_end"]
            ws_62[f"I{row}"] = grand_total_60_62["total_credit_end"]
            
            
            row += 1
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                ws_62[i].font = TOTAL_FONT
                ws_62[i].alignment = RIGHT_ALIGN
                ws_62[i].fill = GRAY_FILL_1
                ws_62[i].border = THIN_BORDER
                
            ws_62.merge_cells(f"A{row}:C{row}")
            ws_62[f"A{row}"].alignment = LEFT_ALIGN
            ws_62[f"A{row}"] = "Всего:"
            
            ws_62[f"D{row}"].number_format = PRICE_FMT
            ws_62[f"E{row}"].number_format = PRICE_FMT
            ws_62[f"F{row}"].number_format = PRICE_FMT
            ws_62[f"G{row}"].number_format = PRICE_FMT
            ws_62[f"H{row}"].number_format = PRICE_FMT
            ws_62[f"I{row}"].number_format = PRICE_FMT
            
            saldo_start = grand_total_60_62["total_debit_start"] - grand_total_60_62["total_credit_start"]
            ws_62[f"D{row}"] = saldo_start if saldo_start > 0 else 0
            ws_62[f"E{row}"] = abs(saldo_start) if saldo_start < 0 else 0
            
            saldo_turnover = grand_total_60_62["total_debit_turnover"] - grand_total_60_62["total_credit_turnover"]
            ws_62[f"F{row}"] = saldo_turnover if saldo_turnover > 0 else 0
            ws_62[f"G{row}"] = abs(saldo_turnover) if saldo_turnover < 0 else 0
            
            saldo_end = grand_total_60_62["total_debit_end"] - grand_total_60_62["total_credit_end"]
            ws_62[f"H{row}"] = saldo_end if saldo_end > 0 else 0
            ws_62[f"I{row}"] = abs(saldo_end) if saldo_end < 0 else 0
            
            
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"buh_oborot_klient_{day_start_str}_{day_end_str}.xlsx"
    
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    print("EXCEL RESPONSE SENT")
    return response

        
        
    
        
        
    # data = []
    # return Response(data)
    
   
    