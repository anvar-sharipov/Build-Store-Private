
from django.http import JsonResponse, HttpResponse
# from django.views.decorators.http import require_GET, require_POST
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Q, Sum, Count, F, DecimalField, ExpressionWrapper, Prefetch
from datetime import datetime, timedelta
# from django.utils.dateparse import parse_date

# from django.db.models import F
# from django.contrib.postgres.search import TrigramSimilarity
from ..models import Invoice, InvoiceItem
from icecream import ic
# import json
# from collections import defaultdict

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from django.http import FileResponse
# import os
# from io import BytesIO
# from django.core.files.base import ContentFile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import CellIsRule
# from openpyxl.worksheet.page import PageMargins

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db.models.functions import Coalesce





# ================== СТИЛИ ==================

# Шапка
HEADER_FILL = PatternFill(fill_type="solid",fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN_WRAP = Alignment(vertical="center",horizontal="left",wrap_text=True)

# Категория
CATEGORY_FILL = PatternFill(fill_type="solid", fgColor="E7E6E6")
CATEGORY_FONT = Font(bold=True)

# Обычные ячейки
NORMAL_FONT = Font(color="000000")
LEFT_ALIGN = Alignment(vertical="center", horizontal="left")
RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

# Границы
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

LEFT_BORDER = Border(left=Side(style="thin"))
RIGHT_BORDER = Border(right=Side(style="thin"))
TOP_BORDER = Border(top=Side(style="thin"))
BOTTOM_BORDER = Border(bottom=Side(style="thin"))


# Форматы чисел
PRICE_FMT = '#,##0.00'
QTY_FMT = '#,##0.###'

# Ширина колонок
COLUMN_WIDTHS = {
    "A": 25,
    "B": 40,
    "C": 10,
    "D": 12,
    "E": 8, "F": 8,
    "G": 8, "H": 8,
    "I": 8, "J": 8,
    "K": 8, "L": 8,
    "M": 8, "N": 8,
}

TOTAL_FILL = PatternFill(fill_type="solid",fgColor="D9D9D9")

TOTAL_FONT = Font(bold=True)

GRAY_FILL_0 = PatternFill(fill_type="solid", fgColor="F5F5F5")
GRAY_FILL_1 = PatternFill(fill_type="solid", fgColor="EDEDED")
GRAY_FILL_2 = PatternFill( fill_type="solid", fgColor="DCDCDC")
GRAY_FILL_3 = PatternFill(fill_type="solid", fgColor="C8C8C8")
GRAY_FILL_4 = PatternFill(fill_type="solid", fgColor="B0B0B0")

GREEN_FILL_0 = PatternFill(fill_type="solid", fgColor="E2F0D9")
GREEN_FILL_1 = PatternFill(fill_type="solid", fgColor="C6EFCE")
GREEN_FILL_2 = PatternFill(fill_type="solid", fgColor="92D050")
GREEN_FILL_3 = PatternFill(fill_type="solid", fgColor="006100")

RED_FONT = Font(color="FF0000")
GREEN_FONT = Font(color="006400")
BLUE_FONT = Font(color="0000FF")

def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)




@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_excel_fakturs_diapazon(request):
    
    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")
    wozwrat_or_prihod = request.GET.get("wozwrat_or_prihod") # wozwrat, prihod, rashod, transfer
    partner_id = request.GET.get("partner_id") # po id
    selectedEntry = request.GET.get("selectedEntry") # entried, notEntried, canceled
    sortInvoice = request.GET.get("sortInvoice") # asc, desc po total_ptice
    
    
    
    # ic(wozwrat_or_prihod)
    # ic(sortInvoice)
    # ic(selectedEntry)
    # ic(partner_id)
    # ic(request)
    

    if not date_from or not date_to:
        return Response(
            {"error": "choose diapazon date"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        day_start = datetime.strptime(date_from, "%Y-%m-%d").date()
        day_end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return Response(
            {"error": "invalid diapazon date format"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if day_start > day_end:
        return Response(
            {"error": "date start must be <= date end"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    day_start_str = day_start.strftime("%d.%m.%Y")
    day_end_str = day_end.strftime("%d.%m.%Y")
    
    # ic(day_start)
    # ic(day_end)
    
    #  entry_created_at_handle created_at_handle updated_at_handle invoice_date
    qs = Invoice.objects.filter(
        invoice_date__date__gte=day_start,
        invoice_date__date__lte=day_end,
    )
    
    if wozwrat_or_prihod:
        qs = qs.filter(wozwrat_or_prihod=wozwrat_or_prihod)
        
    if partner_id:
        qs = qs.filter(partner_id=partner_id)
        
    if selectedEntry == "entried":
        qs = qs.filter(is_entry=True, canceled_at__isnull=True)

    elif selectedEntry == "notEntried":
        qs = qs.filter(is_entry=False)

    elif selectedEntry == "canceled":
        qs = qs.filter(canceled_at__isnull=False)
        
    qs = qs.annotate(
            total_sum=Sum(
                ExpressionWrapper(
                    F("items__selected_quantity") * F("items__selected_price"),
                    output_field=DecimalField(max_digits=14, decimal_places=3),
                )
            )
        )
    
    if sortInvoice == "asc":
        qs = qs.order_by("total_sum")
    elif sortInvoice == "desc":
        qs = qs.order_by("-total_sum")
    else:
        qs = qs.order_by("-invoice_date")
        
        
    qs = qs.select_related(
            "partner",
            "warehouse",
            "warehouse2",
            "created_by",
            "entry_created_by",
        ).prefetch_related(
            Prefetch(
                "items",
                queryset=InvoiceItem.objects.select_related(
                    "product",
                    "base_unit_obj",
                ).prefetch_related(
                    "free_items",
                    "units",
                    "product__images",
                )
            )
        )
        
        
        
        
    # ic(qs.count())

    # invoice = qs.first()
    # if invoice:
    #     ic(invoice.id, invoice.partner, invoice.items.count())
    
    # ic(qs)
    
    
    # ================= EXCEL =================

    wb = Workbook()
    ws = wb.active
    ws.title = "Фактуры"
    
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Все Фактуры за {day_start_str} - {day_end_str}"
    ws["A1"].alignment = CENTER_ALIGN
    ws["A1"].font = Font(size=16, bold=True)
    
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 11
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    
    
    row = 3
    count = 0
    grand_total = Decimal("0")
    
   
    for q in qs:
        count += 1
        
        invoice_date = q.invoice_date.strftime("%d.%m.%Y")
        partner_name = q.partner.name if q.partner else ""
        price_type = "Оптовая" if q.type_price == "wholesale_price" else "Розница"
        awto = q.awto.name if q.awto else ""
        warehouse_name = q.warehouse.name if q.warehouse else ""
        warehouse_name2 = q.warehouse2.name if q.warehouse2 else ""
    
        faktura_type = "Черновик"
        if q.wozwrat_or_prihod == "rashod":
            faktura_type = "Расход"
        elif q.wozwrat_or_prihod == "prihod":
            faktura_type = "Приход"
        elif q.wozwrat_or_prihod == "wozwrat":
            faktura_type = "Возврат"
        elif q.wozwrat_or_prihod == "transfer":
            faktura_type = "Со склада на склад"
            
        
        ws.merge_cells(f"A{row}:F{row}")
        if faktura_type == "Черновик":
            ws[f"A{row}"] = f"{count}. {faktura_type} Фактура"
        else:
            ws[f"A{row}"] = f"{count}. {faktura_type} Фактура № {q.id}"
            
        ws[f"A{row}"].font = Font(size=14, bold=True)
        ws[f"A{row}"].alignment = CENTER_ALIGN
            
        row += 1
        
        ws[f"A{row}"] = "Дата:"
        ws[f"B{row}"] = invoice_date
        
        row += 1
        
        ws[f"A{row}"] = "Партнёр:"
        ws[f"A{row}"].alignment = Alignment(vertical="center")
        ws[f"B{row}"] = partner_name
        # ws[f"B{row}"].font = Font(size=14, bold=True)
        
        row += 1
        
        if q.wozwrat_or_prihod != 'transfer':
            ws[f"A{row}"] = "Склад"
            ws[f"B{row}"] = warehouse_name
        else:
            ws[f"A{row}"] = "Склад"
            ws[f"B{row}"] = f"со склада: {warehouse_name} на склад: {warehouse_name2}"
            
         
            
        
        
        row += 1
        
        # ws[f"A{row}"] = "Тип цены:"
        # ws[f"B{row}"] = price_type
        
        # row += 1
        
        ws[f"A{row}"] = "Авто:"
        ws[f"B{row}"] = awto
        
        row += 1
        
        ws[f"A{row}"] = "№     ✓"
        ws[f"B{row}"] = "Наименование"
        ws[f"C{row}"] = "Кол-во"
        ws[f"D{row}"] = "Ед. изм."
        ws[f"E{row}"] = "Цена за шт."
        ws[f"F{row}"] = "Общая цена"
        
        for i in 'ABCDEF':
            ws[f"{i}{row}"].alignment = CENTER_ALIGN
            ws[f"{i}{row}"].font = HEADER_FONT
            ws[f"{i}{row}"].fill = HEADER_FILL
        
        
        row += 1
        
        
        
        item_index = 1
        invoice_total = Decimal("0")
        
        for item in q.items.all():
            product = item.product
            qty = item.selected_quantity or Decimal("0")
            unit = item.unit_name_on_selected_warehouses
            price = item.selected_price or Decimal("0")
            total = qty * price
            invoice_total += total
            
            
            
            for i in 'ABCDEF':
                ws[f"{i}{row}"].border = THIN_BORDER
                
            ws[f"A{row}"] = f" {item_index}"
            ws[f"A{row}"].alignment = LEFT_ALIGN
            
            ws[f"B{row}"] = item.product.name if item.product else ""
            # ws[f"B{row}"].font = Font(size=14)
            ws[f"B{row}"].alignment = LEFT_ALIGN_WRAP
            
            ws[f"C{row}"] = qty
            # ws[f"C{row}"].font = Font(size=14)
            ws[f"C{row}"].number_format = QTY_FMT
            
            ws[f"D{row}"] = unit
            ws[f"D{row}"].alignment = CENTER_ALIGN
            
            ws[f"D{row}"] = unit
            ws[f"D{row}"].alignment = CENTER_ALIGN
            
            ws[f"E{row}"] = price
            ws[f"E{row}"].number_format = PRICE_FMT
            
            ws[f"F{row}"] = total
            ws[f"F{row}"].number_format = PRICE_FMT
            
            
            item_index += 1
            row += 1
        
        ws[f"E{row}"] = "ИТОГО:"
        ws[f"F{row}"] = invoice_total
        ws[f"F{row}"].number_format = PRICE_FMT
        for i in 'ABCDEF':
            ws[f"{i}{row}"].border = THIN_BORDER
            ws[f"{i}{row}"].fill = GREEN_FILL_0
            ws[f"B{row}"].font = Font(bold=True)
            
        grand_total += invoice_total
        
        row += 3
        
    # row += 1
    ws[f"A{row}"] = "ИТОГО:"
    ws[f"A{row}"].font = Font(bold=True, size=14)
    ws[f"A{row}"].alignment = RIGHT_ALIGN
    # row += 1
    ws[f"B{row}"] = f"Кол-во фактур: {count}"
    ws[f"B{row}"].font = Font(bold=True, size=12)
    
    
    row += 1
    
    ws[f"B{row}"] = f"Общая цена: {grand_total}"
    ws[f"B{row}"].font = Font(bold=True, size=12)

    
    
    
    # ws.page_setup.paperSize = ws.PAPERSIZE_A4
    # ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # ws.page_setup.fitToWidth = 1      # 1 страница в ширину
    # ws.page_setup.fitToHeight = False # высота свободная
    
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    ws.page_setup.scale = 75  # 🔥 РЕАЛЬНО РАБОТАЕТ

    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    ws.print_area = f"A1:F{ws.max_row}"
    
        
        
        
        
        

   
   
    # ================= RESPONSE =================


    response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    response["Content-Disposition"] = (
        f'attachment; filename="faktury_{str(day_start_str)}_{str(day_end_str)}.xlsx"'
    )

    wb.save(response)
    return response
    
        
    # data = []
    
    # return Response(data)
    
    
    
    
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_excel_fakturs_diapazon2(request):
    
    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")
    wozwrat_or_prihod = request.GET.get("wozwrat_or_prihod") # wozwrat, prihod, rashod, transfer
    partner_id = request.GET.get("partner_id") # po id
    selectedEntry = request.GET.get("selectedEntry") # entried, notEntried, canceled
    sortInvoice = request.GET.get("sortInvoice") # asc, desc po total_ptice
    
    
    
    # ic(wozwrat_or_prihod)
    # ic(sortInvoice)
    # ic(selectedEntry)
    # ic(partner_id)
    # ic(request)
    

    if not date_from or not date_to:
        return Response(
            {"error": "choose diapazon date"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        day_start = datetime.strptime(date_from, "%Y-%m-%d").date()
        day_end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return Response(
            {"error": "invalid diapazon date format"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if day_start > day_end:
        return Response(
            {"error": "date start must be <= date end"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    day_start_str = day_start.strftime("%d.%m.%Y")
    day_end_str = day_end.strftime("%d.%m.%Y")
    
    
    # ================= BASE QUERY =================

    qs = Invoice.objects.filter(
        invoice_date__date__gte=day_start,
        invoice_date__date__lte=day_end,
    )

    if wozwrat_or_prihod:
        qs = qs.filter(wozwrat_or_prihod=wozwrat_or_prihod)

    if partner_id:
        qs = qs.filter(partner_id=partner_id)

    if selectedEntry == "entried":
        qs = qs.filter(is_entry=True, canceled_at__isnull=True)

    elif selectedEntry == "notEntried":
        qs = qs.filter(is_entry=False)

    elif selectedEntry == "canceled":
        qs = qs.filter(canceled_at__isnull=False)

    # ================= EXPRESSIONS =================
    # используем историческую цену из InvoiceItem

    total_expr = F("items__selected_quantity") * F("items__selected_price")

    pribyl_expr = F("items__selected_quantity") * (
        F("items__selected_price") - F("items__purchase_price")
    )

    raznisa_expr = F("items__selected_quantity") * (
        F("items__selected_price") - F("items__wholesale_price")
    )

    # ================= ANNOTATE (для отображения по накладным) =================

    qs = qs.annotate(
        total_sum=Coalesce(
            Sum(total_expr, output_field=DecimalField(max_digits=14, decimal_places=3)),
            Decimal("0")
        ),
        pribyl=Coalesce(
            Sum(pribyl_expr, output_field=DecimalField(max_digits=14, decimal_places=3)),
            Decimal("0")
        ),
        raznisa=Coalesce(
            Sum(raznisa_expr, output_field=DecimalField(max_digits=14, decimal_places=3)),
            Decimal("0")
        ),
    )

    # ================= SORT =================

    if sortInvoice == "asc":
        qs = qs.order_by("total_sum")
    elif sortInvoice == "desc":
        qs = qs.order_by("-total_sum")
    else:
        qs = qs.order_by("id")

    # ================= SEPARATE QUERYSETS =================

    rashod_qs = qs.filter(wozwrat_or_prihod="rashod")
    prihod_qs = qs.filter(wozwrat_or_prihod="prihod")
    wozwrat_qs = qs.filter(wozwrat_or_prihod="wozwrat")
    transfer_qs = qs.filter(wozwrat_or_prihod="transfer")

    # ================= TOTALS FUNCTION =================

    def calculate_totals(queryset):
        return queryset.aggregate(
            total_sum=Coalesce(
                Sum(total_expr, output_field=DecimalField(max_digits=14, decimal_places=3)),
                Decimal("0")
            ),
            total_pribyl=Coalesce(
                Sum(pribyl_expr, output_field=DecimalField(max_digits=14, decimal_places=3)),
                Decimal("0")
            ),
            total_raznisa=Coalesce(
                Sum(raznisa_expr, output_field=DecimalField(max_digits=14, decimal_places=3)),
                Decimal("0")
            ),
        )

    # ================= TOTALS PER TYPE =================

    rashod_totals = calculate_totals(rashod_qs)
    prihod_totals = calculate_totals(prihod_qs)
    wozwrat_totals = calculate_totals(wozwrat_qs)
    transfer_totals = calculate_totals(transfer_qs)

    # ================= GRAND TOTAL =================
    # ВАЖНО: не используем Sum("total_sum")

    grand_totals = calculate_totals(qs)

    # ================= DEBUG =================

    # for q in rashod_qs:
    #     print(
    #         q.id,
    #         q.total_sum,
    #         q.pribyl,
    #         q.raznisa,
    #         q.partner.name if q.partner else None
    #     )

    # ================= RESPONSE (пример) =================
        
   
    # ================= EXCEL =================

    wb = Workbook()
    ws = wb.active
    ws.title = "Фактуры"
    
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Все Фактуры за {day_start_str} - {day_end_str}"
    ws["A1"].alignment = CENTER_ALIGN
    ws["A1"].font = Font(size=16, bold=True)
    
    ws.merge_cells("A3:I3")
    # ws["A3"] = f"Расходные Фактуры (кол-во: {rashod_qs.count()}, общая цена: {rashod_totals['total_sum']})"
    ws["A3"] = f"Расходные Фактуры"
    ws["A3"].alignment = CENTER_ALIGN
    ws["A3"].font = Font(size=14, bold=True, color="0000FF")
    
    ws["A5"] = "№"
    ws["B5"] = "Дата"
    ws["C5"] = "Фактура №"
    ws["D5"] = "Партнёр"
    ws["E5"] = "Сумма"
    ws["F5"] = "Прибыль"
    ws["G5"] = "Разница"
    ws["H5"] = "Статус"
    ws["I5"] = "Автомобиль"
    
    for i in 'ABCDEFGHI':
        ws[f"{i}5"].alignment = CENTER_ALIGN
        ws[f"{i}5"].font = Font(bold=True)
        ws[f"{i}5"].fill = GRAY_FILL_1
        ws[f"{i}5"].border = THIN_BORDER
    
    
    
    
    
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 35
    
    
    row = 6
    count = 0
    # grand_total = Decimal("0")
    
   
    for q in rashod_qs:
        count += 1
        
        invoice_date = q.invoice_date.strftime("%d.%m.%Y")
        faktura_nomer = q.id
        partner_name = q.partner.name if q.partner else ""
        faktura_sum = q.total_sum
        faktura_pribyl = q.pribyl
        faktura_raznisa = q.raznisa
        faktura_status = "Проведено" if q.is_entry else "Черновик"
        awto = q.awto.name if q.awto else ""
        
        # ic(invoice_date, faktura_nomer, partner_name, faktura_sum, faktura_pribyl, faktura_raznisa, faktura_status)
        for i in 'ABCDEFGHI':
            ws[f"{i}{row}"].border = THIN_BORDER
        ws[f"A{row}"] = count
        ws[f"A{row}"].alignment = CENTER_ALIGN
        ws[f"B{row}"] = invoice_date
        ws[f"C{row}"] = faktura_nomer
        ws[f"C{row}"].alignment = CENTER_ALIGN
        ws[f"D{row}"] = partner_name
        ws[f"E{row}"] = faktura_sum
        ws[f"E{row}"].number_format = PRICE_FMT
        ws[f"F{row}"] = faktura_pribyl
        ws[f"F{row}"].number_format = PRICE_FMT
        ws[f"G{row}"] = faktura_raznisa
        ws[f"G{row}"].number_format = PRICE_FMT
        ws[f"H{row}"] = faktura_status
        ws[f"H{row}"].alignment = CENTER_ALIGN
        ws[f"I{row}"] = awto
        
        row += 1
    ws[f"D{row}"] = f"ИТОГО:"     
    ws[f"D{row}"].alignment = RIGHT_ALIGN   
    ws[f"D{row}"].font = Font(bold=True)
    ws[f"E{row}"] = rashod_totals['total_sum']
    ws[f"E{row}"].number_format = PRICE_FMT
    ws[f"F{row}"] = rashod_totals['total_pribyl']
    ws[f"F{row}"].number_format = PRICE_FMT
    ws[f"G{row}"] = rashod_totals['total_raznisa']
    ws[f"G{row}"].number_format = PRICE_FMT
    for i in 'ABCDEFGHI':
        ws[f"{i}{row}"].border = THIN_BORDER
        ws[f"{i}{row}"].fill = GRAY_FILL_2
        ws[f"{i}{row}"].font = Font(bold=True)
        
    row += 3
        
        
    # wozwrat
    
    ws.merge_cells(f"A{row}:I{row}")
    # ws["A3"] = f"Расходные Фактуры (кол-во: {rashod_qs.count()}, общая цена: {rashod_totals['total_sum']})"
    ws[f"A{row}"] = f"Возвратные Фактуры"
    ws[f"A{row}"].alignment = CENTER_ALIGN
    ws[f"A{row}"].font = Font(size=14, bold=True, color="FF0000")
    
    row += 2
    
    ws[f"A{row}"] = "№"
    ws[f"B{row}"] = "Дата"
    ws[f"C{row}"] = "Фактура №"
    ws[f"D{row}"] = "Партнёр"
    ws[f"E{row}"] = "Сумма"
    ws[f"F{row}"] = "Прибыль"
    ws[f"G{row}"] = "Разница"
    ws[f"H{row}"] = "Статус"
    ws[f"I{row}"] = "Автомобиль"
    
    for i in 'ABCDEFGHI':
        ws[f"{i}{row}"].alignment = CENTER_ALIGN
        ws[f"{i}{row}"].font = Font(bold=True)
        ws[f"{i}{row}"].fill = GRAY_FILL_1
        ws[f"{i}{row}"].border = THIN_BORDER
        
        
    row += 1
    count = 0
    
    for q in wozwrat_qs:
        count += 1
        
        invoice_date = q.invoice_date.strftime("%d.%m.%Y")
        faktura_nomer = q.id
        partner_name = q.partner.name if q.partner else ""
        faktura_sum = q.total_sum
        faktura_pribyl = q.pribyl
        faktura_raznisa = q.raznisa
        faktura_status = "Проведено" if q.is_entry else "Черновик"
        awto = q.awto.name if q.awto else ""
        
        # ic(invoice_date, faktura_nomer, partner_name, faktura_sum, faktura_pribyl, faktura_raznisa, faktura_status)
        for i in 'ABCDEFGHI':
            ws[f"{i}{row}"].border = THIN_BORDER
        ws[f"A{row}"] = count
        ws[f"A{row}"].alignment = CENTER_ALIGN
        ws[f"B{row}"] = invoice_date
        ws[f"C{row}"] = faktura_nomer
        ws[f"C{row}"].alignment = CENTER_ALIGN
        ws[f"D{row}"] = partner_name
        ws[f"E{row}"] = -faktura_sum if faktura_sum > 0 else abs(faktura_sum)
        ws[f"E{row}"].number_format = PRICE_FMT
        ws[f"F{row}"] = -faktura_pribyl if faktura_pribyl > 0 else abs(faktura_pribyl)
        ws[f"F{row}"].number_format = PRICE_FMT
        ws[f"G{row}"] = -faktura_raznisa if faktura_raznisa > 0 else abs(faktura_raznisa)
        ws[f"G{row}"].number_format = PRICE_FMT
        ws[f"H{row}"] = faktura_status
        ws[f"H{row}"].alignment = CENTER_ALIGN
        ws[f"I{row}"] = awto
        
        row += 1
        
    ws[f"D{row}"] = f"ИТОГО:"     
    ws[f"D{row}"].alignment = RIGHT_ALIGN    
    ws[f"D{row}"].font = Font(bold=True)
    ws[f"E{row}"] = -wozwrat_totals['total_sum'] if wozwrat_totals['total_sum'] > 0 else abs(wozwrat_totals['total_sum'])
    ws[f"E{row}"].number_format = PRICE_FMT
    ws[f"F{row}"] = -wozwrat_totals['total_pribyl'] if wozwrat_totals['total_pribyl'] > 0 else abs(wozwrat_totals['total_pribyl'])
    ws[f"F{row}"].number_format = PRICE_FMT
    ws[f"G{row}"] = -wozwrat_totals['total_raznisa'] if wozwrat_totals['total_raznisa'] > 0 else abs(wozwrat_totals['total_raznisa'])
    ws[f"G{row}"].number_format = PRICE_FMT
    for i in 'ABCDEFGHI':
        ws[f"{i}{row}"].border = THIN_BORDER
        ws[f"{i}{row}"].fill = GRAY_FILL_2
        ws[f"{i}{row}"].font = Font(bold=True)
        
    row += 3
    
    # prihod
    
    ws.merge_cells(f"A{row}:I{row}")
    # ws["A3"] = f"Расходные Фактуры (кол-во: {rashod_qs.count()}, общая цена: {rashod_totals['total_sum']})"
    ws[f"A{row}"] = f"Приходные Фактуры"
    ws[f"A{row}"].alignment = CENTER_ALIGN
    ws[f"A{row}"].font = Font(size=14, bold=True, color="006400")
    
    row += 2
    
    ws[f"A{row}"] = "№"
    ws[f"B{row}"] = "Дата"
    ws[f"C{row}"] = "Фактура №"
    ws[f"D{row}"] = "Партнёр"
    ws[f"E{row}"] = "Сумма"
    ws[f"F{row}"] = "Прибыль"
    ws[f"G{row}"] = "Разница"
    ws[f"H{row}"] = "Статус"
    ws[f"I{row}"] = "Автомобиль"
    
    for i in 'ABCDEFGHI':
        ws[f"{i}{row}"].alignment = CENTER_ALIGN
        ws[f"{i}{row}"].font = Font(bold=True)
        ws[f"{i}{row}"].fill = GRAY_FILL_1
        ws[f"{i}{row}"].border = THIN_BORDER
        
        
    row += 1
    count = 0
    
    for q in prihod_qs:
        count += 1
        
        invoice_date = q.invoice_date.strftime("%d.%m.%Y")
        faktura_nomer = q.id
        partner_name = q.partner.name if q.partner else ""
        faktura_sum = q.total_sum
        faktura_pribyl = q.pribyl
        faktura_raznisa = q.raznisa
        faktura_status = "Проведено" if q.is_entry else "Черновик"
        awto = q.awto.name if q.awto else ""
        
        # ic(invoice_date, faktura_nomer, partner_name, faktura_sum, faktura_pribyl, faktura_raznisa, faktura_status)
        for i in 'ABCDEFGHI':
            ws[f"{i}{row}"].border = THIN_BORDER
        ws[f"A{row}"] = count
        ws[f"A{row}"].alignment = CENTER_ALIGN
        ws[f"B{row}"] = invoice_date
        ws[f"C{row}"] = faktura_nomer
        ws[f"C{row}"].alignment = CENTER_ALIGN
        ws[f"D{row}"] = partner_name
        ws[f"E{row}"] = -faktura_sum
        ws[f"E{row}"].number_format = PRICE_FMT
        ws[f"F{row}"] = "❌"
        ws[f"F{row}"].alignment = CENTER_ALIGN
        ws[f"G{row}"] = "❌"
        ws[f"G{row}"].alignment = CENTER_ALIGN
        ws[f"H{row}"] = faktura_status
        ws[f"H{row}"].alignment = CENTER_ALIGN
        ws[f"I{row}"] = awto
        
        row += 1
        
    ws[f"D{row}"] = f"ИТОГО:"     
    ws[f"D{row}"].alignment = RIGHT_ALIGN  
    ws[f"D{row}"].font = Font(bold=True)
    ws[f"E{row}"] = -prihod_totals['total_sum']
    ws[f"E{row}"].number_format = PRICE_FMT
    ws[f"F{row}"] = "❌"
    ws[f"F{row}"].alignment = CENTER_ALIGN
    ws[f"G{row}"] = "❌"
    ws[f"G{row}"].alignment = CENTER_ALIGN
    for i in 'ABCDEFGHI':
        ws[f"{i}{row}"].border = THIN_BORDER
        ws[f"{i}{row}"].fill = GRAY_FILL_2
        ws[f"{i}{row}"].font = Font(bold=True)
        
    row += 3
    
    
    # grand total
    
    grand_total_money = (
        rashod_totals['total_sum']
        - wozwrat_totals['total_sum']
        - prihod_totals['total_sum']
    )
    
    grand_total_pribyl = (
        rashod_totals['total_pribyl']
        - wozwrat_totals['total_pribyl']
    )
    
    grand_total_raznisa = (
        rashod_totals['total_raznisa']
        - wozwrat_totals['total_raznisa']
    )
    
    ws[f"D{row}"] = "ОБЩИЙ ИТОГ:"
    ws[f"D{row}"].font = Font(bold=True, size=14)
    ws[f"D{row}"].alignment = RIGHT_ALIGN
    ws[f"E{row}"] = grand_total_money
    ws[f"E{row}"].number_format = PRICE_FMT
    ws[f"E{row}"].font = Font(bold=True)

    ws[f"F{row}"] = grand_total_pribyl
    ws[f"F{row}"].number_format = PRICE_FMT
    ws[f"F{row}"].font = Font(bold=True)

    ws[f"G{row}"] = grand_total_raznisa
    ws[f"G{row}"].number_format = PRICE_FMT
    ws[f"G{row}"].font = Font(bold=True)
    

    # ================= RESPONSE =================


    response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    response["Content-Disposition"] = (
        f'attachment; filename="faktury_{str(day_start_str)}_{str(day_end_str)}.xlsx"'
    )

    wb.save(response)
    return response
    
        
    # data = []
    
    # return Response(data)
   
    
    
    