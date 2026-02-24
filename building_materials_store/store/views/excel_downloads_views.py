
from django.http import JsonResponse, HttpResponse
# from django.views.decorators.http import require_GET, require_POST
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
from decimal import Decimal, ROUND_HALF_UP
from django.db.models import Q, Sum, Count
from datetime import datetime, timedelta
# from django.utils.dateparse import parse_date

# from django.db.models import F
# from django.contrib.postgres.search import TrigramSimilarity
from ..models import Account, Entry, DayClosing, PartnerBalanceSnapshot, Partner, WarehouseAccount, Product, InvoiceItem
from icecream import ic
# import json
# from collections import defaultdict

from openpyxl import Workbook
# from openpyxl.utils import get_column_letter
# from django.http import FileResponse
# import os
# from io import BytesIO
# from django.core.files.base import ContentFile
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# from openpyxl.formatting.rule import CellIsRule
# from openpyxl.worksheet.page import PageMargins

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status


from ..my_func.get_unit_map import get_unit_map 
from ..my_func.get_unit_and_cf import get_unit_and_cf 


# ================== СТИЛИ ==================

# Шапка
HEADER_FILL = PatternFill(fill_type="solid",fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Категория
CATEGORY_FILL = PatternFill(fill_type="solid", fgColor="E7E6E6")
CATEGORY_FONT = Font(bold=True)

# Обычные ячейки
NORMAL_FONT = Font(color="000000")
LEFT_ALIGN = Alignment(vertical="center", horizontal="left")
RIGHT_ALIGN = Alignment(vertical="center", horizontal="right")

# Границы
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Форматы чисел
PRICE_FMT = '#,##0.00'
QTY_FMT = '#,##0.###'

# Ширина колонок
COLUMN_WIDTHS = {
    "A": 25,
    "B": 40,
    "C": 10,
    "D": 12,
    "E": 8, "F": 8,
    "G": 8, "H": 8,
    "I": 8, "J": 8,
    "K": 8, "L": 8,
    "M": 8, "N": 8,
}

TOTAL_FILL = PatternFill(fill_type="solid",fgColor="D9D9D9")

TOTAL_FONT = Font(bold=True)

GRAY_FILL_0 = PatternFill(fill_type="solid", fgColor="F5F5F5")
GRAY_FILL_1 = PatternFill(fill_type="solid", fgColor="EDEDED")
GRAY_FILL_2 = PatternFill( fill_type="solid", fgColor="DCDCDC")
GRAY_FILL_3 = PatternFill(fill_type="solid", fgColor="C8C8C8")
GRAY_FILL_4 = PatternFill(fill_type="solid", fgColor="B0B0B0")

GREEN_FILL_0 = PatternFill(fill_type="solid", fgColor="E2F0D9")
GREEN_FILL_1 = PatternFill(fill_type="solid", fgColor="C6EFCE")
GREEN_FILL_2 = PatternFill(fill_type="solid", fgColor="92D050")
GREEN_FILL_3 = PatternFill(fill_type="solid", fgColor="006100")

RED_FONT = Font(color="FF0000")
GREEN_FONT = Font(color="006400")
BLUE_FONT = Font(color="0000FF")

def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_osw_excel(request):
    user = request.user

    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")
    

    if not date_from or not date_to:
        return Response(
            {"error": "choose diapazon date"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        day_start = datetime.strptime(date_from, "%Y-%m-%d").date()
        day_end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return Response(
            {"error": "invalid diapazon date format"},
            status=status.HTTP_400_BAD_REQUEST
        )

    if day_start > day_end:
        return Response(
            {"error": "date start must be <= date end"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    day_start_str = day_start.strftime("%d.%m.%Y")
    day_end_str = day_end.strftime("%d.%m.%Y")

    last_closing = DayClosing.objects.filter(
        date__lt=day_start
    ).order_by("-date").first()
    
    accounts_OSW = {}
    
    for account in Account.objects.all():
        accounts_OSW[account.id] = {
            'id': account.id,
            'number': account.number,
            'name': account.name,
            'parent_id': account.parent_id,
            'debit_start': Decimal('0'),
            'credit_start': Decimal('0'),
            'opening_balance': Decimal('0'),
            'debit_turnover': Decimal('0'),
            'credit_turnover': Decimal('0'),
            'closing_balance': Decimal('0'),
        }
        
    opening_OSW = (
        Entry.objects
        .filter(transaction__date__lt=day_start)
        .values('account')
        .annotate(
            debit_sum=Sum('debit'),
            credit_sum=Sum('credit')
        )
    )
    
    for o in opening_OSW:
        debit = o['debit_sum'] or Decimal('0')
        credit = o['credit_sum'] or Decimal('0')

        acc = accounts_OSW[o['account']]
        acc['debit_start'] = debit
        acc['credit_start'] = credit
        acc['opening_balance'] = debit - credit
        
    turnover_OSW = (
        Entry.objects
        .filter(transaction__date__gte=day_start, transaction__date__lte=day_end)
        .values('account')
        .annotate(
            debit_sum=Sum('debit'),
            credit_sum=Sum('credit')
        )
    )
    
    for t in turnover_OSW:
        debit = t['debit_sum'] or Decimal('0')
        credit = t['credit_sum'] or Decimal('0')

        acc = accounts_OSW[t['account']]
        acc['debit_turnover'] = debit
        acc['credit_turnover'] = credit
        
    for acc in accounts_OSW.values():
        acc['closing_balance'] = (
            acc['opening_balance']
            + acc['debit_turnover']
            - acc['credit_turnover']
        )
        
    total_debit = sum(acc['debit_turnover'] for acc in accounts_OSW.values())
    total_credit = sum(acc['credit_turnover'] for acc in accounts_OSW.values())
    
    if total_debit != total_credit:
        return Response(
            {"error": "double entry violated"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    # сортируем счета так, чтобы дети шли первыми
    accounts_sorted = sorted(
        accounts_OSW.values(),
        key=lambda x: x['number'].count('.'),
        reverse=True
    )
    
    for acc in accounts_sorted:
        parent_id = acc['parent_id']
        if parent_id:
            parent = accounts_OSW[parent_id]

            parent['debit_start'] += acc['debit_start']
            parent['credit_start'] += acc['credit_start']
            parent['opening_balance'] += acc['opening_balance']

            parent['debit_turnover'] += acc['debit_turnover']
            parent['credit_turnover'] += acc['credit_turnover']
            parent['closing_balance'] += acc['closing_balance']
            
    wb_OSW = Workbook()
    wb_OSW.remove(wb_OSW.active)
    
    ws = wb_OSW.create_sheet(title="ОСВ")
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws["A1"] = "Оборотно-сальдовая ведомость"
    ws["A2"] = f"{str(day_start_str)} - {str(day_end_str)}"
    
    ws["A1"].alignment = CENTER_ALIGN
    ws["A2"].alignment = CENTER_ALIGN
    
    ws["A1"].font = CATEGORY_FONT
    ws["A2"].font = CATEGORY_FONT
    
    ws["A4"] = "Счёт"
    ws["B4"] = "Название счёта"
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    
    ws.merge_cells("C4:D4")
    ws["C4"] = "Сальдо на начало"
    
    ws.merge_cells("E4:F4")
    ws["E4"] = "Обороты за период"
    
    ws.merge_cells("G4:H4")
    ws["G4"] = "Сальдо на конец"
    
    ws["C5"] = "Дебет"
    ws["D5"] = "Кредит"
    
    ws["E5"] = "Дебет"
    ws["F5"] = "Кредит"
    
    ws["G5"] = "Дебет"
    ws["H5"] = "Кредит"
    
    for i in ["A4", "B4", "C4", "D4", "E4", "F4", "G4", "H4"]:
        ws[i].alignment = CENTER_ALIGN
        ws[i].font = CATEGORY_FONT
        ws[i].fill = GRAY_FILL_1
        ws[i].border = THIN_BORDER
        
    for i in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5"]:
        ws[i].alignment = CENTER_ALIGN
        ws[i].font = CATEGORY_FONT
        ws[i].fill = GRAY_FILL_1
        ws[i].border = THIN_BORDER
        
    row = 6
    
    # сначала найдём все parent_id
    parent_ids = {
        acc['parent_id']
        for acc in accounts_OSW.values()
        if acc['parent_id']
    }

    # листовые счета = те, кто не родитель
    leaf_accounts = [
        acc for acc_id, acc in accounts_OSW.items()
        if acc_id not in parent_ids
    ]
    
    detail_sheets = {}
    
    for acc in leaf_accounts:
        title=f"{acc['number']} {acc['name'][:20]}"
        ws_detail = wb_OSW.create_sheet(title=title)
        detail_sheets[acc["id"]] = ws_detail

        ws_detail["A1"] = "← ОСВ"
        ws_detail["A1"].font = Font(bold=True, color="0000FF")
        ws_detail["A1"].hyperlink = f"#'ОСВ'!A1"
        
        if acc['number'].startswith(("40", "42")):
            ws_detail.merge_cells("A2:N2")
            ws_detail.merge_cells("A3:N3")
        else:
            ws_detail.merge_cells("A2:H2")
            ws_detail.merge_cells("A3:H3")
        ws_detail["A2"] = f"Счёт {acc['number']} — {acc['name']}"
        ws_detail["A2"].alignment = CENTER_ALIGN
        ws_detail["A2"].font = TOTAL_FONT
        
       
        ws_detail["A3"] = f"{str(day_start_str)} - {str(day_end_str)}"
        ws_detail["A3"].alignment = CENTER_ALIGN
        ws_detail["A3"].font = TOTAL_FONT
        
    for account_id, data in accounts_OSW.items():
        cell = ws[f"B{row}"]
        cell.value = data["number"]
        
        title=f"{data['number']} {data['name'][:20]}"
        if data["id"] in detail_sheets:
            cell.hyperlink = f"#'{title}'!A1"
            cell.style = "Hyperlink"
        
        for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}"]:
            ws[i].border = THIN_BORDER
            
        ws[f"C{row}"].number_format = PRICE_FMT
        ws[f"D{row}"].number_format = PRICE_FMT
        ws[f"E{row}"].number_format = PRICE_FMT
        ws[f"F{row}"].number_format = PRICE_FMT
        ws[f"G{row}"].number_format = PRICE_FMT
        ws[f"H{row}"].number_format = PRICE_FMT

     
            
        ws[f"A{row}"] = data["number"]
        ws[f"B{row}"] = data["name"]
        
        start_osw_debit = Decimal("0.00")
        start_osw_credit = Decimal("0.00")
        if (data["debit_start"] - data["credit_start"]) > 0:
            start_osw_debit = data["debit_start"] - data["credit_start"]
        elif (data["debit_start"] - data["credit_start"]) < 0:
            start_osw_credit = abs(data["debit_start"] - data["credit_start"])
        
        ws[f"C{row}"] = start_osw_debit
        ws[f"D{row}"] = start_osw_credit
        
        ws[f"E{row}"] = data["debit_turnover"]
        ws[f"F{row}"] = data["credit_turnover"]
        ws[f"E{row}"].font = GREEN_FONT
        ws[f"F{row}"].font = RED_FONT
        
        
        if data["closing_balance"] >= 0:
            ws[f"G{row}"] = data["closing_balance"]   # дебет
            ws[f"H{row}"] = Decimal("0.00") 
        else:
            ws[f"G{row}"] = Decimal("0.00")
            ws[f"H{row}"] = abs(data["closing_balance"])  # кредит
            
        row += 1
        
    grand_total = {
        'debit_start': Decimal('0'),
        'credit_start': Decimal('0'),
        'debit_turnover': Decimal('0'),
        'credit_turnover': Decimal('0'),
        'closing_debit': Decimal('0'),
        'closing_credit': Decimal('0'),
    }
    
    for acc in accounts_OSW.values():
        grand_total['debit_start'] += acc['debit_start']
        grand_total['credit_start'] += acc['credit_start']
        grand_total['debit_turnover'] += acc['debit_turnover']
        grand_total['credit_turnover'] += acc['credit_turnover']

        if acc['closing_balance'] >= 0:
            grand_total['closing_debit'] += acc['closing_balance']
        else:
            grand_total['closing_credit'] += abs(acc['closing_balance'])
            
    grand_total = {
        'debit_start': sum(a['debit_start'] for a in leaf_accounts),
        'credit_start': sum(a['credit_start'] for a in leaf_accounts),
        'debit_turnover': sum(a['debit_turnover'] for a in leaf_accounts),
        'credit_turnover': sum(a['credit_turnover'] for a in leaf_accounts),
        'closing_debit': sum(
            a['closing_balance'] for a in leaf_accounts if a['closing_balance'] >= 0
        ),
        'closing_credit': sum(
            abs(a['closing_balance']) for a in leaf_accounts if a['closing_balance'] < 0
        ),
    }
    
    ws[f"A{row}"] = "ИТОГО"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    ws[f"C{row}"] = grand_total['debit_start']
    ws[f"D{row}"] = grand_total['credit_start']
    ws[f"E{row}"] = grand_total['debit_turnover']
    ws[f"F{row}"] = grand_total['credit_turnover']
    ws[f"G{row}"] = grand_total['closing_debit']
    ws[f"H{row}"] = grand_total['closing_credit']

    for col in "ABCDEFGH":
        ws[f"{col}{row}"].font = CATEGORY_FONT
        ws[f"{col}{row}"].border = THIN_BORDER
        ws[f"{col}{row}"].fill = GRAY_FILL_1
        
    for col in "CDEFGH":
        ws[f"{col}{row}"].number_format = PRICE_FMT
        
    row += 1
    
    ws[f"A{row}"] = "САЛЬДО"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    saldo_start = grand_total['debit_start'] - grand_total['credit_start']
    saldo_turnover = grand_total['debit_turnover'] - grand_total['credit_turnover']
    saldo_end = grand_total['closing_debit'] - grand_total['closing_credit']
    if saldo_start > 0:
        ws[f"C{row}"] = saldo_start
    else:
        ws[f"D{row}"] = abs(saldo_start)
    if saldo_turnover > 0:
        ws[f"E{row}"] = saldo_turnover
    else:
        ws[f"F{row}"] = abs(saldo_turnover)
    if saldo_end > 0:    
        ws[f"G{row}"] = saldo_end
    else:
        ws[f"H{row}"] = abs(saldo_end)

    for col in "ABCDEFGH":
        ws[f"{col}{row}"].font = CATEGORY_FONT
        ws[f"{col}{row}"].border = THIN_BORDER
        ws[f"{col}{row}"].fill = GRAY_FILL_1
        
    for col in "CDEFGH":
        ws[f"{col}{row}"].number_format = PRICE_FMT
        
    partner_snap_map = {}
    
    if last_closing:
        partner_balance_snapshots = PartnerBalanceSnapshot.objects.filter(closing=last_closing).select_related("partner")
        partner_snap_map = {
            (s.partner.id): {
                    "balance_60_usd": s.balance_60_usd, 
                    "balance_62_tmt": s.balance_62_tmt, 
                    "balance_75_usd": s.balance_75_usd, 
                    "balance_76_tmt": s.balance_76_tmt,
                    
                    "balance_60_usd_credit": s.balance_60_usd_credit,
                    "balance_60_usd_debit": s.balance_60_usd_debit,
                    
                    "balance_62_tmt_credit": s.balance_62_tmt_credit,
                    "balance_62_tmt_debit": s.balance_62_tmt_debit,
                    
                    "balance_75_usd_credit": s.balance_75_usd_credit,
                    "balance_75_usd_debit": s.balance_75_usd_debit,
                    
                    "balance_76_tmt_credit": s.balance_76_tmt_credit,
                    "balance_76_tmt_debit": s.balance_76_tmt_debit,

                }
            for s in partner_balance_snapshots
        }
        
    for acc in leaf_accounts:
        account_id = acc["id"]
        account_number = acc["number"]
        

        if account_number.startswith("60") or account_number.startswith("62"):
        
            partners = Partner.objects.all().select_related("agent")
            
            account_60_62 = {}
            for p in partners:
                partner = {
                    "id": p.id,
                    "name": p.name,
                    "type": p.type,
                }
                
                if p.agent:
                    agent = {
                        "id": p.agent.id,
                        "name": p.agent.name,
                    }
                else:
                    agent = {}
                    
                account_60_62[p.id] = {
                    "partner": partner,
                    "agent": agent,
                    "debit_start": Decimal("0.00"),
                    "credit_start": Decimal("0.00"),
                    "debit_turnover": Decimal("0.00"),
                    "credit_turnover": Decimal("0.00"),
                    "debit_end": Decimal("0.00"),
                    "credit_end": Decimal("0.00"),
                }
                
            if last_closing:
                for p_id, value in partner_snap_map.items():
                    if p_id not in account_60_62:
                        continue
                    if account_number == "60":
                        account_60_62[p_id]["debit_start"] += value["balance_60_usd_debit"]
                        account_60_62[p_id]["credit_start"] += value["balance_60_usd_credit"]
                    else:
                        account_60_62[p_id]["debit_start"] += value["balance_62_tmt_debit"]
                        account_60_62[p_id]["credit_start"] += value["balance_62_tmt_credit"]
                        
            else:
                entries_start = (
                    Entry.objects
                    .filter(
                        account_id=account_id,
                        transaction__date__lt=day_end
                    )
                    .select_related(
                        'transaction',
                        'partner__agent',
                        'product',
                        'warehouse',
                        'transaction__invoice',
                    )
                    .order_by('transaction__date', 'id')
                )
                
                for e in entries_start:
                    if not e.partner:
                        continue
                    partner_id = e.partner.id
                    
                    
                    account_60_62[partner_id]["debit_start"] += money(e.debit)
                    account_60_62[partner_id]["credit_start"] += money(e.credit)
                    
            entries_turnover = (
                Entry.objects
                .filter(
                    account_id=account_id,
                    transaction__date__gte=day_start,
                    transaction__date__lte=day_end,
                )
                .select_related(
                    'transaction',
                    'partner__agent',
                    'product',
                    'warehouse',
                    'transaction__invoice',
                )
                .order_by('transaction__date', 'id')
            )
            
            for e in entries_turnover:
                    if not e.partner:
                        continue
                    partner_id = e.partner.id
                    
                    account_60_62[partner_id]["debit_turnover"] += money(e.debit)
                    account_60_62[partner_id]["credit_turnover"] += money(e.credit)
                    
            total_credit_start = Decimal("0.00")        
            total_debit_start = Decimal("0.00")        
            total_credit_turnover = Decimal("0.00")        
            total_debit_turnover = Decimal("0.00")        
            total_credit_end = Decimal("0.00")        
            total_debit_end = Decimal("0.00")
            
            
            for p_id, row in account_60_62.items():
                debit_end_raw = row["debit_start"] + row["debit_turnover"]
                credit_end_raw = row["credit_start"] + row["credit_turnover"]

                saldo = debit_end_raw - credit_end_raw

                if saldo >= 0:
                    row["debit_end"] = saldo
                    row["credit_end"] = Decimal("0.00")
                else:
                    row["debit_end"] = Decimal("0.00")
                    row["credit_end"] = -saldo
                    
                # 👉 ИТОГИ СЧИТАЕМ ЗДЕСЬ
                total_debit_start += row["debit_start"]
                total_credit_start += row["credit_start"]

                total_debit_turnover += row["debit_turnover"]
                total_credit_turnover += row["credit_turnover"]

                total_debit_end += row["debit_end"]
                total_credit_end += row["credit_end"]
                
            grand_total_60_62 =  {
                "total_credit_start": total_credit_start,
                "total_debit_start": total_debit_start,
                "total_credit_turnover": total_credit_turnover,
                "total_debit_turnover": total_debit_turnover,
                "total_credit_end": total_credit_end,
                "total_debit_end": total_debit_end,
            }
            
            ws_detail = detail_sheets[account_id]
            ws_detail.freeze_panes = "A7"
            ws_detail["A5"] = "№"
            ws_detail.column_dimensions["A"].width = 5
            ws_detail["B5"] = "Agent"
            ws_detail.column_dimensions["B"].width = 20
            ws_detail["C5"] = "Субконто"
            ws_detail.column_dimensions["C"].width = 45
            
            ws_detail.merge_cells("D5:E5")
            ws_detail["D5"] = "Сальдо на начало"

            
            ws_detail.merge_cells("F5:G5")
            ws_detail["F5"] = "Обороты за период"
         
            
            ws_detail.merge_cells("H5:I5")
            ws_detail["H5"] = "Сальдо на конец"
            
            
            for i in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", 
                      "A6", "B6", "C6", "D6", "E6", "F6", "G6", "H6", "I6"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = CENTER_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
                
                
            for i in ["D", "E", "F", "G", "H", "I"]:
                ws_detail.column_dimensions[i].width = 15
                
                
            ws_detail["D6"] = "Дебит"
            ws_detail["E6"] = "Кредит"
            
            ws_detail["F6"] = "Дебит"
            ws_detail["G6"] = "Кредит"
            
            ws_detail["H6"] = "Дебит"
            ws_detail["I6"] = "Кредит"
            
            row = 7
            count = 1
            for p_id, v in account_60_62.items():
                
                ws_detail[f"D{row}"].number_format = PRICE_FMT
                ws_detail[f"E{row}"].number_format = PRICE_FMT
                ws_detail[f"F{row}"].number_format = PRICE_FMT
                ws_detail[f"G{row}"].number_format = PRICE_FMT
                ws_detail[f"H{row}"].number_format = PRICE_FMT
                ws_detail[f"I{row}"].number_format = PRICE_FMT
                
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                    ws_detail[i].border = THIN_BORDER
                
        
        
                ws_detail[f"A{row}"] = count
                ws_detail[f"B{row}"] = v["agent"]["name"] if v["agent"] else ""
                ws_detail[f"C{row}"] = v["partner"]["name"] if v["partner"] else ""
                
                start_saldo_debit = Decimal("0")
                start_saldo_credit = Decimal("0")
                saldo_start_row = v["debit_start"] - v["credit_start"]
                if saldo_start_row > 0:
                    start_saldo_debit = saldo_start_row
                elif saldo_start_row < 0:
                    start_saldo_credit = abs(saldo_start_row)
                ws_detail[f"D{row}"] = start_saldo_debit
                ws_detail[f"E{row}"] = start_saldo_credit
                
                ws_detail[f"F{row}"] = v["debit_turnover"]
                ws_detail[f"G{row}"] = v["credit_turnover"]
                
                ws_detail[f"H{row}"] = v["debit_end"]
                ws_detail[f"I{row}"] = v["credit_end"]
                
                count += 1
                row += 1
                
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = RIGHT_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
            
            
            ws_detail.merge_cells(f"A{row}:C{row}")
            ws_detail[f"A{row}"].alignment = LEFT_ALIGN
            ws_detail[f"A{row}"] = "Итого развернутое:"
            
            ws_detail[f"D{row}"].number_format = PRICE_FMT
            ws_detail[f"E{row}"].number_format = PRICE_FMT
            ws_detail[f"F{row}"].number_format = PRICE_FMT
            ws_detail[f"G{row}"].number_format = PRICE_FMT
            ws_detail[f"H{row}"].number_format = PRICE_FMT
            ws_detail[f"I{row}"].number_format = PRICE_FMT
            
            ws_detail[f"D{row}"] = grand_total_60_62["total_debit_start"]
            ws_detail[f"E{row}"] = grand_total_60_62["total_credit_start"]
            
            ws_detail[f"F{row}"] = grand_total_60_62["total_debit_turnover"]
            ws_detail[f"G{row}"] = grand_total_60_62["total_credit_turnover"]
            
            ws_detail[f"H{row}"] = grand_total_60_62["total_debit_end"]
            ws_detail[f"I{row}"] = grand_total_60_62["total_credit_end"]
            
            
            row += 1
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = RIGHT_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
                
            ws_detail.merge_cells(f"A{row}:C{row}")
            ws_detail[f"A{row}"].alignment = LEFT_ALIGN
            ws_detail[f"A{row}"] = "Всего:"
            
            ws_detail[f"D{row}"].number_format = PRICE_FMT
            ws_detail[f"E{row}"].number_format = PRICE_FMT
            ws_detail[f"F{row}"].number_format = PRICE_FMT
            ws_detail[f"G{row}"].number_format = PRICE_FMT
            ws_detail[f"H{row}"].number_format = PRICE_FMT
            ws_detail[f"I{row}"].number_format = PRICE_FMT
            
            saldo_start = grand_total_60_62["total_debit_start"] - grand_total_60_62["total_credit_start"]
            ws_detail[f"D{row}"] = saldo_start if saldo_start > 0 else 0
            ws_detail[f"E{row}"] = abs(saldo_start) if saldo_start < 0 else 0
            
            saldo_turnover = grand_total_60_62["total_debit_turnover"] - grand_total_60_62["total_credit_turnover"]
            ws_detail[f"F{row}"] = saldo_turnover if saldo_turnover > 0 else 0
            ws_detail[f"G{row}"] = abs(saldo_turnover) if saldo_turnover < 0 else 0
            
            saldo_end = grand_total_60_62["total_debit_end"] - grand_total_60_62["total_credit_end"]
            ws_detail[f"H{row}"] = saldo_end if saldo_end > 0 else 0
            ws_detail[f"I{row}"] = abs(saldo_end) if saldo_end < 0 else 0
            
        elif account_number.startswith("40") or account_number.startswith("42"):
            # w_acc = WarehouseAccount.objects.get(account_id=account_id)
            account_40_42 = {}
            w_acc = (
                WarehouseAccount.objects
                .select_related("warehouse")
                .get(account_id=account_id)
            )
            
            products = Product.objects.filter(
                warehouse_products__warehouse_id=w_acc.warehouse_id
            ).distinct()

            
            unit_map = get_unit_map()
            
            for p in products:
                if not p.category:
                    continue
                unit, cf = get_unit_and_cf(unit_map, p)
                category_id = p.category.id
                category_name = p.category.name
       
                if category_id not in account_40_42:
                    account_40_42[category_id] = {
                        "category": {
                            "id": category_id,
                            "name": category_name,
                        },
                        "totals": {
                            "start_qty": Decimal("0.00"),
                            "start_price": Decimal("0.00"),

                            "prihod_qty": Decimal("0.00"),
                            "prihod_price": Decimal("0.00"),

                            "wozwrat_qty": Decimal("0.00"),
                            "wozwrat_price": Decimal("0.00"),

                            "rashod_qty": Decimal("0.00"),
                            "rashod_price": Decimal("0.00"),

                            "end_qty": Decimal("0.00"),
                            "end_price": Decimal("0.00"),
                        },
                        "products": {}
                    }

                account_40_42[category_id]["products"][p.id] = {
                    "product": {
                        "id": p.id,
                        "name": p.name,
                        "unit": unit,
                        "cf": Decimal(cf),
                        "wholsale_price": Decimal(p.wholesale_price),
                    },

                    "start_qty": Decimal("0.00"),
                    "start_price": Decimal("0.00"),

                    "prihod_qty": Decimal("0.00"),
                    "prihod_price": Decimal("0.00"),

                    "wozwrat_qty": Decimal("0.00"),
                    "wozwrat_price": Decimal("0.00"),

                    "rashod_qty": Decimal("0.00"),
                    "rashod_price": Decimal("0.00"),

                    "end_qty": Decimal("0.00"),
                    "end_price": Decimal("0.00"),
                }
                
            start_items = InvoiceItem.objects.filter(
                invoice__entry_created_at_handle__lt=day_start,
                invoice__canceled_at__isnull=True
            ).filter(
                Q(invoice__warehouse_id=w_acc.warehouse_id) |
                Q(invoice__warehouse2_id=w_acc.warehouse_id)
            ).select_related(
                "product", "product__category", "product__base_unit", "invoice"
            )
            
            grand_total = {
                "start_qty": Decimal("0.00"),
                "start_price": Decimal("0.00"),

                "prihod_qty": Decimal("0.00"),
                "prihod_price": Decimal("0.00"),

                "wozwrat_qty": Decimal("0.00"),
                "wozwrat_price": Decimal("0.00"),

                "rashod_qty": Decimal("0.00"),
                "rashod_price": Decimal("0.00"),

                "end_qty": Decimal("0.00"),
                "end_price": Decimal("0.00"),
            }
            
            for item in start_items:
                p = item.product
                inv = item.invoice
                
                # Товар уже инициализирован выше
                if not p.category:
                    continue
                if (p.category.id not in account_40_42 or p.id not in account_40_42[p.category.id]["products"]):
                    continue
                cat_totals = account_40_42[p.category.id]["totals"]
                prod = account_40_42[p.category.id]["products"][p.id]
                cf = prod["product"]["cf"]
                qty = Decimal(item.selected_quantity) / cf
                calculated_price = qty * Decimal(p.wholesale_price)
                # Проверяем принадлежность к выбранным складам
                if inv.wozwrat_or_prihod == "prihod":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["start_qty"] += qty
                        prod["start_price"] += calculated_price
                        
                        cat_totals["start_qty"] += qty
                        cat_totals["start_price"] += calculated_price
                        
                        grand_total["start_qty"] += qty
                        grand_total["start_price"] += calculated_price

                elif inv.wozwrat_or_prihod == "rashod":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["start_qty"] -= qty
                        prod["start_price"] -= calculated_price
                        
                        cat_totals["start_qty"] -= qty
                        cat_totals["start_price"] -= calculated_price
                        
                        grand_total["start_qty"] -= qty
                        grand_total["start_price"] -= calculated_price

                elif inv.wozwrat_or_prihod == "wozwrat":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["start_qty"] += qty
                        prod["start_price"] += calculated_price
                        
                        cat_totals["start_qty"] += qty
                        cat_totals["start_price"] += calculated_price
                        
                        grand_total["start_qty"] += qty
                        grand_total["start_price"] += calculated_price
                    
                elif inv.wozwrat_or_prihod == "transfer":
                    # если со склада (и склад в выбранных)
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["start_qty"] -= qty
                        prod["start_price"] -= calculated_price
                        
                        cat_totals["start_qty"] -= qty
                        cat_totals["start_price"] -= calculated_price
                        
                        grand_total["start_qty"] -= qty
                        grand_total["start_price"] -= calculated_price
                        
                    # если на склад (и склад в выбранных)
                    elif inv.warehouse2_id == w_acc.warehouse_id:
                        prod["start_qty"] += qty
                        prod["start_price"] += calculated_price
                        
                        cat_totals["start_qty"] += qty
                        cat_totals["start_price"] += calculated_price
                        
                        grand_total["start_qty"] += qty
                        grand_total["start_price"] += calculated_price
                        
            turnover_items = InvoiceItem.objects.filter(
                invoice__entry_created_at_handle__gte=day_start,
                invoice__entry_created_at_handle__lte=day_end,
                invoice__canceled_at__isnull=True
            ).filter(
                Q(invoice__warehouse_id=w_acc.warehouse_id) |
                Q(invoice__warehouse2_id=w_acc.warehouse_id)
            ).select_related(
                "product", "product__category", "product__base_unit", "invoice"
            ).order_by(
                "product__category__name", "product__name"
            )
            
            for item in turnover_items:
                p = item.product
                inv = item.invoice
                
                # Товар уже инициализирован выше
                if not p.category:
                    continue
                if (p.category.id not in account_40_42 or p.id not in account_40_42[p.category.id]["products"]):
                    continue
                cat_totals = account_40_42[p.category.id]["totals"]
                prod = account_40_42[p.category.id]["products"][p.id]
                cf = prod["product"]["cf"]
                qty = Decimal(item.selected_quantity) / cf
                calculated_price = qty * Decimal(item.selected_price)
                # Проверяем принадлежность к выбранным складам
                if inv.wozwrat_or_prihod == "prihod":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["prihod_qty"] += qty
                        prod["prihod_price"] += calculated_price
                        
                        cat_totals["prihod_qty"] += qty
                        cat_totals["prihod_price"] += calculated_price
                        
                        grand_total["prihod_qty"] += qty
                        grand_total["prihod_price"] += calculated_price

                elif inv.wozwrat_or_prihod == "rashod":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["rashod_qty"] += qty
                        prod["rashod_price"] += calculated_price
                        
                        cat_totals["rashod_qty"] += qty
                        cat_totals["rashod_price"] += calculated_price
                        
                        grand_total["rashod_qty"] += qty
                        grand_total["rashod_price"] += calculated_price

                elif inv.wozwrat_or_prihod == "wozwrat":
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["wozwrat_qty"] += qty
                        prod["wozwrat_price"] += calculated_price
                        
                        cat_totals["wozwrat_qty"] += qty
                        cat_totals["wozwrat_price"] += calculated_price
                        
                        grand_total["wozwrat_qty"] += qty
                        grand_total["wozwrat_price"] += calculated_price
                    
                elif inv.wozwrat_or_prihod == "transfer":
                    # если со склада (и склад в выбранных)
                    if inv.warehouse_id == w_acc.warehouse_id:
                        prod["rashod_qty"] += qty
                        prod["rashod_price"] += calculated_price
                        
                        cat_totals["rashod_qty"] += qty
                        cat_totals["rashod_price"] += calculated_price
                        
                        grand_total["rashod_qty"] += qty
                        grand_total["rashod_price"] += calculated_price
                        
                    # если на склад (и склад в выбранных)
                    elif inv.warehouse2_id == w_acc.warehouse_id:
                        prod["prihod_qty"] += qty
                        prod["prihod_price"] += calculated_price
                        
                        cat_totals["prihod_qty"] += qty
                        cat_totals["prihod_price"] += calculated_price
                        
                        grand_total["prihod_qty"] += qty
                        grand_total["prihod_price"] += calculated_price
                        
            for cat in account_40_42.values():
                for prod in cat["products"].values():
                    prod["end_qty"] = (
                        prod["start_qty"]
                        + prod["prihod_qty"]
                        - prod["rashod_qty"]
                        + prod["wozwrat_qty"]
                    )

                    # prod["end_price"] = (
                    #     prod["start_price"]
                    #     + prod["prihod_price"]
                    #     - prod["rashod_price"]
                    #     + prod["wozwrat_price"]
                    # )
                    wholesale_price = Decimal(prod["product"]["wholsale_price"])
                    prod["end_price"] = prod["end_qty"] * wholesale_price
                    
            for cat in account_40_42.values():
                totals = cat["totals"]
                totals["end_qty"] = (
                    totals["start_qty"]
                    + totals["prihod_qty"]
                    - totals["rashod_qty"]
                    + totals["wozwrat_qty"]
                )
                # totals["end_price"] = (
                #     totals["start_price"]
                #     + totals["prihod_price"]
                #     - totals["rashod_price"]
                #     + totals["wozwrat_price"]
                # )
                totals["end_price"] = sum(
                    p["end_price"] for p in cat["products"].values()
                )
                
            grand_total["end_qty"] = (
                grand_total["start_qty"]
                + grand_total["prihod_qty"]
                - grand_total["rashod_qty"]
                + grand_total["wozwrat_qty"]
            )
            # grand_total["end_price"] = (
            #     grand_total["start_price"]
            #     + grand_total["prihod_price"]
            #     - grand_total["rashod_price"]
            #     + grand_total["wozwrat_price"]
            # )
            grand_total["end_price"] = sum(
                cat["totals"]["end_price"] for cat in account_40_42.values()
            )
            
            ws_detail = detail_sheets[account_id]
            ws_detail.freeze_panes = "A7"
            
      
            ws_detail["A5"] = "№"
            ws_detail.column_dimensions["A"].width = 8
            ws_detail["B5"] = "Наименование товара"
            ws_detail.column_dimensions["B"].width = 55
            ws_detail["C5"] = "Ед."
            ws_detail.column_dimensions["C"].width = 8
            
            ws_detail["D5"] = "Цена"
            ws_detail.column_dimensions["D"].width = 8
            ws_detail[f"D5"].number_format = PRICE_FMT
            
            ws_detail.merge_cells("E5:F5")
            ws_detail["E5"] = "Остаток на начало"
   

            
            ws_detail.merge_cells("G5:H5")
            ws_detail["G5"] = "Приход"
         
            
            ws_detail.merge_cells("I5:J5")
            ws_detail["I5"] = "Возврат"
            
            ws_detail.merge_cells("K5:L5")
            ws_detail["K5"] = "Расход"
            
            ws_detail.merge_cells("M5:N5")
            ws_detail["M5"] = "Остаток на конец"
            
            for i in ["A5", "B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5", "L5", "M5", "N5", 
                      "A6", "B6", "C6", "D6", "E6", "F6", "G6", "H6", "I6", "J6", "K6", "L6", "M6", "N6"]:
                ws_detail[i].font = TOTAL_FONT
                ws_detail[i].alignment = CENTER_ALIGN
                ws_detail[i].fill = GRAY_FILL_1
                ws_detail[i].border = THIN_BORDER
                
                
            for i in ["E", "F",  "G", "H", "I", "J", "K", "L", "M", "N"]:
                ws_detail.column_dimensions[i].width = 11
                
            ws_detail["E6"] = "Кол-во"
            ws_detail["F6"] = "Всего"
            
            ws_detail["G6"] = "Кол-во"
            ws_detail["H6"] = "Всего"
            
            ws_detail["I6"] = "Кол-во"
            ws_detail["J6"] = "Всего"
            
            ws_detail["K6"] = "Кол-во"
            ws_detail["L6"] = "Всего"
            
            ws_detail["M6"] = "Кол-во"
            ws_detail["N6"] = "Всего"
            
            row = 7 
            count = 1
            
            for cat_id, values in sorted(
                account_40_42.items(),
                key=lambda item: item[1]["category"]["name"].lower()
            ):
                cat_name = values["category"]["name"]
                totals = values["totals"]
                ws_detail.merge_cells(f"A{row}:N{row}")
                ws_detail[f"A{row}"].fill = GREEN_FILL_0
                ws_detail[f"A{row}"].font = TOTAL_FONT
                ws_detail[f"A{row}"].alignment = LEFT_ALIGN
                ws_detail[f"A{row}"] = cat_name
                
                
                
                for col in "FHJLN":
                    ws_detail[f"{col}{row}"].number_format = PRICE_FMT
                    
                for col in "EGIKM":
                    ws_detail[f"{col}{row}"].number_format = QTY_FMT
                
                row += 1
                
                for products, value in values["products"].items():
                    product = value["product"]
                    unit = product["unit"]
                    wholesale_price = product["wholsale_price"]
                    
                    start_qty = value["start_qty"]
                    start_price = value["start_price"]
                    
                    prihod_qty = value["prihod_qty"]
                    prihod_price = value["prihod_price"]
                    
                    
                    wozwrat_qty = value["wozwrat_qty"]
                    wozwrat_price = value["wozwrat_price"]
                    
                    rashod_qty = value["rashod_qty"]
                    rashod_price = value["rashod_price"]
                    
                    end_qty = value["end_qty"]
                    end_price = value["end_price"] 
                    
                    ws_detail[f"G{row}"].font = GREEN_FONT
                    ws_detail[f"H{row}"].font = GREEN_FONT
                    ws_detail[f"I{row}"].font = RED_FONT
                    ws_detail[f"J{row}"].font = RED_FONT
                    ws_detail[f"K{row}"].font = BLUE_FONT
                    ws_detail[f"L{row}"].font = BLUE_FONT
                    
                    for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                        ws_detail[i].border = THIN_BORDER
                        
                    for col in "FHJLN":
                        ws_detail[f"{col}{row}"].number_format = PRICE_FMT
                        
                    for col in "EGIKM":
                        ws_detail[f"{col}{row}"].number_format = QTY_FMT
                
                    ws_detail[f"A{row}"] = count
                    ws_detail[f"B{row}"] = product["name"]
                    ws_detail[f"C{row}"] = unit
                    ws_detail[f"D{row}"] = wholesale_price
                    
                    ws_detail[f"E{row}"] = start_qty
                    ws_detail[f"F{row}"] = start_price
                    
                    ws_detail[f"G{row}"] = prihod_qty
                    ws_detail[f"H{row}"] = prihod_price
                    
                    ws_detail[f"I{row}"] = wozwrat_qty
                    ws_detail[f"J{row}"] = wozwrat_price
                    
                    ws_detail[f"K{row}"] = rashod_qty
                    ws_detail[f"L{row}"] = rashod_price
                    
                    ws_detail[f"M{row}"] = end_qty
                    ws_detail[f"N{row}"] = end_price
                    
                
                    count += 1
                    row += 1
            
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                        ws_detail[i].border = THIN_BORDER
                        
                ws_detail.merge_cells(f"A{row}:D{row}")
                for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                        ws_detail[i].fill = GRAY_FILL_0
                ws_detail[f"A{row}"].font = TOTAL_FONT
                ws_detail[f"A{row}"].alignment = RIGHT_ALIGN
                ws_detail[f"A{row}"] = f"Итого по категории: {cat_name}"
                
              
                ws_detail[f"E{row}"] = totals["start_qty"]
                ws_detail[f"F{row}"] = totals["start_price"]
                ws_detail[f"G{row}"] = totals["prihod_qty"]
                ws_detail[f"H{row}"] = totals["prihod_price"]
                ws_detail[f"I{row}"] = totals["wozwrat_qty"]
                ws_detail[f"J{row}"] = totals["wozwrat_price"]
                ws_detail[f"K{row}"] = totals["rashod_qty"]
                ws_detail[f"L{row}"] = totals["rashod_price"]
                ws_detail[f"M{row}"] = totals["end_qty"]
                ws_detail[f"N{row}"] = totals["end_price"]
                                
                row += 1
                
            ws_detail[f"E{row}"] = grand_total["start_qty"]
            ws_detail[f"F{row}"] = grand_total["start_price"]
            ws_detail[f"G{row}"] = grand_total["prihod_qty"]
            ws_detail[f"H{row}"] = grand_total["prihod_price"]
            ws_detail[f"I{row}"] = grand_total["wozwrat_qty"]
            ws_detail[f"J{row}"] = grand_total["wozwrat_price"]
            ws_detail[f"K{row}"] = grand_total["rashod_qty"]
            ws_detail[f"L{row}"] = grand_total["rashod_price"]
            ws_detail[f"M{row}"] = grand_total["end_qty"]
            ws_detail[f"N{row}"] = grand_total["end_price"]
            
            
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws_detail[i].border = THIN_BORDER
                
            for col in "FHJLN":
                ws_detail[f"{col}{row}"].number_format = PRICE_FMT
                
            for col in "EGIKM":
                ws_detail[f"{col}{row}"].number_format = QTY_FMT
                        
            ws_detail.merge_cells(f"A{row}:D{row}")
            for i in [f"A{row}", f"B{row}", f"C{row}", f"D{row}", f"E{row}", f"F{row}", f"G{row}", f"H{row}", f"I{row}", f"J{row}", f"K{row}", f"L{row}", f"M{row}", f"N{row}"]:
                ws_detail[i].fill = GREEN_FILL_1
       
            ws_detail[f"A{row}"].font = TOTAL_FONT
            ws_detail[f"A{row}"].alignment = RIGHT_ALIGN
            ws_detail[f"A{row}"] = "ВСЕГО:"
            
        else:
            # poka drugie account continue potom dodelaem
            continue


        
        
        
        
        
        
        
        
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="osw_{str(day_start_str)}_{str(day_end_str)}.xlsx"'
    )

    wb_OSW.save(response)
    return response
    
    
        
    
