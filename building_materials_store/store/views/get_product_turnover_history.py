from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta
from rest_framework.decorators import api_view, permission_classes
from ..models import Invoice, InvoiceItem, ProductUnit, Product, StockSnapshot, ProductImage, Warehouse
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.views.decorators.csrf import csrf_exempt
from collections import defaultdict
from django.db.models import Sum, F, Q, Case, When, Value, IntegerField
from django.http import JsonResponse
from icecream import ic
from django.db import transaction as db_transaction
import pandas as pd
from datetime import date
from django.shortcuts import get_object_or_404
import time
from django.db.models import Sum, F
from rest_framework import status
from django.http import JsonResponse, HttpResponse


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import numbers
from openpyxl.styles import PatternFill
from openpyxl.worksheet.page import PageMargins



HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
SECTION_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")


def clean_number(val):
    val = float(val)
    return val if val != 0 else None



@api_view(['GET'])
def get_product_turnover_history(request):
    product_id = request.GET.get("product_id")
    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")
    warheousesId = request.GET.get("warheousesId")
    
    # ic(product_id, date_from, date_to, warheousesId)
    
    if not date_from or not date_to:
        return Response(
            {"error": "choose diapazon date"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    try:
        day_start = datetime.strptime(date_from, "%Y-%m-%d").date()
        day_end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return Response(
            {"error": "invalid diapazon date format"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    if day_start > day_end:
        return Response(
            {"error": "date start must be <= date end"},
            status=status.HTTP_400_BAD_REQUEST
        )
        
    day_start_str = day_start.strftime("%d.%m.%Y")
    day_end_str = day_end.strftime("%d.%m.%Y")
    warehouse_ids = []
    if warheousesId:
        warehouse_ids = warheousesId.split(",")
        warehouse_ids = [int(id) for id in warehouse_ids]
        
    if warehouse_ids:
        pass
    #     # Новый формат: несколько складов
    #     try:
    #         warehouse_ids = [int(id.strip()) for id in warheousesId.split(",") if id.strip().isdigit()]
    #     except ValueError:
    #         return JsonResponse({"status": "error", "message": "Invalid warehouse IDs"}, status=400)
    # elif warheousesId and warheousesId.isdigit():
    #     # Старый формат: один склад
    #     warehouse_ids = [int(warehouse_id_param)]
    else:
        pass
        # return JsonResponse(
        #     {"status": "error", "message": "choose correct warehouse(s)"},
        #     status=400
        # )
    
    if not product_id:
        return JsonResponse(
            {"status": "error", "message": "choose correct product"},
            status=400
        )

    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return JsonResponse(
            {"status": "error", "message": "product not found"},
            status=404
        )

    # базовая единица по умолчанию
    unit_name = product.base_unit.name
    conversion_factor = 1

    default_unit = ProductUnit.objects.filter(
        product=product,
        is_default_for_sale=True
    ).select_related('unit').first()

    if default_unit:
        unit_name = default_unit.unit.name
        conversion_factor = float(default_unit.conversion_factor)

    # Данные для каждого склада
    warehouses_data = {}
    all_rows = []
    
    if not warehouse_ids:
        warehouse_ids = list(Warehouse.objects.values_list("id", flat=True))
    
    if warehouse_ids:
        for warehouse_id in warehouse_ids:
            try:
                warehouse_obj = Warehouse.objects.get(id=warehouse_id)
            except Warehouse.DoesNotExist:
                continue
            
            items = (
                InvoiceItem.objects
                .filter(
                    product_id=product.id,
                    invoice__entry_created_at_handle__lt=day_start
                )
                .filter(
                    Q(invoice__warehouse_id=warehouse_id) |
                    Q(invoice__warehouse2_id=warehouse_id)
                )
                .select_related("invoice")
            )
            
            start_quantity = Decimal("0")
            cf = Decimal(str(conversion_factor))
            
            for item in items:
                inv = item.invoice
                if inv.canceled_at != None:
                    continue
                qty = Decimal(item.selected_quantity) / cf

                if inv.wozwrat_or_prihod == "prihod":
                    if inv.warehouse_id == warehouse_id:
                        start_quantity += qty

                elif inv.wozwrat_or_prihod == "rashod":
                    if inv.warehouse_id == warehouse_id:
                        start_quantity -= qty

                elif inv.wozwrat_or_prihod == "wozwrat":
                    if inv.warehouse_id == warehouse_id:
                        start_quantity += qty

                elif inv.wozwrat_or_prihod == "transfer":
                    if inv.warehouse_id == warehouse_id:
                        start_quantity -= qty
                    elif inv.warehouse2_id == warehouse_id:
                        start_quantity += qty
            
            period_items = (
                InvoiceItem.objects
                .filter(
                    product_id=product.id,
                    invoice__entry_created_at_handle__gte=day_start,
                    invoice__entry_created_at_handle__lte=day_end
                )
                .filter(
                    Q(invoice__warehouse_id=warehouse_id) |
                    Q(invoice__warehouse2_id=warehouse_id)
                )
                .select_related("invoice", "invoice__partner")
                .order_by("invoice__entry_created_at_handle", "invoice_id")
            )
            
            rows = []
            balance_qty = start_quantity
            cf = Decimal(str(conversion_factor))

            for item in period_items:
                inv = item.invoice
                if inv.canceled_at != None:
                    continue
                qty = Decimal(item.selected_quantity) / cf
                price = Decimal(item.selected_price)
                income_qty = Decimal("0")
                outcome_qty = Decimal("0")
                return_qty = Decimal("0")

                # ПРИХОД
                if inv.wozwrat_or_prihod == "prihod" and inv.warehouse_id == warehouse_id:
                    income_qty = qty
                    balance_qty += qty

                # РАСХОД
                elif inv.wozwrat_or_prihod == "rashod" and inv.warehouse_id == warehouse_id:
                    outcome_qty = qty
                    balance_qty -= qty

                # ВОЗВРАТ
                elif inv.wozwrat_or_prihod == "wozwrat" and inv.warehouse_id == warehouse_id:
                    return_qty = qty
                    balance_qty += qty

                # ПЕРЕМЕЩЕНИЕ
                elif inv.wozwrat_or_prihod == "transfer":
                    if inv.warehouse_id == warehouse_id:
                        outcome_qty = qty
                        balance_qty -= qty
                    elif inv.warehouse2_id == warehouse_id:
                        income_qty = qty
                        balance_qty += qty
                
                row_data = {
                    "date": inv.entry_created_at_handle,
                    "invoice_id": inv.id,
                    "partner": inv.partner.name if inv.partner else "",
                    "operation": inv.wozwrat_or_prihod,
                    "text": f"{inv.get_wozwrat_or_prihod_display()} №{inv.id} ({inv.comment})",

                    "price": price,

                    "income_qty": income_qty,
                    "income_sum": income_qty * price,

                    "outcome_qty": outcome_qty,
                    "outcome_sum": outcome_qty * price,

                    "return_qty": return_qty,
                    "return_sum": return_qty * price,

                    "balance_qty": balance_qty,
                    "balance_sum": balance_qty * price,

                    "warehouse_id": warehouse_id,
                }
                rows.append(row_data)
                all_rows.append(row_data)
            
            total_income_qty = Decimal("0")
            total_income_sum = Decimal("0")
            total_outcome_qty = Decimal("0")
            total_outcome_sum = Decimal("0")
            total_return_qty = Decimal("0")
            total_return_sum = Decimal("0")

            for r in rows:
                total_income_qty += r["income_qty"]
                total_income_sum += r["income_sum"]

                total_outcome_qty += r["outcome_qty"]
                total_outcome_sum += r["outcome_sum"]

                total_return_qty += r["return_qty"]
                total_return_sum += r["return_sum"]

            end_quantity = (
                start_quantity
                + total_income_qty
                + total_return_qty
                - total_outcome_qty
            )
            end_sum = end_quantity * product.wholesale_price
            
            warehouses_data[str(warehouse_id)] = {
                "id": warehouse_id,
                "name": warehouse_obj.name,
                "start_quantity": start_quantity,
                "income_qty": total_income_qty,
                "income_sum": total_income_sum,
                "outcome_qty": total_outcome_qty,
                "outcome_sum": total_outcome_sum,
                "return_qty": total_return_qty,
                "return_sum": total_return_sum,
                "end_quantity": end_quantity,
                "end_sum": end_sum,
            }
        
    # Сортируем общие строки по дате
    all_rows.sort(key=lambda x: x["date"])
    
    # Агрегируем общие данные
    total_start_quantity = sum(wh["start_quantity"] for wh in warehouses_data.values())
    total_income_qty = sum(wh["income_qty"] for wh in warehouses_data.values())
    total_income_sum = sum(wh["income_sum"] for wh in warehouses_data.values())
    total_outcome_qty = sum(wh["outcome_qty"] for wh in warehouses_data.values())
    total_outcome_sum = sum(wh["outcome_sum"] for wh in warehouses_data.values())
    total_return_qty = sum(wh.get("return_qty", 0) for wh in warehouses_data.values())
    total_return_sum = sum(wh.get("return_sum", 0) for wh in warehouses_data.values())
    total_end_quantity = sum(wh["end_quantity"] for wh in warehouses_data.values())
    total_end_sum = total_end_quantity * product.wholesale_price
    
    product_image = (
        ProductImage.objects
        .filter(product=product)
        .only("image")
        .first()
    )

    image_url = product_image.image.url if product_image else None
                    
    data = {
        "product_id": product.id,
        "product_name": product.name,
        "product_unit": unit_name,
        "product_wholesale_price": product.wholesale_price,
        "product_retail_price": product.retail_price,
        "warehouses": warehouses_data,
        "start_quantity": total_start_quantity,
        "image": image_url,
        "turnover": {
            "income_qty": total_income_qty,
            "income_sum": total_income_sum,
            "outcome_qty": total_outcome_qty,
            "outcome_sum": total_outcome_sum,
            "return_qty": total_return_qty,
            "return_sum": total_return_sum,
        },
        "end": {
            "quantity": total_end_quantity,
            "sum": total_end_sum
        },
        "rows": all_rows,
    }
    
    return JsonResponse({"data": data})






def build_product_turnover_data(product, day_start, day_end, warheousesId):
    
    warehouse_ids = []

    if warheousesId:
        warehouse_ids = [int(id) for id in warheousesId.split(",") if id.isdigit()]
        
    if not warehouse_ids:
        warehouse_ids = list(Warehouse.objects.values_list("id", flat=True))
    
        # базовая единица по умолчанию
    unit_name = product.base_unit.name
    conversion_factor = 1

    default_unit = ProductUnit.objects.filter(
        product=product,
        is_default_for_sale=True
    ).select_related('unit').first()

    if default_unit:
        unit_name = default_unit.unit.name
        conversion_factor = float(default_unit.conversion_factor)

    # Данные для каждого склада
    warehouses_data = {}
    all_rows = []
    
    if not warehouse_ids:
        warehouse_ids = list(Warehouse.objects.values_list("id", flat=True))
    
    if warehouse_ids:
        for warehouse_id in warehouse_ids:
            try:
                warehouse_obj = Warehouse.objects.get(id=warehouse_id)
            except Warehouse.DoesNotExist:
                continue
            
            items = (
                InvoiceItem.objects
                .filter(
                    product_id=product.id,
                    invoice__entry_created_at_handle__lt=day_start
                )
                .filter(
                    Q(invoice__warehouse_id=warehouse_id) |
                    Q(invoice__warehouse2_id=warehouse_id)
                )
                .select_related("invoice")
            )
            
            start_quantity = Decimal("0")
            cf = Decimal(str(conversion_factor))
            
            for item in items:
                inv = item.invoice
                if inv.canceled_at != None:
                    continue
                qty = Decimal(item.selected_quantity) / cf

                if inv.wozwrat_or_prihod == "prihod":
                    if inv.warehouse_id == warehouse_id:
                        start_quantity += qty

                elif inv.wozwrat_or_prihod == "rashod":
                    if inv.warehouse_id == warehouse_id:
                        start_quantity -= qty

                elif inv.wozwrat_or_prihod == "wozwrat":
                    if inv.warehouse_id == warehouse_id:
                        start_quantity += qty

                elif inv.wozwrat_or_prihod == "transfer":
                    if inv.warehouse_id == warehouse_id:
                        start_quantity -= qty
                    elif inv.warehouse2_id == warehouse_id:
                        start_quantity += qty
            
            period_items = (
                InvoiceItem.objects
                .filter(
                    product_id=product.id,
                    invoice__entry_created_at_handle__gte=day_start,
                    invoice__entry_created_at_handle__lte=day_end
                )
                .filter(
                    Q(invoice__warehouse_id=warehouse_id) |
                    Q(invoice__warehouse2_id=warehouse_id)
                )
                .select_related("invoice", "invoice__partner")
                .order_by("invoice__entry_created_at_handle", "invoice_id")
            )
            
            rows = []
            balance_qty = start_quantity
            cf = Decimal(str(conversion_factor))

            for item in period_items:
                inv = item.invoice
                if inv.canceled_at != None:
                    continue
                qty = Decimal(item.selected_quantity) / cf
                price = Decimal(item.selected_price)
                income_qty = Decimal("0")
                outcome_qty = Decimal("0")
                return_qty = Decimal("0")

                # ПРИХОД
                if inv.wozwrat_or_prihod == "prihod" and inv.warehouse_id == warehouse_id:
                    income_qty = qty
                    balance_qty += qty

                # РАСХОД
                elif inv.wozwrat_or_prihod == "rashod" and inv.warehouse_id == warehouse_id:
                    outcome_qty = qty
                    balance_qty -= qty

                # ВОЗВРАТ
                elif inv.wozwrat_or_prihod == "wozwrat" and inv.warehouse_id == warehouse_id:
                    return_qty = qty
                    balance_qty += qty

                # ПЕРЕМЕЩЕНИЕ
                elif inv.wozwrat_or_prihod == "transfer":
                    if inv.warehouse_id == warehouse_id:
                        outcome_qty = qty
                        balance_qty -= qty
                    elif inv.warehouse2_id == warehouse_id:
                        income_qty = qty
                        balance_qty += qty
                
                row_data = {
                    "date": inv.entry_created_at_handle,
                    "invoice_id": inv.id,
                    "partner": inv.partner.name if inv.partner else "",
                    "operation": inv.wozwrat_or_prihod,
                    "text": f"{inv.get_wozwrat_or_prihod_display()} №{inv.id} ({inv.comment})",

                    "price": price,

                    "income_qty": income_qty,
                    "income_sum": income_qty * price,

                    "outcome_qty": outcome_qty,
                    "outcome_sum": outcome_qty * price,

                    "return_qty": return_qty,
                    "return_sum": return_qty * price,

                    "balance_qty": balance_qty,
                    "balance_sum": balance_qty * price,

                    "warehouse_id": warehouse_id,
                }
                rows.append(row_data)
                all_rows.append(row_data)
            
            total_income_qty = Decimal("0")
            total_income_sum = Decimal("0")
            total_outcome_qty = Decimal("0")
            total_outcome_sum = Decimal("0")
            total_return_qty = Decimal("0")
            total_return_sum = Decimal("0")

            for r in rows:
                total_income_qty += r["income_qty"]
                total_income_sum += r["income_sum"]

                total_outcome_qty += r["outcome_qty"]
                total_outcome_sum += r["outcome_sum"]

                total_return_qty += r["return_qty"]
                total_return_sum += r["return_sum"]

            end_quantity = (
                start_quantity
                + total_income_qty
                + total_return_qty
                - total_outcome_qty
            )
            end_sum = end_quantity * product.wholesale_price
            
            warehouses_data[str(warehouse_id)] = {
                "id": warehouse_id,
                "name": warehouse_obj.name,
                "start_quantity": start_quantity,
                "income_qty": total_income_qty,
                "income_sum": total_income_sum,
                "outcome_qty": total_outcome_qty,
                "outcome_sum": total_outcome_sum,
                "return_qty": total_return_qty,
                "return_sum": total_return_sum,
                "end_quantity": end_quantity,
                "end_sum": end_sum,
            }
        
    # Сортируем общие строки по дате
    all_rows.sort(key=lambda x: x["date"])
    
    # Агрегируем общие данные
    total_start_quantity = sum(wh["start_quantity"] for wh in warehouses_data.values())
    total_income_qty = sum(wh["income_qty"] for wh in warehouses_data.values())
    total_income_sum = sum(wh["income_sum"] for wh in warehouses_data.values())
    total_outcome_qty = sum(wh["outcome_qty"] for wh in warehouses_data.values())
    total_outcome_sum = sum(wh["outcome_sum"] for wh in warehouses_data.values())
    total_return_qty = sum(wh.get("return_qty", 0) for wh in warehouses_data.values())
    total_return_sum = sum(wh.get("return_sum", 0) for wh in warehouses_data.values())
    total_end_quantity = sum(wh["end_quantity"] for wh in warehouses_data.values())
    total_end_sum = total_end_quantity * product.wholesale_price
    
    product_image = (
        ProductImage.objects
        .filter(product=product)
        .only("image")
        .first()
    )

    image_url = product_image.image.url if product_image else None
                    
    data = {
        "product_id": product.id,
        "product_name": product.name,
        "product_unit": unit_name,
        "product_wholesale_price": product.wholesale_price,
        "product_retail_price": product.retail_price,
        "warehouses": warehouses_data,
        "start_quantity": total_start_quantity,
        "image": image_url,
        "turnover": {
            "income_qty": total_income_qty,
            "income_sum": total_income_sum,
            "outcome_qty": total_outcome_qty,
            "outcome_sum": total_outcome_sum,
            "return_qty": total_return_qty,
            "return_sum": total_return_sum,
        },
        "end": {
            "quantity": total_end_quantity,
            "sum": total_end_sum
        },
        "rows": all_rows,
    }
    return data
    


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_product_turnover_excel(request):

    product_id = request.GET.get("product_id")
    date_from = request.GET.get("dateFrom")
    date_to = request.GET.get("dateTo")
    warheousesId = request.GET.get("warheousesId")

    if not product_id or not date_from or not date_to:
        return Response({"error": "missing params"}, status=400)

    # ✅ Парсим даты
    try:
        day_start = datetime.strptime(date_from, "%Y-%m-%d").date()
        day_end = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return Response({"error": "invalid date format"}, status=400)

    # ✅ Получаем продукт
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        return Response({"error": "product not found"}, status=404)

    # ✅ Передаём правильные данные
    data = build_product_turnover_data(
        product,
        day_start,
        day_end,
        warheousesId
    )

    

    wb = Workbook()
    ws = wb.active
    ws.title = "Оборот товара"

    THIN = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1

    # =====================================================
    # 🔷 ЗАГОЛОВОК
    # =====================================================
    ws.merge_cells("A1:M1")
    ws["A1"] = f"{data['product_name']} | {date_from} - {date_to}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    row += 2


    # =====================================================
    # 🔷 ПЕРВАЯ СТРОКА ШАПКИ (с объединениями)
    # =====================================================
    ws.cell(row=row, column=1).value = "№"
    ws.cell(row=row, column=2).value = "Дата"
    ws.cell(row=row, column=3).value = "Партнёр"
    ws.cell(row=row, column=4).value = "Комментарий"
    ws.cell(row=row, column=5).value = "Цена"

    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    ws.cell(row=row, column=6).value = "Приход"

    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    ws.cell(row=row, column=8).value = "Возврат"

    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
    ws.cell(row=row, column=10).value = "Расход"

    ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=13)
    ws.cell(row=row, column=12).value = "Остаток"

    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
        cell.fill = HEADER_FILL

    row += 1


    # =====================================================
    # 🔷 ВТОРАЯ СТРОКА ШАПКИ
    # =====================================================
    sub_headers = [
        "", "", "", "", "",
        "Кол-во", "Сумма",
        "Кол-во", "Сумма",
        "Кол-во", "Сумма",
        "Кол-во", "Сумма",
    ]
    
    

    for col, header in enumerate(sub_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
        cell.fill = HEADER_FILL

    row += 1


    # =====================================================
    # 🔷 OPENING BALANCE
    # =====================================================
    opening_row = row
    ws.cell(row=row, column=2).value = f"Остаток на начало {date_from}"
    ws.cell(row=row, column=12).value = float(data["start_quantity"])
    ws.cell(row=row, column=13).value = float(
        data["start_quantity"] * data["product_wholesale_price"]
    )

    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN
        cell.font = Font(bold=True, size=13)
        cell.fill = SECTION_FILL

    row += 1
    
    


    # =====================================================
    # 🔷 ОСНОВНЫЕ СТРОКИ
    # =====================================================
    for i, r in enumerate(data["rows"], start=1):

        values = [
            i,
            r["date"].strftime("%d.%m.%Y"),
            r["partner"] or "-",
            r["text"] or "-",
            clean_number(r["price"]),
            clean_number(r["income_qty"]),
            clean_number(r["income_sum"]),
            clean_number(r["return_qty"]),
            clean_number(r["return_sum"]),
            clean_number(r["outcome_qty"]),
            clean_number(r["outcome_sum"]),
            clean_number(r["balance_qty"]),
            clean_number(r["balance_sum"]),
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = val
            cell.border = THIN
            cell.font = Font(size=12)

            # if col >= 5:
            #     cell.number_format = '#,##0.00'
            #     cell.alignment = Alignment(horizontal="right")
            
            if col >= 5 and val not in (None, 0):
                cell.number_format = '#,##0.00'

                # Приход (6,7) — зелёный
                if col in (6, 7) and val:
                    # cell.font = Font(color="008000")  # green
                    cell.font = Font(size=12, color="008000") # green

                # Возврат (8,9) — красный
                elif col in (8, 9) and val:
                    cell.font = Font(size=12, color="FF0000")  # red

                # Расход (10,11) — синий
                elif col in (10, 11) and val:
                    cell.font = Font(size=12, color="0000FF")  # blue

        row += 1
        
        # # Приход
        # ws.cell(row=row, column=6).font = Font(bold=True, color="008000")
        # ws.cell(row=row, column=7).font = Font(bold=True, color="008000")

        # # Возврат
        # ws.cell(row=row, column=8).font = Font(bold=True, color="FF0000")
        # ws.cell(row=row, column=9).font = Font(bold=True, color="FF0000")

        # # Расход
        # ws.cell(row=row, column=10).font = Font(bold=True, color="0000FF")
        # ws.cell(row=row, column=11).font = Font(bold=True, color="0000FF")


    # =====================================================
    # 🔷 TOTAL TURNOVER
    # =====================================================
    total_row = row
    ws.cell(row=row, column=2).value = "Итого оборот"

    ws.cell(row=row, column=6).value = float(data["turnover"]["income_qty"])
    ws.cell(row=row, column=7).value = float(data["turnover"]["income_sum"])
    ws.cell(row=row, column=8).value = float(data["turnover"]["return_qty"])
    ws.cell(row=row, column=9).value = float(data["turnover"]["return_sum"])
    ws.cell(row=row, column=10).value = float(data["turnover"]["outcome_qty"])
    ws.cell(row=row, column=11).value = float(data["turnover"]["outcome_sum"])

    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN
        cell.font = Font(bold=True, size=13)
        cell.fill = SECTION_FILL

    row += 1


    # =====================================================
    # 🔷 CLOSING BALANCE
    # =====================================================
    closing_row = row
    ws.cell(row=row, column=2).value = f"Остаток на конец {date_to}"
    ws.cell(row=row, column=12).value = float(data["end"]["quantity"])
    ws.cell(row=row, column=13).value = float(data["end"]["sum"])

    for col in range(1, 14):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN
        cell.font = Font(bold=True, size=13)
        cell.fill = SECTION_FILL


    # =====================================================
    # 🔷 ШИРИНА КОЛОНОК (чтобы красиво)
    # =====================================================
    # widths = [5, 12, 20, 35, 10, 12, 14, 12, 14, 12, 14, 12, 14]
    widths = [6, 14, 28, 35, 10, 12, 14, 12, 14, 12, 14, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
        
    # Включаем перенос строки для колонок A и B
    # Включаем перенос строки для колонок A, B, C, D
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):

        current_row = r[0].row

        # A и B — центр
        for col in (1, 2):
            cell = r[col - 1]

            # ❌ без переноса для служебных строк
            if current_row in (opening_row, closing_row, total_row):
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=False
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        # C и D — перенос только для обычных строк
        for col in (3, 4):
            cell = r[col - 1]

            if current_row in (opening_row, closing_row, total_row):
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=False
                )
            else:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True
                )


    # =====================================================
    # 🔷 RESPONSE
    # =====================================================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="product_turnover_{product_id}.xlsx"'
    )
    
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToPage = True
    
    ws.page_margins = PageMargins(
        left=0.3,
        right=0.2,
        top=0.2,
        bottom=0.2,
    )

    wb.save(response)
    return response

