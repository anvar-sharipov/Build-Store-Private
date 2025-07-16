from django.shortcuts import render
import time
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from icecream import ic
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter
from collections import OrderedDict


from rest_framework import viewsets, status, filters
from .models import *
from .serializers import *

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, BasePermission, IsAuthenticated, SAFE_METHODS
from rest_framework.decorators import api_view, action

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import render, get_object_or_404
from rest_framework.generics import CreateAPIView
from django_filters.rest_framework import DjangoFilterBackend
from .filters import ProductFilter
from django.views.decorators.http import require_GET
from django.http import JsonResponse
from django.contrib.postgres.search import TrigramSimilarity
from django.db.models import Q
from django.utils.dateparse import parse_datetime, parse_date
from django.db.models import Sum, F, Count
from openpyxl.styles import Font
from rest_framework.exceptions import PermissionDenied
from django.db import transaction
from datetime import datetime

from rest_framework.pagination import PageNumberPagination
# swoy pagination 
class CustomPageNumberPagination(PageNumberPagination):
    page_size = 2              # Кол-во элементов на странице по умолчанию
    page_size_query_param = 'page_size'  # Позволяет клиенту указать page_size в запросе (?page_size=50)
    max_page_size = 100          # Максимальное кол-во элементов на странице, чтобы не перегружать сервер



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user(request):
    user = request.user
    return Response({
        "username": user.username,
        "photo": request.build_absolute_uri(user.photo.url) if user.photo else None
    })


# swoyo razgranichenie dostupa
class IsInAdminOrWarehouseGroup(BasePermission):
    """
    Чтение — всем авторизованным.
    Изменение — только группам 'admin' и 'warehouse_manager'.
    """
    def has_permission(self, request, view):
        user = request.user
        # Сначала проверяем – аутентификация выполнена?
        if not user or not user.is_authenticated:
            return False
        # Разрешаем чтение всем аутентифицированным
        if request.method in SAFE_METHODS:
            return True # GET, HEAD, OPTIONS
        # Для мутации — проверяем группы пользователя
        allowed = user.groups.filter(name__in=['admin', 'warehouse_manager']).exists()
        if not allowed:
            raise PermissionDenied(detail="accessOnlyForAdmin")
        return allowed





class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer









# стал полнофункциональным ModelViewSet (CRUD),
# работал с одной точки входа (/api/products/),
# использовал групповое разграничение доступа:
# GET, HEAD, OPTIONS → всем авторизованным,
# POST, PUT, PATCH, DELETE → только тем, кто в группе admin или warehouse_manager.

class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    pagination_class = CustomPageNumberPagination

    filter_backends = [DjangoFilterBackend]
    filterset_class = ProductFilter

    def get_permissions(self):
        return [IsInAdminOrWarehouseGroup()]

    def get_queryset(self):
        qs = Product.objects.all()
        qs = qs.select_related('category', 'base_unit', 'brand', 'model')
        qs = qs.prefetch_related('units__unit', 'images', 'batches')
        return qs.distinct()

    def list(self, request, *args, **kwargs):
        # time.sleep(1)  # для теста задержка
        queryset = self.filter_queryset(self.get_queryset())

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        # time.sleep(1)  # задержка для теста
        return super().retrieve(request, *args, **kwargs)
    
    def update(self, request, *args, **kwargs):
        time.sleep(2)
        return super().update(request, *args, **kwargs)




class PartnerEntriesView(APIView):
    
    def get(self, request, partner_id):
        entries = Entry.objects.filter(transaction__partner_id=partner_id).order_by('transaction__date', 'id')

        running_balance = Decimal('0.00')
        result = []

        for entry in entries:
            entry_data = EntrySerializer(entry).data

            debit = Decimal(entry_data.get('debit') or '0')
            credit = Decimal(entry_data.get('credit') or '0')

            account_number = entry_data['account']['number']
            # print('account_number', account_number)

            # Считаем сальдо только по счету покупателя 62.*
            if account_number.startswith('62'):
                running_balance += debit - credit
                print('running_balance', running_balance)
                entry_data['running_balance'] = str(running_balance)
            else:
                entry_data['running_balance'] = ''  # или None

            result.append(entry_data)

        print('running_balance', running_balance)


        return Response(result)


class PartnerViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Partner.objects.all().order_by('-pk')
    serializer_class = PartnerSerializer

    def get_permissions(self):
        return [IsInAdminOrWarehouseGroup()]

    def get_queryset(self):
        queryset = super().get_queryset()
        agent_id = self.request.query_params.get('agent_id')
        if agent_id:
            queryset = queryset.filter(agent_id=agent_id)
        return queryset

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # time.sleep(2)
        self.perform_destroy(instance)
        return Response(
            {"message": "partnerDeleted"},
            status=status.HTTP_204_NO_CONTENT
        )
    

    def update(self, request, *args, **kwargs):
        # time.sleep(2)
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data) # res.data
    

        
class WarehouseViewSet(viewsets.ModelViewSet):
    queryset = Warehouse.objects.all()
    serializer_class = WarehouseSerializer


class RegisterView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Пользователь создан'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    

class GroupViewSet(viewsets.ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializers

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(
            {
                'message': 'groupAdded',
                'group': serializer.data
            },
            status=status.HTTP_201_CREATED
        )



class AgentViewSet(viewsets.ModelViewSet):
    queryset = Agent.objects.all().order_by('-pk')
    serializer_class = AgentSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        return [IsInAdminOrWarehouseGroup()]

    def list(self, request, *args, **kwargs):
        # time.sleep(1)  # задержка 2 секунды
        return super().list(request, *args, **kwargs)

    
    def destroy(self, request, *args, **kwargs):
        # return Response(
        #         {'message': 'supplierHasSupplies'},
        #         status=status.HTTP_400_BAD_REQUEST
        #     )
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({'message': 'AgentDeleted'}, status=status.HTTP_200_OK)
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response({'data': serializer.data, "type": "success"}, status=status.HTTP_201_CREATED)

 


class EmployeeViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = Employee.objects.all().order_by('-pk')
    serializer_class = EmployeeSerializer

    def get_permissions(self):
        return [IsInAdminOrWarehouseGroup()]

    # Фильтрация и поиск
    filter_backends = [filters.SearchFilter]
    search_fields = ['name'] # сюда поля для поиска
    # ordering_fields = ['name', 'email', 'id']  # опцио

    def create(self, request, *args, **kwargs):
        # Можно вставить задержку или лог
        # time.sleep(2)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        return Response(
            {
                'message': 'employeeAdded',
                'employee': serializer.data
            },
            status=status.HTTP_201_CREATED
        )

    def destroy(self, request, *args, **kwargs):
        # time.sleep(2)
        # return Response(
        #         {'message': 'employeeNotDeleted'},
        #         status=status.HTTP_400_BAD_REQUEST
        #     )
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response(status=status.HTTP_200_OK)
    
    def update(self, request, *args, **kwargs):
        # time.sleep(2)
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(serializer.data)
    

    @action(detail=False, methods=['get'], url_path='export_excel')
    def export_excel(self, request):
        # применяем поиск/фильтры, как и в list()
        queryset = self.filter_queryset(self.get_queryset())

        # создаем Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"

        # Заголовки
        headers = ["№", "Işgär"]
        ws.append(headers)

        # Строки
        for index, emp in enumerate(queryset, 1):
            ws.append([index, emp.name])

        # Автоширина колонок
        for i, column in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        # Отдаём как файл
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
        wb.save(response)
        return response


# @api_view(['GET', 'POST'])
# def suppliers_list(request):
#     if request.method == 'GET':
#         suppliers = Supplier.objects.all()
#         serializer = SupplierSerializer(suppliers, many=True)
#         return Response(serializer.data)
    
#     elif request.method == 'POST':
#         serializer = SupplierSerializer(data=request.data)
#         if serializer.is_valid():
#             serializer.save()
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




class MySecureView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        if user.groups.filter(name="admin").exists():
            return Response({"authUser": f"{request.user}", "authGroup": "admin"})
        elif user.groups.filter(name="worker").exists():
            return Response({"authUser": f"{request.user}", "authGroup": "worker"})
        else:
            return Response({"message": "Нет доступа"}, status=403)
        



class AssignPartnersToAgentView(APIView):
    def post(self, request):
        # time.sleep(2)
        partners_id = request.data.get("partners_id")
        agent_id = request.data.get("igent_id")

        if agent_id is None:
            return Response({"error": "Invalid data"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            agent = Agent.objects.get(id=agent_id)
        except Agent.DoesNotExist:
            return Response({"error": "Agent not found"}, status=404)

        # Обнулить старые связи
        Partner.objects.filter(agent=agent).update(agent=None)

        # Назначить новые связи
        if isinstance(partners_id, list):
            Partner.objects.filter(id__in=partners_id).update(agent=agent)

        return Response({"message": "partnerSuccessUpdated"}, status=200)
    


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


class BrandViewSet(viewsets.ModelViewSet):
    queryset = Brand.objects.all()
    serializer_class = BrandSerializer
    permission_classes = [IsAuthenticated]


class ModelViewSet(viewsets.ModelViewSet):
    queryset = Model.objects.all()
    serializer_class = ModelSerializer
    permission_classes = [IsAuthenticated]


class TagViewSet(viewsets.ModelViewSet):
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    permission_classes = [IsAuthenticated]


class ProductImageViewSet(viewsets.ModelViewSet):
    queryset = ProductImage.objects.all()
    serializer_class = ProductImageSerializer
    permission_classes = [IsAuthenticated]  




class ProductUnitViewSet(viewsets.ModelViewSet):
    queryset = ProductUnit.objects.all()
    serializer_class = ProductUnitSerializer
    permission_classes = [IsAuthenticated]  


    
    # def destroy(self, request, *args, **kwargs):
    #     # return Response(
    #     #         {'message': 'supplierHasSupplies'},
    #     #         status=status.HTTP_400_BAD_REQUEST
    #     #     )
    #     instance = self.get_object()
    #     self.perform_destroy(instance)
    #     return Response({'message': 'AgentDeleted'}, status=status.HTTP_200_OK)
    
    # def create(self, request, *args, **kwargs):
    #     serializer = self.get_serializer(data=request.data)
    #     serializer.is_valid(raise_exception=True)
    #     self.perform_create(serializer)
    #     return Response({'data': serializer.data, "type": "success"}, status=status.HTTP_201_CREATED)




















class UnitOfMeasurementViewSet(viewsets.ModelViewSet):
    queryset = UnitOfMeasurement.objects.all()
    serializer_class = UnitOfMeasurementSerializer

    def create(self, request, *args, **kwargs):
        # time.sleep(2)
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        # time.sleep(2)
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        # time.sleep(2)
        return super().partial_update(request, *args, **kwargs)
    


@require_GET
def check_name_unique(request):
    name = request.GET.get('name', '').strip()
    exists = Product.objects.filter(name__iexact=name).exists()
    return JsonResponse({'exists': exists})


# dlya poiska producta for free add
@api_view(["GET"])
def search_products(request):
    query = request.GET.get("q", "")
    results = Product.objects.annotate(
        similarity=TrigramSimilarity("name", query)
    ).filter(similarity__gt=0.1).order_by("-similarity")[:10]

    return Response(ProductSerializer(results, many=True).data)







VALID_PRICE_TYPES = ['purchase', 'retail', 'wholesale', 'discount']

class PriceChangeReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        price_type = request.query_params.get('price_type')  # 'purchase', 'retail', и т.п.

        if not start_date or not end_date:
            return Response({"detail": "start_date и end_date обязательны"}, status=status.HTTP_400_BAD_REQUEST)

        start = parse_date(start_date)
        end = parse_date(end_date)

        if not start or not end:
            return Response({"detail": "Неверный формат даты. Используйте YYYY-MM-DD"}, status=status.HTTP_400_BAD_REQUEST)

        queryset = PriceChangeHistory.objects.filter(
            changed_at__date__range=(start, end)
        )

        if price_type:
            if price_type not in VALID_PRICE_TYPES:
                return Response({"detail": f"price_type должен быть одним из {VALID_PRICE_TYPES}"}, status=status.HTTP_400_BAD_REQUEST)
            queryset = queryset.filter(price_type=price_type)

        queryset = queryset.select_related('product__base_unit')
        serializer = PriceChangeReportSerializer(queryset, many=True)
        return Response(serializer.data)


class PriceChangeExcelDownloadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        price_type = request.query_params.get('price_type')

        if not start_date or not end_date:
            return HttpResponse("start_date и end_date обязательны", status=400)

        if price_type and price_type not in VALID_PRICE_TYPES:
            return HttpResponse(f"price_type должен быть одним из {VALID_PRICE_TYPES}", status=400)

        start = parse_date(start_date)
        end = parse_date(end_date)

        if not start or not end:
            return HttpResponse("Неверный формат даты. Используйте YYYY-MM-DD", status=400)

        queryset = PriceChangeHistory.objects.filter(changed_at__date__range=(start, end))

        if price_type:
            queryset = queryset.filter(price_type=price_type)

        queryset = queryset.select_related("product__base_unit")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Изменения цен"

        header_text = f"Отчёт по изменениям цен с {start_date} по {end_date}"
        if price_type:
            if price_type == 'purchase':
                price_word = 'Цена закупки'
            elif price_type == 'retail':
                price_word = 'Розничная цена'
            elif price_type == 'wholesale':
                price_word = 'Оптовая цена'
            header_text += f" ({price_word})"

        ws.merge_cells('A1:L1')
        cell = ws['A1']
        cell.value = header_text
        cell.font = Font(size=14, bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        headers = [
            "#", "Продукт", "Ед. изм.", "Старая цена", "Кол-во", "Сумма старая",
            "Новая цена", "Кол-во", "Сумма новая", "Прибыль", "Убыток", "Дата"
        ]
        ws.append([])
        ws.append(headers)

        for idx, record in enumerate(queryset, start=1):
            old_price = float(record.old_price)
            new_price = float(record.new_price)
            quantity = float(record.quantity_at_change)
            old_total = old_price * quantity
            new_total = new_price * quantity
            profit = max(record.difference, 0)
            loss = min(record.difference, 0)

            ws.append([
                idx,
                record.product.name,
                record.product.base_unit.name,
                old_price,
                quantity,
                round(old_total, 2),
                new_price,
                quantity,
                round(new_total, 2),
                round(profit, 2),
                round(loss, 2),
                record.changed_at.strftime("%Y-%m-%d %H:%M"),
            ])

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="price_changes_report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    



# download prodcts excel
# class ProductExportExcelView(APIView):
#     permission_classes = [IsAuthenticated]  # Uncomment this line

#     def post(self, request):
#         try:
#             product_ids = request.data.get("product_ids", [])
#             if not product_ids:
#                 return Response({"error": "product_ids не переданы"}, status=400)

#             # Check if product_ids is a list and contains valid integers
#             if not isinstance(product_ids, list) or not all(isinstance(pid, int) for pid in product_ids):
#                 return Response({"error": "product_ids должен быть списком целых чисел"}, status=400)

#             queryset = Product.objects.filter(id__in=product_ids).select_related('base_unit', 'category', 'brand', 'model')
            
#             if not queryset.exists():
#                 return Response({"error": "Продукты не найдены"}, status=404)

#             wb = openpyxl.Workbook()
#             ws = wb.active
#             ws.title = "Товары"

#             headers = [
#                 "ID", "Наименование", "Категория", "Ед. изм.", "Артикул", "Количество",
#                 "Цена закупки", "Розничная цена", "Оптовая цена", "Цена со скидкой",
#                 "Бренд", "Модель", "Вес (кг)", "Объём (м³)"
#             ]
#             ws.append(headers)

#             for product in queryset:
#                 ws.append([
#                     product.id,
#                     product.name,
#                     product.category.name if product.category else "",
#                     product.base_unit.name if product.base_unit else "",
#                     product.sku or "",
#                     float(product.quantity) if product.quantity else 0,
#                     float(product.purchase_price) if product.purchase_price else 0,
#                     float(product.retail_price) if product.retail_price else 0,
#                     float(product.wholesale_price) if product.wholesale_price else 0,
#                     float(product.discount_price) if product.discount_price else 0,
#                     product.brand.name if product.brand else "",
#                     product.model.name if product.model else "",
#                     float(product.weight) if product.weight else 0,
#                     float(product.volume) if product.volume else 0,
#                 ])

#             # Автоматическая ширина колонок
#             for col in ws.columns:
#                 max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
#                 ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

#             response = HttpResponse(
#                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#             )
#             response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
#             response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            
#             wb.save(response)
#             return response
            
#         except Exception as e:
#             return Response({"error": f"Внутренняя ошибка сервера: {str(e)}"}, status=500)



class ProductExportExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            # Получаем фильтры или product_ids
            filters = request.data.get("filters", {})
            product_ids = request.data.get("product_ids", [])
            
            if product_ids:
                # Старый способ - по ID товаров (ограниченный пагинацией)
                queryset = Product.objects.filter(id__in=product_ids)
            elif filters:
                # Новый способ - применяем все фильтры для получения всех товаров
                queryset = Product.objects.all()
                print('filters', filters)
                # print('filters', filters)
                
                # Применяем поиск
                if 'search' in filters and filters['search']:
                    search_query = filters['search']
                    queryset = queryset.filter(
                        Q(name__icontains=search_query) |
                        Q(sku__icontains=search_query) |
                        Q(category__name__icontains=search_query) |
                        Q(brand__name__icontains=search_query)
                    )
                
                # Применяем фильтр по категории
                if 'categories' in filters and filters['categories']:
                    queryset = queryset.filter(category_id=filters['categories'])
                
                # Применяем фильтр по бренду
                if 'brands' in filters and filters['brands']:
                    queryset = queryset.filter(brand_id=filters['brands'])
                
                # Применяем фильтр по модели
                if 'models' in filters and filters['models']:
                    queryset = queryset.filter(model_id=filters['models'])

                if 'tags' in filters and filters['tags']:
                    queryset = queryset.filter(model_id=filters['tags'])


                if 'wholesale_price_max' in filters or "wholesale_price_min" in filters:
                    wholesale_max = filters.get('wholesale_price_max', None)
                    wholesale_min = filters.get('wholesale_price_min', None)

                    if wholesale_min is not None:
                        queryset = queryset.filter(wholesale_price__gte=wholesale_min)
                    if wholesale_max is not None:
                        queryset = queryset.filter(wholesale_price__lte=wholesale_max)


                if 'retail_price_max' in filters or "retail_price_min" in filters:
                    retail_max = filters.get('retail_price_max', None)
                    retail_min = filters.get('retail_price_min', None)

                    if retail_min is not None:
                        queryset = queryset.filter(retail_price__gte=retail_min)
                    if retail_max is not None:
                        queryset = queryset.filter(retail_price__lte=retail_max)


                if 'quantity_max' in filters or "quantity_min" in filters:
                    quantity_max = filters.get('quantity_max', None)
                    quantity_min = filters.get('quantity_min', None)

                    if quantity_min is not None:
                        queryset = queryset.filter(quantity__gte=quantity_min)
                    if quantity_max is not None:
                        queryset = queryset.filter(quantity__lte=quantity_max)


                if 'is_active' in filters:
                    if filters['is_active'] == 'true':
                        queryset = queryset.filter(is_active=True)
                    else:
                        queryset = queryset.filter(is_active=False)



                
                
                # Применяем другие фильтры по необходимости
                # Добавьте свои фильтры здесь
                
            else:
                # Если ничего не передано, экспортируем все товары
                queryset = Product.objects.all()

            # Оптимизируем запрос
            queryset = queryset.select_related('base_unit', 'category', 'brand', 'model').order_by('id')
            
            total_count = queryset.count()
            if total_count == 0:
                return Response({"error": "Товары не найдены"}, status=404)
            
            # Ограничиваем количество товаров для экспорта (защита от перегрузки)
            MAX_EXPORT_LIMIT = 10000  # Максимум 10,000 товаров
            if total_count > MAX_EXPORT_LIMIT:
                return Response({
                    "error": f"Слишком много товаров для экспорта. Максимум: {MAX_EXPORT_LIMIT}, найдено: {total_count}"
                }, status=400)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Товары"

            headers = [
                "№","ID", "Наименование", "Категория", "Ед. изм.", "Артикул", "Количество",
                "Цена закупки", "Розничная цена", "Оптовая цена", "Цена со скидкой",
                "Бренд", "Модель", "Вес (кг)", "Объём (м³)"
            ]
            ws.append(headers)

            # Обрабатываем товары партиями для больших объемов
            batch_size = 1000
            row_number = 1
            for i in range(0, total_count, batch_size):
                batch_queryset = queryset[i:i + batch_size]
                
                for product in batch_queryset:
                    ws.append([
                        row_number,
                        product.id,
                        product.name,
                        product.category.name if product.category else "",
                        product.base_unit.name if product.base_unit else "",
                        product.sku or "",
                        float(product.quantity) if product.quantity else 0,
                        float(product.purchase_price) if product.purchase_price else 0,
                        float(product.retail_price) if product.retail_price else 0,
                        float(product.wholesale_price) if product.wholesale_price else 0,
                        float(product.discount_price) if product.discount_price else 0,
                        product.brand.name if product.brand else "",
                        product.model.name if product.model else "",
                        float(product.weight) if product.weight else 0,
                        float(product.volume) if product.volume else 0,
                    ])
                    row_number += 1

            # Автоматическая ширина колонок
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="products_export_{total_count}_items.xlsx"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({"error": f"Внутренняя ошибка сервера: {str(e)}"}, status=500)






######################################################################################################################### Faktura START


class PurchaseInvoiceViewSet(viewsets.ModelViewSet):
    queryset = PurchaseInvoice.objects.all()  # ОБЯЗАТЕЛЬНО добавить!
    serializer_class = PurchaseInvoiceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return PurchaseInvoice.objects.select_related(
            'supplier', 'created_by', 'canceled_by'
        ).prefetch_related('items__product').all()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @transaction.atomic
    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel(self, request, pk=None):
        invoice = self.get_object()
        if invoice.is_canceled:
            return Response({'detail': 'Накладная уже отменена'}, status=status.HTTP_400_BAD_REQUEST)
        reason = request.data.get('cancel_reason')
        if not reason:
            return Response({'cancel_reason': 'Это поле обязательно'}, status=status.HTTP_400_BAD_REQUEST)

        invoice.is_canceled = True
        invoice.canceled_at = timezone.now()
        invoice.cancel_reason = reason
        invoice.canceled_by = request.user
        invoice.save()

        serializer = self.get_serializer(invoice)
        return Response(serializer.data)
    

class SalesInvoiceViewSet(viewsets.ModelViewSet):
    

    queryset = SalesInvoice.objects.all()  # ОБЯЗАТЕЛЬНО
    serializer_class = SalesInvoiceSerializer
    permission_classes = [IsAuthenticated]


    def get_queryset(self):
        return SalesInvoice.objects.select_related(
            'buyer', 'delivered_by', 'created_by', 'canceled_by'
        ).prefetch_related('items__product').all()

    # def perform_create(self, serializer):
    #     serializer.save(created_by=self.request.user)

    @transaction.atomic
    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel(self, request, pk=None):
        invoice = self.get_object()
        if invoice.is_canceled:
            return Response({'detail': 'Накладная уже отменена'}, status=status.HTTP_400_BAD_REQUEST)
        reason = request.data.get('cancel_reason')
        if not reason:
            return Response({'cancel_reason': 'Это поле обязательно'}, status=status.HTTP_400_BAD_REQUEST)

        invoice.is_canceled = True
        invoice.canceled_at = timezone.now()
        invoice.cancel_reason = reason
        invoice.canceled_by = request.user
        invoice.save()

        serializer = self.get_serializer(invoice)
        return Response(serializer.data)
    

class PurchaseReturnInvoiceViewSet(viewsets.ModelViewSet):
    queryset = PurchaseReturnInvoice.objects.all()  # ОБЯЗАТЕЛЬНО
    serializer_class = PurchaseReturnInvoiceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return PurchaseReturnInvoice.objects.select_related(
            'original_invoice', 'created_by'
        ).prefetch_related('items__product').all()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


# class SalesReturnInvoiceViewSet(viewsets.ModelViewSet):
#     queryset = SalesReturnInvoice.objects.all()  # ОБЯЗАТЕЛЬНО
#     serializer_class = SalesReturnInvoiceSerializer
#     permission_classes = [IsAuthenticated]

#     def get_queryset(self):
#         return SalesReturnInvoice.objects.select_related(
#             'original_invoice', 'created_by'
#         ).prefetch_related('items__product').all()

#     def perform_create(self, serializer):
#         serializer.save(created_by=self.request.user)


######################################################################################################################### Faktura END



class CurrencyViewSet(viewsets.ModelViewSet):
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer
    permission_classes = [IsAuthenticated]




##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
####################################################################################################################### Entries START

class CurrencyViewSet(viewsets.ModelViewSet):
    queryset = Currency.objects.all()
    serializer_class = CurrencySerializer


class CurrencyRateViewSet(viewsets.ModelViewSet):
    queryset = CurrencyRate.objects.all()
    serializer_class = CurrencyRateSerializer


class AccountViewSet(viewsets.ModelViewSet):
    queryset = Account.objects.all()
    serializer_class = AccountSerializer


class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all().order_by('-date')
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = {
        'partner': ['exact'],
        'date': ['gte', 'lte'],
    }

    def get_serializer_class(self):
        if self.action == 'create':
            return TransactionWriteSerializer
        return TransactionSerializer


class EntryViewSet(viewsets.ModelViewSet):
    queryset = Entry.objects.all()
    serializer_class = EntrySerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = {
        'account': ['exact'],
        'transaction__date': ['gte', 'lte'],
    }
####################################################################################################################### Entries END
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################
##################################################################################################################################################################################################################################################